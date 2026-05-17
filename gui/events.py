# -*- coding: utf-8 -*-
"""
GUI 事件与按钮回调（v36 抽取）
⚠️ 本文件仅包含 GUI 回调，不修改任何业务逻辑
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys as _sys
import glob as _glob
import pandas as pd
from widgets import C, STEPS
from storage import storage
from analysis.analyzer import do_analysis_v2
from core.decorators import with_feedback
from core.state_store import get_state
from core.rule_engine import RuleEngine
from core.task_manager import TaskManager
from core.exporter import ExcelExporter
from openpyxl import Workbook, load_workbook
from domain.alt_material.alt_manager import save_alt_pairs, load_alt_pairs
import time
import datetime
from datetime import datetime
import threading
import traceback
import json
import csv
import calendar

class EventsMixIn:
    # ── audit_tree 列头排序映射 ─────────────────────────
    _COL_DISPLAY = {
        'idx': '序号', 'excel_row': '原表行号', 'factory': '工厂名称',
        'admin': '生产管理员', 'order_date': '订单日期', 'order_no': '流程订单',
        'code': '物料号', 'name': '物料描述', 'quota': '定额', 'actual': '实际',
        'dev_rate': '偏差率%', 'is_alt': '替代料', 'status': '状态',
        'remark': '备注', 'batch_remark': '批量备注',
        'audit_result': '审核结果', 'AI建议': 'AI建议',
        'audit_status': '审核状态', 'audit_source': '审核来源',
        'deviation_amount': '偏差金额',
    }
    _COL_TO_DF = {
        'idx': None, 'excel_row': '原表行号', 'factory': '工厂',
        'admin': '车间', 'order_date': '订单日期', 'order_no': '流程订单',
        'code': '物料编码', 'name': '物料名称', 'quota': '定额', 'actual': '实际',
        'dev_rate': '偏差率(%)', 'is_alt': '替代料', 'status': '状态原因',
        'remark': '备注原因', 'batch_remark': '批量备注',
        'audit_result': 'audit_result', 'AI建议': 'AI建议',
        'audit_status': 'audit_result', 'audit_source': 'audit_result',
        'deviation_amount': '偏差金额',
    }


    """包含所有 GUI 事件处理方法，供 ZPP011Beautiful 继承"""

    def _save_audit_back(self):
        """保存审核结果到原 Excel，同时同步到本地数据库 """
        # ── 1. 前置检查 ──
        self.log("💾 正在保存审核结果...", "info")
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showwarning("提示", "没有审核数据可保存")
            return
        src_path = self.input_file.get()
        if not src_path or not os.path.exists(src_path):
            messagebox.showerror("错误", "原始文件不存在，请先选择正确的输入文件")
            return

        try:
            # ── 2. 备份原文件 ──
            self.log("📦 正在备份原文件...", "info")
            backup_dir = os.path.dirname(src_path)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{os.path.splitext(os.path.basename(src_path))[0]}_审核前备份_{ts}.xlsx"
            backup_path = os.path.join(backup_dir, backup_name)
            shutil.copy2(src_path, backup_path)
            self.log(f"📦 已备份原文件：{backup_name}", "info")
        except Exception as e:
            self.log(f"⚠ 备份失败：{e}", "warning")
            if not messagebox.askyesno("备份失败", f"无法备份原文件：{e}\n是否不备份直接保存？"):
                return

        # ── 3. 打开 Excel 写入审核结果 ──
        try:
            self.log("📂 正在打开 Excel 文件...", "info")
            wb = load_workbook(src_path)
            ws = wb['Data']
            self.log(f"   Excel 已打开，共 {ws.max_row} 行 x {ws.max_column} 列", "info")
        except Exception as e:
            self.log(f"❌ 打开 Excel 失败：{e}", "error")
            messagebox.showerror("保存失败", f"无法打开原始文件：\n{e}\n请关闭 Excel 后重试。")
            return

        # ── 4. 构建待写入的数据列表 ──
        save_list = []
        # 订单列查找（支持多种列名）
        order_col = None
        for possible in ['流程订单', '订单号', '订单编号', '订单号码', '订单No', 'Order No', '生产订单']:
            if possible in self.audit_data.columns:
                order_col = possible
                break
        if order_col is None:
            wb.close()
            self.log(f"❌ 审核数据中缺少订单列，实际列名: {list(self.audit_data.columns)[:10]}", "error")
            messagebox.showerror("保存失败", f"审核数据中缺少订单号列，无法定位原表行。\n实际列名: {list(self.audit_data.columns)[:10]}")
            return

        for _, row in self.audit_data.iterrows():
            work_date = str(row.get('订单日期', ''))[:10]
            order_no = str(row.get(order_col, ''))
            mat_code = str(row.get('组件物料号', ''))
            if not work_date or not order_no or not mat_code:
                continue
            # 取最终备注：优先取已有备注，其次取 AI 建议
            remark = str(row.get('备注原因', '') or row.get('AI建议', '') or '').strip()
            save_list.append((work_date, order_no, mat_code, remark))
        self.log(f"   待保存记录：{len(save_list)} 条", "info")

        # ── 5. 匹配 Excel 表头列索引 ──
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(1, col_idx).value
            if val:
                headers[val.strip()] = col_idx

        date_col = headers.get('订单开始日期') or headers.get('订单日期')
        order_col_excel = headers.get('流程订单') or headers.get('订单号')
        mat_col = headers.get('组件物料号') or headers.get('物料编码')

        if not all([date_col, order_col_excel, mat_col]):
            wb.close()
            missing = []
            if not date_col: missing.append('日期')
            if not order_col_excel: missing.append('订单号')
            if not mat_col: missing.append('物料编码')
            self.log(f"❌ Excel 中缺少关键列：{', '.join(missing)}", "error")
            messagebox.showerror("列匹配失败", f"原始文件中未找到以下关键列：\n{', '.join(missing)}\n无法定位写入位置。")
            return

        self.log(f"   定位到的列：日期={date_col}, 订单={order_col_excel}, 物料={mat_col}", "info")

        # ── 6. 写入审核状态列 ──
        audit_cols = {'审核状态': None, '审核备注': None, '审核人': None, '审核时间': None}
        next_col = ws.max_column + 1
        for col_name in audit_cols:
            if col_name in headers:
                audit_cols[col_name] = headers[col_name]
            else:
                ws.cell(1, next_col, col_name).font = Font(bold=True)
                audit_cols[col_name] = next_col
                next_col += 1
        self.log(f"   审核列：审核状态={audit_cols['审核状态']}, 审核备注={audit_cols['审核备注']}", "info")

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        auditor = os.getlogin()
        saved_count = 0
        match_count = 0

        for r_idx in range(2, ws.max_row + 1):
            r_date = str(ws.cell(r_idx, date_col).value)[:10] if ws.cell(r_idx, date_col).value else ''
            r_order = str(ws.cell(r_idx, order_col_excel).value).strip()
            r_mat = str(ws.cell(r_idx, mat_col).value).strip()

            match = next((item for item in save_list
                          if item[0] == r_date and item[1] == r_order and item[2] == r_mat), None)
            if match:
                match_count += 1
                _, _, _, remark = match
                ws.cell(r_idx, audit_cols['审核状态'], '已备注' if remark else '未审核')
                ws.cell(r_idx, audit_cols['审核备注'], remark)
                ws.cell(r_idx, audit_cols['审核人'], auditor)
                ws.cell(r_idx, audit_cols['审核时间'], now_str)
                saved_count += 1

        self.log(f"   匹配到 {match_count} 行，待写入 {saved_count} 行", "info")

        # ── 7. 保存 Excel 文件 ──
        try:
            self.log("💾 正在保存 Excel 文件...", "info")
            wb.save(src_path)
            wb.close()
            self.log("   Excel 保存成功", "info")
        except PermissionError:
            wb.close()
            self.log("❌ 文件被占用，无法保存。请关闭 Excel 后重试。", "error")
            messagebox.showerror("保存失败", "文件被其他程序占用，请关闭 Excel 后重试。")
            return
        except Exception as e:
            wb.close()
            self.log(f"❌ 写入 Excel 失败：{e}", "error")
            messagebox.showerror("保存失败", f"写入文件时出错：\n{e}")
            return
        finally:
            try:
                wb.close()
            except:
                pass

        # ── 8. 同步到 SQLite 审核数据库 ──
        try:
            self.log("📊 正在同步到审核数据库...", "info")
            # 构造适合 storage 模块的 DataFrame（确保有订单号列）
            save_df = self.audit_data.copy()
            if '订单号' not in save_df.columns and '流程订单' in save_df.columns:
                save_df['订单号'] = save_df['流程订单']
            if '订单日期' not in save_df.columns:
                save_df['订单日期'] = save_df.get('订单日期', '')
            storage.save_audit_to_db(save_df, auditor=auditor, log_cb=self.log)
            self.log("   数据库同步成功", "info")
        except Exception as e:
            self.log(f"⚠ 审核数据库同步失败（不影响 Excel 保存）：{e}", "warn")

        # ── 9. 最终反馈 ──
        self.log(f"✅ 审核结果已保存：Excel 写入 {saved_count} 行，备份 {backup_name}", "success")
        messagebox.showinfo("保存成功",
                            f"审核结果已写入原始文件 {saved_count} 行。\n"
                            f"新增列：审核状态 / 审核备注 / 审核人 / 审核时间\n"
                            f"原始备份：{backup_name}")


    def _export_audit_backup(self):
        default_name = f"审核记录备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("ZIP 压缩文件", "*.zip")],
            initialfile=default_name,
            title="导出审核记录备份"
        )
        if not file_path:
            return
        try:
            storage.export_audit_backup(file_path, log_cb=self.log)
            self.log(f"✅ 审核记录已导出：{file_path}", "success")
            messagebox.showinfo("导出成功", f"备份文件已保存到：\n{file_path}")
        except FileNotFoundError:
            self.log("❌ 导出失败：审核数据库不存在", "error")
            messagebox.showerror("导出失败", "审核数据库不存在，请先保存审核记录后再导出。")
        except Exception as e:
            self.log(f"❌ 导出失败：{e}", "error")
            messagebox.showerror("导出失败", str(e))


    def _import_audit_backup(self):
        file_path = filedialog.askopenfilename(
            title="选择审核记录备份文件",
            filetypes=[("ZIP 压缩文件", "*.zip"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        try:
            storage.import_audit_backup(file_path, log_cb=self.log)
            self.log("✅ 审核记录已从备份恢复，下次加载时生效", "success")
            messagebox.showinfo("导入成功", "审核记录已恢复。\n重新加载数据时将自动匹配历史审核。")
        except Exception as e:
            self.log(f"❌ 导入失败：{e}", "error")
            messagebox.showerror("导入失败", str(e))

    # ── 版本日志已迁移到 utils/version_history.py ──────────
    # 原 _CHANGELOG_EMBEDDED 硬编码已删除，通过导入动态读取
    _CHANGELOG_EMBEDDED = None  # 占位，保持向后兼容
    # ── 公共函数：获取 changelog.json 路径 ──────────

    def _show_shortcuts_help(self):
        """显示快捷键说明对话框"""
        import tkinter as tk
        from tkinter import ttk

        win = tk.Toplevel(self.root)
        win.title("快捷键说明")
        win.geometry("420x260")
        win.resizable(False, False)

        text = tk.Text(win, font=("微软雅黑", 10), wrap=tk.WORD, padx=10, pady=10)
        text.pack(fill=tk.BOTH, expand=True)

        shortcuts = [
            ("Ctrl+S", "保存审核结果"),
            ("Ctrl+E", "导出Excel"),
            ("Ctrl+A", "AI审核备注"),
            ("F1",     "显示本帮助"),
            ("Ctrl+Q", "退出程序"),
        ]
        for key, desc in shortcuts:
            text.insert(tk.END, f"{key:12} {desc}\n")
        text.config(state=tk.DISABLED)

        ttk.Button(win, text="关闭", command=win.destroy).pack(pady=10)

    def _show_about(self):
        from utils.version_history import (
            get_current_version, get_version_display,
            get_version_history_text, APP_NAME, AUTHOR
        )
        _ver_str = get_current_version()
        _ver_date = ''
        # 从 VERSION_HISTORY 获取日期
        from utils.version_history import VERSION_HISTORY
        if VERSION_HISTORY:
            _ver_date = VERSION_HISTORY[0].get('date', '')
        changelog = get_version_history_text()
        info = (
            f"{APP_NAME}\n"
            f"版本：{_ver_str}\n"
            f"作者：{AUTHOR}\n"
            f"日期：{_ver_date}\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "功能：SAP生产订单偏差自动分析\n"
            "特点：多条件筛选 + 批量操作 + 替代料管理\n"
            "新增：AI审核建议、审核进度条、日志导出\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        d = tk.Toplevel(self.root)
        d.title("关于")
        d.geometry("480x580")
        d.resizable(False, False)
        d.transient(self.root)
        d.grab_set()
        container = tk.Frame(d, bg="#1a1a2e")
        container.pack(fill="both", expand=True)
        tk.Label(container, text="关于", font=("Microsoft YaHei", 14, "bold"),
                 fg="#e94560", bg="#1a1a2e").pack(pady=(15, 5))
        info_lbl = tk.Label(container, text=info, font=("Microsoft YaHei", 10),
                            fg="#eaeaea", bg="#16213e", justify="left",
                            anchor="w", padx=20, pady=12)
        info_lbl.pack(fill="x", padx=30, ipady=4)
        tk.Label(container, text="📋 版本日志", font=("Microsoft YaHei", 11, "bold"),
                 fg="#0f3460", bg="#1a1a2e", anchor="w",
                 padx=30).pack(fill="x", pady=(10, 0))
        log_frame = tk.Frame(container, bg="#f8f9fa")
        log_frame.pack(fill="both", expand=True, padx=30, pady=(5, 15))
        tb = tk.Text(log_frame, font=("Microsoft YaHei", 9),
                     fg="#1f2328", bg="#ffffff",
                     relief="flat", state="disabled", wrap="word",
                     height=18)
        tb.pack(fill="both", expand=True)
        tb.configure(state="normal")
        tb.insert("1.0", changelog.strip())
        tb.configure(state="disabled")
        tk.Button(container, text="确定", font=("Microsoft YaHei", 10),
                  command=d.destroy, bg="#e94560", fg="white",
                  relief="flat", padx=20, pady=5).pack(pady=(0, 12))

    # ── 界面构建 ────────────────

    def open_output(self):
        """打开输出目录，若目录不存在则打开桌面"""
        output_path = getattr(self, 'output_path', None)
        if output_path and os.path.exists(output_path):
            os.startfile(os.path.dirname(output_path))
        elif self.output_dir.get() and os.path.exists(self.output_dir.get()):
            os.startfile(self.output_dir.get())
        else:
            os.startfile(os.path.expanduser('~'))


    def generate_ppt(self):
        """选择 Excel 分析结果，生成 PPT 报告"""
        excel_path = filedialog.askopenfilename(
            title="选择 zpp011 偏差分析 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not excel_path:
            return
        out_dir = self.output_dir.get() or os.path.dirname(excel_path)
        base = os.path.splitext(os.path.basename(excel_path))[0]
        output_path = os.path.join(out_dir, base + ".pptx")
        self.log(f"📊 开始生成 PPT：{os.path.basename(excel_path)}", "info")
        self.ppt_btn.configure(state="disabled", text="生成中...")
        self.status_lbl.configure(text="正在生成 PPT...", fg=C['purple'])

        def worker():
            try:
                ppt_generator.run_ppt_generation(excel_path, output_path, log_cb=self.log)
                self.root.after(0, lambda: self._on_ppt_done(output_path))
            except Exception as e:
                self.root.after(0, lambda: self._on_ppt_error(str(e)))

        threading.Thread(target=worker, daemon=True).start()


    def _on_ppt_done(self, output_path):
        self.ppt_btn.configure(state="normal", text="📊 生成PPT")
        self.status_lbl.configure(text=f"PPT 已生成 — {os.path.basename(output_path)}", fg=C['green'])
        self.log(f"✅ PPT 已保存：{output_path}", "success")
        try:
            os.startfile(output_path)
        except Exception:
            pass


    def _on_ppt_error(self, msg):
        self.ppt_btn.configure(state="normal", text="📊 生成PPT")
        self.status_lbl.configure(text="PPT 生成出错", fg=C['danger'])
        self.log(f"❌ PPT 生成失败：{msg}", "error")
        messagebox.showerror("PPT 生成出错", msg)


    def generate_excel_direct(self):
        """弹出另存为对话框，生成 Excel 分析表格"""
        input_path = self.input_file.get()
        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("错误", "请先选择输入文件！")
            return

        # ── 尝试从原始数据中获取日期范围，用于默认文件名 ──
        date_tag = ""
        try:
            df_raw = pd.read_excel(input_path, sheet_name='Data')
            # 查找可能的日期列
            date_col = None
            for col in ['订单开始日期', '订单日期', '日期']:
                if col in df_raw.columns:
                    date_col = col
                    break
            if date_col is not None:
                dates = pd.to_datetime(df_raw[date_col], errors='coerce').dropna()
                if not dates.empty:
                    d_min = dates.min().strftime('%Y%m%d')
                    d_max = dates.max().strftime('%m%d')
                    date_tag = f"{d_min}-{d_max}"
        except Exception:
            pass

        # 如果从数据中获取失败，回退到手动输入或当前月份
        if not date_tag:
            start_str = self.start_date.get().strip()
            end_str = self.end_date.get().strip()
            if start_str and end_str:
                date_tag = f"{start_str.replace('-', '')[:8]}-{end_str.replace('-', '')[-4:]}"
            else:
                now = datetime.now()
                first_day = now.replace(day=1).strftime('%Y%m%d')
                # 使用 calendar 计算月末
                import calendar
                last_day_num = calendar.monthrange(now.year, now.month)[1]
                last_day = now.replace(day=last_day_num).strftime('%Y%m%d')
                date_tag = f"{first_day}-{last_day[-4:]}"

        # ── 构建默认路径和文件名 ──
        default_dir = r"E:\zpp011_dev\ZPP011偏差分析"
        os.makedirs(default_dir, exist_ok=True)
        default_name = f"ZPP011偏差分析_{date_tag}.xlsx"

        # ── 弹出另存为对话框 ──
        file_path = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            title="保存偏差分析表格"
        )
        if not file_path:
            return  # 用户取消

        self.excel_btn.configure(state="disabled", text="生成中...")
        self.status_lbl.configure(text="正在生成表格...", fg=C['purple'])
        self.log("📋 开始生成偏差分析表格...", "info")

        threading.Thread(target=self._generate_excel_thread,
                       args=(file_path,), daemon=True).start()

    # ── 树形视图（v31 新增）────────────────────────

    def _show_tree_view(self):
        """在新窗口中以树形结构展示审核数据"""
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showwarning("提示", "没有可展示的数据")
            return

        # 创建新窗口
        win = tk.Toplevel(self.root)
        win.title("偏差数据 - 树形视图")
        win.geometry("1000x650")
        win.transient(self.root)
        win.grab_set()

        # Treeview
        tree = ttk.Treeview(win, show="tree headings",
                            columns=("code", "name", "order_date", "quota", "actual", "dev_rate", "status", "remark"),
                            height=25)
        tree.heading("#0", text="工厂 / 车间 / 物料分类")
        tree.heading("code", text="物料号")
        tree.heading("name", text="物料描述")
        tree.heading("order_date", text="订单日期")
        tree.heading("quota", text="定额")
        tree.heading("actual", text="实际")
        tree.heading("dev_rate", text="偏差率%")
        tree.heading("status", text="状态")
        tree.heading("remark", text="备注")

        tree.column("#0", width=220)
        tree.column("code", width=70, anchor="center")
        tree.column("name", width=100, anchor="w")
        tree.column("order_date", width=70, anchor="center")
        tree.column("quota", width=50, anchor="e")
        tree.column("actual", width=50, anchor="e")
        tree.column("dev_rate", width=55, anchor="center")
        tree.column("status", width=55, anchor="center")
        tree.column("remark", width=80, anchor="w")

        scroll = ttk.Scrollbar(win, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scroll.pack(side="right", fill="y", pady=10)

        # 灵活确定列名
        groupby_factory = next((c for c in ['工厂', '工厂名称'] if c in self.audit_data.columns), None)
        groupby_workshop = next((c for c in ['车间', '生产管理员描述'] if c in self.audit_data.columns), None)
        groupby_category = next((c for c in ['物料类型', '物料分类'] if c in self.audit_data.columns), None)
        code_col = next((c for c in ['物料编码', '组件物料号'] if c in self.audit_data.columns), '物料编码')
        name_col = next((c for c in ['物料名称', '组件物料描述'] if c in self.audit_data.columns), '物料名称')
        quota_col = next((c for c in ['定额', '数量-定额'] if c in self.audit_data.columns), '定额')
        actual_col = next((c for c in ['实际', '数量-实际'] if c in self.audit_data.columns), '实际')
        dev_col = next((c for c in ['偏差率', '偏差率(%)'] if c in self.audit_data.columns), '偏差率')
        remark_col = next((c for c in ['备注', '备注原因'] if c in self.audit_data.columns), '备注')

        if not groupby_factory:
            messagebox.showwarning("提示", "数据中未找到工厂列，无法生成树形视图")
            win.destroy()
            return

        for factory, f_grp in self.audit_data.groupby(groupby_factory):
            factory_id = tree.insert('', 'end', text=factory, open=False)

            if groupby_workshop and groupby_category:
                for (workshop, mat_cat), w_grp in f_grp.groupby([groupby_workshop, groupby_category]):
                    workshop_id = tree.insert(factory_id, 'end',
                                              text=f"{workshop} - {mat_cat}  ({len(w_grp)}条)",
                                              open=False)
                    self._insert_tree_rows(tree, workshop_id, w_grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col)
            elif groupby_workshop:
                for workshop, w_grp in f_grp.groupby(groupby_workshop):
                    workshop_id = tree.insert(factory_id, 'end',
                                              text=f"{workshop}  ({len(w_grp)}条)",
                                              open=False)
                    self._insert_tree_rows(tree, workshop_id, w_grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col)
            else:
                for _, row in f_grp.iterrows():
                    self._insert_tree_rows(tree, factory_id, f_grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col)
                break

        self.log("[OK] tree view opened", "info")

    def _insert_tree_rows(self, tree, parent_id, grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col):
        """向树形视图插入行数据"""
        for _, row in grp.iterrows():
            tree.insert(parent_id, 'end',
                        text=str(row.get(name_col, ''))[:20],
                        values=(
                            str(row.get(code_col, '')),
                            str(row.get(name_col, ''))[:20],
                            str(row.get('订单日期', ''))[:12] if pd.notna(row.get('订单日期')) else '',
                            f"{row.get(quota_col, 0):.3f}" if pd.notna(row.get(quota_col)) else '-',
                            f"{row.get(actual_col, 0):.3f}" if pd.notna(row.get(actual_col)) else '-',
                            f"{row.get(dev_col, 0):.2f}%" if pd.notna(row.get(dev_col)) else '-',
                            row.get('_audit_status', '未审核'),
                            str(row.get(remark_col, ''))[:20]
                        ))

    # ── 以下为 v30 原有的全部方法，必须完整复制到此处 ────────
    # 方法列表（请从原 v30.py 中逐个复制，确保不遗漏）：
    # _refresh_alt_view, _add_alt, _del_alt, _reset_alt,
    # _get_config_path, _save_alt_pairs, _load_alt_pairs, _extract_alt_name,
    # _select_input, _select_output, _auto_find, _preview, log, _set_step,
    # start_analysis, _analysis_thread, _on_progress, _update_timer, request_cancel,
    # _on_done, _on_cancel, _on_error, _show_step_log, open_output,
    # generate_excel_direct, _generate_excel_thread, generate_ppt, _on_ppt_done, _on_ppt_error,
    # _load_audit_data, _apply_row_colors, _edit_audit_remark, _update_audit_stats,
    # _show_audit_context_menu, _filter_audit_tree, _update_filter_options,
    # _on_filter_changed, _reset_all_filters, _refresh_audit_tree,
    # _run_ai_audit, _export_audit_excel, _batch_change_status, _batch_fill_remark,
    # _batch_export, _add_custom_status, _load_custom_statuses, _save_custom_statuses
    # 注：上面这些方法的完整代码请从 v30.py 中全选复制，插入到此处。
    # 复制时注意保持缩进（全部与 _show_tree_view 对齐，class 内部缩进 4 空格）。

    # ── 以下方法从 v30.py 完整移植 ──

    def _refresh_alt_view(self, inner):
        for w in inner.winfo_children():
            w.destroy()
        for a, b in self.alt_pairs:
            fr = tk.Frame(inner, bg=C['surface2'])
            fr.pack(fill="x", pady=1)
            # 解析物料A：支持三元组 (factory, code, name)
            if isinstance(a, (list, tuple)) and len(a) == 3:
                _, a_code, a_name = a
            elif isinstance(a, (list, tuple)) and len(a) == 2:
                a_code, a_name = a
            else:
                a_code, a_name = str(a), ''
            # 解析物料B
            if isinstance(b, (list, tuple)) and len(b) == 3:
                _, b_code, b_name = b
            elif isinstance(b, (list, tuple)) and len(b) == 2:
                b_code, b_name = b
            else:
                b_code, b_name = str(b), ''
            # 显示：编码 + 名称（若名称存在）
            a_disp = f"{a_code} {a_name}" if a_name else a_code
            b_disp = f"{b_code} {b_name}" if b_name else b_code
            tk.Label(fr, text=f"↔ {a_disp}", font=("Consolas", 8), fg=C['text'],
                     bg=C['surface2'], anchor="w").pack(side="left", padx=4)
            tk.Label(fr, text="|", font=("Consolas", 8), fg=C['text_dim'],
                     bg=C['surface2']).pack(side="left")
            tk.Label(fr, text=f"{b_disp}", font=("Consolas", 8), fg=C['purple'],
                     bg=C['surface2'], anchor="w").pack(side="left", padx=4)

    def _run_ai_audit(self):
        """AI审核异步化入口（v4.2）：评估备注质量，支持取消、进度反馈、Mock熔断"""
        # ── 硬锁：防止重复启动 ──────────────────────────────────────────
        if self.is_auditing:
            messagebox.showwarning("提示", "AI审核正在运行，请等待完成或点击取消")
            return
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showwarning("警告", "没有审核数据，请先加载审核数据")
            return

        # 确保必要列存在
        for col in ['audit_result', 'AI建议', '审核来源']:
            if col not in self.audit_data.columns:
                self.audit_data[col] = ''
            self.audit_data[col] = self.audit_data[col].astype(str)

        # 只审核未审核行
        to_audit_mask = (
            self.audit_data['audit_result'].isna() |
            (self.audit_data['audit_result'] == '') |
            (self.audit_data['audit_result'] == 'nan')
        )
        if not to_audit_mask.any():
            messagebox.showinfo("提示", "所有行均已审核")
            return

        # ── 进度条重置（修复模式冲突）──────────────────────────────────
        self.progress_bar.stop()
        self.progress_bar.configure(mode='determinate', maximum=100)
        self.progress_bar['value'] = 0
        self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)
        self.set_status("AI审核中...")

        # ── 取消标志（Worker 直接检查，不经过 task_manager）──────────────
        self._ai_cancel_flag = threading.Event()
        self.is_auditing = True

        # ── 构建 Worker 数据快照（避免线程竞争）──────────────────────────
        audit_indices = self.audit_data[to_audit_mask].index.tolist()
        df_snapshot = self.audit_data.loc[audit_indices].copy()

        # 捕获 cancel_flag 到本地变量，Worker 通过闭包直接引用，
        # 不依赖 task_manager 的自动注入（避免 flag 不匹配的竞态问题）
        _my_cancel_flag = self._ai_cancel_flag

        def _worker(progress_callback):
            """在线程池中执行 AI 审核，每行均检查取消标志"""
            total = len(audit_indices)
            popup_rows = []

            # ── 兜底列名查找（修复硬编码崩溃）──────────────────────────
            remark_col = next(
                (c for c in ['备注原因', '备注', 'remark'] if c in df_snapshot.columns),
                None
            )
            name_col = next(
                (c for c in ['组件物料描述', '物料名称', '物料描述'] if c in df_snapshot.columns),
                None
            )
            rate_col = next(
                (c for c in ['偏差率(%)', '偏差率', 'dev_rate'] if c in df_snapshot.columns),
                None
            )

            for seq, idx in enumerate(audit_indices):
                # ── 取消检查：直接读闭包捕获的 flag ────────────────────
                if _my_cancel_flag.is_set():
                    raise InterruptedError("用户取消了AI审核")

                row = df_snapshot.loc[idx]

                # 安全读取字段
                remark = row[remark_col] if remark_col else ''
                dev_rate_raw = row[rate_col] if rate_col else 0
                try:
                    dev_rate = float(dev_rate_raw or 0)
                except (ValueError, TypeError):
                    dev_rate = 0.0
                material_desc = str(row[name_col]) if name_col else ''

                try:
                    # ── 调用 ai_client（Mock 或真实接口）──────────────
                    result = self.ai_client.audit(remark, dev_rate)
                    audit_result = result.get('result', '需补备注')
                    ai_suggestion = result.get('suggestion', '')
                except (TimeoutError, ConnectionError) as ex:
                    audit_result = '审核失败'
                    ai_suggestion = str(ex)
                except Exception as ex:
                    audit_result = '审核失败'
                    # 完整日志写入文件，UI 只显示简略
                    import logging
                    logging.getLogger("AIWorker").exception("单行审核异常")
                    ai_suggestion = f"异常: {type(ex).__name__}"

                popup_rows.append({
                    'idx': idx,
                    '物料': material_desc,
                    '偏差率': f"{dev_rate:.1f}%",
                    '原备注': str(remark) if not pd.isna(remark) else '',
                    'AI建议': ai_suggestion,
                    '审核结果': audit_result,
                    '_audit_result': audit_result,
                    '_ai_suggestion': ai_suggestion,
                })

                # 上报进度（0-100）
                if progress_callback:
                    progress_callback(int((seq + 1) / total * 100))

            return popup_rows

        def _on_progress(pct):
            """进度回调——由 task_manager 轮询线程调度，但这里直接 after 保证线程安全"""
            self.root.after(0, lambda p=pct: (
                self.progress_bar.configure(value=p),
                self.set_status(f"AI审核中... {p}%")
            ))

        self.task_manager.run(
            _worker,
            callback=self._on_ai_audit_done,
            error_callback=self._on_ai_audit_error,
            progress_callback=_on_progress,
        )

    def _cancel_ai_audit(self):
        """取消当前 AI 审核（直接设置 cancel_flag）"""
        if hasattr(self, '_ai_cancel_flag') and self._ai_cancel_flag is not None:
            self._ai_cancel_flag.set()
        self.set_status("正在取消AI审核...")

    def _on_ai_audit_done(self, popup_rows):
        """AI审核完成回调（在主线程执行）"""
        self.is_auditing = False
        self.unsaved_ai_results = True  # 标记有未保存的 AI 结果（留给后续子任务处理导出拦截）

        # 进度条归位
        self.progress_bar.configure(value=100)
        self.root.after(400, lambda: (
            self.progress_bar.pack_forget(),
            self.progress_bar.configure(mode='indeterminate', value=0)
        ))

        # ── 将 Worker 结果写回 audit_data ──────────────────────────────
        for row_data in popup_rows:
            idx = row_data['idx']
            if idx in self.audit_data.index:
                self.audit_data.at[idx, 'audit_result'] = row_data['_audit_result']
                self.audit_data.at[idx, 'AI建议'] = row_data['_ai_suggestion']
                self.audit_data.at[idx, '审核来源'] = 'AI审核'

        # 更新审核状态列
        self.audit_data['审核状态'] = self.audit_data['audit_result'].apply(
            lambda x: '已审核' if x and str(x).strip() not in ['', 'nan'] else '未审核'
        )

        self._refresh_audit_tree(self.audit_data)
        self.set_status(f"AI审核完成，共 {len(popup_rows)} 条")

        # ── 显示结果弹窗 ────────────────────────────────────────────────
        if not popup_rows:
            return
        result_df = pd.DataFrame(popup_rows, columns=['物料', '偏差率', '原备注', 'AI建议', '审核结果'])
        win = tk.Toplevel(self.root)
        win.title(f"AI审核结果（共{len(popup_rows)}条）")
        win.geometry("900x500")
        win.transient(self.root)
        win.grab_set()
        cols = ['物料', '偏差率', '原备注', 'AI建议', '审核结果']
        tree = ttk.Treeview(win, columns=cols, show='headings')
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=150 if col in ('AI建议', '原备注') else 100)
        tree.pack(fill='both', expand=True, padx=5, pady=5)
        for _, r in result_df.iterrows():
            tree.insert('', 'end', values=[str(v) if pd.notna(v) else '' for v in r])
        sb = ttk.Scrollbar(win, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        btn_fr = tk.Frame(win)
        btn_fr.pack(pady=5)
        def _copy():
            lines = ['\t'.join(cols)]
            for child in tree.get_children():
                lines.append('\t'.join(str(v) for v in tree.item(child)['values']))
            win.clipboard_clear()
            win.clipboard_append('\n'.join(lines))
            messagebox.showinfo("已复制", "结果已复制到剪贴板", parent=win)
        tk.Button(btn_fr, text="复制到剪贴板", command=_copy).pack(side='left', padx=5)
        tk.Button(btn_fr, text="关闭", command=win.destroy).pack(side='left', padx=5)

    def _on_ai_audit_error(self, exc):
        """AI审核错误回调（在主线程执行）"""
        self.is_auditing = False
        self.progress_bar.pack_forget()
        self.progress_bar.configure(mode='indeterminate', value=0)

        if isinstance(exc, InterruptedError):
            self.set_status("AI审核已取消")
            messagebox.showinfo("已取消", "AI审核已被用户取消")
        else:
            self.set_status("AI审核失败")
            import traceback
            self.log(f"❌ AI审核异常：{traceback.format_exc()}", "error")
            messagebox.showerror("AI审核失败", f"审核过程发生异常：\n{str(exc)}")

    @with_feedback("导出成功", "导出失败")
    @with_feedback("Exporting Excel...", show_progress=True)
    def _export_audit_excel(self, cancel_flag=None, progress_callback=None):
        """Async export audit result (launcher)"""
        if self.audit_data is None or len(self.audit_data) == 0:
            messagebox.showwarning("Tip", "No data to export")
            return
        
        out_path = self.output_dir.get() or os.path.dirname(self.input_file.get()) or os.getcwd()
        file_path = filedialog.asksaveasfilename(
            initialdir=out_path, initialfile="ZPP011_Audit_Result.xlsx",
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path:
            return
        
        # Start async task
        self.is_exporting = True
        self.task_manager.run(
            lambda c, p: ExcelExporter.export(
                self.audit_data.copy(),
                file_path,
                cancel_flag=c,
                progress_callback=p
            ),
            task_id="export_audit_excel",
            on_progress=self._on_export_progress,
            on_done=self._on_export_done,
            on_error=self._on_export_error
        )
    
    def _on_export_progress(self, current, total, **kwargs):
        """Export progress callback"""
        if hasattr(self, 'progress_bar') and self.progress_bar:
            self.progress_bar["value"] = current
            self.progress_bar["maximum"] = total
            percent = int(current / total * 100) if total > 0 else 0
            eta_seconds = kwargs.get('eta_seconds', 0)
            eta_str = f"ETA {int(eta_seconds)}s" if eta_seconds > 0 else ""
            self.progress_bar.configure(
                text=f"Exporting {current}/{total} ({percent}%) {eta_str}"
            )
            self.root.update_idletasks()
    
    def _on_export_done(self, result):
        """Export success callback"""
        self.is_exporting = False
        if hasattr(self, 'progress_bar') and self.progress_bar:
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
        
        file_path = result.get('file_path', '')
        if file_path and os.path.exists(file_path):
            if messagebox.askyesno("Success", f"Exported to:\n{file_path}\n\nOpen folder?"):
                os.startfile(os.path.dirname(file_path))
            self.log(f"Audit result exported (async): {file_path}", "success")
        else:
            messagebox.showinfo("Success", "Export completed!")
    
    def _on_export_error(self, error):
        """Export error callback"""
        self.is_exporting = False
        if hasattr(self, 'progress_bar') and self.progress_bar:
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
        
        # Clean temp files
        temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
        if os.path.exists(temp_dir):
            for f in os.listdir(temp_dir):
                if f.endswith('.tmp.xlsx'):
                    try:
                        os.remove(os.path.join(temp_dir, f))
                    except:
                        pass
        
        messagebox.showerror("Error", f"Export failed: {error}")
        self.log(f"Export failed: {error}", "error")

    def _show_step_log(self, idx):
        pass

    def _load_lock_state(self):
        return {}

    def _add_alt(self):
        d = tk.Toplevel(self.root)
        d.title("添加替代料配对")
        d.geometry("580x300")
        d.transient(self.root)
        d.grab_set()
        d.configure(bg=C['bg'])

        tk.Label(d, text="物料A（编码或名称）：", font=("Microsoft YaHei", 10),
                 fg=C['text'], bg=C['bg']).pack(pady=(10, 3), anchor="w", padx=20)

        var_a = tk.StringVar()
        if hasattr(self, 'material_list') and self.material_list:
            cb_a = ttk.Combobox(d, textvariable=var_a, values=self.material_list,
                                font=("Consolas", 9), state="normal", width=65)
            cb_a.pack(padx=20, fill="x")
            cb_a.set("输入关键字或点击下拉选择...")
            def on_focus_a(e):
                if var_a.get() == "输入关键字或点击下拉选择...":
                    var_a.set("")
            cb_a.bind("<FocusIn>", on_focus_a)
        else:
            e_a = tk.Entry(d, font=("Consolas", 9), bg=C['surface2'],
                           fg=C['text'], insertbackground=C['accent'], relief="flat")
            e_a.pack(padx=20, fill="x")
            e_a.insert(0, "（未找到物料列表，请手动输入）")

        tk.Label(d, text="物料B（编码或名称）：", font=("Microsoft YaHei", 10),
                 fg=C['text'], bg=C['bg']).pack(pady=(10, 3), anchor="w", padx=20)

        var_b = tk.StringVar()
        if hasattr(self, 'material_list') and self.material_list:
            cb_b = ttk.Combobox(d, textvariable=var_b, values=self.material_list,
                                font=("Consolas", 9), state="normal", width=65)
            cb_b.pack(padx=20, fill="x")
            cb_b.set("输入关键字或点击下拉选择...")
            def on_focus_b(e):
                if var_b.get() == "输入关键字或点击下拉选择...":
                    var_b.set("")
            cb_b.bind("<FocusIn>", on_focus_b)
        else:
            e_b = tk.Entry(d, font=("Consolas", 9), bg=C['surface2'],
                           fg=C['text'], insertbackground=C['accent'], relief="flat")
            e_b.pack(padx=20, fill="x")
            e_b.insert(0, "（未找到物料列表，请手动输入）")

        # 已有配对列表（用于删除）
        lst = None  # 将在后面定义

        def confirm():
            a = var_a.get().strip()
            b = var_b.get().strip()
            if not a or not b:
                messagebox.showwarning("提示", "物料A和物料B都必须填写！")
                return

            def parse_selection(x):
                if '|' in x:
                    parts = [p.strip() for p in x.split('|')]
                    if len(parts) >= 3:
                        return parts[0], parts[1], parts[2]
                    elif len(parts) == 2:
                        return '', parts[0], parts[1]
                    else:
                        return '', parts[0], ''
                else:
                    # 手动输入时尝试从物料列表匹配
                    for item in getattr(self, 'material_list', []):
                        if x in item:
                            parts = [p.strip() for p in item.split('|')]
                            if len(parts) >= 3:
                                return parts[0], parts[1], parts[2]
                    return '', x, x

            factory_a, a_code, a_name = parse_selection(a)
            factory_b, b_code, b_name = parse_selection(b)

            # 去重检查：是否已存在相同配对
            exact_match = False
            conflict = False
            for (ea, eb) in self.alt_pairs:
                # 提取配对中的编码
                ea_code = ea[1] if isinstance(ea, (list, tuple)) and len(ea) >= 2 else str(ea)
                eb_code = eb[1] if isinstance(eb, (list, tuple)) and len(eb) >= 2 else str(eb)
                if (ea_code == a_code and eb_code == b_code) or (ea_code == b_code and eb_code == a_code):
                    exact_match = True
                    break
                if a_code in (ea_code, eb_code) or b_code in (ea_code, eb_code):
                    conflict = True

            if exact_match:
                msg = f"配对已存在：{a_code} ↔ {b_code}\n是否仍要添加？"
                if not messagebox.askyesno("重复配对", msg):
                    return
            if conflict:
                warn = f"物料 {a_code} 或 {b_code} 已存在于其他配对中，继续添加可能导致冲突。\n是否仍要添加？"
                if not messagebox.askyesno("物料冲突", warn):
                    return

            save_a = (factory_a, a_code, a_name)
            save_b = (factory_b, b_code, b_name)
            self.alt_pairs.append((save_a, save_b))
            try:
                save_alt_pairs(self.alt_pairs, log_cb=self.log)
            except Exception as e:
                messagebox.showerror("错误", f"替代料添加失败：{e}")
                return
            self._refresh_alt_view(self._alt_inner)
            messagebox.showinfo("提示", "替代料添加成功！")
            d.destroy()

        def do_del():
            if lst and lst.curselection():
                idx = lst.curselection()[0]
                del self.alt_pairs[idx]
                save_alt_pairs(self.alt_pairs, log_cb=self.log)
                self._refresh_alt_view(self._alt_inner)
                d.destroy()
            else:
                messagebox.showwarning("提示", "请先选择要删除的配对")

        btn_frame = tk.Frame(d, bg=C['bg'])
        btn_frame.pack(pady=12)
        tk.Button(btn_frame, text="✓ 确定", command=confirm, bg="#4CAF50", fg="white",
                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=8)
        tk.Button(btn_frame, text="✗ 取消", command=d.destroy, bg="#9E9E9E", fg="white",
                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=8)

        if self.alt_pairs:
            tk.Label(d, text="已有配对（点击可删除）：", font=("Microsoft YaHei", 9),
                     fg=C['text_dim'], bg=C['bg']).pack(pady=(10, 3))
            lst = tk.Listbox(d, font=("Consolas", 9), height=4, selectmode='single')
            lst.pack(fill="x", padx=20)
            for a, b in self.alt_pairs:
                # 解析显示
                if isinstance(a, (list, tuple)) and len(a) == 3:
                    _, a_code, a_name = a
                elif isinstance(a, (list, tuple)) and len(a) == 2:
                    a_code, a_name = a
                else:
                    a_code, a_name = str(a), ''
                if isinstance(b, (list, tuple)) and len(b) == 3:
                    _, b_code, b_name = b
                elif isinstance(b, (list, tuple)) and len(b) == 2:
                    b_code, b_name = b
                else:
                    b_code, b_name = str(b), ''
                left = f"{a_code} {a_name}" if a_name else a_code
                right = f"{b_code} {b_name}" if b_name else b_code
                lst.insert('end', left + " ⇄ " + right)
            tk.Button(d, text="删除选中配对", command=do_del, bg=C['danger'], fg="white",
                      font=("Microsoft YaHei", 9), relief="flat").pack(pady=5)



    def _del_alt(self):
        """删除替代料配对"""
        if not self.alt_pairs:
            messagebox.showinfo("提示", "当前没有替代料配对可删除")
            return
        d = tk.Toplevel(self.root)
        d.title("删除替代料配对")
        d.geometry("500x350")
        d.transient(self.root)
        d.grab_set()
        tk.Label(d, text="请选择要删除的配对：", font=("Microsoft YaHei", 10)).pack(pady=8)
        lst = tk.Listbox(d, font=("Consolas", 10), selectmode='single')
        lst.pack(fill="both", expand=True, padx=10)
        for a, b in self.alt_pairs:
            # a, b 都是 (factory, code, name) 元组
            factory_a, code_a, name_a = a if isinstance(a, tuple) else ('', a, '')
            factory_b, code_b, name_b = b if isinstance(b, tuple) else ('', b, '')
            left = f"{factory_a} {code_a} {name_a}"
            right = f"{factory_b} {code_b} {name_b}"
            lst.insert('end', left + " ⇄ " + right)
        if self.alt_pairs:
            lst.select_set(0)
        btn_frame = tk.Frame(d)
        btn_frame.pack(pady=10)
        def do_delete():
            sel = lst.curselection()
            if sel:
                del self.alt_pairs[sel[0]]
                save_alt_pairs(self.alt_pairs, log_cb=self.log)
                self._refresh_alt_view(self._alt_inner)
                d.destroy()
            else:
                messagebox.showwarning("提示", "请先选择要删除的配对")
        tk.Button(btn_frame, text="删除", command=do_delete, bg=C['danger'], fg="white",
                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)
        tk.Button(btn_frame, text="取消", command=d.destroy, bg="#d0d7de",
                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)


    def _reset_alt(self):
        self.alt_pairs = []   # 清空替代料配对
        # 刷新界面显示
        canvas = self.alt_list_frame.winfo_children()[0]
        inner = canvas.winfo_children()[0]
        self._refresh_alt_view(inner)
        # 保存到配置文件
        save_alt_pairs(self.alt_pairs, log_cb=self.log)
        




    def _show_alt_snapshot(self):
        """显示替代料配置的JSON快照（基于 alt_manager 动态路径）"""
        from domain.alt_material.alt_manager import _get_config_path

        config_path = _get_config_path()
        if not os.path.exists(config_path):
            messagebox.showinfo("提示",
                f"未找到配置文件：\n{config_path}\n\n当前使用默认内置替代料数据。")
            return

        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                content = f.read()
            parsed = json.loads(content)
            formatted = json.dumps(parsed, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("读取失败", f"无法解析 JSON 文件：\n{e}")
            return

        win = tk.Toplevel(self.root)
        win.title(f"替代料快照 - {os.path.basename(config_path)}")
        win.geometry("750x550")

        frame = tk.Frame(win)
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        scroll = tk.Scrollbar(frame)
        scroll.pack(side='right', fill='y')

        text = tk.Text(frame, wrap='none', yscrollcommand=scroll.set,
                       font=('Consolas', 10))
        text.pack(side='left', fill='both', expand=True)
        scroll.config(command=text.yview)

        text.insert('end', formatted)
        text.config(state='disabled')

        def copy_to_clip():
            win.clipboard_clear()
            win.clipboard_append(formatted)
            messagebox.showinfo("已复制", "JSON 内容已复制到剪贴板")

        tk.Button(win, text="复制到剪贴板", command=copy_to_clip,
                 font=("Microsoft YaHei", 10), relief="flat",
                 bg=C['accent'], fg="white").pack(pady=8)

    def _select_input(self):
        default_dir = r"E:\zpp011_dev\ZPP011导出文件原数据"
        os.makedirs(default_dir, exist_ok=True)
        p = filedialog.askopenfilename(
            title="选择 ZPP011 数据文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialdir=default_dir)

        if p:
            self.input_file.set(p)

            self._preview()


    def _select_output(self):
        p = filedialog.askdirectory(title="选择输出目录")

        if p:
            self.output_dir.set(p)


    def _auto_find(self):
        for pattern in ['zpp011 *.xlsx', 'ZPP011 *.xlsx']:
            hits = _glob.glob(os.path.join(os.getcwd(), pattern))

            if hits:
                hits.sort(key=os.path.getmtime, reverse=True)

                self.input_file.set(hits[0])

                self._preview()

                return

            hits = _glob.glob(
                os.path.join(
                    os.path.expanduser('~'),
                    'Desktop',
                    pattern))

            if hits:
                hits.sort(key=os.path.getmtime, reverse=True)

                self.input_file.set(hits[0])

                self._preview()

                return


    def _preview(self):
        path = self.input_file.get()

        if not path or not os.path.exists(path):
            self.preview_lbl.configure(text="❌ 文件不存在", fg=C['danger'])

            return

        try:
            df = pd.read_excel(path, sheet_name='Data', nrows=5)

            total = pd.read_excel(path, sheet_name='Data').shape[0]

            cols = df.columns.tolist()

            self.preview_lbl.configure(

                text=f"✅ {os.path.basename(path)}\n"

                f"   总行数：{total:,}  行\n"

                f"   替代料配对：{len(self.alt_pairs)}  组\n"

                f"   列数：{len(cols)}",

                fg=C['green'], justify="left"

            )

        except Exception as e:
            self.preview_lbl.configure(text=f"❌ 读取失败：{e}", fg=C['danger'])

        # 刷新物料列表（供替代料添加对话框下拉使用）
        self._load_material_list()


    def _export_log(self):
        """导出日志到文件"""
        try:
            import datetime as _dt
            ts = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"ZPP011日志_{ts}.txt"
            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                initialfile=default_name,
                title="导出日志"
            )
            if not path:
                return
            with open(path, 'w', encoding='utf-8') as f:
                f.write(self.log_text.get('1.0', 'end'))
            self.log(f"日志已导出：{os.path.basename(path)}", "success")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))


    def _export_changelog(self):
        """导出版本日志"""
        try:
            # 使用公共函数获取 changelog.json 路径
            cl_path = self._get_changelog_path() or ''
            
            # 读取 changelog.json
            cl_data = {}
            if os.path.isfile(cl_path):
                with open(cl_path, 'r', encoding='utf-8') as f:
                    cl_data = json.load(f)
            
            # 格式化日志
            lines = []
            app_name = cl_data.get('app_name', '云南达利ZPP011生产偏差分析器')
            author = cl_data.get('author', '裴盛清')
            versions = cl_data.get('versions', [])
            
            lines.append(app_name)
            lines.append(f"作者：{author}")
            lines.append("=" * 50)
            for v in versions:
                lines.append(f"\n【{v.get('version', '')}】{v.get('date', '')}")
                for feat in v.get('features', []):
                    lines.append(f"  ✦ {feat}")
                for fix in v.get('fixes', []):
                    lines.append(f"  🔧 {fix}")
                for opt in v.get('optimizations', []):
                    lines.append(f"  ⚡ {opt}")
                for les in v.get('lessons', []):
                    lines.append(f"  📌 {les}")
            
            changelog_text = "\n".join(lines)
            
            # 弹出保存对话框
            import datetime as _dt
            ts = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"ZPP011版本日志_{ts}.txt"
            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                initialfile=default_name,
                title="导出版本日志"
            )
            if not path:
                return
            with open(path, 'w', encoding='utf-8') as f:
                f.write(changelog_text)
            self.log(f"版本日志已导出：{os.path.basename(path)}", "success")
            messagebox.showinfo("导出成功", f"版本日志已保存到：\n{path}")
        except Exception as e:
            self.log(f"导出版本日志失败：{e}", "error")
            messagebox.showerror("导出失败", str(e))


    def _show_storyline(self):
        """弹出故事线窗口"""
        text_content = self._generate_storyline()
        if not text_content:
            return
        d = tk.Toplevel(self.root)
        d.title("📖 本周偏差故事线")
        d.geometry("600x520")
        d.transient(self.root)
        d.update_idletasks()
        rx = self.root.winfo_rootx() + (self.root.winfo_width() - 600) // 2
        ry = self.root.winfo_rooty() + (self.root.winfo_height() - 520) // 2
        d.geometry(f"+{rx}+{ry}")
        tk.Label(d, text="📖 本周偏差故事线", font=("Microsoft YaHei", 12, "bold")).pack(pady=10)
        text = tk.Text(d, font=("Microsoft YaHei", 10), wrap="word", height=20)
        text.pack(fill="both", expand=True, padx=10, pady=5)
        text.insert("1.0", text_content)
        text.configure(state="disabled")

        def copy_to_clipboard():
            d.clipboard_clear()
            d.clipboard_append(text_content)
            self.log("📖 故事线已复制到剪贴板", "info")
            messagebox.showinfo("已复制", "故事线已复制到剪贴板")

        tk.Button(d, text="📋 复制到剪贴板", command=copy_to_clipboard,
                  bg="#4a90d9", fg="white", font=("Microsoft YaHei", 10),
                  relief="flat", width=15).pack(pady=8)


    def _show_cleanup_window(self):
        """显示备注清洗标准化窗口"""
        suggestions = self._scan_remark_cleanup()
        if not suggestions:
            messagebox.showinfo("提示", "未发现可清洗的备注，数据已很规范")
            return

        d = tk.Toplevel(self.root)
        d.title("🧹 备注清洗标准化")
        d.geometry("750x480")
        d.transient(self.root)
        d.grab_set()
        d.update_idletasks()
        rx = self.root.winfo_rootx() + (self.root.winfo_width() - 750) // 2
        ry = self.root.winfo_rooty() + (self.root.winfo_height() - 480) // 2
        d.geometry(f"+{rx}+{ry}")

        tk.Label(d, text=f"发现 {len(suggestions)} 条可标准化的备注，勾选后点击执行",
                 font=("Microsoft YaHei", 10)).pack(pady=10)

        # 表格：物料 / 当前备注 / 建议标准化为
        tree_frame = tk.Frame(d)
        tree_frame.pack(fill="both", expand=True, padx=10)
        cols = ("select", "material", "current", "suggested")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15)
        tree.heading("select", text="✓")
        tree.heading("material", text="物料描述")
        tree.heading("current", text="当前备注")
        tree.heading("suggested", text="建议标准化为")
        tree.column("select", width=30, anchor="center")
        tree.column("material", width=150, anchor="w")
        tree.column("current", width=250, anchor="w")
        tree.column("suggested", width=200, anchor="w")

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll_y.set)
        scroll_y.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        # 填充数据
        self._cleanup_checkboxes = {}
        for i, sug in enumerate(suggestions):
            item_id = tree.insert('', 'end', values=('☐', sug['物料'], sug['当前备注'], sug['建议标准化为']))
            self._cleanup_checkboxes[item_id] = (False, sug)

        # 点击行切换勾选状态
        def toggle_check(event):
            item_id = tree.identify_row(event.y)
            if not item_id or item_id not in self._cleanup_checkboxes:
                return
            checked, sug = self._cleanup_checkboxes[item_id]
            new_checked = not checked
            self._cleanup_checkboxes[item_id] = (new_checked, sug)
            tree.set(item_id, "select", '☑' if new_checked else '☐')

        tree.bind("<Button-1>", toggle_check)

        # 底部按钮
        btn_frame = tk.Frame(d)
        btn_frame.pack(pady=10)

        def select_all():
            for item_id in self._cleanup_checkboxes:
                _, sug = self._cleanup_checkboxes[item_id]
                self._cleanup_checkboxes[item_id] = (True, sug)
                tree.set(item_id, "select", '☑')

        def execute_cleanup():
            cleaned = 0
            for item_id, (checked, sug) in self._cleanup_checkboxes.items():
                if checked and self.audit_data is not None:
                    self.audit_data.at[sug['idx'], '备注原因'] = sug['建议标准化为']
                    cleaned += 1
            if cleaned > 0:
                self._refresh_audit_tree(self.audit_data)
                self._update_audit_stats()
                self._update_filter_options()
                self.log(f"🧹 备注清洗完成：{cleaned} 条已标准化", "success")
                messagebox.showinfo("完成", f"已标准化 {cleaned} 条备注")
            d.destroy()

        tk.Button(btn_frame, text="全选", command=select_all,
                  bg="#d0d7de", font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)
        tk.Button(btn_frame, text="执行选中", command=execute_cleanup,
                  bg="#4CAF50", fg="white", font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)
        tk.Button(btn_frame, text="取消", command=d.destroy,
                  bg="#9E9E9E", fg="white", font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)


    def _set_step(self, idx, done=True):
        fr_data = self.step_frames.get(idx)

        if not fr_data:
            return

        fr = fr_data['frame']

        lbl = fr.winfo_children()[0]

        color = C['green'] if done else (
            C['accent'] if idx == self._current_step else C['text_dim'])

        lbl.configure(bg=color, fg="white" if done else C['text_dim'])

        fr.configure(bg=color, relief="flat" if done else "flat")


    def start_analysis(self):
        if self.running:
            return

        path = self.input_file.get()

        if not path or not os.path.exists(path):
            messagebox.showerror("错误", "请先选择有效的输入文件！")

            return

        self.running = True

        self.cancel_req = False

        self.run_btn.configure(state="disabled", text="⏳ 分析中...")

        self.cancel_btn.configure(state="normal")

        self.open_btn.configure(state="disabled")

        self.status_lbl.configure(text="分析中...", fg=C['warn'])

        self.log_text.configure(state="normal")

        self.log_text.delete("1.0", "end")

        self.log_text.configure(state="disabled")

        self.progress_var.set(0)

        self._current_step = -1

        for i in self.step_frames:
            self._set_step(i, False)

        self.start_time = time.time()

        self._update_timer()

        t = threading.Thread(target=self._analysis_thread, daemon=True)

        t.start()


    def _analysis_thread(self):
        import traceback
        import os

        _log = os.path.join(os.environ.get('TEMP', '.'), 'zpp011_debug.log')

        try:
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(
                    f"\n=== {__import__('datetime').datetime.now()} 线程启动 ===\n")

                _f.write(f"input_file={self.input_file.get()}\n")

                _f.write(f"output_dir={self.output_dir.get()}\n")

            self.output_path = do_analysis_v2(
                input_file=self.input_file.get(),
                output_dir=self.output_dir.get(),
                alt_pairs=self.alt_pairs,
                progress_callback=self._on_progress,
                cancel_check=lambda: self.cancel_req,
                start_date=self.start_date.get() or None,
                end_date=self.end_date.get() or None,
                material_search=self.material_search.get() or None,
            )

            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(
                    f"{__import__('datetime').datetime.now()} do_analysis_v2 正常返回\n")

            self.root.after(0, self._on_done)

        except KeyboardInterrupt:
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(
                    f"{__import__('datetime').datetime.now()} KeyboardInterrupt\n")

            self.root.after(0, self._on_cancel)

        except Exception as e:
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(
                    f"{__import__('datetime').datetime.now()} Exception: {e}\n")

                _f.write(traceback.format_exc() + "\n")

            self.root.after(0, lambda e=e: self._on_error(str(e)))

        except BaseException as e:
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(
                    f"{__import__('datetime').datetime.now()} BaseException: {e}\n")

                _f.write(traceback.format_exc() + "\n")

            self.root.after(0, lambda e=e: self._on_error(f"BaseException: {e}"))


    def _on_progress(self, step_idx, step_name, percent):
        if self.cancel_req:
            return

        self._current_step = step_idx

        import time as _t

        now = _t.strftime("%H:%M:%S")

        self.root.after(0, lambda step_idx=step_idx, step_name=step_name, percent=percent, now=now: (
            self.progress_var.set(percent),
            self.progress_lbl.configure(text=f"{step_name}  {percent:.0f}%", fg=C['accent']),
            self._set_step(step_idx),
            self.log(f"[{now}] [{step_idx + 1}/{len(STEPS)}] {step_name} {percent:.0f}%", "step")
        ))


    def _update_timer(self):
        if not self.running or not self.start_time:
            return

        elapsed = int(time.time() - self.start_time)

        m, s = elapsed // 60, elapsed % 60

        prog = self.progress_var.get()

        if prog > 5:
            total = int(elapsed / (prog / 100))

            rem = max(0, total - elapsed)

            rm, rs = rem // 60, rem % 60

            self.timer_lbl.configure(
                text=f"⏱ {m:02d}:{s:02d}  |  剩余 ~{rm:02d}:{rs:02d}",
                fg=C['warn'])

        else:
            self.timer_lbl.configure(
                text=f"⏱ {m:02d}:{s:02d}",
                fg=C['text_dim'])

        self.timer_id = self.root.after(1000, self._update_timer)


    def request_cancel(self):
        if self.running:
            self.cancel_req = True

            self.cancel_btn.configure(state="disabled", text="取消中...")

            self.log("⚠ 用户请求取消...", "warn")


    def _on_done(self):
        self.running = False
        self.log(f"🔍 _on_done: output_path = {getattr(self, 'output_path', 'NOT_SET')}", "info")

        # 如果 output_path 为 None，尝试自动查找最新的输出文件
        if not self.output_path:
            try:
                import glob as _glob
                out_dir = self.output_dir.get()
                if not out_dir or not os.path.isdir(out_dir):
                    out_dir = os.path.dirname(self.input_file.get()) if self.input_file.get() else os.path.expanduser('~/Desktop')
                pattern = os.path.join(out_dir, 'ZPP011偏差分析最终版_*.xlsx')
                files = _glob.glob(pattern)
                if files:
                    latest_file = max(files, key=os.path.getmtime)
                    self.output_path = latest_file
                    self.log(f"🔍 自动找到输出文件：{os.path.basename(latest_file)}", "info")
            except Exception as _e:
                self.log(f"🔍 自动查找输出文件失败：{_e}", "warn")

        if self.timer_id:
            self.root.after_cancel(self.timer_id)

        total = int(time.time() - self.start_time)

        m, s = total // 60, total % 60

        self.timer_lbl.configure(text=f"✅ {m:02d}:{s:02d} 完成", fg=C['green'])

        self.run_btn.configure(state="normal", text="▶ 重新分析")

        self.cancel_btn.configure(state="disabled")

        self.open_btn.configure(state="normal")

        self.progress_var.set(100)

        self.progress_lbl.configure(text="✅ 完成 (100%)", fg=C['green'])

        _basename = os.path.basename(self.output_path) if self.output_path else '未知文件'
        self.status_lbl.configure(
            text=f"完成 — {_basename}",
            fg=C['green'])

        for i in self.step_frames:
            self._set_step(i, True)

        self.log(f"\n✅ 分析完成！总用时 {m:02d}:{s:02d}", "success")

        self.status_lbl.configure(text="完成 — 数据已就绪，可加载审核", fg=C['green'])

        # ✅ 新增：自动加载审核数据，联动统计卡片（从分析结果文件加载）
        self._analysis_output_path = self.output_path  # 保存路径供加载使用
        self.log("🔄 准备加载审核数据...", "info")
        self.root.after(100, self._load_audit_data_from_output)
        self.root.after(200, self._run_pre_check)

        # 加载完成后删除自动生成的Excel（用户需要时可手动点击"生成Excel"按钮）
        self.root.after(500, self._cleanup_auto_excel)

        # ✅ 启用"加载审核数据"按钮，让用户手动加载
        if hasattr(self, 'load_audit_btn'):
            self.load_audit_btn.configure(state="normal")
            self.log("✅ 已启用「加载审核数据」按钮", "info")

    def _cleanup_auto_excel(self):
        """删除分析自动生成的Excel文件，用户需要时可手动生成"""
        try:
            if self._analysis_output_path and os.path.exists(self._analysis_output_path):
                fname = os.path.basename(self._analysis_output_path)
                os.remove(self._analysis_output_path)
                self.log(f"🗑️ 已清理自动生成文件: {fname}（需要时可点击「生成Excel」按钮）", "info")
                self._analysis_output_path = None
        except Exception as e:
            self.log(f"清理文件失败: {e}", "warn")


    def _on_cancel(self):
        self.running = False
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
        self.run_btn.configure(state="normal", text="▶ 开始分析")
        self.cancel_btn.configure(state="disabled")
        self.status_lbl.configure(text="已取消", fg=C['warn'])
        self.log("\n⚠ 分析已取消", "warn")


    def _on_error(self, msg):
        self.running = False
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
        self.run_btn.configure(state="normal", text="▶ 开始分析")
        self.cancel_btn.configure(state="disabled")
        self.status_lbl.configure(text="出错", fg=C['danger'])
        self.log(f"\n❌ 错误：{msg}", "error")
        messagebox.showerror("分析出错", msg)


    def _filter_by_stat(self, filter_key):
        """根据顶部统计卡片点击，筛选审核表格"""
        if self.audit_data is None or len(self.audit_data) == 0:
            return

        df = self.audit_data.copy()

        if filter_key == 'total':
            filtered = df
        elif filter_key == 'big_dev':
            filtered = df[df['偏差率(%)'].abs() > 10]
        elif filter_key == 'no_note':
            filtered = df[df['备注原因'].isna() | (df['备注原因'].astype(str).str.strip() == '')]
        elif filter_key == 'approved':
            filtered = df[df['备注原因'].notna() & (df['备注原因'].astype(str).str.strip() != '')]
        else:
            return

        # 刷新审核表格
        self._refresh_audit_tree(filtered)

        # 同步更新审核模块内的小统计块
        self._update_audit_stats(filtered)

        # 更新筛选提示文字
        labels = {
            'total': '全部',
            'big_dev': '偏差>10%',
            'no_note': '需补备注',
            'approved': '已审核',
        }
        self.status_filter_label.config(
            text=f"筛选：{labels.get(filter_key, '')} | 共 {len(filtered)} 条"
        )
    

    def _save_lock_state(self, locked):
        """保存列宽锁定状态"""
        try:
            from gui.ui_builder import COLUMN_WIDTHS_FILE
            state_path = COLUMN_WIDTHS_FILE.replace('column_widths.json', 'lock_state.json')
            os.makedirs(os.path.dirname(state_path), exist_ok=True)
            with open(state_path, 'w', encoding='utf-8') as f:
                json.dump({'locked': locked}, f)
        except Exception:
            pass

    def _toggle_column_lock(self):
        """切换列宽锁定状态"""
        self.column_locked = not self.column_locked
        self._save_lock_state(self.column_locked)
        
        # 解绑旧事件
        if hasattr(self, 'audit_tree'):
            try:
                self.audit_tree.unbind("<<TreeviewColumnResized>>")
            except:
                pass
            
            # 绑定新事件
            if self.column_locked:
                self.audit_tree.bind("<<TreeviewColumnResized>>", self._on_column_resized_lock)
            else:
                self.audit_tree.bind("<<TreeviewColumnResized>>", self._on_column_resized_save)
        
        self._apply_column_lock()
        self.lock_btn.configure(
            text="🔒 已锁定" if self.column_locked else "🔓 可调整",
            bg="#f0f0f0" if self.column_locked else "#e8f5e9"
        )
        self.log(f"列宽{'已锁定' if self.column_locked else '已解锁，可拖动调整'}", "info")

    def _on_column_resized_lock(self, event):
        """锁定状态下：阻止列宽变化，恢复为保存的宽度"""
        col = event.column
        if hasattr(self, 'column_widths') and col in self.column_widths:
            try:
                self.audit_tree.column(col, width=self.column_widths[col], stretch=False)
            except:
                pass

    def _on_column_resized_save(self, event):
        """解锁状态下：保存新的列宽"""
        col = event.column
        try:
            new_w = self.audit_tree.column(col, 'width')
            if not hasattr(self, 'column_widths'):
                self.column_widths = {}
            self.column_widths[col] = new_w
            self._save_column_widths()
        except:
            pass

    def _save_column_widths(self):
        """保存列宽配置到文件"""
        import json
        try:
            from gui.ui_builder import COLUMN_WIDTHS_FILE
            os.makedirs(os.path.dirname(COLUMN_WIDTHS_FILE), exist_ok=True)
            with open(COLUMN_WIDTHS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.column_widths, f)
        except:
            pass

    def _load_column_widths(self):
        """从文件加载列宽配置"""
        import json
        try:
            from gui.ui_builder import COLUMN_WIDTHS_FILE, DEFAULT_COL_WIDTHS
            if os.path.exists(COLUMN_WIDTHS_FILE):
                with open(COLUMN_WIDTHS_FILE, 'r', encoding='utf-8') as f:
                    saved = json.load(f)
                    return saved
        except:
            pass
        from gui.ui_builder import DEFAULT_COL_WIDTHS
        return DEFAULT_COL_WIDTHS.copy()

    def _reset_default_widths(self):
        """重置列宽为默认值"""
        from gui.ui_builder import DEFAULT_COL_WIDTHS, COLUMN_WIDTHS_FILE
        self.column_widths = DEFAULT_COL_WIDTHS.copy()
        if hasattr(self, 'audit_tree'):
            for col_id, w in DEFAULT_COL_WIDTHS.items():
                try:
                    min_w = w if getattr(self, 'column_locked', True) else 20
                    self.audit_tree.column(col_id, width=w, minwidth=min_w, stretch=False)
                except Exception:
                    pass
        # 删除保存的列宽文件
        try:
            if os.path.exists(COLUMN_WIDTHS_FILE):
                os.remove(COLUMN_WIDTHS_FILE)
        except Exception:
            pass
        self.log("列宽已重置为默认值", "info")


    def _apply_column_lock(self):
        """将锁定状态应用到审核表格的所有列"""
        if not hasattr(self, 'audit_tree'):
            return
        locked = getattr(self, 'column_locked', True)
        for col in self.audit_tree['columns']:
            try:
                current_width = self.audit_tree.column(col, 'width')
                if locked:
                    self.audit_tree.column(col, stretch=False, minwidth=current_width)
                else:
                    self.audit_tree.column(col, stretch=False, minwidth=20)
            except Exception:
                pass


    def _restore_column_widths(self):
        """从 ConfigManager 恢复保存的列宽"""
        if not hasattr(self, 'audit_tree') or not hasattr(self, 'config'):
            return
        widths = self.config.get('table.column_widths', {})
        if not widths:
            return
        for col, width in widths.items():
            if col in self.audit_tree['columns']:
                try:
                    self.audit_tree.column(col, width=int(width))
                except Exception:
                    pass

    @with_feedback("数据加载成功", "数据加载失败")
    def _load_audit_data(self):
        path = self.input_file.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "请先选择输入文件")
            return
        try:
            df = pd.read_excel(path, sheet_name='Data')
            df['偏差率(%)'] = pd.to_numeric(df['偏差率(%)'], errors='coerce').fillna(0)
            df['excel_row'] = df.index + 2
            
            # 日期范围过滤（如果用户输入了日期）
            start_date_str = self.start_date.get().strip()
            end_date_str = self.end_date.get().strip()
            
            if start_date_str or end_date_str:
                if '订单开始日期' in df.columns:
                    df['订单开始日期'] = pd.to_datetime(df['订单开始日期'], errors='coerce')
                    if start_date_str:
                        try:
                            start_dt = pd.to_datetime(start_date_str)
                            df = df[df['订单开始日期'] >= start_dt]
                            self.log(f"已按开始日期 {start_date_str} 过滤", "info")
                        except:
                            pass
                    if end_date_str:
                        try:
                            end_dt = pd.to_datetime(end_date_str)
                            df = df[df['订单开始日期'] <= end_dt]
                            self.log(f"已按结束日期 {end_date_str} 过滤", "info")
                        except:
                            pass
            
            total = len(df)
            high_dev = df[df['偏差率(%)'].abs() > 10]
            need_note = high_dev[high_dev['备注原因'].isna()]
            ok_note = high_dev[high_dev['备注原因'].notna()]
            
            # ── P1#11：计算偏差金额 ─────────────────────────────
            df['数量-定额'] = pd.to_numeric(df['数量-定额'], errors='coerce').fillna(0)
            df['数量-实际'] = pd.to_numeric(df['数量-实际'], errors='coerce').fillna(0)
            df['金额-实际(含税)'] = pd.to_numeric(df['金额-实际(含税)'], errors='coerce').fillna(0)
            # 计算含税单价（避开除以0）
            df['_unit_price'] = 0.0
            mask = df['数量-实际'] != 0
            df.loc[mask, '_unit_price'] = df.loc[mask, '金额-实际(含税)'] / df.loc[mask, '数量-实际']
            # 偏差金额 = (实际 - 定额) × 含税单价
            df['偏差金额'] = ((df['数量-实际'] - df['数量-定额']) * df['_unit_price']).round(2)
            # 删除临时列
            df.drop(columns=['_unit_price'], inplace=True)
            
            # 更新统计卡片
            self.audit_stat_labels['total'].configure(text=str(total))
            self.audit_stat_labels['high_dev'].configure(text=str(len(high_dev)))
            self.audit_stat_labels['need_note'].configure(text=str(len(need_note)))
            self.audit_stat_labels['ok_note'].configure(text=str(len(ok_note)))
            
            self.audit_data = high_dev.copy()
            # 映射订单日期列（Data sheet 用"订单开始日期"，统一为"订单日期"）
            if '订单开始日期' in self.audit_data.columns:
                self.audit_data['订单日期'] = pd.to_datetime(self.audit_data['订单开始日期'], errors='coerce').dt.strftime('%Y-%m-%d')
            elif '订单日期' not in self.audit_data.columns:
                self.audit_data['订单日期'] = ''
            self._update_filter_options()
            self._refresh_audit_tree(self.audit_data)
            
            self.audit_tree.tag_configure('need_note', background='#fff0e0', foreground='#b04000')
            self.audit_tree.tag_configure('ok_note',   background='#e8f5e9', foreground='#1a6b1a')
            self.audit_tree.tag_configure('ai_gen',    background='#fce4ec', foreground='#880e4f')
            self.audit_ai_btn.configure(state="normal")
            self.audit_export_btn.configure(state="normal")
            # 同时启用偏差明细表的按钮
            if hasattr(self, 'unified_ai_btn'):
                self.unified_ai_btn.configure(state="normal")
            if hasattr(self, 'unified_export_btn'):
                self.unified_export_btn.configure(state="normal")
            if hasattr(self, 'cleanup_btn'):
                self.cleanup_btn.configure(state="normal")
            if hasattr(self, 'save_audit_btn'):
                self.save_audit_btn.configure(state="normal")
            self._apply_row_colors()
            # P0-B4：恢复审核记录（从数据库）
            storage.restore_audit_from_db(self.audit_data, log_cb=self.log)
            self.log(f"智能审核：加载完成 | 共{total}条 | 偏差>10%共{len(high_dev)}条 | 需补备注{len(need_note)}条", "success")
        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败：{e}")
            self.log(f"加载审核数据失败：{e}", "error")


    def _load_audit_data_from_output_click(self, event=None):
        """按钮点击事件：加载审核数据（包装函数，忽略event参数）"""
        self._load_audit_data_from_output()


    @with_feedback("", "加载数据失败")
    @with_feedback("数据加载成功", "数据加载失败")
    def _load_audit_data_from_output(self, file_path=None):
        """分析完成后：异步从输出目录加载偏差明细到审核表格"""
        try:
            # 显示进度条
            if hasattr(self, 'progress_bar'):
                self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)
                self.progress_bar.start(10)
            if hasattr(self, 'set_status'):
                self.set_status("正在加载数据，请稍候...")

            # Capture UI values for worker (no UI access in worker)
            self._pending_output_dir = self.output_dir.get() if hasattr(self, 'output_dir') else None
            self.task_manager.run(
                lambda: self._load_data_worker(file_path),
                callback=self._on_load_done,
                error_callback=self._on_load_error
            )
        except Exception as e:
            if hasattr(self, 'progress_bar'):
                self.progress_bar.stop()
                self.progress_bar.pack_forget()
            raise

    def _load_data_worker(self, file_path=None):
        """纯数据处理：查找文件、读取、清洗、构建audit_df（禁止UI操作）"""
        import pandas as pd
        import glob as _glob

        # 1. 确定输出目录
        if not file_path:
            out_dir = None
            # output_dir 和 input_file 是 UI 变量，在launcher中获取
            if hasattr(self, '_pending_output_dir'):
                out_dir = self._pending_output_dir
            if not out_dir or not os.path.isdir(out_dir):
                out_dir = os.path.expanduser('~/Desktop')
        else:
            out_dir = None

        # 2. 找文件
        if file_path:
            latest_file = file_path
        else:
            pattern = os.path.join(out_dir, 'ZPP011偏差分析最终版_*.xlsx')
            files = _glob.glob(pattern)
            if not files:
                raise FileNotFoundError("未找到任何分析结果文件")
            latest_file = max(files, key=os.path.getmtime)

        # 3. 读取
        dev_df = pd.read_excel(latest_file, sheet_name='完整偏差明细')
        if dev_df.empty:
            raise ValueError("偏差明细工作表为空")

        # 4. 解析偏差率
        def parse_rate(v):
            if isinstance(v, str):
                return float(v.replace('%','').replace('＞','>').replace('>','')) / 100
            return abs(float(v)) if pd.notna(v) else 0
        dev_df['偏差率数值'] = dev_df['偏差率'].apply(parse_rate)

        # 5. 构建 audit_df
        audit_df = dev_df.copy()
        audit_df['excel_row'] = audit_df['原表行号'].apply(lambda x: int(x) if pd.notna(x) else 0)
        audit_df['组件物料号'] = audit_df.get('物料编码', '')
        audit_df['组件物料描述'] = audit_df.get('物料名称', '')
        audit_df['工厂名称'] = audit_df.get('工厂', '')
        audit_df['生产管理员描述'] = audit_df.get('车间', '')
        audit_df['数量-定额'] = audit_df.get('定额', 0)
        audit_df['数量-实际'] = audit_df.get('实际', 0)
        audit_df['偏差率(%)'] = audit_df['偏差率数值'] * 100
        audit_df['备注原因'] = audit_df.get('备注', '')

        # 偏差金额计算
        if '偏差金额' not in audit_df.columns or pd.to_numeric(audit_df['偏差金额'], errors='coerce').abs().max() == 0:
            if '金额-实际(含税)' in dev_df.columns and '数量-实际' in dev_df.columns:
                audit_df['_unit_price'] = 0.0
                m = (dev_df['数量-实际'] != 0) & dev_df['数量-实际'].notna()
                audit_df.loc[m, '_unit_price'] = dev_df.loc[m, '金额-实际(含税)'] / dev_df.loc[m, '数量-实际']
                if '数量-定额' in dev_df.columns:
                    audit_df['数量偏差'] = dev_df['数量-实际'] - dev_df['数量-定额']
                else:
                    audit_df['数量偏差'] = 0.0
                audit_df['偏差金额'] = (audit_df['数量偏差'] * audit_df['_unit_price']).round(2)
                audit_df.drop(columns=['_unit_price'], inplace=True, errors='ignore')
            else:
                audit_df['偏差金额'] = 0.0
        else:
            audit_df['偏差金额'] = pd.to_numeric(audit_df['偏差金额'], errors='coerce').fillna(0).round(2)

        # 订单列查找
        order_col = None
        for possible in ['流程订单', '订单号', '订单编号', '订单号码', '订单No', 'Order No', '生产订单']:
            if possible in audit_df.columns:
                order_col = possible
                break
            if possible in dev_df.columns:
                order_col = possible
                break
        if order_col is None:
            audit_df['流程订单'] = ''
        elif order_col != '流程订单':
            audit_df['流程订单'] = audit_df[order_col]

        # 生成唯一ID
        audit_df['_uid'] = (
            audit_df['订单日期'].astype(str).str[:10] + '_' +
            audit_df['流程订单'].astype(str) + '_' +
            audit_df['组件物料号'].astype(str)
        )
        if '备注来源' not in audit_df.columns:
            audit_df['备注来源'] = ''
        if '原备注' not in audit_df.columns:
            audit_df['原备注'] = audit_df.get('备注原因', '')
        if 'AI建议' not in audit_df.columns:
            audit_df['AI建议'] = ''
        if 'audit_result' not in audit_df.columns:
            audit_df['audit_result'] = ''

        return audit_df

    def _on_load_done(self, result_df):
        """异步加载成功回调：处理所有UI更新"""
        try:
            self.audit_data = result_df
            self._full_dev_df = result_df.copy()

            # 更新筛选选项和行颜色
            self._update_filter_options()
            self._apply_row_colors()
            if hasattr(self, '_update_trend_display'):
                self._update_trend_display()

            # 统计卡片
            total = len(result_df)
            high_dev = len(result_df[result_df['偏差率(%)'] > 10]) if '偏差率(%)' in result_df.columns else 0
            no_note = len(result_df[result_df['备注原因'].isna() | (result_df['备注原因'] == '')]) if '备注原因' in result_df.columns else 0
            approved_note = result_df['备注来源'].isin(['AI审核合格','AI审核待改进','AI生成']).sum() if '备注来源' in result_df.columns else 0

            if hasattr(self, 'audit_stat_labels'):
                for k, v in [('total', total), ('high_dev', high_dev),
                               ('need_note', no_note), ('ok_note', approved_note)]:
                    if k in self.audit_stat_labels:
                        self.audit_stat_labels[k].configure(text=str(v))
            if hasattr(self, 'unified_result_lbl'):
                self.unified_result_lbl.configure(
                    text=f"已加载 {total} 条 | 偏差>10%: {high_dev} | 需补备注: {no_note} | 已审核: {approved_note}")

            # 启用按钮
            for btn_name in [
                'load_audit_btn', 'unified_ai_btn', 'unified_export_btn',
                'save_audit_btn', 'tree_view_btn', 'cleanup_btn',
                'export_db_btn', 'import_db_btn', 'quarantine_btn', 'bom_btn',
            ]:
                btn = getattr(self, btn_name, None)
                if btn:
                    btn.configure(state='normal')

            self.log(f"✅ 审核数据加载完成 | 共 {total} 条 | 偏差>10%: {high_dev} 条", "success")

            # 恢复筛选和排序状态
            self._restore_filters()
            self.root.after(100, lambda: self._on_filter_changed(None) if self.audit_data is not None else None)
            self._restore_sort_state()
            self.root.after(300, lambda: self._show_precheck_report(self.audit_data))
            # BOM 过期提醒
            self.root.after(500, self._check_and_remind_bom)
            # 恢复审核记录
            storage.restore_audit_from_db(self.audit_data, log_cb=self.log)

            # 断点续审提示
            state = self._load_resume_state()
            if state:
                self.resume_btn.configure(state="normal")
                saved_row = state.get('selected_row', None)
                if saved_row is not None:
                    self.log(f"📌 检测到上次审核进度：第 {saved_row} 行", "info")
                    tip = tk.Toplevel(self.root)
                    tip.wm_overrideredirect(True)
                    tip.geometry(f"280x28+{self.root.winfo_rootx()+self.root.winfo_width()-290}+{self.root.winfo_rooty()+self.root.winfo_height()-60}")
                    tk.Label(tip, text=f"💡 上次审核到第 {saved_row} 行，点击「恢复进度」继续",
                            font=("Microsoft YaHei", 9), bg="#1a1a2e", fg="white",
                            padx=8, pady=4).pack()
                    tip.after(4000, tip.destroy)

            # 隐藏进度条
            if hasattr(self, 'progress_bar'):
                self.progress_bar.stop()
                self.progress_bar.pack_forget()
            if hasattr(self, 'set_status'):
                self.set_status(f"加载完成，共 {total} 行")
        except Exception as e:
            if hasattr(self, 'progress_bar'):
                self.progress_bar.stop()
                self.progress_bar.pack_forget()
            raise

    def _on_load_error(self, error):
        """异步加载失败回调"""
        if hasattr(self, 'progress_bar'):
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
        if hasattr(self, 'set_status'):
            self.set_status("加载失败")
        messagebox.showerror("加载错误", str(error))
        try:
            from core.logger import get_logger
            get_logger().error(f"数据加载异步失败: {error}")
        except Exception:
            pass


    def _on_tree_double_click(self, event):
        """双击审核行打开卡片详情（同万能搜索框结果兼容）"""
        self._show_audit_card(event)

    # ── P1：常用备注自动排序 ──

    def _show_audit_card(self, event):
        selection = self.audit_tree.selection()
        if not selection:
            return
        item = selection[0]
        vals = self.audit_tree.item(item, 'values')
        cols = self.audit_tree['columns']
        data = dict(zip(cols, vals))
        if hasattr(self, '_card_win') and self._card_win and self._card_win.winfo_exists():
            self._card_win.destroy()
        self._card_win = tk.Toplevel(self.root)
        self._card_win.title("\u5ba1\u6838\u5361\u7247")
        self._card_win.geometry("360x420")
        self._card_win.transient(self.root)
        self._card_win.attributes("-topmost", True)
        self.root.update_idletasks()
        rx = self.root.winfo_rootx() + self.root.winfo_width() + 5
        ry = self.root.winfo_rooty() + 100
        self._card_win.geometry(f"+{rx}+{ry}")
        card_bg = '#fefefe'
        self._card_win.configure(bg=card_bg)
        tk.Label(self._card_win, text="\U0001f4cb \u5ba1\u6838\u5361\u7247", font=("Microsoft YaHei", 12, "bold"),
            bg=card_bg).pack(pady=(12, 6))
        text = tk.Text(self._card_win, font=("Microsoft YaHei", 10), bg=card_bg,
            relief="flat", wrap="word", height=14)
        text.pack(fill="both", expand=True, padx=12, pady=6)
        parts = []
        for c in cols:
            if c != "idx":
                parts.append(c + "：" + str(data.get(c, "")))
        info = "\n".join(parts)
        # ── 成本换算器 ──
        try:
            dev_amount_raw = data.get('deviation_amount', '0')
            if dev_amount_raw and str(dev_amount_raw).strip() not in ('0', '-', ''):
                dev_amount_clean = str(dev_amount_raw).replace(',', '')
                dev_amount_val = float(dev_amount_clean) if dev_amount_clean else 0
                if abs(dev_amount_val) > 0.001:
                    excel_row = int(data.get('excel_row', 0))
                    unit_price = 0.0
                    unit_name = ''
                    if self.audit_data is not None and excel_row > 0:
                        er_mask = self.audit_data['excel_row'].astype(str) == str(excel_row)
                        if er_mask.any():
                            rd = self.audit_data[er_mask].iloc[0]
                            for pc in ('单价', '_单价'):
                                if pc in rd.index:
                                    try: unit_price = float(rd[pc] or 0); break
                                    except: pass
                            for uc in ('组件单位', '单位'):
                                if uc in rd.index:
                                    unit_name = str(rd[uc] or ''); break
                    if unit_price > 0.001:
                        est_qty = abs(dev_amount_val) / unit_price
                        ud = unit_name if unit_name else '单位'
                        info += f"\n💰 偏差¥{dev_amount_val:,.2f} ≈ {est_qty:.1f} {ud}（单价¥{unit_price:.2f}/{ud})"
                    else:
                        info += f"\n💰 偏差金额：¥{dev_amount_val:,.2f}（无单价数据）"
        except Exception as e:
            self.log(f"成本换算器出错：{e}", "warn")
        text.insert("1.0", info)
        text.configure(state="disabled")
        btn_fr = tk.Frame(self._card_win, bg=card_bg)
        btn_fr.pack(pady=(0, 10))
        
        # P1：预设备注列表（按频率排序）
        preset_remarks = ["系统无定额", "替代料", "已核实", "工艺调整", "其他"]
        sorted_remarks, freq = self._get_sorted_remarks(preset_remarks)
        
        def fill_remark(tag):
            self.audit_tree.set(item, 'remark', tag)
            self.audit_tree.set(item, 'status', '已备注')
            uid = data.get('_uid', '')
            if self.audit_data is not None and uid:
                # 通过唯一ID精确定位，避免误修改其他行
                mask = (self.audit_data['_uid'].astype(str) == str(uid))
                if mask.any():
                    self.audit_data.loc[mask, '备注原因'] = tag
            self._record_remark_freq(tag)
            self._card_win.destroy()
        
        for i, remark in enumerate(sorted_remarks):
            count = freq.get(remark, 0)
            # P1：前3名样式突出
            if i < 3 and count > 0:
                btn_font = ("Microsoft YaHei", 10, "bold")
                btn_bg = "#bbdefb"
            else:
                btn_font = ("Microsoft YaHei", 9)
                btn_bg = "#e3f2fd" if "无定额" in remark else "#fff9c4" if "替代" in remark else "#f5f5f5"
            tk.Button(btn_fr, text=f"{remark}({count})" if count > 0 else remark,
                      command=lambda r=remark: fill_remark(r),
                      bg=btn_bg, font=btn_font, relief="flat", width=12).pack(side="left", padx=3)


    def _batch_change_status(self, event=None):
        """批量更改选中行的状态"""
        selected = self.audit_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要更改状态的行")
            return

        status_options = ["已备注", "需补备注", "已确认", "待处理", "异常"]
        if hasattr(self, 'custom_statuses'):
            status_options += self.custom_statuses

        dialog = tk.Toplevel(self.root)
        dialog.title("批量更改状态")
        dialog.geometry("300x200")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 300) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 200) // 2
        dialog.geometry(f"+{x}+{y}")

        tk.Label(dialog, text=f"选择新状态（将为 {len(selected)} 行数据更改）:",
                 font=("Microsoft YaHei", 10)).pack(pady=15)
        status_var = tk.StringVar(value=status_options[0])
        status_combo = ttk.Combobox(dialog, textvariable=status_var, values=status_options,
                                    state="readonly", width=20, font=("Microsoft YaHei", 10))
        status_combo.pack(pady=10)

        def apply_change():
            new_status = status_var.get()
            for item in selected:
                self.audit_tree.set(item, 'status', new_status)
                excel_row = int(self.audit_tree.set(item, 'excel_row'))
                if self.audit_data is not None:
                    mask = self.audit_data['excel_row'].astype(str) == str(excel_row)
                    if mask.any() and '备注原因' in self.audit_data.columns:
                        self.audit_data.loc[mask, '备注原因'] = new_status
            self._update_audit_stats()
            self._update_filter_options()
            messagebox.showinfo("完成", f"已将 {len(selected)} 行状态更改为「{new_status}」")
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        tk.Button(btn_frame, text="确定", width=10, command=apply_change).pack(side="left", padx=10)
        tk.Button(btn_frame, text="取消", width=10, command=dialog.destroy).pack(side="left", padx=10)


    def _batch_fill_remark(self, event=None):
        """批量填写备注"""
        selected = self.audit_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要填写备注的行")
            return

        # P1：按使用频率排序的预设备注列表
        base_remarks = ["已核实，偏差正常", "已沟通，确认无误", "已调整，请复查", "待进一步确认", "替代料替换", "工艺调整", "库存调整", "其他原因"]
        sorted_remarks, _ = self._get_sorted_remarks(base_remarks)
        remark_options = sorted_remarks + ["(清空备注)"]
        if hasattr(self, 'custom_remarks'):
            remark_options = self.custom_remarks + remark_options

        dialog = tk.Toplevel(self.root)
        dialog.title("批量填写备注")
        dialog.geometry("320x180")
        dialog.transient(self.root)
        dialog.grab_set()
        x = self.root.winfo_x() + (self.root.winfo_width() - 320) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 180) // 2
        dialog.geometry(f"+{x}+{y}")

        tk.Label(dialog, text=f"选择备注（将为 {len(selected)} 行填写）:",
                 font=("Microsoft YaHei", 10)).pack(pady=15)
        remark_var = tk.StringVar()
        remark_combo = ttk.Combobox(dialog, textvariable=remark_var, values=remark_options,
                                    width=28, font=("Microsoft YaHei", 10))
        remark_combo.pack(pady=10)
        remark_combo.focus()

        def apply_remark():
            remark = remark_var.get()
            if remark == "(清空备注)":
                remark = ""
            count = 0
            for item in selected:
                self.audit_tree.set(item, 'batch_remark', remark)  # 写入批量备注列
                excel_row = int(self.audit_tree.set(item, 'excel_row'))
                if self.audit_data is not None:
                    mask = self.audit_data['excel_row'].astype(str) == str(excel_row)
                    if mask.any():
                        # 优先写入批量备注列，兼容 fallback
                        batch_col = None
                        for col in ['批量备注原因', '批量备注']:
                            if col in self.audit_data.columns:
                                batch_col = col
                                break
                        if batch_col:
                            self.audit_data.loc[mask, batch_col] = remark
                            count += 1
                        elif '备注原因' in self.audit_data.columns:
                            self.audit_data.loc[mask, '备注原因'] = remark
                            count += 1
            self._record_remark_freq(remark_var.get())  # P1：记录频率
            self._update_audit_stats()
            self._update_filter_options()
            label = "清空" if not remark else f"「{remark}」"
            messagebox.showinfo("完成", f"已为 {count if count else len(selected)} 行填写备注 {label}")
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="确定", width=10, command=apply_remark).pack(side="left", padx=10)
        tk.Button(btn_frame, text="取消", width=10, command=dialog.destroy).pack(side="left", padx=10)

    @with_feedback("批量备注完成", "批量备注失败")
    def _batch_remark(self, event=None):
        """批量备注：为选中的行添加备注（下拉框选择，可追加换行）"""
        try:
            selected = self.audit_tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请先选中要添加备注的行（可多选）")
                return

            # P1：按使用频率排序的预设备注列表
            base_remarks = ["已核实，偏差正常", "已沟通，确认无误", "已调整，请复查", "待进一步确认", "替代料替换", "工艺调整", "库存调整", "其他原因"]
            sorted_remarks, _ = self._get_sorted_remarks(base_remarks)
            remark_options = sorted_remarks + ["(清空备注)", "——手动输入——"]
            if hasattr(self, 'custom_remarks'):
                remark_options = self.custom_remarks + remark_options

            dialog = tk.Toplevel(self.root)
            dialog.title("批量备注")
            dialog.geometry("360x200")
            dialog.transient(self.root)
            dialog.grab_set()
            x = self.root.winfo_x() + (self.root.winfo_width() - 360) // 2
            y = self.root.winfo_y() + (self.root.winfo_height() - 200) // 2
            dialog.geometry(f"+{x}+{y}")

            tk.Label(dialog, text=f"选择备注（将为 {len(selected)} 行填写）:",
                     font=("Microsoft YaHei", 10)).pack(pady=10)

            remark_var = tk.StringVar()
            remark_combo = ttk.Combobox(dialog, textvariable=remark_var, values=remark_options,
                                        width=36, font=("Microsoft YaHei", 10), state="readonly")
            remark_combo.pack(pady=8)
            remark_combo.focus()

            # 手动输入框（默认隐藏）
            input_frame = tk.Frame(dialog)
            input_frame.pack(pady=4)
            custom_entry = ttk.Entry(input_frame, width=36, font=("Microsoft YaHei", 10))
            custom_entry.pack()
            input_frame.pack_forget()  # 默认隐藏

            # 监听选择变化，切换手动输入框
            def on_combo_change(*args):
                if remark_var.get() == "——手动输入——":
                    input_frame.pack(pady=4)
                    custom_entry.focus()
                else:
                    input_frame.pack_forget()

            remark_combo.bind("<<ComboboxSelected>>", on_combo_change)

            def apply_remark():
                remark = remark_var.get()
                if remark == "(清空备注)":
                    remark = ""
                elif remark == "——手动输入——":
                    remark = custom_entry.get().strip()
                    if not remark:
                        messagebox.showwarning("提示", "请输入备注内容")
                        return
                if not remark:
                    return

                # 确定 DataFrame 中的备注列名（优先批量备注列）
                df_remark_col = None
                for col in ['批量备注原因', '批量备注']:
                    if col in self.audit_data.columns:
                        df_remark_col = col
                        break
                if df_remark_col is None:
                    df_remark_col = "批量备注"
                    self.audit_data[df_remark_col] = ""

                # 使用 excel_row 定位，构建 item->df_idx 映射
                item_to_idx = {}
                for item in selected:
                    excel_row = int(self.audit_tree.set(item, 'excel_row'))
                    mask = self.audit_data['excel_row'].astype(str) == str(excel_row)
                    if mask.any():
                        item_to_idx[item] = mask.idxmax()

                if not item_to_idx:
                    messagebox.showwarning("警告", "无法定位选中行，请刷新重试")
                    return

                # 追加备注（换行连接，区分符 /）
                count = 0
                for item, idx in item_to_idx.items():
                    self.audit_tree.set(item, 'batch_remark', remark)  # 刷新树形列
                    current = self.audit_data.at[idx, df_remark_col]
                    if pd.notna(current) and str(current).strip():
                        self.audit_data.at[idx, df_remark_col] = f"{current}\n/{remark}"
                    else:
                        self.audit_data.at[idx, df_remark_col] = remark
                    count += 1

                self._record_remark_freq(remark)  # P1：记录频率
                self._refresh_audit_tree()
                self._update_audit_stats()
                self._update_filter_options()
                label = "清空" if not remark else f"「{remark}」"
                messagebox.showinfo("完成", f"已为 {count} 行添加备注 {label}")
                dialog.destroy()

            btn_frame = tk.Frame(dialog)
            btn_frame.pack(pady=10)
            tk.Button(btn_frame, text="确定", width=10, command=apply_remark).pack(side="left", padx=10)
            tk.Button(btn_frame, text="取消", width=10, command=dialog.destroy).pack(side="left", padx=10)

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"批量备注失败：{str(e)}")

    def _add_custom_status(self, event=None):
        """添加自定义状态标签"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加状态标签")
        dialog.geometry("280x140")
        dialog.transient(self.root)
        dialog.grab_set()
        x = self.root.winfo_x() + (self.root.winfo_width() - 280) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 140) // 2
        dialog.geometry(f"+{x}+{y}")

        tk.Label(dialog, text="输入新状态标签名称：", font=("Microsoft YaHei", 10)).pack(pady=10)
        name_var = tk.StringVar()
        entry = tk.Entry(dialog, textvariable=name_var, width=20, font=("Microsoft YaHei", 10))
        entry.pack(pady=8)
        entry.focus()

        def do_add():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("提示", "标签名称不能为空")
                return
            if not hasattr(self, 'custom_statuses'):
                self.custom_statuses = []
            if name not in self.custom_statuses:
                self.custom_statuses.append(name)
                self._update_filter_options()
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=12)
        tk.Button(btn_frame, text="添加", width=8, command=do_add).pack(side="left", padx=8)
        tk.Button(btn_frame, text="取消", width=8, command=dialog.destroy).pack(side="left", padx=8)


    def _auto_close(self):
        """自动结案（异步版本）"""
        # 初始化状态变量
        if not hasattr(self, '_is_auto_closing'):
            self._is_auto_closing = False
        if not hasattr(self, '_auto_close_cancel_flag'):
            self._auto_close_cancel_flag = None
        
        if self._is_auto_closing:
            messagebox.showwarning("提示", "自动结案任务进行中，请勿重复操作")
            return
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showwarning("警告", "没有数据可操作")
            return
        
        self._is_auto_closing = True
        self.progress_bar.configure(mode='determinate', maximum=100)
        self.progress_bar['value'] = 0
        self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)
        self.set_status("正在执行自动结案...")
        
        # 双快照：数据和规则引擎（防止规则漂移）
        df_snapshot = self.audit_data.copy(deep=True)
        rule_engine_snapshot = deepcopy(self.rule_engine)
        
        # 包装函数，保存 cancel_flag 引用
        def wrapper(cancel_flag, progress_callback):
            self._auto_close_cancel_flag = cancel_flag
            return AutoCloser.process(
                df_snapshot, rule_engine_snapshot, progress_callback, cancel_flag
            )
        
        self.task_manager.run(
            wrapper,
            callback=self._on_auto_close_done,
            error_callback=self._on_auto_close_error,
            progress_callback=self._on_auto_close_progress
        )
    
    def _on_auto_close_progress(self, current, total, eta):
        """进度回调"""
        percent = int(current / total * 100) if total else 0
        self.progress_bar['value'] = percent
        self.progress_bar.update_idletasks()
        eta_text = f", 剩余约{int(eta)}秒" if eta else ""
        self.set_status(f"自动结案中: {current}/{total} ({percent}%){eta_text}")
    
    def _on_auto_close_done(self, result):
        """成功回调"""
        df, success, fail, fail_rows = result
        self.audit_data = df
        self._refresh_audit_tree()
        self._is_auto_closing = False
        self._auto_close_cancel_flag = None
        self.progress_bar.pack_forget()
        self.set_status(f"自动结案完成，成功 {success} 行，失败 {fail} 行")
        msg = f"自动结案完成\n成功: {success} 行\n失败: {fail} 行"
        if fail > 0:
            msg += f"\n失败行号: {fail_rows[:10]}{'...' if len(fail_rows) > 10 else ''}\n请查看日志获取详情。"
        messagebox.showinfo("完成", msg)
    
    def _on_auto_close_error(self, error):
        """错误回调（取消时丢弃快照，不回写数据）"""
        self._is_auto_closing = False
        self._auto_close_cancel_flag = None
        self.progress_bar.pack_forget()
        
        if isinstance(error, InterruptedError):
            self.set_status("自动结案已取消")
            messagebox.showwarning("已取消", "自动结案操作已取消，数据未变动。")
        else:
            self.set_status("自动结案失败")
            messagebox.showerror("错误", f"自动结案失败: {error}")
            from core.logger import get_logger
            get_logger("Events").error(f"自动结案异步失败: {error}")
    
    def _cancel_auto_close(self):
        """取消自动结案"""
        if self._auto_close_cancel_flag:
            self._auto_close_cancel_flag.set()
            self.set_status("正在取消自动结案...")
        else:
            messagebox.showwarning("提示", "当前没有正在运行的自动结案任务")
    def _move_to_quarantine(self):
        """将选中的行移动到隔离区（标记但不删除）"""
        try:
            selected = self.audit_tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请先选中要隔离的行")
                return

            if 'is_quarantined' not in self.audit_data.columns:
                self.audit_data['is_quarantined'] = False

            count = 0
            for item in selected:
                idx = self.audit_tree.index(item)
                self.audit_data.at[idx, 'is_quarantined'] = True
                count += 1

            self._refresh_audit_tree(self.audit_data)
            messagebox.showinfo("完成", f"已隔离 {count} 行")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"隔离操作失败：{str(e)}")

    def _on_close(self):
        """窗口关闭前保存断点状态"""
        try:
            self._save_resume_state()
            self.log("断点已保存", "success")
        except Exception as e:
            self.log(f"断点保存失败：{e}", "warn")
        # 保存列宽 + 窗口几何
        try:
            if hasattr(self, 'audit_tree'):
                widths = {}
                for col in self.audit_tree['columns']:
                    widths[col] = self.audit_tree.column(col, 'width')
                self.config.set('table.column_widths', widths)
            self.config.save_window_geometry(self.root)
        except Exception:
            pass
        self.root.destroy()


    def _get_resume_state_path(self):
        """获取断点状态文件路径"""
        app_dir = os.path.join(os.path.expanduser('~'), '.zpp011_audit')
        return os.path.join(app_dir, 'resume_state.json')

    def _load_resume_state(self):
        """加载断点状态"""
        path = self._get_resume_state_path()
        if not os.path.exists(path):
            return None
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return None

    def _save_resume_state(self):
        """保存断点状态"""
        app_dir = os.path.join(os.path.expanduser('~'), '.zpp011_audit')
        os.makedirs(app_dir, exist_ok=True)
        path = self._get_resume_state_path()
        
        state = {}
        # 保存选择的行
        if hasattr(self, 'current_row_idx'):
            state['selected_row'] = self.current_row_idx
        # 保存搜索文字
        if hasattr(self, 'search_var'):
            state['search_text'] = self.search_var.get()
        # 保存筛选条件
        if hasattr(self, 'filter_widgets'):
            state['filter_values'] = {}
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(state, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.log(f"保存断点失败：{e}", "warn")


    def _do_resume_state(self):
        """一键恢复上次审核状态"""
        state = self._load_resume_state()
        if not state:
            return
        # 恢复搜索文字
        search_text = state.get('search_text', '')
        if search_text and hasattr(self, 'search_var'):
            self.search_var.set(search_text)
        # 恢复筛选条件
        filters = state.get('filter_values', {})
        for key, value in filters.items():
            if key in self.filter_widgets:
                widget = self.filter_widgets[key]
                if key == 'name':
                    widget.delete(0, tk.END)
                    widget.insert(0, value)
                elif key == 'order_date':
                    if isinstance(widget, tuple) and len(widget) == 2 and isinstance(value, list) and len(value) == 2:
                        widget[0].delete(0, tk.END)
                        widget[0].insert(0, value[0])
                        widget[1].delete(0, tk.END)
                        widget[1].insert(0, value[1])
                else:
                    widget.set(value)
        # 重新触发筛选
        self._on_filter_changed('restore')
        # 恢复选中行
        saved_row = state.get('selected_row')
        if saved_row:
            children = self.audit_tree.get_children()
            if int(saved_row) <= len(children):
                self.audit_tree.selection_set(children[int(saved_row) - 1])
                self.audit_tree.see(children[int(saved_row) - 1])
        self.resume_btn.configure(state="disabled")
        self.log("✅ 已恢复上次审核进度", "success")


    def _import_bom(self):
        """导入 BOM 数据文件，与旧 BOM 比对差异"""
        file_path = filedialog.askopenfilename(
            title="选择 BOM 数据文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        try:
            self.log(f"📦 正在读取 BOM 文件：{os.path.basename(file_path)}", "info")
            new_df = pd.read_excel(file_path, sheet_name=0)
            self.log(f"   共读取 {len(new_df)} 行数据", "info")
        except Exception as e:
            messagebox.showerror("读取失败", f"无法读取 BOM 文件：{e}")
            return

        # ── 查找物料编码键 ──
        key_candidates = ['组件物料号', '物料编码', '物料号', '物料', 'code', 'material_code']
        key_col = None
        for col in key_candidates:
            if col in new_df.columns:
                key_col = col
                break
        if not key_col:
            cols_str = '、'.join(new_df.columns[:10].tolist())
            messagebox.showerror("列缺失", f"未找到物料编码列（候选：{key_candidates}）\n当前列：{cols_str}")
            return
        self.log(f"   物料编码键列：{key_col}", "info")

        # 清洗新数据
        new_df[key_col] = new_df[key_col].astype(str).str.strip()
        new_records = new_df.to_dict('records')
        new_keys = set(new_df[key_col].dropna().unique())

        # ── 加载旧 BOM ──
        old_path = self._get_bom_stored_path()
        old_records = []
        old_keys = set()
        old_key_col = key_col  # 默认使用新文件的key列
        added_count, removed_count, modified_count = 0, 0, 0
        detail_lines = []

        if os.path.exists(old_path):
            try:
                with open(old_path, 'r', encoding='utf-8') as f:
                    old_data = json.load(f)
                if isinstance(old_data, dict) and 'records' in old_data:
                    old_records = old_data['records']
                    # 使用旧BOM保存时的key列名（兼容新旧格式不一致的情况）
                    old_key_col = old_data.get('key_column', key_col)
                elif isinstance(old_data, list):
                    old_records = old_data
                    old_key_col = key_col
                else:
                    old_key_col = key_col
                old_keys = {str(r.get(old_key_col, '')).strip() for r in old_records if r.get(old_key_col)}
                self.log(f"   旧 BOM 共 {len(old_records)} 条记录", "info")
            except Exception as e:
                self.log(f"   旧 BOM 读取失败（将视为全新导入）：{e}", "warning")

        # ── 比对 ──
        if not old_keys:
            added_count = len(new_keys)
            detail_lines.append(f"🆕 全新导入，共 {added_count} 条新记录")
        else:
            added_keys = new_keys - old_keys
            removed_keys = old_keys - new_keys
            modified_count = 0

            # 快速修改检测（旧记录转dict）
            old_map = {str(r.get(old_key_col, '')).strip(): r for r in old_records if r.get(old_key_col)}
            for key in new_keys & old_keys:
                new_row = new_df[new_df[key_col] == key].iloc[0]
                old_row = old_map.get(key, {})
                changed = False
                for c in new_df.columns:
                    if c == key_col:
                        continue
                    new_val = str(new_row.get(c, '')).strip()
                    old_val = str(old_row.get(c, '')).strip()
                    if new_val != old_val:
                        changed = True
                        break
                if changed:
                    modified_count += 1

            added_count = len(added_keys)
            removed_count = len(removed_keys)

            if added_count > 0:
                sample_added = list(added_keys)[:5]
                detail_lines.append(f"🆕 新增 {added_count} 条（示例：{', '.join(sample_added)}）")
            if removed_count > 0:
                sample_removed = list(removed_keys)[:5]
                detail_lines.append(f"➖ 删除 {removed_count} 条（示例：{', '.join(sample_removed)}）")
            if modified_count > 0:
                detail_lines.append(f"✏️ 修改 {modified_count} 条")

        # ── 保存新 BOM ──
        bom_store = {
            'version': '1.0',
            'imported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'key_column': key_col,
            'total_rows': len(new_df),
            'records': new_records
        }
        try:
            with open(old_path, 'w', encoding='utf-8') as f:
                json.dump(bom_store, f, ensure_ascii=False, indent=2)
            self.log(f"   BOM 已保存至：{old_path}", "info")
        except Exception as e:
            messagebox.showerror("保存失败", f"无法保存 BOM 数据：{e}")
            return

        # ── 展示对比报告 ──
        report_win = Toplevel(self.root)
        report_win.title("BOM 导入报告")
        report_win.geometry("600x400")
        report_win.transient(self.root)
        report_win.grab_set()
        center_window(report_win, 600, 400)

        bg_frame = Frame(report_win, bg='#1e1e1e')
        bg_frame.pack(fill='both', expand=True)

        title_label = Label(bg_frame, text="📦 BOM 导入报告", font=('微软雅黑', 14, 'bold'),
                            bg='#1e1e1e', fg='#ffffff')
        title_label.pack(pady=(20, 10))

        summary_text = f"文件：{os.path.basename(file_path)}\n"
        summary_text += f"总行数：{len(new_df)}　｜　"
        summary_text += f"🆕新增 {added_count}　｜　"
        summary_text += f"➖删除 {removed_count}　｜　"
        summary_text += f"✏️修改 {modified_count}"
        summary_label = Label(bg_frame, text=summary_text, font=('微软雅黑', 10),
                              bg='#1e1e1e', fg='#cccccc', justify='left', anchor='w')
        summary_label.pack(padx=20, fill='x')

        canvas = Canvas(bg_frame, bg='#1e1e1e', highlightthickness=0)
        scrollbar = Scrollbar(bg_frame, orient='vertical', command=canvas.yview)
        content_frame = Frame(canvas, bg='#1e1e1e')

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=10)
        scrollbar.pack(side='right', fill='y', pady=10, padx=(0, 20))

        canvas_window = canvas.create_window((0, 0), window=content_frame, anchor='nw')
        content_frame.bind('<Configure>',
                           lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>',
                    lambda e: canvas.itemconfig(canvas_window, width=e.width))

        if detail_lines:
            for line in detail_lines:
                color = '#1a7f37' if '新增' in line else ('#d29922' if '删除' in line else '#79c0ff')
                lbl = Label(content_frame, text=line, font=('微软雅黑', 10),
                            bg='#1e1e1e', fg=color, anchor='w', justify='left')
                lbl.pack(anchor='w', pady=2)
        else:
            Label(content_frame, text="✓ 无变更（数据完全一致）", font=('微软雅黑', 10),
                  bg='#1e1e1e', fg='#1a7f37', anchor='w').pack(anchor='w', pady=2)

        close_btn = Button(bg_frame, text="关闭", font=('微软雅黑', 10),
                           bg='#3a3a3a', fg='#ffffff', activebackground='#555555',
                           command=report_win.destroy, cursor='hand2')
        close_btn.pack(pady=15)

        self.log(f"📦 BOM 导入完成：新增 {added_count} / 删除 {removed_count} / 修改 {modified_count}", "info")
        self.log(f"   BOM 数据已保存，下次过期检查将自动对比差异", "info")


    def _batch_export(self, event=None):
        """批量导出选中行为Excel或CSV"""
        selected = self.audit_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要导出的行")
            return

        file_path = filedialog.asksaveasfilename(
            title="导出选中行",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        columns = self.audit_tree["columns"]
        export_data = []
        for item in selected:
            row_values = [self.audit_tree.set(item, col) for col in columns]
            export_data.append(row_values)

        try:
            if file_path.endswith('.csv'):
                import csv
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(columns)
                    writer.writerows(export_data)
            else:
                pd.DataFrame(export_data, columns=columns).to_excel(file_path, index=False, engine='openpyxl')
            messagebox.showinfo("导出成功", f"已成功导出 {len(export_data)} 行数据到\n{file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出时发生错误：{str(e)}")


    def _copy_wechat_draft(self):
        """将选中行生成微信草稿并复制到剪贴板"""
        selected = self.audit_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要生成草稿的行")
            return

        cols = self.audit_tree['columns']
        lines = ["【料控指令】", ""]
        for i, item in enumerate(selected, 1):
            vals = self.audit_tree.item(item, 'values')
            data = dict(zip(cols, vals))
            order_no = data.get('order_date', '')
            mat_name = data.get('name', '')
            dev_rate = data.get('dev_rate', '')
            status = data.get('status', '')
            remark = data.get('remark', '')
            code = data.get('code', '')

            # 简洁格式：序号. 物料 偏差率 状态
            line = f"{i}. {mat_name}（{code}）偏差{dev_rate} {status}"
            if remark and remark.strip():
                line += f"\n   备注：{remark}"
            lines.append(line)

        lines.append("")
        lines.append(f"共 {len(selected)} 条 | 请确认后处理")

        draft = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(draft)
        self.log(f"📋 已复制 {len(selected)} 条微信草稿到剪贴板", "info")
        messagebox.showinfo("复制成功", f"已复制 {len(selected)} 条指令到剪贴板，可直接粘贴到微信")


    @with_feedback("移入隔离区成功", "移入隔离区失败")
    def _quarantine_selected(self):
        """将选中行移入隔离区"""
        selected = self.audit_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要移入隔离区的行")
            return

        quarantine_list = self._load_quarantine()
        removed_indices = []

        for item in selected:
            values = self.audit_tree.item(item, 'values')
            cols = self.audit_tree['columns']
            record = dict(zip(cols, values))
            record['_quarantined_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            quarantine_list.append(record)

            # 从底层 DataFrame 中移除对应行
            excel_row = int(record.get('excel_row', 0))
            if self.audit_data is not None:
                mask = self.audit_data['excel_row'].astype(str) == str(excel_row)
                removed_indices.extend(self.audit_data[mask].index.tolist())
                self.audit_data = self.audit_data[~mask]

        # 保存隔离区文件
        self._save_quarantine(quarantine_list)

        # 从表格中移除选中行
        for item in selected:
            self.audit_tree.delete(item)

        # 刷新统计和筛选
        self._update_audit_stats()
        self._update_filter_options()
        self._apply_row_colors()

        self.log(f"🔒 已将 {len(selected)} 条记录移入隔离区（累计 {len(quarantine_list)} 条）", "info")
        messagebox.showinfo("完成", f"已移入隔离区 {len(selected)} 条记录。\n隔离区累计：{len(quarantine_list)} 条")


    def _open_quarantine(self):
        """打开隔离区窗口"""
        quarantine_list = self._load_quarantine()

        win = tk.Toplevel(self.root)
        win.title("📦 异常隔离区")
        win.geometry("900x500")
        win.transient(self.root)

        # ── 表格区域 ──
        tree_frame = tk.Frame(win)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(10, 5))

        # 复用审核表格的列定义
        cols = self.audit_tree['columns']
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15,
                            selectmode="extended")
        _QUAR_COL_DISPLAY = getattr(self, '_COL_DISPLAY', {})
        for col in cols:
            tree.heading(col, text=_QUAR_COL_DISPLAY.get(col, col))
            tree.column(col, width=self.audit_tree.column(col, 'width'), anchor="w")

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        tree.pack(side="left", fill="both", expand=True)

        # 填充数据
        for i, rec in enumerate(quarantine_list):
            vals = [rec.get(col, '') for col in cols]
            tag = 'row_even' if i % 2 == 0 else 'row_odd'
            tree.insert('', 'end', values=vals, tags=(tag,))
        tree.tag_configure('row_even', background='#f5f7fa')
        tree.tag_configure('row_odd', background='#ffffff')

        # ── 操作按钮 ──
        btn_frame = tk.Frame(win, bg=C['bg'])
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def restore_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要恢复的行")
                return
            restored = []
            for item in selected:
                vals = tree.item(item, 'values')
                rec = dict(zip(cols, vals))
                quarantine_list.remove(rec)
                restored.append(rec)

                # 重新插入审核表格（如果有 audit_data）
                if self.audit_data is not None:
                    new_row = pd.DataFrame([{
                        'excel_row': int(rec.get('excel_row', 0)),
                        '物料编码': rec.get('物料编码', ''),
                        '物料名称': rec.get('物料名称', ''),
                        '工厂名称': rec.get('工厂名称', ''),
                        '车间': rec.get('车间', ''),
                        '订单日期': rec.get('订单日期', ''),
                        '定额': float(rec.get('定额', 0)) if rec.get('定额') not in ('', '-', None) else 0,
                        '实际': float(rec.get('实际', 0)) if rec.get('实际') not in ('', '-', None) else 0,
                        '偏差率(%)': float(str(rec.get('偏差率(%)', '0')).rstrip('%')),
                        '偏差数量': float(rec.get('偏差数量', 0)) if rec.get('偏差数量') not in ('', '-', None) else 0,
                        '备注原因': rec.get('备注原因', ''),
                        '备注来源': rec.get('备注来源', ''),
                        '组件物料号': rec.get('组件物料号', ''),
                        '组件物料描述': rec.get('组件物料描述', ''),
                        '数量-定额': float(rec.get('数量-定额', 0)) if rec.get('数量-定额') not in ('', '-', None) else 0,
                        '数量-实际': float(rec.get('数量-实际', 0)) if rec.get('数量-实际') not in ('', '-', None) else 0,
                        '生产管理员描述': rec.get('生产管理员描述', ''),
                    }])
                    self.audit_data = pd.concat([self.audit_data, new_row], ignore_index=True)

                tree.delete(item)

            self._save_quarantine(quarantine_list)
            self._refresh_audit_tree(self.audit_data)
            self._update_audit_stats()
            self._update_filter_options()
            self.log(f"📤 已从隔离区恢复 {len(restored)} 条记录", "info")
            messagebox.showinfo("完成", f"已恢复 {len(restored)} 条记录到审核表格")

        def clear_all():
            if not quarantine_list:
                messagebox.showinfo("提示", "隔离区已空")
                return
            if messagebox.askyesno("确认", f"确定清空所有 {len(quarantine_list)} 条隔离记录？此操作不可恢复。"):
                quarantine_list.clear()
                self._save_quarantine(quarantine_list)
                for item in tree.get_children():
                    tree.delete(item)
                self.log("🗑️ 隔离区已清空", "info")

        def export_quarantine():
            if not quarantine_list:
                messagebox.showwarning("提示", "没有可导出的数据")
                return
            file_path = filedialog.asksaveasfilename(
                title="导出隔离区", defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("CSV 文件", "*.csv")]
            )
            if not file_path:
                return
            try:
                export_df = pd.DataFrame(quarantine_list)
                if file_path.endswith('.csv'):
                    export_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                else:
                    export_df.to_excel(file_path, index=False, engine='openpyxl')
                self.log(f"📤 隔离区已导出：{file_path}", "success")
                messagebox.showinfo("导出成功", f"已导出 {len(quarantine_list)} 条记录")
            except Exception as e:
                messagebox.showerror("导出失败", str(e))

        tk.Button(btn_frame, text="⬆ 恢复选中", command=restore_selected,
                  bg="#10b981", fg="white", font=("Microsoft YaHei", 10), relief="flat",
                  width=12).pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="🗑️ 清空隔离区", command=clear_all,
                  bg="#ef4444", fg="white", font=("Microsoft YaHei", 10), relief="flat",
                  width=12).pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="📤 导出", command=export_quarantine,
                  bg="#3b82f6", fg="white", font=("Microsoft YaHei", 10), relief="flat",
                  width=12).pack(side="left")


    def _update_audit_stats(self, filtered_data=None):
        """更新统计卡片（联动筛选后的数据）"""
        if self.audit_data is None or len(self.audit_data) == 0:
            return
        data = filtered_data if filtered_data is not None else self.audit_data
        total = len(data)
        high_dev = len(data[data['偏差率(%)'].abs() > 10])
        need_note = data[data['备注原因'].isna() | (data['备注原因'].astype(str).str.strip() == '')]
        ok_note = data[data['备注原因'].notna() & (data['备注原因'].astype(str).str.strip() != '')]
        # 更新四个统计卡片
        self.audit_stat_labels['total'].configure(text=str(total))
        self.audit_stat_labels['high_dev'].configure(text=str(high_dev))
        self.audit_stat_labels['need_note'].configure(text=str(len(need_note)))
        self.audit_stat_labels['ok_note'].configure(text=str(len(ok_note)))
        # 同步更新统一按钮行状态标签
        if hasattr(self, 'unified_result_lbl'):
            self.unified_result_lbl.configure(text=f"筛选结果：{total} 条 | 偏差>10%: {high_dev} | 需补备注: {len(need_note)}")


    def _show_audit_context_menu(self, event):
        """显示审核表格右键菜单"""
        # 选中点击位置的项
        item = self.audit_tree.identify_row(event.y)
        if item:
            # 如果点击的项不在已选中项中，清除其他选择并选中当前项
            if item not in self.audit_tree.selection():
                self.audit_tree.selection_set(item)
        # 显示菜单
        try:
            self.audit_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.audit_context_menu.grab_release()


    def _filter_audit_tree(self, filter_type=None):
        if self.audit_data is None or len(self.audit_data) == 0:
            return
        for key in self.audit_stat_cards:
            self.audit_stat_cards[key].configure(bg=C['surface2'])
            for child in self.audit_stat_cards[key].winfo_children():
                child.configure(bg=C['surface2'])
        if filter_type:
            self.audit_stat_cards[filter_type].configure(bg='#e8f0fe')
            for child in self.audit_stat_cards[filter_type].winfo_children():
                child.configure(bg='#e8f0fe')
        for row in self.audit_tree.get_children():
            self.audit_tree.delete(row)
        if filter_type == 'need_note':
            filtered_data = self.audit_data[self.audit_data['备注原因'].isna()]
        elif filter_type == 'ok_note':
            filtered_data = self.audit_data[self.audit_data['备注原因'].notna()]
        # ── P1#13：颜色筛选 ──
        elif filter_type == '_color':
            color_val = self.filter_vars.get('_color', tk.StringVar(value='全部')).get()
            if color_val and color_val != '全部':
                color_map = {'🔴 红': '红', '🟠 橙': '橙', '🟡 黄': '黄', '🟢 绿': '绿'}
                target_color = color_map.get(color_val, color_val)
                filtered_data = self.audit_data[self.audit_data['_priority_label'] == target_color]
            else:
                filtered_data = self.audit_data
        else:
            filtered_data = self.audit_data
        # 联动更新顶部统计卡片
        self._update_audit_stats(filtered_data)
        for i, (_, row) in enumerate(filtered_data.iterrows(), 1):
            dev_rate = row.get('偏差率', row.get('偏差率(%)', 0))
            has_note = pd.notna(row.get('备注', row.get('备注原因', ''))) and str(row.get('备注', row.get('备注原因', ''))).strip() != ''
            status = "已备注" if has_note else "需补备注"
            remark = str(row.get('备注', row.get('备注原因', '')))[:30] if pd.notna(row.get('备注', row.get('备注原因', ''))) else ''
            batch_remark = str(row.get('批量备注', row.get('批量备注原因', '')))[:30] if pd.notna(row.get('批量备注', row.get('批量备注原因', ''))) else ''
            is_alt = '是' if row.get('备注来源') == '替代料' else '否'
            # audit_status 和 audit_source
            audit_result_val = str(row.get('audit_result', ''))[:30]
            audit_status_val = '已审核' if audit_result_val and audit_result_val.strip() not in ('', 'nan') else '未审核'
            audit_source_val = str(row.get('审核来源', ''))
            if audit_source_val in ('nan', 'NaN', 'None', ''):
                # 尝试从备注来源推断
                note_source = str(row.get('备注来源', ''))
                if note_source in ('AI审核合格', 'AI审核待改进', 'AI生成'):
                    audit_source_val = 'AI'
                elif note_source == '人工填写':
                    audit_source_val = '手动'
                else:
                    audit_source_val = '系统'   # 默认来源
            # 流程订单
            order_no_val = ''
            for _col in ['流程订单', '订单号', '订单编号']:
                if _col in row.index and pd.notna(row.get(_col)):
                    order_no_val = str(row.get(_col))
                    break
            # 灵活列名
            factory_val = str(row.get('工厂', row.get('工厂名称', '')))
            admin_val = str(row.get('车间', row.get('生产管理员描述', '')))
            code_val = str(row.get('物料编码', row.get('组件物料号', '')))
            name_val = str(row.get('物料名称', row.get('组件物料描述', '')))[:20]
            quota_val = f"{row.get('定额', row.get('数量-定额', 0)):.3f}"
            actual_val = f"{row.get('实际', row.get('数量-实际', 0)):.3f}"
            dev_rate_str = f"{dev_rate:.2f}%" if isinstance(dev_rate, (int, float)) else str(dev_rate)
            item = self.audit_tree.insert('', 'end', values=(
                i,  # idx
                int(row.get('原表行号', row.get('excel_row', 0))) if pd.notna(row.get('原表行号', row.get('excel_row'))) else '',  # excel_row
                factory_val,  # factory
                admin_val,  # admin
                str(row.get('订单日期', ''))[:12] if pd.notna(row.get('订单日期')) else '',  # order_date
                order_no_val,  # order_no
                code_val,  # code
                name_val,  # name
                quota_val,  # quota
                actual_val,  # actual
                dev_rate_str,  # dev_rate
                is_alt,  # is_alt
                status,  # status
                remark,  # remark
                batch_remark,  # batch_remark
                audit_result_val,  # audit_result
                str(row.get('AI建议', ''))[:50],  # AI建议
                audit_status_val,  # audit_status
                audit_source_val,  # audit_source
                f"{row.get('偏差金额', 0):,.2f}",  # deviation_amount
            ))
            priority_label = row.get('_priority_label', '绿')
            priority_tag = 'priority_red' if priority_label == '红' else ('priority_yellow' if priority_label == '黄' else 'priority_green')
            self.audit_tree.item(item, tags=(priority_tag,))
        self.log(f"筛选完成：显示 {len(filtered_data)} 条记录", "info")

    # ── 智能审核筛选栏方法 ─────────────────────────────


    def _update_filter_options(self):
        """根据当前 audit_data 更新筛选下拉框的值列表"""
        if self.audit_data is None or len(self.audit_data) == 0:
            return
        # 确保 status_temp 列存在
        if 'status_temp' not in self.audit_data.columns:
            self.audit_data['status_temp'] = self.audit_data['备注原因'].apply(
                lambda x: '已备注' if pd.notna(x) and str(x).strip() != '' else '需补备注'
            )

        col_map = {
            'factory': '工厂名称',
            'admin': '生产管理员描述',
            'name': '组件物料描述',
            'status': 'status_temp',
            'dev_rate': '偏差率(%)',
            'is_alt': None,  # 特殊处理：替代料筛选
            'remark': '备注原因',
        }

        dev_rate_presets = ['全部', '>10%', '>20%', '>30%', '绝对值>10%', '<-10%', '<-20%']
        is_alt_presets = ['全部', '是', '否']

        for key, cb in self.filter_widgets.items():
            # name 是 Entry 控件，不需要更新下拉选项
            if key == 'name':
                continue
            if key == 'dev_rate':
                cb['values'] = dev_rate_presets
                if cb.get() not in dev_rate_presets:
                    cb.set('全部')
                continue
            if key == 'remark':
                # 备注筛选：已填写的值 + "为空" 选项
                str_vals = sorted([str(v) for v in self.audit_data['备注原因'].dropna() if v != '' and pd.notna(v)])
                seen = set()
                unique_str = []
                for v in str_vals:
                    if v not in seen:
                        seen.add(v)
                        unique_str.append(v)
                cb['values'] = ['全部', '为空', '不为空'] + unique_str
                if cb.get() not in cb['values']:
                    cb.set('全部')
                continue
            if key == 'ai_result':
                src_vals = self.audit_data['备注来源'].dropna().unique()
                opts = ['全部']
                if any(v == 'AI审核合格' for v in src_vals): opts.append('合格')
                if any(v == 'AI审核待改进' for v in src_vals): opts.append('需改进')
                if any(v.startswith('AI建议') for v in src_vals): opts.append('AI建议')
                if any(v not in ['AI审核合格','AI审核待改进'] and not v.startswith('AI建议') and v != '' for v in src_vals):
                    opts.append('未处理')
                cb['values'] = opts
                if cb.get() not in opts:
                    cb.set('全部')
                continue
            if key == 'is_alt':
                # 替代料筛选：固定选项
                cb['values'] = is_alt_presets
                if cb.get() not in is_alt_presets:
                    cb.set('全部')
                continue
            df_col = col_map.get(key)
            if df_col is None or df_col not in self.audit_data.columns:
                continue
            unique_vals = self.audit_data[df_col].dropna()
            if key == 'status':
                unique_vals = [v for v in unique_vals if v != '']
            str_vals = sorted([str(v) for v in unique_vals if v != '' and pd.notna(v)])
            # 去重保持顺序
            seen = set()
            unique_str = []
            for v in str_vals:
                if v not in seen:
                    seen.add(v)
                    unique_str.append(v)
            cb['values'] = ['全部'] + unique_str
            if cb.get() not in cb['values']:
                cb.set('全部')


    def _collect_filters(self):
        """收集所有筛选控件的当前值"""
        filters = {}
        try:
            for key, widget in self.filter_widgets.items():
                if isinstance(widget, tuple):
                    # 日期范围
                    filters[key] = (widget[0].get().strip(), widget[1].get().strip())
                elif isinstance(widget, tk.Entry):
                    val = widget.get().strip()
                    if val and val != "输入任意关键词，实时过滤全部列...":
                        filters[key] = val
                else:
                    val = widget.get()
                    if val and val != '全部':
                        filters[key] = val
        except Exception:
            pass
        return filters

    def _show_precheck_report(self, df):
        """F6 预检报告弹窗"""
        if df is None or df.empty:
            return
        total = len(df)
        rate_col = None
        for col in ['偏差率(%)', '偏差率', '偏差率%', 'dev_rate', 'rate']:
            if col in df.columns:
                rate_col = col
                break
        if rate_col:
            clean_rates = pd.to_numeric(df[rate_col], errors='coerce').dropna()
            abnormal = (clean_rates.abs() >= 10).sum()
            warning = ((clean_rates.abs() >= 5) & (clean_rates.abs() < 10)).sum()
            normal = (clean_rates.abs() < 5).sum()
        else:
            abnormal = warning = normal = 0
        status_col = None
        for col in ['审核状态', 'audit_status', '状态']:
            if col in df.columns:
                status_col = col
                break
        if status_col:
            reviewed = (df[status_col] == '已审核').sum()
            un_reviewed = total - reviewed
        else:
            reviewed = un_reviewed = 0
        msg = (
            f"数据加载完成\n\n"
            f"总行数：{total}\n"
            f"偏差异常（≥10%）：{abnormal}\n"
            f"偏差关注（5%-10%）：{warning}\n"
            f"偏差正常（<5%）：{normal}\n\n"
            f"已审核：{reviewed}\n"
            f"未审核：{un_reviewed}"
        )
        messagebox.showinfo("预检报告", msg)

    def _restore_filters(self):
        """从 StateStore 恢复筛选条件到 UI 控件"""
        if not hasattr(self, 'state'):
            return
        saved = self.state.get('filters', 'all', {})
        if not saved:
            return
        try:
            for key, val in saved.items():
                if key not in self.filter_widgets:
                    continue
                widget = self.filter_widgets[key]
                if isinstance(widget, tuple) and isinstance(val, (list, tuple)) and len(val) == 2:
                    try:
                        widget[0].delete(0, 'end')
                        widget[0].insert(0, str(val[0]))
                        widget[1].delete(0, 'end')
                        widget[1].insert(0, str(val[1]))
                    except Exception:
                        pass
                elif isinstance(widget, tk.Entry):
                    try:
                        widget.delete(0, 'end')
                        widget.insert(0, str(val))
                    except Exception:
                        pass
                elif hasattr(widget, 'set'):
                    try:
                        widget.set(str(val))
                    except Exception:
                        pass
        except Exception:
            pass

    def _restore_sort_state(self):
        """从 StateStore 恢复排序状态"""
        if not hasattr(self, 'state'):
            return
        saved = self.state.get('sort', 'columns', [])
        if not saved or not isinstance(saved, list):
            return
        try:
            if not hasattr(self, 'audit_tree'):
                return
            tree_cols = list(self.audit_tree['columns'])
            valid = [(col, asc) for col, asc in saved if col in tree_cols]
            if valid:
                self.sort_columns = valid
                self._on_tree_sort.__wrapped__(self) if hasattr(self._on_tree_sort, '__wrapped__') else None
                # Fallback: just refresh display
                self._update_sort_indicators()
            if hasattr(self, 'state'):
                self.state.set('sort', 'columns', self.sort_columns, auto_save=True)
        except Exception:
            pass


    def _on_filter_changed(self, col_key):
        """任一筛选下拉框变化时
        if not hasattr(self, 'state'):
            self.state = get_state()
        if not hasattr(self, 'rule_engine'):
            self.rule_engine = RuleEngine()
，组合所有筛选条件并刷新表格"""
        if self.audit_data is None or len(self.audit_data) == 0:
            return

        # P1-1-3 万能搜索框：跨列关键词过滤
        search_text = self.search_var.get().strip()
        if search_text and search_text != "输入任意关键词，实时过滤全部列...":
            df_filtered = self.audit_data.copy()
            mask = pd.Series(False, index=df_filtered.index)
            for col in df_filtered.columns:
                mask |= df_filtered[col].astype(str).str.contains(
                    search_text, case=False, na=False)
            df_filtered = df_filtered[mask]
        else:
            df_filtered = self.audit_data.copy()

        col_map = {
            'factory': '工厂名称',
            'admin': '生产管理员描述',
            'name': '组件物料描述',
            'status': 'status_temp',
            'dev_rate': '偏差率(%)',
            'is_alt': None,  # 特殊处理
            'remark': '备注原因',
        }

        # 构建替代料名称集合（用于筛选）
        alt_all_descs = set()
        for a, b in getattr(self, 'alt_pairs', []):
            if a:
                alt_all_descs.add(a)
            if b:
                alt_all_descs.add(b)

        # ── P1：异常突变检测 ──
        self.mutation_materials = set()
        if '订单日期' in df_filtered.columns and '组件物料号' in df_filtered.columns:
            df_dates = pd.to_datetime(df_filtered['订单日期'].astype(str).str[:10], errors='coerce')
            unique_dates = sorted(df_dates.dropna().unique())
            if len(unique_dates) >= 5:
                split_idx = int(len(unique_dates) * 0.7)
                early_dates = set(unique_dates[:split_idx])
                recent_dates = set(unique_dates[split_idx:])
                df_temp = df_filtered.copy()
                df_temp['_date'] = df_dates
                df_temp['_dev_num'] = pd.to_numeric(df_temp['偏差率(%)'], errors='coerce').abs()
                for code, grp in df_temp.groupby('组件物料号'):
                    early_avg = grp[grp['_date'].isin(early_dates)]['_dev_num'].mean()
                    recent_avg = grp[grp['_date'].isin(recent_dates)]['_dev_num'].mean()
                    if pd.notna(early_avg) and pd.notna(recent_avg) and early_avg <= 5 and recent_avg >= 15:
                        self.mutation_materials.add(code)

        # 构建筛选条件
        # df_filtered 已在上方初始化
        # 先处理日期筛选（特殊处理：元组类型）
        if 'order_date' in self.filter_widgets:
            date_widgets = self.filter_widgets['order_date']
            if isinstance(date_widgets, tuple) and len(date_widgets) == 2:
                start_date = date_widgets[0].get().strip()
                end_date = date_widgets[1].get().strip()
                if start_date or end_date:
                    # 转换日期格式，支持多种输入
                    def parse_date_str(s):
                        if not s:
                            return None
                        # 尝试多种格式
                        for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d', '%Y%m%d']:
                            try:
                                return pd.to_datetime(s, format=fmt)
                            except:
                                continue
                        # 最后尝试 pandas 自动解析
                        try:
                            return pd.to_datetime(s)
                        except:
                            return None
                    
                    start_dt = parse_date_str(start_date)
                    end_dt = parse_date_str(end_date)
                    
                    # 筛选日期列（兼容字符串和Timestamp两种格式）
                    if '订单日期' in df_filtered.columns:
                        date_series = df_filtered['订单日期']
                        # 统一转为字符串比较（避免Timestamp精度问题）
                        if start_dt:
                            start_str = start_dt.strftime('%Y-%m-%d')
                            df_filtered = df_filtered[date_series.astype(str).str[:10] >= start_str]
                        if end_dt:
                            end_str = end_dt.strftime('%Y-%m-%d')
                            df_filtered = df_filtered[date_series.astype(str).str[:10] <= end_str]
                        self.log(f"日期筛选: {start_date or '..'} ~ {end_date or '..'} → {len(df_filtered)}条", "info")
        
        for key, cb in self.filter_widgets.items():
            if key == 'order_date':
                continue  # 日期已经在上面处理过了
            if isinstance(cb, tuple):
                continue  # 跳过元组类型的控件
            selected = cb.get()
            if not selected or selected == '全部':
                continue

            # ── 特殊处理的筛选键（在 df_col 检查之前）──
            if key == 'is_alt':
                # 替代料筛选：根据 alt_pairs 匹配
                if selected == '是':
                    df_filtered = df_filtered[
                        df_filtered['组件物料描述'].astype(str).str.strip().isin(alt_all_descs)
                    ]
                elif selected == '否':
                    df_filtered = df_filtered[
                        ~df_filtered['组件物料描述'].astype(str).str.strip().isin(alt_all_descs)
                    ]
                continue  # 处理完直接下一个

            # AI审核结果筛选
            if key == 'ai_result':
                if selected == '合格':
                    df_filtered = df_filtered[df_filtered['备注来源'] == 'AI审核合格']
                elif selected == '需改进':
                    df_filtered = df_filtered[df_filtered['备注来源'] == 'AI审核待改进']
                elif selected == 'AI建议':
                    df_filtered = df_filtered[df_filtered['备注来源'].str.startswith('AI建议', na=False)]
                elif selected == '未处理':
                    df_filtered = df_filtered[~df_filtered['备注来源'].isin(
                        ['AI审核合格','AI审核待改进','AI建议','AI建议（小偏差）']) |
                        (df_filtered['备注来源'].isna())]
                continue

            df_col = col_map.get(key)
            if df_col is None:
                continue

            if key == 'dev_rate':
                # 偏差率预设阈值筛选
                if selected == '>10%':
                    df_filtered = df_filtered[df_filtered['偏差率(%)'] > 10]
                elif selected == '>20%':
                    df_filtered = df_filtered[df_filtered['偏差率(%)'] > 20]
                elif selected == '>30%':
                    df_filtered = df_filtered[df_filtered['偏差率(%)'] > 30]
                elif selected == '绝对值>10%':
                    df_filtered = df_filtered[df_filtered['偏差率(%)'].abs() > 10]
                elif selected == '<-10%':
                    df_filtered = df_filtered[df_filtered['偏差率(%)'] < -10]
                elif selected == '<-20%':
                    df_filtered = df_filtered[df_filtered['偏差率(%)'] < -20]
            elif key == 'status':
                if selected == '已备注':
                    df_filtered = df_filtered[df_filtered['备注原因'].notna() &
                                             df_filtered['备注原因'].apply(lambda x: str(x).strip() != '')]
                elif selected == '需补备注':
                    df_filtered = df_filtered[df_filtered['备注原因'].isna() |
                                             df_filtered['备注原因'].apply(lambda x: str(x).strip() == '')]
            elif key == 'remark':
                # 备注筛选："为空" 只显示未填备注的记录，"不为空"只显示已填写的
                if selected == '为空':
                    df_filtered = df_filtered[df_filtered['备注原因'].isna() |
                                             df_filtered['备注原因'].apply(lambda x: str(x).strip() == '')]
                elif selected == '不为空':
                    df_filtered = df_filtered[df_filtered['备注原因'].notna() &
                                             df_filtered['备注原因'].apply(lambda x: str(x).strip() != '')]
                elif selected and selected != '全部':
                    df_filtered = df_filtered[df_filtered['备注原因'].astype(str) == selected]
            elif key == 'name':
                # 物料描述：模糊搜索（包含关键词）
                if selected and selected.strip():
                    keyword = selected.strip().lower()
                    df_filtered = df_filtered[df_filtered[df_col].astype(str).str.lower().str.contains(keyword, na=False)]
            else:
                df_filtered = df_filtered[df_filtered[df_col].astype(str) == selected]

        if hasattr(self, 'state'):
            self.state.set('filters', 'all', self._collect_filters(), auto_save=True)

        self._refresh_audit_tree(df_filtered)

        # 更新状态提示
        active_filters = []
        for key, cb in self.filter_widgets.items():
            # order_date 的 filter_widgets 是 (start_entry, end_entry) tuple
            if isinstance(cb, tuple):
                v_start = cb[0].get() if cb[0].get() else ''
                v_end = cb[1].get() if cb[1].get() else ''
                if v_start or v_end:
                    active_filters.append(f"日期: {v_start or '...'} ~ {v_end or '...'}")
                continue
            v = cb.get()
            if v and v != '全部':
                col_labels = {'factory': '工厂', 'admin': '车间', 'name': '物料', 'status': '状态', 'dev_rate': '偏差率', 'is_alt': '替代料', 'remark': '备注'}
                active_filters.append(f"{col_labels.get(key, key)}={v}")
        filter_text = f"🎯 当前筛选：{' | '.join(active_filters)} | 共 {len(df_filtered)} 条" if active_filters else f"📋 显示全部 | 共 {len(df_filtered)} 条"
        self._apply_row_colors()
        if self.filter_status_lbl:
            self.filter_status_lbl.configure(text=f"筛选: {', '.join(active_filters)}" if active_filters else "")
        if hasattr(self, 'status_filter_label'):
            self.status_filter_label.configure(text=filter_text)


    def _reset_all_filters(self):
        """重置所有筛选条件"""
        for key, widget in self.filter_widgets.items():
            if key == 'name':
                # Entry 控件：清空文本
                widget.delete(0, tk.END)
            elif key == 'order_date':
                # 日期控件：清空两个输入框
                if isinstance(widget, tuple) and len(widget) == 2:
                    widget[0].delete(0, tk.END)
                    widget[1].delete(0, tk.END)
            else:
                # Combobox 控件：重置为"全部"
                widget.set('全部')
        if self.audit_data is not None:
            self._refresh_audit_tree(self.audit_data)
        if self.filter_status_lbl:
            self.filter_status_lbl.configure(text="")
        if hasattr(self, 'status_filter_label'):
            self.status_filter_label.configure(text=f"📋 显示全部 | 共 {len(self.audit_data)} 条")
        # P1-1-4 重置万能搜索框
        self.search_var.set("")
        self.search_entry.delete(0, "end")
        self.search_entry.insert(0, "输入任意关键词，实时过滤全部列...")
        self.log("已重置所有筛选条件", "info")


    def _refresh_audit_tree(self, df, skip_auto_sort=False):
        """用给定的 DataFrame 刷新智能审核表格"""
        # 确保 audit_result 和 AI建议 列存在（在副本上操作，不影响原始数据）
        df = df.copy()
        for col in ('audit_result', 'AI建议'):
            if col not in df.columns:
                df[col] = ''
        for row in self.audit_tree.get_children():
            self.audit_tree.delete(row)
        if df is None or len(df) == 0:
            return

        # ── 性能优化：暂停绘制，批量插入后再恢复 ──
        try:
            self.audit_tree.configure(displaycolumns=[])  # 隐藏所有列以暂停绘制
        except Exception:
            pass

        # ── P1：智能优先级标记 ──
        def calc_priority(row):
            if not hasattr(self, 'rule_engine'):
                self.rule_engine = RuleEngine()
            dev = abs(float(row.get('偏差率(%)', 0) or 0))
            has_note = pd.notna(row.get('备注原因')) and str(row.get('备注原因', '')).strip() != ''
            level = self.rule_engine.get_level_for_deviation_rate(dev)
            # level: info/warning/error -> map to label
            if level == 'error':
                label = '红'
                order = 0 if not has_note else 1
            elif level == 'warning':
                label = '黄'
                order = 2 if not has_note else 3
            else:
                label = '绿'
                order = 4
            return order, label

        df = df.copy()
        df[['_priority_order', '_priority_label']] = df.apply(
            lambda r: pd.Series(calc_priority(r)), axis=1)
        if not skip_auto_sort:
            df = df.sort_values('_priority_order')

        # ── P1：金额排名着色 ──
        amount_col = None
        if '偏差金额' in df.columns:
            amount_col = '偏差金额'
        elif '偏差金额(含税)' in df.columns:
            amount_col = '偏差金额(含税)'
        else:
            amount_col = '_dev_qty_abs'
            df = df.copy()
            df['_dev_qty_abs'] = pd.to_numeric(df['偏差数量'], errors='coerce').abs()

        rank_dict = {}
        if '物料分类' in df.columns and amount_col:
            for cat, grp in df.groupby('物料分类'):
                grp_sorted = grp.sort_values(amount_col, ascending=False, key=abs)
                top3 = grp_sorted.head(3).index.tolist()
                next7 = grp_sorted.iloc[3:10].index.tolist() if len(grp_sorted) > 3 else []
                for idx in top3:
                    rank_dict[idx] = 'amt_rank_1'
                for idx in next7:
                    rank_dict[idx] = 'amt_rank_2'

        # 构建替代料名称集合
        alt_all_descs = set()
        for a, b in getattr(self, 'alt_pairs', []):
            if a:
                alt_all_descs.add(a)
            if b:
                alt_all_descs.add(b)

        # ── 填充 Tree（按优先级排序） ──
        from widgets import C as _C
        for i, (_, row) in enumerate(df.iterrows(), 1):
            dev_rate = row.get('偏差率(%)', 0) or 0
            # 原备注
            orig_remark = str(row.get('备注原因', '')) if pd.notna(row.get('备注原因')) else ''
            if orig_remark in ('nan', 'NaN', 'None'):
                orig_remark = ''
            # 批量备注
            batch_remark = str(row.get('批量备注原因', '')) if pd.notna(row.get('批量备注原因')) else ''
            if batch_remark in ('nan', 'NaN', 'None'):
                batch_remark = ''
            # 备注来源
            note_src = str(row.get('备注来源', ''))
            if note_src in ('nan', 'NaN', 'None'):
                note_src = ''
            has_note = (orig_remark.strip() != '' or batch_remark.strip() != '')
            # 状态（简单状态，不含审核结论）
            if note_src == 'AI生成':
                dev_dir = float(dev_rate or 0)
                status = f'AI生成{"↑" if dev_dir > 0 else "↓"}'
            elif note_src in ('人工填写', '系统无定额(广宣)', '自动填充', '替代料'):
                status = "已备注"
            else:
                # 包括AI审核合格/待改进/建议等，统一按备注内容显示状态
                status = "已备注" if has_note else "需补备注"
            remark = orig_remark[:30]
            mat_desc = str(row.get('组件物料描述', row.get('物料名称', ''))).strip()
            is_alt = "是" if mat_desc and mat_desc in alt_all_descs else ""
            # audit_status, audit_source, order_no
            audit_result_val = str(row.get('audit_result', ''))
            audit_status_val = '已审核' if audit_result_val and audit_result_val.strip() not in ('', 'nan') else '未审核'
            audit_source_val = str(row.get('审核来源', ''))
            if audit_source_val in ('nan', 'NaN', 'None', ''):
                # 尝试从备注来源推断
                note_source = str(row.get('备注来源', ''))
                if note_source in ('AI审核合格', 'AI审核待改进', 'AI生成'):
                    audit_source_val = 'AI'
                elif note_source == '人工填写':
                    audit_source_val = '手动'
                else:
                    audit_source_val = '系统'   # 默认来源
            order_no_val = ''
            for _col in ['流程订单', '订单号', '订单编号']:
                if _col in row.index and pd.notna(row.get(_col)):
                    order_no_val = str(row.get(_col))
                    break
            item = self.audit_tree.insert('', 'end', values=(
                i,
                int(row.get('原表行号', row.get('excel_row', 0))) if pd.notna(row.get('原表行号', row.get('excel_row'))) else '',
                str(row.get('工厂', row.get('工厂名称', ''))),
                str(row.get('车间', row.get('生产管理员描述', ''))),
                str(row.get('订单日期', ''))[:12] if pd.notna(row.get('订单日期')) else '',
                order_no_val,
                str(row.get('物料编码', row.get('组件物料号', ''))),
                mat_desc[:20],
                f"{row.get('定额', row.get('数量-定额', 0)):.3f}",
                f"{row.get('实际', row.get('数量-实际', 0)):.3f}",
                f"{dev_rate:.2f}%",
                is_alt,
                status,
                remark,
                batch_remark[:30],
                audit_result_val,
                str(row.get('AI建议', '')),
                audit_status_val,
                audit_source_val,
                f"{row.get('偏差金额', 0):,.2f}",
            ))
            # RuleEngine 偏差率颜色
            color_hex = None
            try:
                if hasattr(self, 'rule_engine'):
                    color_hex = self.rule_engine.get_color_for_deviation_rate(dev_rate)
                    if color_hex not in self.audit_tree.tag_names():
                        self.audit_tree.tag_configure(color_hex, background=color_hex)
                else:
                    color_hex = None
            except Exception:
                color_hex = None
            # 颜色标签
            if note_src == '自动结案':
                tag = ('auto_closed',)
            elif note_src == 'AI生成':
                tag = ('ai_gen',)
            elif note_src in ('AI审核合格', '人工填写', '系统无定额(广宣)', '自动填充', '替代料'):
                tag = ('ok_note',)
            elif note_src == 'AI审核待改进':
                tag = ('need_note',)
            elif not has_note:
                tag = ('need_note',)
            else:
                tag = ('ok_note',)
            mat_code = str(row.get('组件物料号', ''))
            if hasattr(self, 'mutation_materials') and mat_code in self.mutation_materials:
                tag = ('mutation_alert',) + (tag if isinstance(tag, tuple) else (tag,))
            rank_tag = rank_dict.get(_, None)  # _ 是当前行的原始索引
            if rank_tag:
                tag = (rank_tag,) + (tag if isinstance(tag, tuple) else (tag,))
            # 金额颜色区分：正偏差红色，负偏差绿色
            if dev_rate > 0:
                tag = tag + ('over_amount',)
            elif dev_rate < 0:
                tag = tag + ('under_amount',)
            final_tag = (color_hex,) + tag if color_hex and isinstance(color_hex, str) else tag
            self.audit_tree.item(item, tags=final_tag)

        # ── 恢复绘制 ──
        try:
            self.audit_tree.configure(displaycolumns=[c for c in self.audit_tree['columns']])
            self.root.update_idletasks()
        except Exception:
            pass




    # ==================== 隔离区辅助方法 ====================
    # ==================== 列显示名映射 ====================
    _COL_DISPLAY = {
        "idx": "序号", "excel_row": "原表行号", "factory": "工厂名称",
        "admin": "生产管理员", "order_date": "订单日期", "order_no": "流程订单",
        "code": "物料号", "name": "物料描述", "quota": "定额", "actual": "实际",
        "dev_rate": "偏差率%", "deviation_amount": "偏差金额",
        "is_alt": "替代料", "status": "状态", "remark": "备注",
        "batch_remark": "批量备注", "audit_result": "审核结果",
        "AI建议": "AI建议", "audit_status": "审核状态", "audit_source": "审核来源",
    }

    def _get_quarantine_path(self):
        d = os.path.join(os.path.expanduser('~'), '.zpp011_audit')
        os.makedirs(d, exist_ok=True)
        return os.path.join(d, 'quarantine.json')

    def _load_quarantine(self):
        p = self._get_quarantine_path()
        if not os.path.exists(p):
            return []
        try:
            with open(p, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return []

    def _save_quarantine(self, data):
        with open(self._get_quarantine_path(), 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # ==================== 隔离区窗口 ====================
    def _open_quarantine(self):
        quarantine_list = self._load_quarantine()
        win = tk.Toplevel(self.root)
        win.title("异常隔离区")
        win.geometry("900x500")
        win.transient(self.root)
        tree_frame = tk.Frame(win)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(10, 5))
        cols = self.audit_tree['columns']
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15, selectmode="extended")
        _QUAR_COL_DISPLAY = getattr(self, '_COL_DISPLAY', {})
        for col in cols:
            tree.heading(col, text=_QUAR_COL_DISPLAY.get(col, col))
            tree.column(col, width=self.audit_tree.column(col, 'width'), anchor="w")
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        tree.pack(side="left", fill="both", expand=True)
        for i, rec in enumerate(quarantine_list):
            vals = [rec.get(col, '') for col in cols]
            tag = 'row_even' if i % 2 == 0 else 'row_odd'
            tree.insert('', 'end', values=vals, tags=(tag,))
        tree.tag_configure('row_even', background='#f5f7fa')
        tree.tag_configure('row_odd', background='#ffffff')
        btn_frame = tk.Frame(win, bg=C['bg'])
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def restore_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要恢复的行")
                return
            restored = []
            for item in selected:
                vals = tree.item(item, 'values')
                rec = dict(zip(cols, vals))
                if rec in quarantine_list:
                    quarantine_list.remove(rec)
                restored.append(rec)
                if self.audit_data is not None:
                    try:
                        new_row = pd.DataFrame([{
                            'excel_row': int(rec.get('excel_row', 0)),
                            '物料编码': rec.get('物料编码', ''),
                            '物料名称': rec.get('物料名称', ''),
                            '工厂名称': rec.get('工厂名称', ''),
                            '车间': rec.get('车间', ''),
                            '订单日期': rec.get('订单日期', ''),
                            '定额': float(rec.get('定额', 0)) if rec.get('定额') not in ('', '-', None) else 0,
                            '实际': float(rec.get('实际', 0)) if rec.get('实际') not in ('', '-', None) else 0,
                            '偏差率(%)': float(str(rec.get('偏差率(%)', '0')).rstrip('%')),
                            '偏差数量': float(rec.get('偏差数量', 0)) if rec.get('偏差数量') not in ('', '-', None) else 0,
                            '备注原因': rec.get('备注原因', ''),
                            '备注来源': rec.get('备注来源', ''),
                            '组件物料号': rec.get('组件物料号', ''),
                            '组件物料描述': rec.get('组件物料描述', ''),
                            '数量-定额': float(rec.get('数量-定额', 0)) if rec.get('数量-定额') not in ('', '-', None) else 0,
                            '数量-实际': float(rec.get('数量-实际', 0)) if rec.get('数量-实际') not in ('', '-', None) else 0,
                            '生产管理员描述': rec.get('生产管理员描述', ''),
                        }])
                        self.audit_data = pd.concat([self.audit_data, new_row], ignore_index=True)
                    except Exception:
                        pass
                tree.delete(item)
            self._save_quarantine(quarantine_list)
            self._refresh_audit_tree(self.audit_data)
            self._update_audit_stats()
            self._update_filter_options()
            self.log(f"已从隔离区恢复 {len(restored)} 条记录", "info")
            messagebox.showinfo("完成", f"已恢复 {len(restored)} 条记录到审核表格")

        def clear_all():
            if not quarantine_list:
                messagebox.showinfo("提示", "隔离区已空")
                return
            if messagebox.askyesno("确认", f"确定清空所有 {len(quarantine_list)} 条隔离记录？"):
                quarantine_list.clear()
                self._save_quarantine(quarantine_list)
                for item in tree.get_children():
                    tree.delete(item)
                self.log("隔离区已清空", "info")

        def export_quarantine():
            if not quarantine_list:
                messagebox.showwarning("提示", "没有可导出的数据")
                return
            file_path = filedialog.asksaveasfilename(
                title="导出隔离区", defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("CSV 文件", "*.csv")]
            )
            if not file_path:
                return
            try:
                export_df = pd.DataFrame(quarantine_list)
                if file_path.endswith('.csv'):
                    export_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                else:
                    export_df.to_excel(file_path, index=False, engine='openpyxl')
                self.log(f"隔离区已导出：{file_path}", "success")
                messagebox.showinfo("导出成功", f"已导出 {len(quarantine_list)} 条记录")
            except Exception as e:
                messagebox.showerror("导出失败", str(e))

        tk.Button(btn_frame, text="恢复选中", command=restore_selected,
                  bg="#10b981", fg="white", font=("Microsoft YaHei", 10), relief="flat",
                  width=12).pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="清空隔离区", command=clear_all,
                  bg="#ef4444", fg="white", font=("Microsoft YaHei", 10), relief="flat",
                  width=12).pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="导出", command=export_quarantine,
                  bg="#3b82f6", fg="white", font=("Microsoft YaHei", 10), relief="flat",
                  width=12).pack(side="left")

    # ==================== 防抖搜索方法 ====================
    def _on_search_delayed(self, event):
        if hasattr(self, '_search_timer'):
            self.root.after_cancel(self._search_timer)
        self._search_timer = self.root.after(300, lambda: self._on_filter_changed('search'))

    # ==================== 预检报告弹窗 ====================
    def _run_pre_check(self):
        """完整预检报告：列完整性 + 重复订单 + 数值异常 + 替代料配置 + 弹窗"""
        try:
            input_path = self.input_file.get()
            if not input_path or not os.path.exists(input_path):
                self.log("预检失败：输入文件不存在", "error")
                messagebox.showwarning("预检失败", "请先选择输入文件")
                return

            df = pd.read_excel(input_path, sheet_name='Data', nrows=1000)
            results = []

            # 1. 列完整性检查（黄金模板）
            golden_cols = [
                '流程订单', '订单开始日期', '组件物料号', '组件物料描述',
                '组件数量', '单位', '工厂名称', '生产管理员描述'
            ]
            missing_cols = [col for col in golden_cols if col not in df.columns]
            if missing_cols:
                results.append(('严重', f'缺失标准列：{", ".join(missing_cols)}'))
            else:
                results.append(('通过', '黄金模板列完整'))

            # 2. 重复订单检查（日期 + 流程订单 + 物料编码 联合去重）
            date_col = None
            for col in ['订单开始日期', '订单日期', '日期']:
                if col in df.columns:
                    date_col = col
                    break
            # 确定物料编码列名
            mat_col = None
            for col in ['物料编码', '组件物料号', '物料号', '编码']:
                if col in df.columns:
                    mat_col = col
                    break
            if date_col and '流程订单' in df.columns and mat_col:
                df['_check_date'] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y-%m-%d')
                dup_mask = df.duplicated(subset=['_check_date', '流程订单', mat_col], keep=False)
                dup_orders = df[dup_mask]
                if not dup_orders.empty:
                    dup_groups = dup_orders.groupby(['_check_date', '流程订单', mat_col]).ngroups
                    results.append(('警告', f'发现 {len(dup_orders)} 条重复记录（{dup_groups} 组重复），请检查 SAP 导出是否重复'))
                else:
                    results.append(('通过', '无重复记录（日期+订单+物料）'))
                df.drop(columns=['_check_date'], inplace=True)
            else:
                missing = []
                if not date_col: missing.append('日期列')
                if '流程订单' not in df.columns: missing.append('流程订单')
                if not mat_col: missing.append('物料编码/组件物料号')
                results.append(('警告', f'缺少必要列（{", ".join(missing)}），无法检测重复订单'))

            # 3. 数值异常检查
            if '组件数量' in df.columns:
                neg_quota = df[df['组件数量'] < 0]
                if not neg_quota.empty:
                    results.append(('警告', f'发现 {len(neg_quota)} 行定额为负数'))
                else:
                    results.append(('通过', '定额无负数'))

            # 4. 替代料配置检查
            try:
                from domain.alt_material.alt_manager import _get_config_path, load_alt_pairs
                alt_path = _get_config_path()
                if not os.path.exists(alt_path):
                    results.append(('警告', '替代料配置文件不存在，将使用内置配对'))
                else:
                    alt_pairs = load_alt_pairs(log_cb=self.log)
                    results.append(('通过', f'替代料配置加载成功，共 {len(alt_pairs)} 组配对'))
            except Exception as e:
                results.append(('严重', f'替代料配置加载失败：{e}'))

            # 5. 黄金模板对比
            try:
                gold_cols = self._load_golden_columns()
                if gold_cols:
                    actual_cols = set(df.columns)
                    template_cols = set(gold_cols)
                    missing = template_cols - actual_cols
                    extra = actual_cols - template_cols
                    if missing:
                        results.append(('严重', f'黄金模板缺失列：{", ".join(sorted(missing))}'))
                    if extra:
                        results.append(('警告', f'黄金模板多出列：{", ".join(sorted(extra))}'))
                    if not missing and not extra:
                        results.append(('通过', '黄金模板列结构完全匹配'))
                else:
                    results.append(('警告', '尚未设置黄金模板，跳过列结构对比'))
            except Exception as e:
                results.append(('警告', f'黄金模板对比失败：{e}'))

            # 汇总到日志
            self.log("📋 数据预检报告：", "info")
            for severity, msg in results:
                self.log(f" [{severity}] {msg}", severity if severity in ("通过", "警告", "严重") else "info")

            # 弹窗显示
            self._show_pre_check_report(results)

        except Exception as e:
            self.log(f"预检失败：{e}", "error")
            messagebox.showerror("预检错误", f"预检过程中发生错误：{str(e)}")

    def _load_golden_columns(self):
        """加载黄金模板列名（从配置文件，若不存在返回None）"""
        try:
            gold_path = os.path.join(os.path.expanduser('~'), '.zpp011_audit', 'golden_columns.json')
            if not os.path.exists(gold_path):
                return None
            with open(gold_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return None

    def _show_pre_check_report(self, results):
        win = tk.Toplevel(self.root)
        win.title("数据预检报告")
        win.geometry("600x500")
        win.transient(self.root)
        win.grab_set()
        win.update_idletasks()
        x = (win.winfo_screenwidth() // 2) - (600 // 2)
        y = (win.winfo_screenheight() // 2) - (500 // 2)
        win.geometry(f"+{x}+{y}")
        text = tk.Text(win, wrap="word", font=("Consolas", 10), padx=10, pady=10)
        text.pack(fill="both", expand=True)
        text.tag_configure("严重", foreground="#cf222e")
        text.tag_configure("警告", foreground="#d29922")
        text.tag_configure("通过", foreground="#1a7f37")
        for level, msg in results:
            text.insert("end", msg + "\n", level)
        text.configure(state="disabled")
        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)
        if hasattr(self, '_duplicate_records') and hasattr(self._duplicate_records, 'empty') and not self._duplicate_records.empty:
            tk.Button(btn_frame, text="导出重复数据", command=self._export_duplicate_records,
                      bg="#f0f0f0", font=("Microsoft YaHei", 9)).pack(side="left", padx=5)
        tk.Button(btn_frame, text="关闭", command=win.destroy,
                  bg="#f0f0f0", font=("Microsoft YaHei", 9)).pack(side="left", padx=5)

    # ==================== 导出重复数据 ====================
    def _export_duplicate_records(self):
        if not hasattr(self, '_duplicate_records') or not hasattr(self._duplicate_records, 'empty') or self._duplicate_records.empty:
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="重复订单记录.xlsx"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
            df = self._duplicate_records.copy()
            if '流程订单' in df.columns and '物料编码' in df.columns:
                df = df.sort_values(['流程订单', '物料编码'])
            wb = Workbook()
            ws = wb.active
            ws.title = "重复记录"
            headers = list(df.columns)
            for col_idx, h in enumerate(headers, 1):
                ws.cell(1, col_idx, h)
            colors = ["FFCCCC", "FFE5CC", "FFFFCC", "CCFFCC", "CCE5FF", "E5CCFF", "FFCCE5", "E5E5E5"]
            group_colors = {}
            group_idx = 0
            last_group = None
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                if '流程订单' in df.columns and '物料编码' in df.columns:
                    group_key = (row['流程订单'], row['物料编码'])
                    if group_key != last_group:
                        group_idx += 1
                        last_group = group_key
                        group_colors[group_key] = colors[(group_idx - 1) % len(colors)]
                    fill_color = group_colors.get(group_key, "FFFFFF")
                else:
                    fill_color = "FFFFFF"
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row_idx, col_idx, val)
                    if fill_color != "FFFFFF":
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            wb.save(file_path)
            self.log(f"重复数据已导出：{file_path}", "success")
            messagebox.showinfo("导出成功", "已导出到：" + file_path)
        except Exception as e:
            self.log(f"导出重复数据失败：{e}", "error")
            messagebox.showerror("导出失败", str(e))

    # ==================== 菜单初始化 ====================
    def _init_menu(self):
        try:
            menubar = tk.Menu(self.root)
            self.root.config(menu=menubar)
            help_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="帮助", menu=help_menu)
            help_menu.add_command(label="查看历史源码", command=self._open_source_backup)
            help_menu.add_separator()
            help_menu.add_command(label="关于", command=self._show_about)
        except Exception as e:
            print(f"[WARN] 菜单栏初始化失败: {e}")

    # ==================== 历史源码查看 ====================
    def _open_source_backup(self):
        backup_dir = os.path.join(os.path.expanduser('~'), '.zpp011_audit', 'source_backups')
        if os.path.exists(backup_dir):
            os.startfile(backup_dir)
        else:
            messagebox.showinfo("提示", "还没有任何源码备份，请先执行打包操作")

    # ── 多列联动排序 ─────────────────────────────────────────────
    def _init_sort_columns(self):
        for col_id in self.audit_tree['columns']:
            self.audit_tree.heading(col_id, command=lambda cid=col_id: self._on_tree_sort(cid))
        self.sort_columns = []

    def _on_tree_sort(self, col_id):
        remaining = [(cid, asc) for cid, asc in self.sort_columns if cid != col_id]
        if col_id in [cid for cid, _ in self.sort_columns]:
            old_asc = next(asc for cid, asc in self.sort_columns if cid == col_id)
            self.sort_columns = [(col_id, not old_asc)] + remaining
        else:
            # 新列追加到末尾
            self.sort_columns = self.sort_columns + [(col_id, True)]
        self._apply_sort_and_refresh()

    def _apply_sort_and_refresh(self):
        if self.audit_data is None or self.audit_data.empty:
            return
        if not self.sort_columns:
            self._refresh_audit_tree(self.audit_data)
            return
        valid = []
        for col_id, asc in self.sort_columns:
            df_col = self._COL_TO_DF.get(col_id)
            if df_col and df_col in self.audit_data.columns:
                valid.append((df_col, asc))
        if not valid:
            self._refresh_audit_tree(self.audit_data)
            return
        by = [col for col, _ in valid]
        asc_list = [asc for _, asc in valid]
        df_sorted = self.audit_data.copy()
        for col in by:
            if col in ('偏差率(%)', '偏差金额', '数量-定额', '数量-实际'):
                df_sorted[col] = pd.to_numeric(df_sorted[col], errors='coerce').fillna(0)
        df_sorted = df_sorted.sort_values(by=by, ascending=asc_list, na_position='last')
        self._refresh_audit_tree(df_sorted, skip_auto_sort=True)
        self._update_sort_indicators()

    def _update_sort_indicators(self):
        if not hasattr(self, 'audit_tree'):
            return
        for col_id in self.audit_tree['columns']:
            base = self._COL_DISPLAY.get(col_id, col_id)
            self.audit_tree.heading(col_id, text=base)
        for idx, (col_id, asc) in enumerate(self.sort_columns, start=1):
            base = self._COL_DISPLAY.get(col_id, col_id)
            arrow = '↑' if asc else '↓'
            new_text = base + ' ' + arrow if idx == 1 else f'{base} [{idx}]{arrow}'
            self.audit_tree.heading(col_id, text=new_text)



# ── S01 库存检查方法 ──────────────────────────────
    def _s01_start_inventory_check(self, data: pd.DataFrame) -> None:
        """启动库存流程检查（异步），线程创建失败时释放锁"""
        if self._is_s01_processing:
            messagebox.showwarning("提示", "已有库存检查任务进行中")
            return
        if data is None or data.empty:
            messagebox.showwarning("警告", "无数据可检查")
            return

        self._is_s01_processing = True
        try:
            self._s01_disable_ui()
            self.progress_bar.configure(mode='determinate', maximum=100)
            self.progress_bar['value'] = 0
            self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)
            self.set_status("正在执行库存检查...")

            data_copy = data.copy(deep=True)
            self._s01_cancel_flag = threading.Event()

            self._s01_thread = threading.Thread(
                target=self._s01_inventory_worker,
                args=(data_copy, self._s01_cancel_flag)
            )
            self._s01_thread.daemon = True
            self._s01_thread.start()
        except Exception as e:
            self._s01_cleanup()
            messagebox.showerror("错误", f"任务启动失败: {e}")

    def _s01_inventory_worker(self, df, cancel_flag):
        """异步任务：纯数据处理，使用 itertuples，每 50 行检查取消并回调进度"""
        total = len(df)
        status_col = None
        for col in ['库存状态', 'inventory_status']:
            if col in df.columns:
                status_col = col
                break
        if status_col is None:
            df['库存状态'] = ''
            status_col = '库存状态'

        try:
            for idx, row in enumerate(df.itertuples(index=False)):
                if cancel_flag.is_set():
                    self._s01_clean_temp_files()
                    raise InterruptedError("操作已取消")
                row_dict = row._asdict()
                # 业务逻辑待补充
                if (idx + 1) % 50 == 0 or idx + 1 == total:
                    self.root.after(0, self._s01_progress_callback, idx + 1, total)
            self.root.after(0, self._s01_on_success, df)
        except InterruptedError as e:
            self.root.after(0, self._s01_on_error, e)
        except Exception as e:
            self.root.after(0, self._s01_on_error, e)

    def _s01_progress_callback(self, current: int, total: int) -> None:
        """进度回调"""
        percent = int(current / total * 100) if total else 0
        self.progress_bar['value'] = percent
        self.set_status(f"库存检查中: {current}/{total} ({percent}%)")

    def _s01_on_success(self, result_df: pd.DataFrame) -> None:
        """成功回调：调用高亮版 _s01_populate_table"""
        self.audit_data = result_df
        self._s01_populate_table(result_df)
        self._s01_cleanup()
        messagebox.showinfo("完成", "库存检查完成")

    def _s01_on_error(self, error: Exception) -> None:
        """错误/取消回调"""
        if isinstance(error, InterruptedError):
            self.set_status("操作已取消")
            messagebox.showwarning("已���消", str(error))
        else:
            self.set_status("操作失败")
            messagebox.showerror("错误", f"库存检查失败: {error}")
        self._s01_cleanup()

    def _s01_cleanup(self) -> None:
        """统一清理"""
        self._is_s01_processing = False
        self._s01_cancel_flag = None
        self._s01_thread = None
        self._s01_enable_ui()
        self.progress_bar['value'] = 0
        self.progress_bar.pack_forget()
        self.set_status("就绪")
        self._s01_clean_temp_files()

    def _s01_cancel_inventory_check(self) -> None:
        if self._s01_cancel_flag:
            self._s01_cancel_flag.set()

    def _s01_disable_ui(self) -> None:
        """禁用相关按钮"""
        pass

    def _s01_enable_ui(self) -> None:
        """恢复按钮"""
        pass

    # ── S01 异常高亮方法 ─────────────────────────────────────────────────
    def _evaluate_condition(self, row: dict, condition_str: str) -> bool:
        """安全评估条件表达式（仅用于 S01 高亮）"""
        dangerous = ['__', 'import', 'exec', 'eval', 'compile', 'open', 'file']
        if any(d in condition_str for d in dangerous):
            return False
        allowed = {k: v for k, v in row.items() if isinstance(v, (int, float, str, bool))}
        try:
            return bool(eval(condition_str, {"__builtins__": {}}, allowed))
        except Exception:
            return False

    def _s01_setup_treeview_tags(self, tree):
        """为 Treeview 配置高亮 tag（仅首次）"""
        cache_key = f"_s01_tags_configured_{id(tree)}"
        if getattr(self, cache_key, False):
            return
        config = getattr(self, 's01_display_config', {})
        for rule in config.get('rules', []):
            tag = rule.get('tag')
            color = rule.get('color')
            if tag and color:
                try:
                    tree.tag_configure(tag, background=color)
                except Exception:
                    pass
        setattr(self, cache_key, True)

    def _s01_populate_table(self, df):
        """将 DataFrame 填充到 audit_tree，并根据 s01_display_config 应用高亮"""
        tree = getattr(self, 'audit_tree', None)
        if tree is None:
            self._refresh_audit_tree(df)
            return

        self._s01_setup_treeview_tags(tree)
        config = getattr(self, 's01_display_config', {})
        rules = config.get('rules', [])
        default_color = config.get('default_color', '#FFFFFF')
        try:
            tree.tag_configure('s01_normal', background=default_color)
        except Exception:
            pass

        # 清空旧数据
        for item in tree.get_children():
            tree.delete(item)

        cols = list(tree['columns']) if tree['columns'] else (list(df.columns) if not df.empty else [])

        for _, row in df.iterrows():
            row_dict = row.to_dict()
            tag = 's01_normal'
            for rule in rules:
                condition = rule.get('condition', '')
                if condition and self._evaluate_condition(row_dict, condition):
                    tag = rule.get('tag', 's01_normal')
                    break
            values = [row_dict.get(c, '') for c in cols]
            tree.insert('', 'end', values=values, tags=(tag,))
