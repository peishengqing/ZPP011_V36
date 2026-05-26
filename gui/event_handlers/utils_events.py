# -*- coding: utf-8 -*-
"""鏁呬簨绾裤€佹浛浠ｆ枡銆丅OM銆侀妫€銆佹爲瑙嗗浘銆佹柇鐐圭瓑鏉傞」浜嬩欢"""

import shutil
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from widgets import C
import os, sys as _sys, glob as _glob
import pandas as pd
from storage import storage
from core.state_store import get_state
from core.rule_engine import RuleEngine
from core.task_manager import TaskManager
from modules.audit.models.audit_model import AuditModel
from openpyxl import Workbook, load_workbook
from copy import deepcopy
from domain.alt_material.alt_manager import save_alt_pairs, load_alt_pairs
import time, datetime, threading, traceback, json, csv, calendar


class UtilsEvents:
    """鏁呬簨绾裤€佹浛浠ｆ枡銆丅OM銆侀妫€銆佹爲瑙嗗浘銆佹柇鐐圭瓑鏉傞」浜嬩欢"""
    def _show_storyline(self):
        """寮瑰嚭鏁呬簨绾跨獥鍙ｏ紙鐩存帴鍖呭惈缁熻閫昏緫锛?""
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showinfo("鎻愮ず", "鏃犳暟鎹敓鎴愭晠浜嬬嚎")
            return

        lines = []
        lines.append(f"馃搮 缁熻鍛ㄦ湡锛歿self.start_date.get()} 鑷?{self.end_date.get()}")
        lines.append("")

        if '鍋忓樊閲戦' in self.audit_data.columns:
            total_amount = self.audit_data['鍋忓樊閲戦'].sum()
            lines.append(f"馃挵 鎬诲亸宸噾棰濓細楼{total_amount:,.2f}")
        else:
            lines.append("馃挵 鎬诲亸宸噾棰濓細鏃犳暟鎹?)

        over_col = next((c for c in self.audit_data.columns if '澶氳€? in c or '瓒呰€? in c), None)
        under_col = next((c for c in self.audit_data.columns if '灏戣€? in c or '鑺傜害' in c), None)
        if over_col and under_col:
            lines.append(f"馃搱 澶氳€楁€婚锛毬self.audit_data[over_col].sum():,.2f}")
            lines.append(f"馃搲 灏戣€楁€婚锛毬self.audit_data[under_col].sum():,.2f}")

        if '鍋忓樊鐜?%)' in self.audit_data.columns:
            high = (self.audit_data['鍋忓樊鐜?%)'].abs() > 10).sum()
            medium = ((self.audit_data['鍋忓樊鐜?%)'].abs() >= 5) & (self.audit_data['鍋忓樊鐜?%)'].abs() <= 10)).sum()
            low = (self.audit_data['鍋忓樊鐜?%)'].abs() < 5).sum()
            lines.append("")
            lines.append(f"馃敶 鍋忓樊鐜?>10%锛歿high} 琛?)
            lines.append(f"馃煛 鍋忓樊鐜?5%-10%锛歿medium} 琛?)
            lines.append(f"馃煝 鍋忓樊鐜?<5%锛歿low} 琛?)

        if '澶囨敞鍘熷洜' in self.audit_data.columns:
            total = len(self.audit_data)
            has_remark = self.audit_data['澶囨敞鍘熷洜'].notna().sum()
            lines.append("")
            lines.append(f"馃摑 澶囨敞瑕嗙洊鐜囷細{has_remark}/{total} ({has_remark/total*100:.1f}%)")

        text_content = chr(10).join(lines)

        # 鍒涘缓寮圭獥
        d = tk.Toplevel(self.root)
        d.title("馃摉 鏈懆鍋忓樊鏁呬簨绾?)
        d.geometry("600x520")
        d.transient(self.root)
        d.grab_set()
        d.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 600) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 520) // 2
        d.geometry(f"+{x}+{y}")

        tk.Label(d, text="馃摉 鏈懆鍋忓樊鏁呬簨绾?, font=("Microsoft YaHei", 12, "bold")).pack(pady=10)
        text = tk.Text(d, font=("Microsoft YaHei", 10), wrap="word", height=20)
        text.pack(fill="both", expand=True, padx=10, pady=5)
        text.insert("1.0", text_content)
        text.configure(state="disabled")

        def copy_to_clip():
            d.clipboard_clear()
            d.clipboard_append(text_content)
            messagebox.showinfo("宸插鍒?, "鏁呬簨绾垮凡澶嶅埗鍒板壀璐存澘")

        tk.Button(d, text="馃搵 澶嶅埗鍒板壀璐存澘", command=copy_to_clip,
                  bg="#4a90d9", fg="white", relief="flat", width=15).pack(pady=8)


    def _scan_remark_cleanup(self):
        """鎵弿澶囨敞鏁版嵁骞剁敓鎴愭竻娲楀缓璁紙濡傚瓨鍦級"""
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showinfo("鎻愮ず", "鏃犳暟鎹彲鎵弿")
            return
        # 纭畾鐗╂枡鎻忚堪鍒?
        mat_col = None
        for col in self.audit_data.columns:
            if '鐗╂枡鎻忚堪' in col or '缁勪欢鎻忚堪' in col:
                mat_col = col
                break





        if mat_col is None:





            mat_col = '缁勪欢鐗╂枡鎻忚堪'

        # 纭畾澶囨敞鍒?
        remark_col = next((c for c in ['澶囨敞', '澶囨敞鍘熷洜'] if c in self.audit_data.columns), '澶囨敞')











        # 娓呮礂瑙勫垯





        cleanup_rules = {





            '棰濆畾涓嶈冻': '瀹氶鍋忎綆', '瀹氶浣?: '瀹氶鍋忎綆',





            '瀹氶楂?: '瀹氶鍋忛珮', '璁鹃珮': '瀹氶鍋忛珮',





            '璁句綆': '瀹氶鍋忎綆', '娌℃湁瀹氶': '绯荤粺鏃犲畾棰?,





            '鏃犲畾棰?: '绯荤粺鏃犲畾棰?, '鏈畾棰?: '绯荤粺鏃犲畾棰?,





            '娌＄敤杩?: '鏈～鍐?, '涓嶇煡閬?: '鏈～鍐?,





            '寰呮煡': '鏈～鍐?, '娴嬭瘯': '璇曚骇',





            '璇曟満': '璇曚骇', '鎵撴牱': '璇曚骇',





            '鏍锋澘': '璇曚骇', '鍫?: '鍫垫枡',





            '鍗?: '鍗℃枡', '鐮?: '鐮存崯', '鐑?: '鐮存崯',





        }











        suggestions = []





        for idx, row in self.audit_data.iterrows():





            remark = str(row.get(remark_col, '')).strip()





            if not remark or remark in ('nan', 'NaN', 'None'):





                continue





            for keyword, standard in cleanup_rules.items():





                if keyword in remark and remark != standard:





                    suggestions.append({





                        'idx': idx,





                        '鐗╂枡': str(row.get(mat_col, ''))[:25],





                        '褰撳墠澶囨敞': remark,





                        '寤鸿鏍囧噯鍖栦负': standard,





                        'excel_row': int(row.get('excel_row', 0)),





                        'remark_col': remark_col,





                    })





                    break





        return suggestions











    def _show_cleanup_window(self):





        """鏄剧ず澶囨敞娓呮礂鏍囧噯鍖栫獥鍙?""





        suggestions = self._scan_remark_cleanup()





        if not suggestions:





            messagebox.showinfo("鎻愮ず", "鏈彂鐜板彲娓呮礂鐨勫娉紝鏁版嵁宸插緢瑙勮寖")





            return











        d = tk.Toplevel(self.root)





        d.title("馃Ч 澶囨敞娓呮礂鏍囧噯鍖?)





        d.geometry(self.config.get("ui.cleanup_window_size", "750x480"))





        d.transient(self.root)





        d.grab_set()





        d.update_idletasks()





        rx = self.root.winfo_rootx() + (self.root.winfo_width() - 750) // 2





        ry = self.root.winfo_rooty() + (self.root.winfo_height() - 480) // 2





        d.geometry(f"+{rx}+{ry}")











        tk.Label(d, text=f"鍙戠幇 {len(suggestions)} 鏉″彲鏍囧噯鍖栫殑澶囨敞锛屽嬀閫夊悗鐐瑰嚮鎵ц",





                 font=("Microsoft YaHei", 10)).pack(pady=10)











        # 琛ㄦ牸锛氱墿鏂?/ 褰撳墠澶囨敞 / 寤鸿鏍囧噯鍖栦负





        tree_frame = tk.Frame(d)





        tree_frame.pack(fill="both", expand=True, padx=10)





        cols = ("select", "material", "current", "suggested")





        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15)





        tree.heading("select", text="鉁?)





        tree.heading("material", text="鐗╂枡鎻忚堪")





        tree.heading("current", text="褰撳墠澶囨敞")





        tree.heading("suggested", text="寤鸿鏍囧噯鍖栦负")





        tree.column("select", width=30, anchor="center")





        tree.column("material", width=150, anchor="w")





        tree.column("current", width=250, anchor="w")





        tree.column("suggested", width=200, anchor="w")











        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)





        tree.configure(yscrollcommand=scroll_y.set)





        scroll_y.pack(side="right", fill="y")





        tree.pack(side="left", fill="both", expand=True)











        # 濉厖鏁版嵁





        self._cleanup_checkboxes = {}





        for i, sug in enumerate(suggestions):





            item_id = tree.insert('', 'end', values=('鈽?, sug['鐗╂枡'], sug['褰撳墠澶囨敞'], sug['寤鸿鏍囧噯鍖栦负']))





            self._cleanup_checkboxes[item_id] = (False, sug)











        # 鐐瑰嚮琛屽垏鎹㈠嬀閫夌姸鎬?





        def toggle_check(event):





            item_id = tree.identify_row(event.y)





            if not item_id or item_id not in self._cleanup_checkboxes:





                return





            checked, sug = self._cleanup_checkboxes[item_id]





            new_checked = not checked





            self._cleanup_checkboxes[item_id] = (new_checked, sug)





            tree.set(item_id, "select", '鈽? if new_checked else '鈽?)











        tree.bind("<Button-1>", toggle_check)











        # 搴曢儴鎸夐挳





        btn_frame = tk.Frame(d)





        btn_frame.pack(pady=10)











        def select_all():





            for item_id in self._cleanup_checkboxes:





                _, sug = self._cleanup_checkboxes[item_id]





                self._cleanup_checkboxes[item_id] = (True, sug)





                tree.set(item_id, "select", '鈽?)











        def execute_cleanup():





            cleaned = 0





            for item_id, (checked, sug) in self._cleanup_checkboxes.items():





                if checked and self.audit_data is not None:





                    remark_col = sug.get('remark_col', '澶囨敞鍘熷洜'); self.audit_data.at[sug['idx'], remark_col] = sug['寤鸿鏍囧噯鍖栦负']





                    cleaned += 1





            if cleaned > 0:





                self._refresh_audit_tree(self.audit_data)





            try:





                self.audit_tree['displaycolumns'] = '#all'





                self.root.update_idletasks()





                self.log("[DEBUG] AI瀹℃牳瀹屾垚鍚庡己鍒舵仮澶峝isplaycolumns", "info")





            except Exception as e:





                self.log(f"[ERROR] 鎭㈠displaycolumns澶辫触: {e}", "error")





                self._update_audit_stats()





                self._update_filter_options()





                self.log(f"馃Ч 澶囨敞娓呮礂瀹屾垚锛歿cleaned} 鏉″凡鏍囧噯鍖?, "success")





                messagebox.showinfo("瀹屾垚", f"宸叉爣鍑嗗寲 {cleaned} 鏉″娉?)





            d.destroy()











        tk.Button(btn_frame, text="鍏ㄩ€?, command=select_all,





                  bg="#d0d7de", font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)





        tk.Button(btn_frame, text="鎵ц閫変腑", command=execute_cleanup,





                  bg="#4CAF50", fg="white", font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)





        tk.Button(btn_frame, text="鍙栨秷", command=d.destroy,





                  bg="#9E9E9E", fg="white", font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)

















    def _add_alt(self):





        d = tk.Toplevel(self.root)





        d.title("娣诲姞鏇夸唬鏂欓厤瀵?)





        d.geometry(self.config.get("ui.alt_dialog_size", "580x300"))





        d.transient(self.root)





        d.grab_set()





        d.configure(bg=C['bg'])











        tk.Label(d, text="鐗╂枡A锛堢紪鐮佹垨鍚嶇О锛夛細", font=("Microsoft YaHei", 10),





                 fg=C['text'], bg=C['bg']).pack(pady=(10, 3), anchor="w", padx=20)











        var_a = tk.StringVar()





        if hasattr(self, 'material_list') and self.material_list:





            cb_a = ttk.Combobox(d, textvariable=var_a, values=self.material_list,





                                font=("Consolas", 9), state="normal", width=65)





            cb_a.pack(padx=20, fill="x")





            cb_a.set("杈撳叆鍏抽敭瀛楁垨鐐瑰嚮涓嬫媺閫夋嫨...")





            def on_focus_a(e):





                if var_a.get() == "杈撳叆鍏抽敭瀛楁垨鐐瑰嚮涓嬫媺閫夋嫨...":





                    var_a.set("")





            cb_a.bind("<FocusIn>", on_focus_a)





        else:





            e_a = tk.Entry(d, font=("Consolas", 9), bg=C['surface2'],





                           fg=C['text'], insertbackground=C['accent'], relief="flat")





            e_a.pack(padx=20, fill="x")





            e_a.insert(0, "锛堟湭鎵惧埌鐗╂枡鍒楄〃锛岃鎵嬪姩杈撳叆锛?)











        tk.Label(d, text="鐗╂枡B锛堢紪鐮佹垨鍚嶇О锛夛細", font=("Microsoft YaHei", 10),





                 fg=C['text'], bg=C['bg']).pack(pady=(10, 3), anchor="w", padx=20)











        var_b = tk.StringVar()





        if hasattr(self, 'material_list') and self.material_list:





            cb_b = ttk.Combobox(d, textvariable=var_b, values=self.material_list,





                                font=("Consolas", 9), state="normal", width=65)





            cb_b.pack(padx=20, fill="x")





            cb_b.set("杈撳叆鍏抽敭瀛楁垨鐐瑰嚮涓嬫媺閫夋嫨...")





            def on_focus_b(e):





                if var_b.get() == "杈撳叆鍏抽敭瀛楁垨鐐瑰嚮涓嬫媺閫夋嫨...":





                    var_b.set("")





            cb_b.bind("<FocusIn>", on_focus_b)





        else:





            e_b = tk.Entry(d, font=("Consolas", 9), bg=C['surface2'],





                           fg=C['text'], insertbackground=C['accent'], relief="flat")





            e_b.pack(padx=20, fill="x")





            e_b.insert(0, "锛堟湭鎵惧埌鐗╂枡鍒楄〃锛岃鎵嬪姩杈撳叆锛?)











        # 宸叉湁閰嶅鍒楄〃锛堢敤浜庡垹闄わ級





        lst = None  # 灏嗗湪鍚庨潰瀹氫箟











        def confirm():





            a = var_a.get().strip()





            b = var_b.get().strip()





            if not a or not b:





                messagebox.showwarning("鎻愮ず", "鐗╂枡A鍜岀墿鏂橞閮藉繀椤诲～鍐欙紒")





                return











            def parse_selection(x):





                if '|' in x:





                    parts = [p.strip() for p in x.split('|')]





                    if len(parts) >= 3:





                        # 鏍煎紡: 宸ュ巶 | 缂栫爜 | 鍚嶇О





                        return parts[0], parts[1], parts[2]  # (宸ュ巶, 缂栫爜, 鍚嶇О)





                    elif len(parts) == 2:





                        # 鏍煎紡: 缂栫爜 | 鍚嶇О





                        return '', parts[0], parts[1]  # (绌哄伐鍘? 缂栫爜, 鍚嶇О)





                    else:





                        return '', parts[0], ''  # (绌哄伐鍘? 缂栫爜, 绌哄悕绉?





                else:





                    # 鎵嬪姩杈撳叆鏃跺皾璇曚粠鐗╂枡鍒楄〃鍖归厤





                    for item in getattr(self, 'material_list', []):





                        if x in item:





                            parts = [p.strip() for p in item.split('|')]





                            if len(parts) >= 3:





                                return parts[0], parts[1], parts[2]  # (宸ュ巶, 缂栫爜, 鍚嶇О)





                    return '', x, x  # 鍏滃簳











            factory_a, a_code, a_name = parse_selection(a)





            factory_b, b_code, b_name = parse_selection(b)











            # Bug 6 淇锛氭鏌?A 鍜?B 涓嶈兘鏄悓涓€涓墿鏂?





            if a_code == b_code:





                messagebox.showwarning("鎻愮ず", "鐗╂枡A鍜岀墿鏂橞涓嶈兘鏄悓涓€涓墿鏂欙紒")





                return











            # 鍘婚噸妫€鏌ワ細鏄惁宸插瓨鍦ㄧ浉鍚岄厤瀵癸紙缂栫爜涓虹┖鏃剁敤鎻忚堪鍖归厤锛?





            exact_match = False





            conflict = False





            for (ea, eb) in self.alt_pairs:





                def _extract_code_and_desc(item):





                    if isinstance(item, (list, tuple)):





                        if len(item) >= 3:





                            code = str(item[1]).strip() if item[1] else ''





                            desc = str(item[2]).strip() if item[2] else ''





                            return code if code else desc  # 缂栫爜涓虹┖鐢ㄦ弿杩?





                        if len(item) == 2:





                            code = str(item[0]).strip() if item[0] else ''





                            desc = str(item[1]).strip() if item[1] else ''





                            return code if code else desc





                        return str(item[0]).strip() if item[0] else ''





                    s = str(item).strip()





                    return s if s else ''





                ea_key = _extract_code_and_desc(ea)





                eb_key = _extract_code_and_desc(eb)





                a_key = a_code if a_code else a_name





                b_key = b_code if b_code else b_name





                if (ea_key == a_key and eb_key == b_key) or (ea_key == b_key and eb_key == a_key):





                    exact_match = True





                    break





                if a_key in (ea_key, eb_key) or b_key in (ea_key, eb_key):





                    conflict = True











            if exact_match:





                msg = f"閰嶅宸插瓨鍦細{a_code} 鈫?{b_code}\n鏄惁浠嶈娣诲姞锛?





                if not messagebox.askyesno("閲嶅閰嶅", msg):





                    return





            if conflict:





                warn = f"鐗╂枡 {a_code} 鎴?{b_code} 宸插瓨鍦ㄤ簬鍏朵粬閰嶅涓紝缁х画娣诲姞鍙兘瀵艰嚧鍐茬獊銆俓n鏄惁浠嶈娣诲姞锛?





                if not messagebox.askyesno("鐗╂枡鍐茬獊", warn):





                    return











            save_a = (factory_a, a_code, a_name)





            save_b = (factory_b, b_code, b_name)





            self.alt_pairs.append((save_a, save_b))





            try:





                save_alt_pairs(self.alt_pairs, log_cb=self.log)





            except Exception as e:





                messagebox.showerror("閿欒", f"鏇夸唬鏂欐坊鍔犲け璐ワ細{e}")





                return





            self._refresh_alt_view(self._alt_inner)





            messagebox.showinfo("鎻愮ず", "鏇夸唬鏂欐坊鍔犳垚鍔燂紒")





            d.destroy()











        def do_del():





            if lst and lst.curselection():





                idx = lst.curselection()[0]





                del self.alt_pairs[idx]





                save_alt_pairs(self.alt_pairs, log_cb=self.log)





                self._refresh_alt_view(self._alt_inner)





                d.destroy()





            else:





                messagebox.showwarning("鎻愮ず", "璇峰厛閫夋嫨瑕佸垹闄ょ殑閰嶅")











        btn_frame = tk.Frame(d, bg=C['bg'])





        btn_frame.pack(pady=12)





        tk.Button(btn_frame, text="鉁?纭畾", command=confirm, bg="#4CAF50", fg="white",





                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=8)





        tk.Button(btn_frame, text="鉁?鍙栨秷", command=d.destroy, bg="#9E9E9E", fg="white",





                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=8)











        if self.alt_pairs:





            tk.Label(d, text="宸叉湁閰嶅锛堢偣鍑诲彲鍒犻櫎锛夛細", font=("Microsoft YaHei", 9),





                     fg=C['text_dim'], bg=C['bg']).pack(pady=(10, 3))





            lst = tk.Listbox(d, font=("Consolas", 9), height=4, selectmode='single')





            lst.pack(fill="x", padx=20)





            for a, b in self.alt_pairs:





                # 瑙ｆ瀽鏄剧ず





                if isinstance(a, (list, tuple)) and len(a) == 3:





                    _, a_code, a_name = a





                elif isinstance(a, (list, tuple)) and len(a) == 2:





                    a_code, a_name = a





                else:





                    a_code, a_name = str(a), ''





                if isinstance(b, (list, tuple)) and len(b) == 3:





                    _, b_code, b_name = b





                elif isinstance(b, (list, tuple)) and len(b) == 2:





                    b_code, b_name = b





                else:





                    b_code, b_name = str(b), ''





                left = f"{a_code} {a_name}" if a_name else a_code





                right = f"{b_code} {b_name}" if b_name else b_code





                lst.insert('end', left + " 鈬?" + right)





            tk.Button(d, text="鍒犻櫎閫変腑閰嶅", command=do_del, bg=C['danger'], fg="white",





                      font=("Microsoft YaHei", 9), relief="flat").pack(pady=5)























    def _del_alt(self):





        """鍒犻櫎鏇夸唬鏂欓厤瀵?""





        if not self.alt_pairs:





            messagebox.showinfo("鎻愮ず", "褰撳墠娌℃湁鏇夸唬鏂欓厤瀵瑰彲鍒犻櫎")





            return





        d = tk.Toplevel(self.root)





        d.title("鍒犻櫎鏇夸唬鏂欓厤瀵?)





        d.geometry("500x350")





        d.transient(self.root)





        d.grab_set()





        tk.Label(d, text="璇烽€夋嫨瑕佸垹闄ょ殑閰嶅锛?, font=("Microsoft YaHei", 10)).pack(pady=8)





        lst = tk.Listbox(d, font=("Consolas", 10), selectmode='single')





        lst.pack(fill="both", expand=True, padx=10)





        for a, b in self.alt_pairs:





            # a, b 閮芥槸 (factory, code, name) 鍏冪粍





            factory_a, code_a, name_a = a if isinstance(a, tuple) else ('', a, '')





            factory_b, code_b, name_b = b if isinstance(b, tuple) else ('', b, '')





            left = f"{factory_a} {code_a} {name_a}"





            right = f"{factory_b} {code_b} {name_b}"





            lst.insert('end', left + " 鈬?" + right)





        if self.alt_pairs:





            lst.select_set(0)





        btn_frame = tk.Frame(d)





        btn_frame.pack(pady=10)





        def do_delete():





            sel = lst.curselection()





            if sel:





                del self.alt_pairs[sel[0]]





                save_alt_pairs(self.alt_pairs, log_cb=self.log)





                self._refresh_alt_view(self._alt_inner)





                d.destroy()





            else:





                messagebox.showwarning("鎻愮ず", "璇峰厛閫夋嫨瑕佸垹闄ょ殑閰嶅")





        tk.Button(btn_frame, text="鍒犻櫎", command=do_delete, bg=C['danger'], fg="white",





                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)





        tk.Button(btn_frame, text="鍙栨秷", command=d.destroy, bg="#d0d7de",





                  font=("Microsoft YaHei", 10), relief="flat", width=10).pack(side="left", padx=5)

















    def _reset_alt(self):





        self.alt_pairs = []   # 娓呯┖鏇夸唬鏂欓厤瀵?





        # 鍒锋柊鐣岄潰鏄剧ず





        canvas = self.alt_list_frame.winfo_children()[0]





        inner = canvas.winfo_children()[0]





        self._refresh_alt_view(inner)





        # 淇濆瓨鍒伴厤缃枃浠?





        save_alt_pairs(self.alt_pairs, log_cb=self.log)





        





























    def _refresh_alt_view(self, inner):





        for w in inner.winfo_children():





            w.destroy()





        for a, b in self.alt_pairs:





            # 鏈熸湜 a, b 閮芥槸 (factory, code, name) 鏍煎紡





            if isinstance(a, (list, tuple)) and len(a) >= 3:





                a_factory, a_code, a_name = a[0], a[1], a[2]





            else:





                a_factory, a_code, a_name = '', str(a), ''





            if isinstance(b, (list, tuple)) and len(b) >= 3:





                b_factory, b_code, b_name = b[0], b[1], b[2]





            else:





                b_factory, b_code, b_name = '', str(b), ''





            # 灏?None / 'None' 杞负绌哄瓧绗︿覆





            a_code = '' if a_code in (None, 'None') else str(a_code).strip()





            a_name = '' if a_name in (None, 'None') else str(a_name).strip()





            b_code = '' if b_code in (None, 'None') else str(b_code).strip()





            b_name = '' if b_name in (None, 'None') else str(b_name).strip()





            





            # 鏄剧ず锛氱紪鐮?+ 鍚嶇О锛堣嫢鍚嶇О瀛樺湪锛?





            a_disp = f"{a_code} {a_name}" if a_name else a_code





            b_disp = f"{b_code} {b_name}" if b_name else b_code





            





            fr = tk.Frame(inner, bg=C['surface2'])





            fr.pack(fill="x", pady=1)





            tk.Label(fr, text=f"鈫?{a_disp}", font=("Consolas", 8), fg=C['text'],





                     bg=C['surface2'], anchor="w").pack(side="left", padx=4)





            tk.Label(fr, text="|", font=("Consolas", 8), fg=C['text_dim'],





                     bg=C['surface2']).pack(side="left")





            tk.Label(fr, text=b_disp, font=("Consolas", 8), fg=C['purple'],





                     bg=C['surface2'], anchor="w").pack(side="left", padx=4)

















    def _show_alt_snapshot(self):





        """鏄剧ず鏇夸唬鏂欓厤缃殑JSON蹇収锛堝熀浜?alt_manager 鍔ㄦ€佽矾寰勶級"""





        from domain.alt_material.alt_manager import _get_config_path











        config_path = _get_config_path()





        if not os.path.exists(config_path):





            messagebox.showinfo("鎻愮ず",





                f"鏈壘鍒伴厤缃枃浠讹細\n{config_path}\n\n褰撳墠浣跨敤榛樿鍐呯疆鏇夸唬鏂欐暟鎹€?)





            return











        try:





            with open(config_path, 'r', encoding='utf-8') as f:





                content = f.read()





            parsed = json.loads(content)





            formatted = json.dumps(parsed, indent=2, ensure_ascii=False)





        except Exception as e:





            messagebox.showerror("璇诲彇澶辫触", f"鏃犳硶瑙ｆ瀽 JSON 鏂囦欢锛歕n{e}")





            return











        win = tk.Toplevel(self.root)





        win.title(f"鏇夸唬鏂欏揩鐓?- {os.path.basename(config_path)}")





        win.geometry("750x550")











        frame = tk.Frame(win)





        frame.pack(fill='both', expand=True, padx=10, pady=10)











        scroll = tk.Scrollbar(frame)





        scroll.pack(side='right', fill='y')











        text = tk.Text(frame, wrap='none', yscrollcommand=scroll.set,





                       font=('Consolas', 10))





        text.pack(side='left', fill='both', expand=True)





        scroll.config(command=text.yview)











        text.insert('end', formatted)





        text.config(state='disabled')











        def copy_to_clip():





            win.clipboard_clear()





            win.clipboard_append(formatted)





            messagebox.showinfo("宸插鍒?, "JSON 鍐呭宸插鍒跺埌鍓创鏉?)











        tk.Button(win, text="澶嶅埗鍒板壀璐存澘", command=copy_to_clip,





                 font=("Microsoft YaHei", 10), relief="flat",





                 bg=C['accent'], fg="white").pack(pady=8)











    def _import_bom(self):





        """瀵煎叆 BOM 鏁版嵁鏂囦欢锛屼笌鏃?BOM 姣斿宸紓"""





        file_path = filedialog.askopenfilename(





            title="閫夋嫨 BOM 鏁版嵁鏂囦欢",





            filetypes=[("Excel 鏂囦欢", "*.xlsx *.xls"), ("鎵€鏈夋枃浠?, "*.*")]





        )





        if not file_path:





            return





        try:





            self.log(f"馃摝 姝ｅ湪璇诲彇 BOM 鏂囦欢锛歿os.path.basename(file_path)}", "info")





            new_df = pd.read_excel(file_path, sheet_name=0)





            self.log(f"   鍏辫鍙?{len(new_df)} 琛屾暟鎹?, "info")





        except Exception as e:





            messagebox.showerror("璇诲彇澶辫触", f"鏃犳硶璇诲彇 BOM 鏂囦欢锛歿e}")





            return











        # 鈹€鈹€ 鏌ユ壘鐗╂枡缂栫爜閿?鈹€鈹€





        key_candidates = ['缁勪欢鐗╂枡鍙?, '鐗╂枡缂栫爜', '鐗╂枡鍙?, '鐗╂枡', 'code', 'material_code']





        key_col = None





        for col in key_candidates:





            if col in new_df.columns:





                key_col = col





                break





        if not key_col:





            cols_str = '銆?.join(new_df.columns[:10].tolist())





            messagebox.showerror("鍒楃己澶?, f"鏈壘鍒扮墿鏂欑紪鐮佸垪锛堝€欓€夛細{key_candidates}锛塡n褰撳墠鍒楋細{cols_str}")





            return





        self.log(f"   鐗╂枡缂栫爜閿垪锛歿key_col}", "info")











        # 娓呮礂鏂版暟鎹?





        new_df[key_col] = new_df[key_col].astype(str).str.strip()





        new_records = new_df.to_dict('records')





        new_keys = set(new_df[key_col].dropna().unique())











        # 鈹€鈹€ 鍔犺浇鏃?BOM 鈹€鈹€





        old_path = self._get_bom_stored_path()





        old_records = []





        old_keys = set()





        old_key_col = key_col  # 榛樿浣跨敤鏂版枃浠剁殑key鍒?





        added_count, removed_count, modified_count = 0, 0, 0





        detail_lines = []











        if os.path.exists(old_path):





            try:





                with open(old_path, 'r', encoding='utf-8') as f:





                    old_data = json.load(f)





                if isinstance(old_data, dict) and 'records' in old_data:





                    old_records = old_data['records']





                    # 浣跨敤鏃OM淇濆瓨鏃剁殑key鍒楀悕锛堝吋瀹规柊鏃ф牸寮忎笉涓€鑷寸殑鎯呭喌锛?





                    old_key_col = old_data.get('key_column', key_col)





                elif isinstance(old_data, list):





                    old_records = old_data





                    old_key_col = key_col





                else:





                    old_key_col = key_col





                old_keys = {str(r.get(old_key_col, '')).strip() for r in old_records if r.get(old_key_col)}





                self.log(f"   鏃?BOM 鍏?{len(old_records)} 鏉¤褰?, "info")





            except Exception as e:





                self.log(f"   鏃?BOM 璇诲彇澶辫触锛堝皢瑙嗕负鍏ㄦ柊瀵煎叆锛夛細{e}", "warning")











        # 鈹€鈹€ 姣斿 鈹€鈹€





        if not old_keys:





            added_count = len(new_keys)





            detail_lines.append(f"馃啎 鍏ㄦ柊瀵煎叆锛屽叡 {added_count} 鏉℃柊璁板綍")





        else:





            added_keys = new_keys - old_keys





            removed_keys = old_keys - new_keys





            modified_count = 0











            # 蹇€熶慨鏀规娴嬶紙鏃ц褰曡浆dict锛?





            old_map = {str(r.get(old_key_col, '')).strip(): r for r in old_records if r.get(old_key_col)}





            for key in new_keys & old_keys:





                new_row = new_df[new_df[key_col] == key].iloc[0]





                old_row = old_map.get(key, {})





                changed = False





                for c in new_df.columns:





                    if c == key_col:





                        continue





                    new_val = str(new_row.get(c, '')).strip()





                    old_val = str(old_row.get(c, '')).strip()





                    if new_val != old_val:





                        changed = True





                        break





                if changed:





                    modified_count += 1











            added_count = len(added_keys)





            removed_count = len(removed_keys)











            if added_count > 0:





                sample_added = list(added_keys)[:5]





                detail_lines.append(f"馃啎 鏂板 {added_count} 鏉★紙绀轰緥锛歿', '.join(sample_added)}锛?)





            if removed_count > 0:





                sample_removed = list(removed_keys)[:5]





                detail_lines.append(f"鉃?鍒犻櫎 {removed_count} 鏉★紙绀轰緥锛歿', '.join(sample_removed)}锛?)





            if modified_count > 0:





                detail_lines.append(f"鉁忥笍 淇敼 {modified_count} 鏉?)











        # 鈹€鈹€ 淇濆瓨鏂?BOM 鈹€鈹€





        bom_store = {





            'version': '1.0',





            'imported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),





            'key_column': key_col,





            'total_rows': len(new_df),





            'records': new_records





        }





        try:





            with open(old_path, 'w', encoding='utf-8') as f:





                json.dump(bom_store, f, ensure_ascii=False, indent=2)





            self.log(f"   BOM 宸蹭繚瀛樿嚦锛歿old_path}", "info")





        except Exception as e:





            messagebox.showerror("淇濆瓨澶辫触", f"鏃犳硶淇濆瓨 BOM 鏁版嵁锛歿e}")





            return











        # 鈹€鈹€ 灞曠ず瀵规瘮鎶ュ憡 鈹€鈹€





        report_win = tk.Toplevel(self.root)





        report_win.title("BOM 瀵煎叆鎶ュ憡")





        report_win.geometry(self.config.get("ui.bom_report_size", "600x400"))





        report_win.transient(self.root)





        report_win.grab_set()





        report_win.update_idletasks()





        w, h = 600, 400





        x = (report_win.winfo_screenwidth() - w) // 2





        y = (report_win.winfo_screenheight() - h) // 2





        report_win.geometry(f'{w}x{h}+{x}+{y}')











        bg_frame = Frame(report_win, bg='#1e1e1e')





        bg_frame.pack(fill='both', expand=True)











        title_label = Label(bg_frame, text="馃摝 BOM 瀵煎叆鎶ュ憡", font=('寰蒋闆呴粦', 14, 'bold'),





                            bg='#1e1e1e', fg='#ffffff')





        title_label.pack(pady=(20, 10))











        summary_text = f"鏂囦欢锛歿os.path.basename(file_path)}\n"





        summary_text += f"鎬昏鏁帮細{len(new_df)}銆€锝溿€€"





        summary_text += f"馃啎鏂板 {added_count}銆€锝溿€€"





        summary_text += f"鉃栧垹闄?{removed_count}銆€锝溿€€"





        summary_text += f"鉁忥笍淇敼 {modified_count}"





        summary_label = Label(bg_frame, text=summary_text, font=('寰蒋闆呴粦', 10),





                              bg='#1e1e1e', fg='#cccccc', justify='left', anchor='w')





        summary_label.pack(padx=20, fill='x')











        canvas = Canvas(bg_frame, bg='#1e1e1e', highlightthickness=0)





        scrollbar = Scrollbar(bg_frame, orient='vertical', command=canvas.yview)





        content_frame = Frame(canvas, bg='#1e1e1e')











        canvas.configure(yscrollcommand=scrollbar.set)





        canvas.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=10)





        scrollbar.pack(side='right', fill='y', pady=10, padx=(0, 20))











        canvas_window = canvas.create_window((0, 0), window=content_frame, anchor='nw')





        content_frame.bind('<Configure>',





                           lambda e: canvas.configure(scrollregion=canvas.bbox('all')))





        canvas.bind('<Configure>',





                    lambda e: canvas.itemconfig(canvas_window, width=e.width))











        if detail_lines:





            for line in detail_lines:





                color = '#1a7f37' if '鏂板' in line else ('#d29922' if '鍒犻櫎' in line else '#79c0ff')





                lbl = Label(content_frame, text=line, font=('寰蒋闆呴粦', 10),





                            bg='#1e1e1e', fg=color, anchor='w', justify='left')





                lbl.pack(anchor='w', pady=2)





        else:





            Label(content_frame, text="鉁?鏃犲彉鏇达紙鏁版嵁瀹屽叏涓€鑷达級", font=('寰蒋闆呴粦', 10),





                  bg='#1e1e1e', fg='#1a7f37', anchor='w').pack(anchor='w', pady=2)











        close_btn = Button(bg_frame, text="鍏抽棴", font=('寰蒋闆呴粦', 10),





                           bg='#3a3a3a', fg='#ffffff', activebackground='#555555',





                           command=report_win.destroy, cursor='hand2')





        close_btn.pack(pady=15)











        self.log(f"馃摝 BOM 瀵煎叆瀹屾垚锛氭柊澧?{added_count} / 鍒犻櫎 {removed_count} / 淇敼 {modified_count}", "info")





        self.log(f"   BOM 鏁版嵁宸蹭繚瀛橈紝涓嬫杩囨湡妫€鏌ュ皢鑷姩瀵规瘮宸紓", "info")

















    def _run_pre_check(self):





        """瀹屾暣棰勬鎶ュ憡锛氬垪瀹屾暣鎬?+ 閲嶅璁㈠崟 + 鏁板€煎紓甯?+ 鏇夸唬鏂欓厤缃?+ 寮圭獥"""





        try:





            input_path = self.input_file.get()





            if not input_path or not os.path.exists(input_path):





                self.log("棰勬澶辫触锛氳緭鍏ユ枃浠朵笉瀛樺湪", "error")





                messagebox.showwarning("棰勬澶辫触", "璇峰厛閫夋嫨杈撳叆鏂囦欢")





                return











            df = pd.read_excel(input_path, sheet_name='Data', nrows=self.config.get('precheck.sample_rows', 1000))





            results = []











            # 1. 鍒楀畬鏁存€ф鏌ワ紙榛勯噾妯℃澘锛?





            golden_cols = self.config.get('precheck.golden_columns', [





                '娴佺▼璁㈠崟', '璁㈠崟寮€濮嬫棩鏈?, '缁勪欢鐗╂枡鍙?, '缁勪欢鐗╂枡鎻忚堪',





                '缁勪欢鏁伴噺', '鍗曚綅', '宸ュ巶鍚嶇О', '鐢熶骇绠＄悊鍛樻弿杩?





            ])





            missing_cols = [col for col in golden_cols if col not in df.columns]





            if missing_cols:





                results.append(('涓ラ噸', f'缂哄け鏍囧噯鍒楋細{", ".join(missing_cols)}'))





            else:





                results.append(('閫氳繃', '榛勯噾妯℃澘鍒楀畬鏁?))











            # 2. 閲嶅璁㈠崟妫€鏌ワ紙鏃ユ湡 + 娴佺▼璁㈠崟 + 鐗╂枡缂栫爜 鑱斿悎鍘婚噸锛?





            date_col = None





            for col in ['璁㈠崟寮€濮嬫棩鏈?, '璁㈠崟鏃ユ湡', '鏃ユ湡']:





                if col in df.columns:





                    date_col = col





                    break





            # 纭畾鐗╂枡缂栫爜鍒楀悕





            mat_col = None





            for col in ['鐗╂枡缂栫爜', '缁勪欢鐗╂枡鍙?, '鐗╂枡鍙?, '缂栫爜']:





                if col in df.columns:





                    mat_col = col





                    break





            if date_col and '娴佺▼璁㈠崟' in df.columns and mat_col:





                df['_check_date'] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y-%m-%d')





                dup_mask = df.duplicated(subset=['_check_date', '娴佺▼璁㈠崟', mat_col], keep=False)





                dup_orders = df[dup_mask]





                if not dup_orders.empty:





                    dup_groups = dup_orders.groupby(['_check_date', '娴佺▼璁㈠崟', mat_col]).ngroups





                    results.append(('璀﹀憡', f'鍙戠幇 {len(dup_orders)} 鏉￠噸澶嶈褰曪紙{dup_groups} 缁勯噸澶嶏級锛岃妫€鏌?SAP 瀵煎嚭鏄惁閲嶅'))





                else:





                    results.append(('閫氳繃', '鏃犻噸澶嶈褰曪紙鏃ユ湡+璁㈠崟+鐗╂枡锛?))





                df.drop(columns=['_check_date'], inplace=True)





            else:





                missing = []





                if not date_col: missing.append('鏃ユ湡鍒?)





                if '娴佺▼璁㈠崟' not in df.columns: missing.append('娴佺▼璁㈠崟')





                if not mat_col: missing.append('鐗╂枡缂栫爜/缁勪欢鐗╂枡鍙?)





                results.append(('璀﹀憡', f'缂哄皯蹇呰鍒楋紙{", ".join(missing)}锛夛紝鏃犳硶妫€娴嬮噸澶嶈鍗?))











            # 3. 鏁板€煎紓甯告鏌?





            if '缁勪欢鏁伴噺' in df.columns:





                neg_quota = df[df['缁勪欢鏁伴噺'] < 0]





                if not neg_quota.empty:





                    results.append(('璀﹀憡', f'鍙戠幇 {len(neg_quota)} 琛屽畾棰濅负璐熸暟'))





                else:





                    results.append(('閫氳繃', '瀹氶鏃犺礋鏁?))











            # 4. 鏇夸唬鏂欓厤缃鏌?





            try:





                from domain.alt_material.alt_manager import _get_config_path, load_alt_pairs





                alt_path = _get_config_path()





                if not os.path.exists(alt_path):





                    results.append(('璀﹀憡', '鏇夸唬鏂欓厤缃枃浠朵笉瀛樺湪锛屽皢浣跨敤鍐呯疆閰嶅'))





                else:





                    alt_pairs = load_alt_pairs(log_cb=self.log)





                    results.append(('閫氳繃', f'鏇夸唬鏂欓厤缃姞杞芥垚鍔燂紝鍏?{len(alt_pairs)} 缁勯厤瀵?))





            except Exception as e:





                results.append(('涓ラ噸', f'鏇夸唬鏂欓厤缃姞杞藉け璐ワ細{e}'))











            # 5. 榛勯噾妯℃澘瀵规瘮





            try:





                gold_cols = self._load_golden_columns()





                if gold_cols:





                    actual_cols = set(df.columns)





                    template_cols = set(gold_cols)





                    missing = template_cols - actual_cols





                    extra = actual_cols - template_cols





                    if missing:





                        results.append(('涓ラ噸', f'榛勯噾妯℃澘缂哄け鍒楋細{", ".join(sorted(missing))}'))





                    if extra:





                        results.append(('璀﹀憡', f'榛勯噾妯℃澘澶氬嚭鍒楋細{", ".join(sorted(extra))}'))





                    if not missing and not extra:





                        results.append(('閫氳繃', '榛勯噾妯℃澘鍒楃粨鏋勫畬鍏ㄥ尮閰?))





                else:





                    results.append(('璀﹀憡', '灏氭湭璁剧疆榛勯噾妯℃澘锛岃烦杩囧垪缁撴瀯瀵规瘮'))





            except Exception as e:





                results.append(('璀﹀憡', f'榛勯噾妯℃澘瀵规瘮澶辫触锛歿e}'))











            # 姹囨€诲埌鏃ュ織





            self.log("馃搵 鏁版嵁棰勬鎶ュ憡锛?, "info")





            for severity, msg in results:





                self.log(f" [{severity}] {msg}", severity if severity in ("閫氳繃", "璀﹀憡", "涓ラ噸") else "info")











            # 寮圭獥鏄剧ず





            self._show_pre_check_report(results)











        except Exception as e:





            self.log(f"棰勬澶辫触锛歿e}", "error")





            messagebox.showerror("棰勬閿欒", f"棰勬杩囩▼涓彂鐢熼敊璇細{str(e)}")











    def _load_golden_columns(self):





        """鍔犺浇榛勯噾妯℃澘鍒楀悕锛堜粠閰嶇疆鏂囦欢锛岃嫢涓嶅瓨鍦ㄨ繑鍥濶one锛?""





        try:





            gold_path = os.path.join(os.path.expanduser('~'), '.zpp011_audit', 'golden_columns.json')





            if not os.path.exists(gold_path):





                return None





            with open(gold_path, 'r', encoding='utf-8') as f:





                return json.load(f)





        except Exception:





            return None











    def _show_pre_check_report(self, results):
        """鏄剧ず棰勬鎶ュ憡锛堢郴缁熸鏌?+ 鏁版嵁缁熻锛夛紝閰嶇疆鍖栧垪鍚嶏紝闃插尽缂哄け鍒?""
        win = tk.Toplevel(self.root)
        win.title("鏁版嵁棰勬鎶ュ憡")
        win.geometry("600x500")
        win.transient(self.root)
        # 闈炴ā鎬侊細涓嶈皟鐢?grab_set()锛屼笉闃诲涓荤獥鍙?
        win.attributes('-topmost', True)  # 缃《鏄剧ず锛屼絾涓嶅己鍒朵氦浜?

        text = tk.Text(win, wrap="word", font=("Consolas", 10), padx=10, pady=10)
        text.pack(fill="both", expand=True)

        text.tag_configure("涓ラ噸", foreground="#cf222e")
        text.tag_configure("璀﹀憡", foreground="#d29922")
        text.tag_configure("閫氳繃", foreground="#1a7f37")
        text.tag_configure("鏍囬", font=("Consolas", 12, "bold"))

        # 1. 绯荤粺妫€鏌ョ粨鏋?
        text.insert("end", "绯荤粺妫€鏌ョ粨鏋淺n", "鏍囬")
        for level, msg in results:
            tag = "涓ラ噸" if level == "涓ラ噸" else ("璀﹀憡" if level == "璀﹀憡" else "閫氳繃")
            text.insert("end", msg + "\n", tag)

        # 2. 鏁版嵁缁熻淇℃伅锛堥厤缃寲鍒楀悕 + 闃插尽锛?
        if self.audit_data is not None and not self.audit_data.empty:
            text.insert("end", "\n鏁版嵁缁熻\n", "鏍囬")

            # 浠庨厤缃鍙栧垪鍚?
            bias_rate_col = self.config.get('columns.bias_rate', '鍋忓樊鐜?%)')
            audit_status_col = self.config.get('columns.audit_status', '瀹℃牳鐘舵€?)

            try:
                # 妫€鏌ュ繀瑕佸垪鏄惁瀛樺湪
                if bias_rate_col not in self.audit_data.columns:
                    text.insert("end", f"璀﹀憡锛氱己灏戝垪 '{bias_rate_col}'锛屾棤娉曠粺璁″亸宸垎甯冦€俓n", "璀﹀憡")
                    stats_available = False
                else:
                    stats_available = True
            except Exception as e:
                text.insert("end", f"缁熻璁＄畻鍑洪敊锛歿e}\n", "涓ラ噸")
                stats_available = False

            if stats_available:
                total = len(self.audit_data)
                high_dev = (self.audit_data[bias_rate_col].abs() > 10).sum()
                mid_dev = ((self.audit_data[bias_rate_col].abs() >= 5) & (self.audit_data[bias_rate_col].abs() <= 10)).sum()
                low_dev = (self.audit_data[bias_rate_col].abs() < 5).sum()

                # 瀹℃牳鐘舵€佸垪鍙兘涓嶅瓨鍦紝灏濊瘯鑾峰彇
                if audit_status_col in self.audit_data.columns:
                    reviewed = (self.audit_data[audit_status_col] == '宸插鏍?).sum()
                    unreviewed = total - reviewed
                else:
                    reviewed = unreviewed = 0
                    text.insert("end", "璀﹀憡锛氱己灏戝鏍哥姸鎬佸垪锛屽鏍哥粺璁℃樉绀轰负0銆俓n", "璀﹀憡")

                stats = f"鎬昏鏁帮細{total}\n"
                stats += f"鍋忓樊寮傚父锛堚墺10%锛夛細{high_dev}\n"
                stats += f"鍋忓樊鍏虫敞锛?%-10%锛夛細{mid_dev}\n"
                stats += f"鍋忓樊姝ｅ父锛?5%锛夛細{low_dev}\n"
                stats += f"宸插鏍革細{reviewed}\n"
                stats += f"鏈鏍革細{unreviewed}"
                text.insert("end", stats + "\n", "閫氳繃")
        else:
            text.insert("end", "\n鏁版嵁缁熻\n鏃犳暟鎹甛n", "鏍囬")

        text.configure(state="disabled")

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)
        if hasattr(self, '_duplicate_records') and hasattr(self._duplicate_records, 'empty') and not self._duplicate_records.empty:
            tk.Button(btn_frame, text="瀵煎嚭閲嶅鏁版嵁", command=self._export_duplicate_records,
                      bg="#f0f0f0", font=("Microsoft YaHei", 9)).pack(side="left", padx=5)
        tk.Button(btn_frame, text="鍏抽棴", command=win.destroy, width=10).pack(side="left", padx=5)
    def _export_duplicate_records(self):





        if not hasattr(self, '_duplicate_records') or not hasattr(self._duplicate_records, 'empty') or self._duplicate_records.empty:





            return





        file_path = filedialog.asksaveasfilename(





            defaultextension=".xlsx",





            filetypes=[("Excel 鏂囦欢", "*.xlsx")],





            initialfile="閲嶅璁㈠崟璁板綍.xlsx"





        )





        if not file_path:





            return





        try:





            from openpyxl import Workbook





            from openpyxl.styles import PatternFill





            df = self._duplicate_records.copy()





            if '娴佺▼璁㈠崟' in df.columns and '鐗╂枡缂栫爜' in df.columns:





                df = df.sort_values(['娴佺▼璁㈠崟', '鐗╂枡缂栫爜'])





            wb = Workbook()





            ws = wb.active





            ws.title = "閲嶅璁板綍"





            headers = list(df.columns)





            for col_idx, h in enumerate(headers, 1):





                ws.cell(1, col_idx, h)





            colors = ["FFCCCC", "FFE5CC", "FFFFCC", "CCFFCC", "CCE5FF", "E5CCFF", "FFCCE5", "E5E5E5"]





            group_colors = {}





            group_idx = 0





            last_group = None





            for row_idx, (_, row) in enumerate(df.iterrows(), 2):





                if '娴佺▼璁㈠崟' in df.columns and '鐗╂枡缂栫爜' in df.columns:





                    group_key = (row['娴佺▼璁㈠崟'], row['鐗╂枡缂栫爜'])





                    if group_key != last_group:





                        group_idx += 1





                        last_group = group_key





                        group_colors[group_key] = colors[(group_idx - 1) % len(colors)]





                    fill_color = group_colors.get(group_key, "FFFFFF")





                else:





                    fill_color = "FFFFFF"





                for col_idx, val in enumerate(row, 1):





                    cell = ws.cell(row_idx, col_idx, val)





                    if fill_color != "FFFFFF":





                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")





            wb.save(file_path)





            self.log(f"閲嶅鏁版嵁宸插鍑猴細{file_path}", "success")





            messagebox.showinfo("瀵煎嚭鎴愬姛", "宸插鍑哄埌锛? + file_path)





        except Exception as e:





            self.log(f"瀵煎嚭閲嶅鏁版嵁澶辫触锛歿e}", "error")





            messagebox.showerror("瀵煎嚭澶辫触", str(e))











    # ==================== 鑿滃崟鍒濆鍖?====================





    def _show_tree_view(self):





        """鍦ㄦ柊绐楀彛涓互鏍戝舰缁撴瀯灞曠ず瀹℃牳鏁版嵁"""





        if self.audit_data is None or self.audit_data.empty:





            messagebox.showwarning("鎻愮ず", "娌℃湁鍙睍绀虹殑鏁版嵁")





            return











        # 鍒涘缓鏂扮獥鍙?





        win = tk.Toplevel(self.root)





        win.title("鍋忓樊鏁版嵁 - 鏍戝舰瑙嗗浘")





        win.geometry("1000x650")





        win.transient(self.root)





        win.grab_set()











        # Treeview





        tree = ttk.Treeview(win, show="tree headings",





                            columns=("code", "name", "order_date", "quota", "actual", "dev_rate", "status", "remark"),





                            height=25)





        tree.heading("#0", text="宸ュ巶 / 杞﹂棿 / 鐗╂枡鍒嗙被")





        tree.heading("code", text="鐗╂枡鍙?)





        tree.heading("name", text="鐗╂枡鎻忚堪")





        tree.heading("order_date", text="璁㈠崟鏃ユ湡")





        tree.heading("quota", text="瀹氶")





        tree.heading("actual", text="瀹為檯")





        tree.heading("dev_rate", text="鍋忓樊鐜?")





        tree.heading("status", text="鐘舵€?)





        tree.heading("remark", text="澶囨敞")











        tree.column("#0", width=220)





        tree.column("code", width=70, anchor="center")





        tree.column("name", width=100, anchor="w")





        tree.column("order_date", width=70, anchor="center")





        tree.column("quota", width=50, anchor="e")





        tree.column("actual", width=50, anchor="e")





        tree.column("dev_rate", width=55, anchor="center")





        tree.column("status", width=55, anchor="center")





        tree.column("remark", width=80, anchor="w")











        scroll = ttk.Scrollbar(win, command=tree.yview)





        tree.configure(yscrollcommand=scroll.set)





        tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)





        scroll.pack(side="right", fill="y", pady=10)











        # 鐏垫椿纭畾鍒楀悕





        groupby_factory = next((c for c in ['宸ュ巶', '宸ュ巶鍚嶇О'] if c in self.audit_data.columns), None)





        groupby_workshop = next((c for c in ['杞﹂棿', '鐢熶骇绠＄悊鍛樻弿杩?] if c in self.audit_data.columns), None)





        groupby_category = next((c for c in ['鐗╂枡绫诲瀷', '鐗╂枡鍒嗙被'] if c in self.audit_data.columns), None)





        code_col = next((c for c in ['鐗╂枡缂栫爜', '缁勪欢鐗╂枡鍙?] if c in self.audit_data.columns), '鐗╂枡缂栫爜')





        name_col = next((c for c in ['鐗╂枡鍚嶇О', '缁勪欢鐗╂枡鎻忚堪'] if c in self.audit_data.columns), '鐗╂枡鍚嶇О')





        quota_col = next((c for c in ['瀹氶', '鏁伴噺-瀹氶'] if c in self.audit_data.columns), '瀹氶')





        actual_col = next((c for c in ['瀹為檯', '鏁伴噺-瀹為檯'] if c in self.audit_data.columns), '瀹為檯')





        dev_col = next((c for c in ['鍋忓樊鐜?, '鍋忓樊鐜?%)'] if c in self.audit_data.columns), '鍋忓樊鐜?)





        remark_col = next((c for c in ['澶囨敞', '澶囨敞鍘熷洜'] if c in self.audit_data.columns), '澶囨敞')











        if not groupby_factory:





            messagebox.showwarning("鎻愮ず", "鏁版嵁涓湭鎵惧埌宸ュ巶鍒楋紝鏃犳硶鐢熸垚鏍戝舰瑙嗗浘")





            win.destroy()





            return











        for factory, f_grp in self.audit_data.groupby(groupby_factory):





            factory_id = tree.insert('', 'end', text=factory, open=False)











            if groupby_workshop and groupby_category:





                for (workshop, mat_cat), w_grp in f_grp.groupby([groupby_workshop, groupby_category]):





                    workshop_id = tree.insert(factory_id, 'end',





                                              text=f"{workshop} - {mat_cat}  ({len(w_grp)}鏉?",





                                              open=False)





                    self._insert_tree_rows(tree, workshop_id, w_grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col)





            elif groupby_workshop:





                for workshop, w_grp in f_grp.groupby(groupby_workshop):





                    workshop_id = tree.insert(factory_id, 'end',





                                              text=f"{workshop}  ({len(w_grp)}鏉?",





                                              open=False)





                    self._insert_tree_rows(tree, workshop_id, w_grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col)





            else:





                for _, row in f_grp.iterrows():





                    self._insert_tree_rows(tree, factory_id, f_grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col)





                break











        self.log("[OK] tree view opened", "info")











    def _insert_tree_rows(self, tree, parent_id, grp, code_col, name_col, quota_col, actual_col, dev_col, remark_col):





        """鍚戞爲褰㈣鍥炬彃鍏ヨ鏁版嵁"""





        for _, row in grp.iterrows():





            tree.insert(parent_id, 'end',





                        text=str(row.get(name_col, ''))[:20],





                        values=(





                            str(row.get(code_col, '')),





                            str(row.get(name_col, ''))[:20],





                            str(row.get('璁㈠崟鏃ユ湡', ''))[:12] if pd.notna(row.get('璁㈠崟鏃ユ湡')) else '',





                            f"{row.get(quota_col, 0):.3f}" if pd.notna(row.get(quota_col)) else '-',





                            f"{row.get(actual_col, 0):.3f}" if pd.notna(row.get(actual_col)) else '-',





                            f"{row.get(dev_col, 0):.2f}%" if pd.notna(row.get(dev_col)) else '-',





                            row.get('_audit_status', '鏈鏍?),





                            str(row.get(remark_col, ''))[:20]





                        ))











    # 鈹€鈹€ 浠ヤ笅涓?v30 鍘熸湁鐨勫叏閮ㄦ柟娉曪紝蹇呴』瀹屾暣澶嶅埗鍒版澶?鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€





    # 鏂规硶鍒楄〃锛堣浠庡師 v30.py 涓€愪釜澶嶅埗锛岀‘淇濅笉閬楁紡锛夛細





    # _refresh_alt_view, _add_alt, _del_alt, _reset_alt,





    # _get_config_path, _save_alt_pairs, _load_alt_pairs, _extract_alt_name,





    # _select_input, _select_output, _auto_find, _preview, log, _set_step,





    # start_analysis, _analysis_thread, _on_progress, _update_timer, request_cancel,





    # _on_done, _on_cancel, _on_error, _show_step_log, open_output,





    # generate_excel_direct, _generate_excel_thread, generate_ppt, _on_ppt_done, _on_ppt_error,





    # _load_audit_data, _apply_row_colors, _edit_audit_remark, _update_audit_stats,





    # _show_audit_context_menu, _filter_audit_tree, _update_filter_options,





    # _on_filter_changed, _reset_all_filters, _refresh_audit_tree,





    # _run_ai_audit, _export_audit_excel, _batch_change_status, _batch_fill_remark,





    # _batch_export, _add_custom_status, _load_custom_statuses, _save_custom_statuses





    # 娉細涓婇潰杩欎簺鏂规硶鐨勫畬鏁翠唬鐮佽浠?v30.py 涓叏閫夊鍒讹紝鎻掑叆鍒版澶勩€?





    # 澶嶅埗鏃舵敞鎰忎繚鎸佺缉杩涳紙鍏ㄩ儴涓?_show_tree_view 瀵归綈锛宑lass 鍐呴儴缂╄繘 4 绌烘牸锛夈€?











    # 鈹€鈹€ 浠ヤ笅鏂规硶浠?v30.py 瀹屾暣绉绘 鈹€鈹€











    def _get_resume_state_path(self):





        """鑾峰彇鏂偣鐘舵€佹枃浠惰矾寰?""





        app_dir = os.path.join(os.path.expanduser('~'), '.zpp011_audit')





        return os.path.join(app_dir, 'resume_state.json')











    def _load_resume_state(self):





        """鍔犺浇鏂偣鐘舵€?""





        path = self._get_resume_state_path()





        if not os.path.exists(path):





            return None





        try:





            with open(path, 'r', encoding='utf-8') as f:





                return json.load(f)





        except Exception:





            return None











    def _save_resume_state(self):





        """淇濆瓨鏂偣鐘舵€?""





        app_dir = os.path.join(os.path.expanduser('~'), '.zpp011_audit')





        os.makedirs(app_dir, exist_ok=True)





        path = self._get_resume_state_path()





        





        state = {}





        # 淇濆瓨閫夋嫨鐨勮





        if hasattr(self, 'current_row_idx'):





            state['selected_row'] = self.current_row_idx





        # 淇濆瓨鎼滅储鏂囧瓧





        if hasattr(self, 'search_var'):





            state['search_text'] = self.search_var.get()





        # 淇濆瓨绛涢€夋潯浠?





        if hasattr(self, 'filter_widgets'):





            state['filter_values'] = {}





        





        try:





            with open(path, 'w', encoding='utf-8') as f:





                json.dump(state, f, ensure_ascii=False, indent=2)





        except Exception as e:





            self.log(f"淇濆瓨鏂偣澶辫触锛歿e}", "warn")

















    def _do_resume_state(self):





        """涓€閿仮澶嶄笂娆″鏍哥姸鎬?""





        state = self._load_resume_state()





        if not state:





            return





        # 鎭㈠鎼滅储鏂囧瓧





        search_text = state.get('search_text', '')





        if search_text and hasattr(self, 'search_var'):





            self.search_var.set(search_text)





        # 鎭㈠绛涢€夋潯浠?





        filters = state.get('filter_values', {})





        for key, value in filters.items():





            if key in self.filter_widgets:





                widget = self.filter_widgets[key]





                if key == 'name':





                    widget.delete(0, tk.END)





                    widget.insert(0, value)





                elif key == 'order_date':





                    if isinstance(widget, tuple) and len(widget) == 2 and isinstance(value, list) and len(value) == 2:





                        widget[0].delete(0, tk.END)





                        widget[0].insert(0, value[0])





                        widget[1].delete(0, tk.END)





                        widget[1].insert(0, value[1])





                else:





                    widget.set(value)





        # 閲嶆柊瑙﹀彂绛涢€?





        self._on_filter_changed('restore')





        # 鎭㈠閫変腑琛?





        saved_row = state.get('selected_row')





        if saved_row:





            children = self.audit_tree.get_children()





            if int(saved_row) <= len(children):





                self.audit_tree.selection_set(children[int(saved_row) - 1])





                self.audit_tree.see(children[int(saved_row) - 1])





        self.resume_btn.configure(state="disabled")





        self.log("鉁?宸叉仮澶嶄笂娆″鏍歌繘搴?, "success")

















    def _select_input(self):





        default_dir = os.path.join(os.path.expanduser("~"), "ZPP011瀵煎嚭鏂囦欢鍘熸暟鎹?)





        os.makedirs(default_dir, exist_ok=True)





        p = filedialog.askopenfilename(





            title="閫夋嫨 ZPP011 鏁版嵁鏂囦欢",





            filetypes=[("Excel鏂囦欢", "*.xlsx"), ("鎵€鏈夋枃浠?, "*.*")],





            initialdir=default_dir)











        if p:





            self.input_file.set(p)











            self._preview()

















    def _select_output(self):





        p = filedialog.askdirectory(title="閫夋嫨杈撳嚭鐩綍")











        if p:





            self.output_dir.set(p)

















    def _auto_find(self):





        for pattern in ['zpp011 *.xlsx', 'ZPP011 *.xlsx']:





            hits = _glob.glob(os.path.join(os.getcwd(), pattern))











            if hits:





                hits.sort(key=os.path.getmtime, reverse=True)











                self.input_file.set(hits[0])











                self._preview()











                return











            hits = _glob.glob(





                os.path.join(





                    os.path.expanduser('~'),





                    'Desktop',





                    pattern))











            if hits:





                hits.sort(key=os.path.getmtime, reverse=True)











                self.input_file.set(hits[0])











                self._preview()











                return

















    def _preview(self):





        path = self.input_file.get()











        if not path or not os.path.exists(path):





            self.preview_lbl.configure(text="鉂?鏂囦欢涓嶅瓨鍦?, fg=C['danger'])











            return











        try:





            df = pd.read_excel(path, sheet_name='Data', nrows=5)











            total = pd.read_excel(path, sheet_name='Data').shape[0]











            cols = df.columns.tolist()











            self.preview_lbl.configure(











                text=f"鉁?{os.path.basename(path)}\n"











                f"   鎬昏鏁帮細{total:,}  琛孿n"











                f"   鏇夸唬鏂欓厤瀵癸細{len(self.alt_pairs)}  缁刓n"











                f"   鍒楁暟锛歿len(cols)}",











                fg=C['green'], justify="left"











            )











        except Exception as e:





            self.preview_lbl.configure(text=f"鉂?璇诲彇澶辫触锛歿e}", fg=C['danger'])











        # 鍒锋柊鐗╂枡鍒楄〃锛堜緵鏇夸唬鏂欐坊鍔犲璇濇涓嬫媺浣跨敤锛?





        self._load_material_list()

















    def _set_step(self, idx, done=True):





        fr_data = self.step_frames.get(idx)











        if not fr_data:





            return











        fr = fr_data['frame']











        lbl = fr.winfo_children()[0]











        color = C['green'] if done else (





            C['accent'] if idx == self._current_step else C['text_dim'])











        lbl.configure(bg=color, fg="white" if done else C['text_dim'])











        fr.configure(bg=color, relief="flat" if done else "flat")

















    def _show_step_log(self, idx):





        pass











    def _load_lock_state(self):





        return {}











    def _save_lock_state(self, locked):





        """淇濆瓨鍒楀閿佸畾鐘舵€?""





        try:





            from gui.ui_builder import COLUMN_WIDTHS_FILE





            state_path = COLUMN_WIDTHS_FILE.replace('column_widths.json', 'lock_state.json')





            os.makedirs(os.path.dirname(state_path), exist_ok=True)





            with open(state_path, 'w', encoding='utf-8') as f:





                json.dump({'locked': locked}, f)





        except Exception:





            pass











    def _on_close(self):





        """绐楀彛鍏抽棴鍓嶄繚瀛樻柇鐐圭姸鎬?""





        try:





            self._save_resume_state()





            self.log("鏂偣宸蹭繚瀛?, "success")





        except Exception as e:





            self.log(f"鏂偣淇濆瓨澶辫触锛歿e}", "warn")





        # 淇濆瓨鍒楀 + 绐楀彛鍑犱綍





        try:





            if hasattr(self, 'audit_tree'):





                widths = {}





                for col in self.audit_tree['columns']:





                    widths[col] = self.audit_tree.column(col, 'width')





                self.config.set('table.column_widths', widths)





            self.config.save_window_geometry(self.root)





        except Exception:





            pass





        self.root.destroy()

















    def update_progress(self, percent: int, message: str = ''):





        """鏇存柊杩涘害鏉★紙渚?Presenter 璋冪敤锛?""





        self.root.after(0, lambda: [





            self.progress_var.set(percent),





            self.progress_lbl.configure(text=message) if hasattr(self, 'progress_lbl') else None





        ])











    def refresh_treeview(self, df=None):





        """鍒锋柊 Treeview锛堜緵 Presenter 璋冪敤锛?""





        if df is not None:





            self.audit_data = df





        self.root.after(0, lambda: self._refresh_audit_tree(self.audit_data) if hasattr(self, '_refresh_audit_tree') else None)











    def show_error(self, msg: str, title: str = '閿欒'):





        """鏄剧ず閿欒瀵硅瘽妗嗭紙渚?Presenter 璋冪敤锛?""





        from tkinter import messagebox





        self.root.after(0, lambda: messagebox.showerror(title, msg))











    def show_info(self, msg: str, title: str = '鎻愮ず'):





        """鏄剧ず淇℃伅瀵硅瘽妗嗭紙渚?Presenter 璋冪敤锛?""





        from tkinter import messagebox





        self.root.after(0, lambda: messagebox.showinfo(title, msg))











    def log_message(self, msg: str, level: str = 'info'):





        """璁板綍鏃ュ織锛堜緵 Presenter 璋冪敤锛?""





        if hasattr(self, 'log'):





            self.log(msg, level)











    def set_buttons_state(self, state: str = 'normal', buttons: list = None):





        """璁剧疆鎸夐挳鐘舵€侊紙渚?Presenter 璋冪敤锛?""





        if buttons is None:





            buttons = ['run_btn', 'save_btn', 'export_btn', 'ai_btn', 'auto_close_btn']





        for btn_name in buttons:





            btn = getattr(self, btn_name, None)





            if btn:





                try:





                    btn.configure(state=state)





                except:





                    pass











    def enable_buttons(self, buttons: list = None):





        """鍚敤鎸夐挳"""





        self.set_buttons_state('normal', buttons)











    def disable_buttons(self, buttons: list = None):





        """绂佺敤鎸夐挳"""





        self.set_buttons_state('disabled', buttons)











    def get_audit_data(self):





        """杩斿洖褰撳墠瀹℃牳鏁版嵁 DataFrame锛堜緵 Presenter 璋冪敤锛?""





        return self.audit_data











    def get_output_path(self):





        """杩斿洖褰撳墠杈撳嚭璺緞锛堜緵 Presenter 璋冪敤锛?""





        return self.output_dir.get() if hasattr(self, 'output_dir') else ''











    def get_alt_pairs(self):





        """杩斿洖鏇夸唬鏂欓厤瀵瑰垪琛紙渚?Presenter 璋冪敤锛?""





        return getattr(self, 'alt_pairs', [])

