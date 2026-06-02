# -*- coding: utf-8 -*-
"""æ°æ®åæãå è½½ãè¿åº¦æ§å¶äºä»¶"""
import tkinter as tk
from tkinter import messagebox
import os
import threading
import time
import pandas as pd
from widgets import C
from analysis.analyzer import do_analysis_v2
from openpyxl import load_workbook
import traceback
import calendar
import json
import sys
import re


def _safe_for_gbk(text: str) -> str:
    """移除 GBK 无法编码的字符（emoji 等），用于 Windows messagebox"""
    if not text:
        return text
    return "".join(c for c in text if ord(c) <= 0xFF)
from core.backup_manager import BackupManager


class AnalysisEvents:
    """æ°æ®åæãå è½½ãè¿åº¦æ§å¶äºä»¶"""

    # ââ è¿åº¦é¶æ®µå¸¸éï¼Task 2ï¼è¿åº¦ç»åï¼ââââââââââââââââââââââ
    STAGES = [
        "1/5 æ­£å¨è¯»å Excel æä»¶",
        "2/5 æ­£å¨è§£æçäº§æ°æ®",
        "3/5 æ­£å¨è®¡ç®åå·®éé¢ååå·®ç",
        "4/5 æ­£å¨å¹éæ¿ä»£æä¿¡æ¯",
        "5/5 æ­£å¨çæå®¡æ ¸è¡¨æ ¼",
    ]
    _ANALYSIS_TIMEOUT_SEC = 300  # 5 åéè¶æ¶çæ­

    def start_analysis(self):
        # ââ Task 1ï¼é²éå¤ç¹å»ï¼Lock æ¿ä»£å¸å°ï¼ââââââââââââââââââââ
        if not hasattr(self, 'analysis_lock'):
            self.analysis_lock = threading.Lock()
        if not self.analysis_lock.acquire(blocking=False):
            self.log("â  åæä»»å¡è¿è¡ä¸­ï¼è¯·å¿éå¤ç¹å»", "warn")
            return
        try:
            if self.running:
                self.analysis_lock.release()
                self.run_btn.configure(state="normal", text="â¶  å¼å§åæ")
                return
            path = self.input_file.get()
            if not path or not os.path.exists(path):
                messagebox.showerror(_safe_for_gbk("éè¯¯"), _safe_for_gbk("è¯·åéæ©ææçè¾å¥æä»¶ï¼"))
                self.analysis_lock.release()
                self.run_btn.configure(state="normal", text="â¶  å¼å§åæ")
                return
            self.running = True
            self.cancel_req = False
            self.run_btn.configure(state="disabled", text="â³ åæä¸­...")
            self.cancel_btn.configure(state="normal")
            self.open_btn.configure(state="disabled")
            self.status_lbl.configure(text="åæä¸­...", fg=C['warn'])
            self.log_text.configure(state="normal")
            self.log_text.delete("1.0", "end")
            self.log_text.configure(state="disabled")
            self.progress_var.set(0)
            self._current_step = -1
            for i in self.step_frames:
                self._set_step(i, False)
            self.start_time = time.time()
            self._update_timer()

            # ââ Task 003ï¼åæåå¼æ­¥å¤ä»½ ââ
            _backup_mgr = BackupManager()
            _backup_mgr.backup_before_analysis_async(
                input_excel_path=path,
                progress_callback=None,
                done_callback=lambda meta, err: self.root.after(
                    0, lambda: self._on_backup_done(meta, err)
                ),
            )

            t = threading.Thread(target=self._analysis_thread, daemon=True)
            t.start()  # å¯å¨çº¿ç¨
        except Exception as e:
            import traceback
            traceback.print_exc()
            from tkinter import messagebox
            messagebox.showerror(_safe_for_gbk("éè¯¯"), _safe_for_gbk(f"åæå¯å¨å¤±è´¥: {e}"))
            self.running = False
            try:
                self.run_btn.configure(state="normal", text="â¶  å¼å§åæ")
            except:
                pass
            finally:
                self.analysis_lock.release()

    def _on_backup_done(self, meta, error):
        """å¼æ­¥å¤ä»½å®æåè°ï¼ä¸»çº¿ç¨å®å¨ï¼"""
        if error:
            self.log(f"â  åæåå¤ä»½å¤±è´¥ï¼{error}", "warn")
        else:
            self.log("â åæåå¤ä»½å®æ", "info")

    def _analysis_thread(self):
        """åæçº¿ç¨ï¼å§æç» Presenterï¼"""
        try:
            self.output_path = self.audit_presenter.start_analysis(
                input_file=self.input_file.get(),
                alt_pairs=self.alt_pairs,
                start_date=self.start_date.get() or None,
                end_date=self.end_date.get() or None,
                material_search=self.material_search.get() or None,
                progress_callback=self._on_progress,
                cancel_check=lambda: self.cancel_req,
            )
            self.root.after(self.config.get(
                'analysis.cleanup_delay_ms', 0), self._on_done)
        except KeyboardInterrupt:
            self.root.after(self.config.get(
                'analysis.cleanup_delay_ms', 0), self._on_cancel)
        except Exception as e:
            self.root.after(self.config.get(
                'analysis.cleanup_delay_ms', 0), lambda e=e: self._on_error(str(e)))
            import traceback
            traceback.print_exc()
        except BaseException as e:
            self.root.after(self.config.get('analysis.cleanup_delay_ms', 0),
                            lambda e=e: self._on_error(f"BaseException: {e}"))

    def _on_progress(self, step_idx, step_name, percent):
        if self.cancel_req:
            return
        # ââ Task 2ï¼è¶æ¶çæ­ï¼5åéï¼ââââââââââââââââââââââââââââ
        if hasattr(self, 'start_time') and self.start_time:
            if time.time() - self.start_time > self._ANALYSIS_TIMEOUT_SEC:
                self.log("â  åæè¶æ¶ï¼>5åéï¼ï¼è¯·èç³»è£´å¥", "warn")
                self.status_lbl.configure(text="åæè¶æ¶ï¼è¯·èç³»è£´å¥", fg=C['danger'])
                self.cancel_req = True
                return
        self._current_step = step_idx
        now = time.strftime("%H:%M:%S")
        total_stages = len(self.STAGES)

        def update():
            self.progress_var.set(percent)
            self.progress_lbl.configure(
                text=f"{step_name}  {percent:.0f}%", fg=C['accent'])
            self._set_step(step_idx)
            self.log(
                f"[{now}] [{step_idx + 1}/{total_stages}] {step_name} {percent:.0f}%", "step")
            self.root.update_idletasks()
        self.root.after(self.config.get(
            'analysis.cleanup_delay_ms', 0), update)

    def _update_timer(self):
        if not self.running or not self.start_time:
            return
        elapsed = int(time.time() - self.start_time)
        m, s = elapsed // 60, elapsed % 60
        prog = self.progress_var.get()
        if prog > 5:
            total = int(elapsed / (prog / 100))
            rem = max(0, total - elapsed)
            rm, rs = rem // 60, rem % 60
            self.timer_lbl.configure(
                text=f"â± {m:02d}:{s:02d}  |  å©ä½ ~{rm:02d}:{rs:02d}",
                fg=C['warn'])
        else:
            self.timer_lbl.configure(
                text=f"â± {m:02d}:{s:02d}",
                fg=C['text_dim'])
        self.timer_id = self.root.after(self.config.get(
            'analysis.cleanup_delay_ms', 1000), self._update_timer)

    def request_cancel(self):
        if self.running:
            self.cancel_req = True
            self.cancel_btn.configure(state="disabled", text="åæ¶ä¸­...")
            self.log("â  ç¨æ·è¯·æ±åæ¶...", "warn")

    def _on_done(self):
        self.running = False
        # ââ Task 1ï¼éæ¾é + æ¥å¿ ââââââââââââââââââââââââââââââ
        if hasattr(self, 'analysis_lock') and self.analysis_lock.locked():
            self.analysis_lock.release()
        self.log("åæä»»å¡ç»æï¼æé®å·²æ¢å¤", "info")
        self.log(
            f"ð _on_done: output_path = {getattr(self, 'output_path', 'NOT_SET')}", "info")
        # å¦æ output_path ä¸º Noneï¼å°è¯èªå¨æ¥æ¾ææ°çè¾åºæä»¶
        if not self.output_path:
            try:
                import glob as _glob
                out_dir = self.output_dir.get()
                if not out_dir or not os.path.isdir(out_dir):
                    out_dir = os.path.dirname(self.input_file.get()) if self.input_file.get(
                    ) else os.path.expanduser('~/Desktop')
                pattern = os.path.join(out_dir, 'ZPP011åå·®åææç»ç_*.xlsx')
                files = _glob.glob(pattern)
                if files:
                    latest_file = max(files, key=os.path.getmtime)
                    self.output_path = latest_file
                    self.log(
                        f"ð èªå¨æ¾å°è¾åºæä»¶ï¼{os.path.basename(latest_file)}", "info")
            except Exception as _e:
                self.log(f"ð èªå¨æ¥æ¾è¾åºæä»¶å¤±è´¥ï¼{_e}", "warn")
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
        total = int(time.time() - self.start_time)
        m, s = total // 60, total % 60
        self.timer_lbl.configure(text=f"â {m:02d}:{s:02d} å®æ", fg=C['green'])
        self.run_btn.configure(state="normal", text="â¶ éæ°åæ")
        self.cancel_btn.configure(state="disabled")
        self.open_btn.configure(state="normal")
        self.progress_var.set(100)
        self.progress_lbl.configure(text="â å®æ (100%)", fg=C['green'])
        _basename = os.path.basename(
            self.output_path) if self.output_path else 'æªç¥æä»¶'
        self.status_lbl.configure(
            text=f"å®æ â {_basename}",
            fg=C['green'])
        for i in self.step_frames:
            self._set_step(i, True)
        self.log(f"\nâ åæå®æï¼æ»ç¨æ¶ {m:02d}:{s:02d}", "success")
        self.status_lbl.configure(text="å®æ â æ°æ®å·²å°±ç»ªï¼å¯å è½½å®¡æ ¸", fg=C['green'])
        # â æ°å¢ï¼èªå¨å è½½å®¡æ ¸æ°æ®ï¼èå¨ç»è®¡å¡çï¼ä»åæç»ææä»¶å è½½ï¼
        self._analysis_output_path = self.output_path  # ä¿å­è·¯å¾ä¾å è½½ä½¿ç¨
        self.log("ð åå¤å è½½å®¡æ ¸æ°æ®...", "info")
        self.root.after(self.config.get('analysis.cleanup_delay_ms', 100),
                        lambda: self._load_audit_data_from_output(self._analysis_output_path))
        self.root.after(self.config.get(
            'analysis.cleanup_delay_ms', 200), self._run_pre_check)
        # ä¸åèªå¨å é¤ä¸´æ¶æä»¶ï¼æ¹ä¸ºå¨å è½½æååå é¤ï¼è§ _on_load_doneï¼
        # â å¯ç¨"å è½½å®¡æ ¸æ°æ®"æé®ï¼è®©ç¨æ·æå¨å è½½
        if hasattr(self, 'load_audit_btn'):
            self.load_audit_btn.configure(state="normal")
            self.log("â å·²å¯ç¨ãå è½½å®¡æ ¸æ°æ®ãæé®", "info")

        # ââ Task 003ï¼åææåï¼æ¸é¤æ¢å¤æ è®° ââ
        try:
            _backup_mgr = BackupManager()
            _backup_mgr._clear_recovery_flag()
        except Exception:
            pass

    def _cleanup_auto_excel(self):
        """å é¤åæèªå¨çæçExcelæä»¶ï¼ç¨æ·éè¦æ¶å¯æå¨çæ"""
        try:
            if self._analysis_output_path and os.path.exists(self._analysis_output_path):
                fname = os.path.basename(self._analysis_output_path)
                os.remove(self._analysis_output_path)
                self.log(f"ðï¸ å·²æ¸çèªå¨çææä»¶: {fname}ï¼éè¦æ¶å¯ç¹å»ãçæExcelãæé®ï¼", "info")
                self._analysis_output_path = None
        except Exception as e:
            self.log(f"æ¸çæä»¶å¤±è´¥: {e}", "warn")

    def _on_cancel(self):
        self.running = False
        # ââ Task 1ï¼éæ¾é + æ¥å¿ ââââââââââââââââââââââââââââââ
        if hasattr(self, 'analysis_lock') and self.analysis_lock.locked():
            self.analysis_lock.release()
        self.log("åæä»»å¡å·²åæ¶ï¼æé®å·²æ¢å¤", "info")
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
        self.run_btn.configure(state="normal", text="â¶ å¼å§åæ")
        self.cancel_btn.configure(state="disabled")
        self.status_lbl.configure(text="å·²åæ¶", fg=C['warn'])
        self.log("\nâ  åæå·²åæ¶", "warn")

    def _on_error(self, msg):
        self.running = False
        # ââ Task 1ï¼éæ¾é ââââââââââââââââââââââââââââââââââââââ
        if hasattr(self, 'analysis_lock') and self.analysis_lock.locked():
            self.analysis_lock.release()
        # ââ Task 3ï¼éè¯¯åå¥½å ââââââââââââââââââââââââââââââââââ
        friendly_msg = self._get_error_message(msg)
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
        self.run_btn.configure(state="normal", text="â¶ å¼å§åæ")
        self.cancel_btn.configure(state="disabled")
        self.status_lbl.configure(text="åºé", fg=C['danger'])
        self.log(f"\nâ éè¯¯ï¼{friendly_msg}", "error")
        messagebox.showerror(_safe_for_gbk("åæåºé"), _safe_for_gbk(friendly_msg))

    # ââ Task 3ï¼éè¯¯åå¥½å âââââââââââââââââââââââââââââââââââââ
    _ERROR_MESSAGES = None  # ç¼å­

    def _get_error_message(self, raw_msg: str) -> str:
        """å°åå§éè¯¯ä¿¡æ¯æ å°ä¸ºç¨æ·åå¥½æç¤ºï¼æ¯æ JSON éç½®"""
        if self._ERROR_MESSAGES is None:
            self._load_error_messages()
        for error_type, info in self._ERROR_MESSAGES.items():
            if error_type == 'default':
                continue
            # ç¨å¼å¸¸ç±»åä½ä¸ºå³é®è¯å¹é
            if error_type.lower() in raw_msg.lower():
                reason = info.get('reason', '')
                solution = info.get('solution', '')
                return f"{reason}\næç¤º: {solution}" if solution else reason
        default = self._ERROR_MESSAGES.get('default', {})
        return default.get('reason', raw_msg)

    @classmethod
    def _load_error_messages(cls):
        """å è½½ error_messages.jsonï¼å¼å®¹å¼åä¸ PyInstaller æåç¯å¢ï¼"""
        try:
            if getattr(sys, 'frozen', False):
                base = os.path.dirname(sys.executable)
            else:
                base = os.path.dirname(os.path.dirname(
                    os.path.dirname(os.path.abspath(__file__))))
            json_path = os.path.join(base, 'config', 'error_messages.json')
            with open(json_path, 'r', encoding='utf-8') as f:
                cls._ERROR_MESSAGES = json.load(f)
        except Exception:
            cls._ERROR_MESSAGES = {
                'default': {'reason': 'åæè¿ç¨åºéï¼è¯·æ£æ¥æä»¶æ ¼å¼åéè¯ã',
                            'solution': 'å¦é®é¢æç»­è¯·èç³»è£´å¥ã'}
            }

    def _load_audit_data(self):
        path = self.input_file.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_safe_for_gbk("æç¤º"), _safe_for_gbk("è¯·åéæ©è¾å¥æä»¶"))
            return
        try:
            df = pd.read_excel(path, sheet_name='Data')
            df['åå·®ç(%)'] = pd.to_numeric(
                df['åå·®ç(%)'], errors='coerce').fillna(0)
            # ç¨ openpyxl è¯»åçå® Excel è¡å·ï¼é¿å pandas è·³è¿ç©ºè¡å¯¼è´åç§»ï¼
            try:
                from openpyxl import load_workbook
                _wb2 = load_workbook(path, read_only=True, data_only=True)
                _ws2 = _wb2['Data']
                _real_rows2 = []
                _rn2 = 0
                for _row2 in _ws2:
                    _rn2 += 1
                    if _rn2 == 1:
                        continue
                    _real_rows2.append(_rn2)
                _wb2.close()
                if len(_real_rows2) == len(df):
                    df['excel_row'] = _real_rows2
                else:
                    df['excel_row'] = df.index + 2
            except Exception:
                df['excel_row'] = df.index + 2
            # æ¥æèå´è¿æ»¤ï¼å¦æç¨æ·è¾å¥äºæ¥æï¼
            start_date_str = self.start_date.get().strip()
            end_date_str = self.end_date.get().strip()
            if start_date_str or end_date_str:
                if 'è®¢åå¼å§æ¥æ' in df.columns:
                    df['è®¢åå¼å§æ¥æ'] = pd.to_datetime(
                        df['è®¢åå¼å§æ¥æ'], errors='coerce')
                    if start_date_str:
                        try:
                            start_dt = pd.to_datetime(start_date_str)
                            df = df[df['è®¢åå¼å§æ¥æ'] >= start_dt]
                            self.log(f"å·²æå¼å§æ¥æ {start_date_str} è¿æ»¤", "info")
                        except:
                            pass
                    if end_date_str:
                        try:
                            end_dt = pd.to_datetime(end_date_str)
                            df = df[df['è®¢åå¼å§æ¥æ'] <= end_dt]
                            self.log(f"å·²æç»ææ¥æ {end_date_str} è¿æ»¤", "info")
                        except:
                            pass
            total = len(df)
            high_dev = df[df['åå·®ç(%)'].abs() > self.config.get(
                'audit.high_deviation_threshold', 10)]
            need_note = high_dev[high_dev['å¤æ³¨åå '].isna()]
            ok_note = high_dev[high_dev['å¤æ³¨åå '].notna()]
            # ââ P1#11ï¼è®¡ç®åå·®éé¢ âââââââââââââââââââââââââââââ
            df['æ°é-å®é¢'] = pd.to_numeric(df['æ°é-å®é¢'], errors='coerce').fillna(0)
            df['æ°é-å®é'] = pd.to_numeric(df['æ°é-å®é'], errors='coerce').fillna(0)
            df['éé¢-å®é(å«ç¨)'] = pd.to_numeric(df['éé¢-å®é(å«ç¨)'],
                                            errors='coerce').fillna(0)
            # è®¡ç®å«ç¨åä»·ï¼é¿å¼é¤ä»¥0ï¼
            df['_unit_price'] = 0.0
            mask = df['æ°é-å®é'] != 0
            df.loc[mask, '_unit_price'] = df.loc[mask,
                                                 'éé¢-å®é(å«ç¨)'] / df.loc[mask, 'æ°é-å®é']
            # åå·®éé¢ = (å®é - å®é¢) Ã å«ç¨åä»·
            df['åå·®éé¢'] = ((df['æ°é-å®é'] - df['æ°é-å®é¢'])
                          * df['_unit_price']).round(2)
            # _unit_price ä¿çï¼ä¾ææ¬æ¢ç®å¨ä½¿ç¨
            # æ´æ°ç»è®¡å¡ç
            self.audit_stat_labels['total'].configure(text=str(total))
            self.audit_stat_labels['high_dev'].configure(
                text=str(len(high_dev)))
            self.audit_stat_labels['need_note'].configure(
                text=str(len(need_note)))
            self.audit_stat_labels['ok_note'].configure(text=str(len(ok_note)))
            self.audit_data = high_dev.copy()
            # æ·»å ç©æå¤§ç±»åï¼å§ç»ç¨ç©æç¼ç åç¼è®¡ç®ï¼ç¡®ä¿å¼åä¸æéé¡¹ä¸è´ï¼
            mat_cat_map = {
                "100": "åè¾æ", "200": "åæ", "400": "é£åè¾æ/é£ååæå",
                "410": "é¥®æè¾æ/é¥®æåæå", "500": "é£åæå", "510": "é¥®ææå",
                "600": "ä¿éå"
            }
            self.audit_data['material_category'] = self.audit_data['ç©æç¼ç '].apply(
                lambda x: mat_cat_map.get(
                    str(x)[:3], str(x)[:3]) if pd.notna(x) else ''
            )
            # ä¿å­å¨éæ°æ®å¯æ¬ï¼å¨æ·»å  material_category åä¹åï¼ç¡®ä¿åå­å¨ï¼
            self.full_audit_data = self.audit_data.copy()
            # å¦ææ°æ®ä¸­ãç©æç±»åãåææ´ç²¾ç¡®çåç±»ï¼å¯ä»¥è¦çï¼å¯éï¼
            # ç®åä¸æä¹±é»è¾ï¼ä¿æç¨ç©æç¼ç åç¼è®¡ç®çç»æ
            print(f'[DEBUG audit_data] å: {list(self.audit_data.columns)}')
            print(
                f'[DEBUG material_category] å¼åå¸: {self.audit_data["material_category"].value_counts().to_dict()}')
            # æ å°è®¢åæ¥æåï¼Data sheet ç¨"è®¢åå¼å§æ¥æ"ï¼ç»ä¸ä¸º"è®¢åæ¥æ"ï¼
            if 'è®¢åå¼å§æ¥æ' in self.audit_data.columns:
                self.audit_data['è®¢åæ¥æ'] = pd.to_datetime(
                    self.audit_data['è®¢åå¼å§æ¥æ'], errors='coerce').dt.strftime('%Y-%m-%d')
            elif 'è®¢åæ¥æ' not in self.audit_data.columns:
                self.audit_data['è®¢åæ¥æ'] = ''
            self._update_filter_options()
            self._refresh_audit_tree(self.audit_data)
            self._update_audit_stats(self.audit_data)
            self._update_summary_row(self.audit_data)  # åè®¡è¡æ´æ°
            self.audit_tree.tag_configure(
                'need_note', background='#fff0e0', foreground='#b04000')
            self.audit_tree.tag_configure(
                'ok_note',   background='#e8f5e9', foreground='#1a6b1a')
            self.audit_tree.tag_configure(
                'ai_gen',    background='#fce4ec', foreground='#880e4f')
            self.audit_ai_btn.configure(state="normal")
            self.audit_export_btn.configure(state="normal")
            # åæ¶å¯ç¨åå·®æç»è¡¨çæé®
            if hasattr(self, 'unified_ai_btn'):
                self.unified_ai_btn.configure(state="normal")
            if hasattr(self, 'unified_export_btn'):
                self.unified_export_btn.configure(state="normal")
            if hasattr(self, 'cleanup_btn'):
                self.cleanup_btn.configure(state="normal")
            if hasattr(self, 'save_audit_btn'):
                self.save_audit_btn.configure(state="normal")
            self._apply_row_colors()
            # P0-B4ï¼æ¢å¤å®¡æ ¸è®°å½ï¼ä»æ°æ®åºï¼
            self.audit_presenter.load_audit_data(self.audit_data)
            self.log(
                f"æºè½å®¡æ ¸ï¼å è½½å®æ | å±{total}æ¡ | åå·®>10%å±{len(high_dev)}æ¡ | éè¡¥å¤æ³¨{len(need_note)}æ¡", "success")
        except Exception as e:
            messagebox.showerror(_safe_for_gbk("éè¯¯"), _safe_for_gbk(f"å è½½æ°æ®å¤±è´¥ï¼{e}"))
            self.log(f"å è½½å®¡æ ¸æ°æ®å¤±è´¥ï¼{e}", "error")

    def _load_audit_data_from_output_click(self, event=None):
        """æé®ç¹å»äºä»¶ï¼å è½½å®¡æ ¸æ°æ®ï¼åè£å½æ°ï¼å¿½ç¥eventåæ°ï¼"""
        self._load_audit_data_from_output()

    def _load_audit_data_from_output(self, file_path=None):
        """åæå®æåï¼ä»è¾åºç®å½å è½½åå·®æç»å°å®¡æ ¸è¡¨æ ¼ï¼åæ­¥å è½½ï¼"""
        try:
            # æ¾ç¤ºè¿åº¦æ¡
            if hasattr(self, 'progress_bar'):
                self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)
                self.progress_bar.start(10)
            if hasattr(self, 'set_status'):
                self.set_status("æ­£å¨å è½½æ°æ®ï¼è¯·ç¨å...")
            # Capture UI values for worker
            self._pending_output_dir = self.output_dir.get(
            ) if hasattr(self, 'output_dir') else None
            # ç´æ¥åæ­¥è°ç¨ï¼8000è¡æ°æ®å¾å¿«ï¼ä¸éè¦çº¿ç¨ï¼
            result_df = self._load_data_worker(file_path)
            self._on_load_done(result_df)
        except Exception as e:
            import traceback
            err_msg = f"{e}\n{traceback.format_exc()}"
            self._on_load_error(err_msg)

    def _load_data_worker(self, file_path=None):
        """çº¯æ°æ®å¤çï¼æ¥æ¾æä»¶ãè¯»åãæ¸æ´ãæå»ºaudit_dfï¼ç¦æ­¢UIæä½ï¼"""
        with open(r"C:\Users\Administrator\Desktop\zpp011_debug.txt", "a", encoding="utf-8") as _lf:
            _lf.write(f"[TRACE] _load_data_worker START\n")
        import pandas as pd
        import glob as _glob
        # 1. ç¡®å®è¾åºç®å½
        if not file_path:
            out_dir = None
            # output_dir å input_file æ¯ UI åéï¼å¨launcherä¸­è·å
            if hasattr(self, '_pending_output_dir'):
                out_dir = self._pending_output_dir
            if not out_dir or not os.path.isdir(out_dir):
                out_dir = os.path.expanduser('~/Desktop')
        else:
            out_dir = None
        # 2. æ¾æä»¶ï¼æ¯æå¤ç§å½åæ ¼å¼ï¼
        if file_path:
            latest_file = file_path
        else:
            # æä¼åçº§å°è¯å¤ç§æä»¶å½åæ ¼å¼
            patterns = [
                os.path.join(out_dir, 'ZPP011åå·®åææç»ç_*.xlsx'),
                os.path.join(out_dir, 'ZPP011åå·®åæç»æå_*.xlsx'),
                os.path.join(out_dir, 'ZPP011åå·®åæ_*.xlsx'),
                os.path.join(out_dir, 'ZPP011*.xlsx'),
            ]
            latest_file = None
            for pattern in patterns:
                files = _glob.glob(pattern)
                if files:
                    latest_file = max(files, key=os.path.getmtime)
                    break
            if not latest_file:
                raise FileNotFoundError("æªæ¾å°ä»»ä½åæç»ææä»¶ï¼å·²å°è¯å¤ç§å½åæ ¼å¼ï¼")
        # 3. è¯»å
        dev_df = pd.read_excel(latest_file, sheet_name='å®æ´åå·®æç»')
        if dev_df.empty:
            raise ValueError("åå·®æç»å·¥ä½è¡¨ä¸ºç©º")
        # 4. è§£æåå·®ç

        def parse_rate(v):
            if isinstance(v, str):
                return float(v.replace('%', '').replace('ï¼', '>').replace('>', '')) / 100
            return abs(float(v)) if pd.notna(v) else 0
        dev_df['åå·®çæ°å¼'] = dev_df['åå·®ç'].apply(parse_rate)
        # 5. æå»º audit_df
        audit_df = dev_df.copy()

        # ===== ä¿çéé¢åæ°éåï¼ç¨äºææ¬æ¢ç® =====
        if 'éé¢-å®é(å«ç¨)' in dev_df.columns and 'æ°é-å®é' in dev_df.columns:
            audit_df['éé¢-å®é(å«ç¨)'] = dev_df['éé¢-å®é(å«ç¨)']
            audit_df['æ°é-å®é'] = dev_df['æ°é-å®é']
            audit_df['_unit_price'] = 0.0
            mask = (audit_df['æ°é-å®é'] != 0) & audit_df['æ°é-å®é'].notna()
            audit_df.loc[mask, '_unit_price'] = audit_df.loc[mask,
                                                             'éé¢-å®é(å«ç¨)'] / audit_df.loc[mask, 'æ°é-å®é']
        else:
            audit_df['_unit_price'] = 0.0
            print("[è­¦å] åå§æ°æ®ç¼ºå°éé¢ææ°éåï¼ææ¬æ¢ç®å°ä¸å¯ç¨")

        # ç¡®ä¿è®¢åç±»ååå­å¨ï¼åå§SAPæ°æ®ä¸­æ"è®¢åç±»å"åï¼å¼ä¸ºZP01ç­ï¼
        if 'è®¢åç±»å' in dev_df.columns:
            audit_df['è®¢åç±»å'] = dev_df['è®¢åç±»å']
        else:
            audit_df['è®¢åç±»å'] = ''

        # ç¡®ä¿ææ¬æ¢ç®å¨æéåå­å¨
        if 'åå·®æ°é' not in audit_df.columns:
            if 'æ°éåå·®' in audit_df.columns:
                audit_df['åå·®æ°é'] = audit_df['æ°éåå·®']
            elif 'ææåå·®' in audit_df.columns:
                audit_df['åå·®æ°é'] = audit_df['ææåå·®']
            else:
                audit_df['åå·®æ°é'] = 0.0
        if 'åä½' not in audit_df.columns:
            # å°è¯ä» dev_df è·å
            if 'ç»ä»¶åä½' in dev_df.columns:
                audit_df['åä½'] = dev_df['ç»ä»¶åä½']
            elif 'åä½' in dev_df.columns:
                audit_df['åä½'] = dev_df['åä½']
            else:
                audit_df['åä½'] = ''
        # ä¿®æ­£åè¡¨è¡å·ï¼ä¼åä½¿ç¨ analyzer è®¡ç®ç _excel_rowï¼openpyxlçå®è¡å·ï¼
        if '_excel_row' in dev_df.columns:
            audit_df['excel_row'] = dev_df['_excel_row'].apply(
                lambda x: int(x) if pd.notna(x) else 0)
            audit_df['åè¡¨è¡å·'] = dev_df['_excel_row']  # åæ­¥ä¿®æ­£æ¾ç¤ºå
        elif 'åè¡¨è¡å·' in dev_df.columns:
            audit_df['excel_row'] = dev_df['åè¡¨è¡å·'].apply(
                lambda x: int(x) if pd.notna(x) else 0)
        else:
            audit_df['excel_row'] = range(2, len(audit_df) + 2)
            audit_df['åè¡¨è¡å·'] = audit_df['excel_row']
        # åæ å°ï¼å¦ææºåå­å¨åèµå¼ï¼å¦åç¨é»è®¤å¼
        for dst, src, default in [
            ('ç»ä»¶ç©æå·', 'ç©æç¼ç ', ''),
            ('ç»ä»¶ç©ææè¿°', 'ç©æåç§°', ''),
            ('å·¥ååç§°', 'å·¥å', ''),
            ('çäº§ç®¡çåæè¿°', 'è½¦é´', ''),
            ('æ°é-å®é¢', 'å®é¢', 0),
            ('æ°é-å®é', 'å®é', 0),
            ('å¤æ³¨åå ', 'å¤æ³¨', ''),
        ]:
            if src in audit_df.columns:
                audit_df[dst] = audit_df[src]
            else:
                audit_df[dst] = default
        audit_df['åå·®ç(%)'] = audit_df['åå·®çæ°å¼'] * 100
        # è°è¯ï¼å¯¹æ¯ è®¢å300354378+ç©æ10000000 çæ°æ®
        _dbg_order = 'æµç¨è®¢å' if 'æµç¨è®¢å' in audit_df.columns else None
        _dbg_mat = 'ç»ä»¶ç©æå·' if 'ç»ä»¶ç©æå·' in audit_df.columns else None
        if _dbg_order and _dbg_mat:
            mask_debug = (audit_df[_dbg_order].astype(str) == '300354378') & (
                audit_df[_dbg_mat].astype(str) == '10000000')
        else:
            mask_debug = pd.Series(dtype=bool)
        if len(mask_debug) > 0 and mask_debug.any():
            d = audit_df[mask_debug].iloc[0]
            self._debug_row = d
            debug_info = f"[DEBUG] è®¢å300354378+ç©æ10000000: "
            debug_info += f"excel_row={d.get('excel_row')}, "
            debug_info += f"å®é¢={d.get('å®é¢')}, å®é={d.get('å®é')}, "
            debug_info += f"æ°é-å®é¢={d.get('æ°é-å®é¢')}, æ°é-å®é={d.get('æ°é-å®é')}, "
            debug_info += f"ç©æç¼ç ={d.get('ç©æç¼ç ')}, ç©æåç§°={d.get('ç©æåç§°')}"
            with open(os.path.join(os.path.expanduser('~'), 'Desktop', 'zpp011_debug.txt'), 'a', encoding='utf-8') as f:
                f.write(debug_info + '\n')
        else:
            self._debug_row = None
        # åå·®éé¢è®¡ç®
        if 'åå·®éé¢' not in audit_df.columns or pd.to_numeric(audit_df['åå·®éé¢'], errors='coerce').abs().max() == 0:
            if 'éé¢-å®é(å«ç¨)' in dev_df.columns and 'æ°é-å®é' in dev_df.columns:
                audit_df['_unit_price'] = 0.0
                m = (dev_df['æ°é-å®é'] != 0) & dev_df['æ°é-å®é'].notna()
                audit_df.loc[m, '_unit_price'] = dev_df.loc[m,
                                                            'éé¢-å®é(å«ç¨)'] / dev_df.loc[m, 'æ°é-å®é']
                if 'æ°é-å®é¢' in dev_df.columns:
                    audit_df['æ°éåå·®'] = dev_df['æ°é-å®é'] - dev_df['æ°é-å®é¢']
                else:
                    audit_df['æ°éåå·®'] = 0.0
                audit_df['åå·®éé¢'] = (audit_df['æ°éåå·®'] *
                                    audit_df['_unit_price']).round(2)
                # ç¡®ä¿"åå·®æ°é"åå­å¨ï¼å¡çææ¬æ¢ç®å¨ä¾èµæ­¤ååï¼
                if 'åå·®æ°é' not in audit_df.columns and 'æ°éåå·®' in audit_df.columns:
                    audit_df['åå·®æ°é'] = audit_df['æ°éåå·®']
                # _unit_price ä¿çï¼ä¾ææ¬æ¢ç®å¨ä½¿ç¨
            else:
                audit_df['åå·®éé¢'] = 0.0
        else:
            audit_df['åå·®éé¢'] = pd.to_numeric(
                audit_df['åå·®éé¢'], errors='coerce').fillna(0).round(2)
        # è®¢ååæ¥æ¾
        order_col = None
        for possible in ['æµç¨è®¢å', 'è®¢åå·', 'è®¢åç¼å·', 'è®¢åå·ç ', 'è®¢åNo', 'Order No', 'çäº§è®¢å']:
            if possible in audit_df.columns:
                order_col = possible
                break
            if possible in dev_df.columns:
                order_col = possible
                break
        if order_col is None:
            audit_df['æµç¨è®¢å'] = ''
        elif order_col != 'æµç¨è®¢å':
            audit_df['æµç¨è®¢å'] = audit_df[order_col]
        # çæå¯ä¸ID
        audit_df['_uid'] = (
            audit_df['è®¢åæ¥æ'].astype(str).str[:10] + '_' +
            audit_df['æµç¨è®¢å'].astype(str) + '_' +
            audit_df['ç»ä»¶ç©æå·'].astype(str)
        )
        if 'å¤æ³¨æ¥æº' not in audit_df.columns:
            audit_df['å¤æ³¨æ¥æº'] = ''
        if 'åå¤æ³¨' not in audit_df.columns:
            audit_df['åå¤æ³¨'] = audit_df['å¤æ³¨åå '] if 'å¤æ³¨åå ' in audit_df.columns else ''
        if 'AIå»ºè®®' not in audit_df.columns:
            audit_df['AIå»ºè®®'] = ''
        if 'audit_result' not in audit_df.columns:
            audit_df['audit_result'] = ''

        # ====== è¡¥å _is_alt åï¼æ¿ä»£ææ è¯ï¼======
        if hasattr(self, 'alt_pairs') and self.alt_pairs:
            alt_materials = set()
            for pair in self.alt_pairs:
                for item in pair:
                    if isinstance(item, (list, tuple)):
                        desc = str(
                            item[-1]).strip() if len(item) > 1 else str(item[0]).strip()
                    else:
                        desc = str(item).strip()
                    if desc:
                        alt_materials.add(desc)
            name_col = None
            for col in ['ç©æåç§°', 'ç©ææè¿°', 'ç»ä»¶ç©ææè¿°']:
                if col in audit_df.columns:
                    name_col = col
                    break
            if name_col:
                audit_df['_is_alt'] = audit_df[name_col].astype(
                    str).str.strip().isin(alt_materials)
                print(f"[ä¸´æ¶è¡¥ä¸] å·²æ è®°æ¿ä»£æï¼{audit_df['_is_alt'].sum()} è¡")
            else:
                print("[ä¸´æ¶è¡¥ä¸] æªæ¾å°ç©æåç§°åï¼æ æ³æ è®°æ¿ä»£æ")
                audit_df['_is_alt'] = False
        else:
            audit_df['_is_alt'] = False

        # ====== è¡¥åå®¡æ ¸æ¥æºï¼audit_sourceï¼======
        if 'audit_source' not in audit_df.columns:
            def infer_audit_source(row):
                src = row.get('å®¡æ ¸æ¥æº', '')
                if src and src not in ('nan', 'None', ''):
                    return src
                note_src = str(row.get('å¤æ³¨æ¥æº', '')).strip()
                if 'AI' in note_src:
                    return 'AI'
                if note_src in ('äººå·¥å¡«å', 'æå¨'):
                    return 'æå¨'
                if note_src == 'æ¿ä»£æ':
                    return 'æ¿ä»£æ'
                if row.get('_is_alt', False):
                    return 'æ¿ä»£æ'
                return 'ç³»ç»'
            audit_df['audit_source'] = audit_df.apply(
                infer_audit_source, axis=1)

        # ====== è¡¥åå®¡æ ¸ç¶æï¼audit_statusï¼======
        if 'audit_status' not in audit_df.columns:
            if 'audit_result' in audit_df.columns:
                audit_df['audit_status'] = audit_df['audit_result'].apply(
                    lambda x: 'å·²å®¡æ ¸' if x and str(
                        x).strip() not in ('', 'nan') else 'æªå®¡æ ¸'
                )
            else:
                audit_df['audit_status'] = audit_df['å¤æ³¨åå '].apply(
                    lambda x: 'å·²å®¡æ ¸' if x and str(
                        x).strip() not in ('', 'nan') else 'æªå®¡æ ¸'
                )

        # Task 007: åè²ä¼åçº§æ ç­¾ï¼ç»ä¸ä½¿ç¨ _calculate_priority_label é»è¾ï¼
        # æ³¨æï¼æ­¤å¤æ æ³è°ç¨ self._calculate_priority_labelï¼å ä¸ºç¼ºå°å¤æ³¨ä¿¡æ¯ï¼ï¼
        # æä»¥ä»åºäºåå·®ççæç®åçæ ç­¾ï¼å®æ´çåè²å¨ _refresh_audit_tree ä¸­è®¡ç®
        if 'åå·®ç(%)' in audit_df.columns:
            def get_priority_label(rate):
                if pd.isna(rate):
                    return 'ç»¿'
                rate = abs(float(rate))
                if rate >= 10:
                    return 'çº¢'  # é»è®¤çº¢è²ï¼æå¤æ³¨çä¼å¨ refresh æ¶æ¹ä¸ºæ©è²
                elif rate >= 5:
                    return 'é»'
                else:
                    return 'ç»¿'
            audit_df['_priority_label'] = audit_df['åå·®ç(%)'].apply(
                get_priority_label)

        with open(r"C:\Users\Administrator\Desktop\zpp011_debug.txt", "a", encoding="utf-8") as _lf:
            _lf.write(
                f"[TRACE] _load_data_worker å®æ: {len(audit_df) if 'audit_df' in dir() else '??'} è¡\n")
        return audit_df

    def _on_load_done(self, result_df):
        """å¼æ­¥å è½½æååè°ï¼å¤çææUIæ´æ°"""
        with open(r"C:\Users\Administrator\Desktop\zpp011_debug.txt", "a", encoding="utf-8") as _lf:
            _lf.write(
                f"[TRACE] _on_load_done START: result_df={len(result_df) if result_df is not None else 'None'} è¡\n")
        try:
            self.audit_data = result_df
            with open(r"C:\Users\Administrator\Desktop\zpp011_debug.txt", "a", encoding="utf-8") as _lf:
                _lf.write(
                    f"[TRACE] audit_data å·²èµå¼: {len(self.audit_data)} è¡\n")
            # ââ é¢è®¡ç®åä»·ï¼éé¢-å®é(å«ç¨) / æ°é-å®éï¼ç¨äºææ¬æ¢ç®å¨ ââ
            if 'éé¢-å®é(å«ç¨)' in self.audit_data.columns and 'æ°é-å®é' in self.audit_data.columns:
                self.audit_data['_unit_price'] = 0.0
                mask = (self.audit_data['æ°é-å®é'].notna()
                        ) & (self.audit_data['æ°é-å®é'] != 0)
                self.audit_data.loc[mask, '_unit_price'] = (
                    self.audit_data.loc[mask,
                                        'éé¢-å®é(å«ç¨)'] / self.audit_data.loc[mask, 'æ°é-å®é']
                )
            else:
                self.audit_data['_unit_price'] = 0.0
            # å¦æå¯ç¨äºæ°ç­éæ ï¼æ´æ°ä¸æéé¡¹
            if hasattr(self, "filter_panel") and self.filter_panel:
                self.filter_panel.update_options(self.audit_data)
            # ç¡®ä¿æ°å¼åä¸ºæ°å¼ç±»åï¼é²æ­¢å­ç¬¦ä¸²å¯¼è´æ¯è¾éè¯¯ï¼
            for col in ["åå·®ç(%)", "ææåå·®", "æ°é-å®é¢", "æ°é-å®é", "åå·®éé¢"]:
                if col in self.audit_data.columns:
                    self.audit_data[col] = pd.to_numeric(
                        self.audit_data[col], errors="coerce").fillna(0)
            self._full_dev_df = result_df.copy()
            # ââ è¡¥å material_category åï¼ä¸ _load_audit_data è·¯å¾ä¿æä¸è´ï¼ ââ
            mat_cat_map = {
                "100": "åè¾æ", "200": "åæ", "400": "é£åè¾æ/é£ååæå",
                "410": "é¥®æè¾æ/é¥®æåæå", "500": "é£åæå", "510": "é¥®ææå",
                "600": "ä¿éå"
            }
            mat_code_col = None
            for mc in ['ç©æç¼ç ', 'ç»ä»¶ç©æå·']:
                if mc in self.audit_data.columns:
                    mat_code_col = mc
                    break
            if mat_code_col:
                # åå é¤å·²æç material_category åï¼é¿åååå²çªå¯¼è´ material_category[2]ï¼
                if 'material_category' in self.audit_data.columns:
                    self.audit_data.drop(
                        columns=['material_category'], inplace=True)
                self.audit_data['material_category'] = self.audit_data[mat_code_col].apply(
                    lambda x: mat_cat_map.get(
                        str(x)[:3], '') if pd.notna(x) else ''
                )
            else:
                self.audit_data['material_category'] = ''
            self.full_audit_data = self.audit_data.copy()
            self.log(f"[DEBUG] _on_load_done: {len(self.audit_data)}è¡", "info")
            # è°è¯ï¼æå°å5è¡ excel_row å åè¡¨è¡å·
            if 'excel_row' in self.audit_data.columns:
                self.log(
                    f"[DEBUG] excel_row å5è¡: {self.audit_data['excel_row'].head().tolist()}", "debug")
            if 'åè¡¨è¡å·' in self.audit_data.columns:
                self.log(
                    f"[DEBUG] åè¡¨è¡å· å5è¡: {self.audit_data['åè¡¨è¡å·'].head().tolist()}", "debug")
            _order_col = 'æµç¨è®¢å' if 'æµç¨è®¢å' in self.audit_data.columns else None
            _mat_col = 'ç»ä»¶ç©æå·' if 'ç»ä»¶ç©æå·' in self.audit_data.columns else None
            if _order_col and _mat_col:
                mask = (self.audit_data[_order_col].astype(str) == '300354378') & (
                    self.audit_data[_mat_col].astype(str) == '10000000')
                if mask.any():
                    debug_row = self.audit_data[mask].iloc[0]
                    self.log(
                        f"[DEBUG] è®¢å300354378+ç©æ10000000: excel_row={debug_row.get('excel_row') if hasattr(debug_row, 'get') else debug_row.get('excel_row', '')}, å®é¢={debug_row.get('æ°é-å®é¢', '')}, å®é={debug_row.get('æ°é-å®é', '')}", "debug")
            # æ´æ°ç­ééé¡¹åè¡é¢è²
            self._update_filter_options()
            self._apply_row_colors()
            if hasattr(self, '_update_trend_display'):
                self._update_trend_display()
            # ç»è®¡å¡ç
            total = len(result_df)
            high_dev = len(result_df[result_df['åå·®ç(%)'] > 10]
                           ) if 'åå·®ç(%)' in result_df.columns else 0
            no_note = len(result_df[result_df['å¤æ³¨åå '].isna() | (
                result_df['å¤æ³¨åå '] == '')]) if 'å¤æ³¨åå ' in result_df.columns else 0
            approved_note = result_df['å¤æ³¨æ¥æº'].isin(
                ['AIå®¡æ ¸åæ ¼', 'AIå®¡æ ¸å¾æ¹è¿', 'AIçæ']).sum() if 'å¤æ³¨æ¥æº' in result_df.columns else 0
            if hasattr(self, 'audit_stat_labels'):
                for k, v in [('total', total), ('high_dev', high_dev),
                             ('need_note', no_note), ('ok_note', approved_note)]:
                    if k in self.audit_stat_labels:
                        self.audit_stat_labels[k].configure(text=str(v))
            if hasattr(self, 'unified_result_lbl'):
                self.unified_result_lbl.configure(
                    text=f"å·²å è½½ {total} æ¡ | åå·®>10%: {high_dev} | éè¡¥å¤æ³¨: {no_note} | å·²å®¡æ ¸: {approved_note}")
            # å¯ç¨æææé®
            enabled = self._enable_audit_buttons()
            self.log(f"[DEBUG] å·²å¯ç¨{enabled}ä¸ªæé®", "info")
            self.log(
                f"â å®¡æ ¸æ°æ®å è½½å®æ | å± {total} æ¡ | åå·®>10%: {high_dev} æ¡", "success")
            # æ¢å¤ç­éåæåºç¶æ
            self._restore_filters()
            self.root.after(self.config.get('analysis.cleanup_delay_ms', 100), lambda: self._on_filter_changed(
                None) if self.audit_data is not None else None)
            self._restore_sort_state()
            self.root.after(self.config.get('analysis.cleanup_delay_ms', 300),
                            lambda: self._show_precheck_report(self.audit_data))
            # BOM è¿ææéï¼åè½å¾å®ç°ï¼ææ³¨éï¼
            # self.root.after(self.config.get('analysis.cleanup_delay_ms', 500), self._check_and_remind_bom)
            # æ¢å¤å®¡æ ¸è®°å½
            self.audit_presenter.load_audit_data(self.audit_data)

            # ââ Task 009ï¼ä¿å­å°åå²æ°æ®åº ââ
            try:
                from core.history_db import save_analysis_result, init_db
                init_db()

                # è®¡ç®ç»è®¡ææ 
                total_rows = len(self.audit_data)

                # åå·®>10% è¡æ°
                dev_col = None
                for col in ['åå·®ç%', 'åå·®ç(%)']:
                    if col in self.audit_data.columns:
                        dev_col = col
                        break
                if dev_col:
                    high_dev = len(
                        self.audit_data[self.audit_data[dev_col].abs() > 10])
                else:
                    high_dev = 0

                # éè¡¥å¤æ³¨è¡æ°
                remark_col = None
                for col in ['å¤æ³¨', 'å¤æ³¨åå ']:
                    if col in self.audit_data.columns:
                        remark_col = col
                        break
                if remark_col:
                    need_note = len(self.audit_data[self.audit_data[remark_col].isna() | (
                        self.audit_data[remark_col].astype(str).str.strip() == '')])
                else:
                    need_note = 0

                # å·²å®¡æ ¸è¡æ°
                status_col = None
                for col in ['å®¡æ ¸ç¶æ', 'audit_status']:
                    if col in self.audit_data.columns:
                        status_col = col
                        break
                if status_col:
                    approved = len(
                        self.audit_data[self.audit_data[status_col] == 'å·²å®¡æ ¸'])
                else:
                    approved = 0

                # è·åç­éæ¡ä»¶ï¼å¦ææï¼
                filter_cond = ''
                if hasattr(self, '_get_current_filter_condition'):
                    filter_cond = self._get_current_filter_condition()

                metadata = {
                    'file_name': os.path.basename(self.input_file.get()) if hasattr(self, 'input_file') else 'unknown',
                    'file_path': self.input_file.get() if hasattr(self, 'input_file') else '',
                    'file_mtime': os.path.getmtime(self.input_file.get()) if hasattr(self, 'input_file') and self.input_file.get() and os.path.exists(self.input_file.get()) else 0,
                    'total_rows': len(self.audit_data),
                    'high_dev_rows': high_dev,
                    'need_note_rows': need_note,
                    'approved_rows': approved,
                    'filter_condition': filter_cond,
                }

                analysis_id = save_analysis_result(metadata, self.audit_data)
                self.log(f"â åæç»æå·²å­å¥åå²åºï¼ID={analysis_id}", "info")
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                self.log(f"â  ä¿å­åå²è®°å½å¤±è´¥ï¼{type(e).__name__}: {e}", "warn")
                print(f"[DEBUG] Exception type: {type(e)}, value: {e!r}")
                print(f"[DEBUG] Traceback:\n{tb}")

            # æ­ç¹ç»­å®¡æç¤º
            state = self._load_resume_state()
            if state:
                self.resume_btn.configure(state="normal")
                saved_row = state.get('selected_row', None)
                if saved_row is not None:
                    self.log(f"ð æ£æµå°ä¸æ¬¡å®¡æ ¸è¿åº¦ï¼ç¬¬ {saved_row} è¡", "info")
                    tip = tk.Toplevel(self.root)
                    tip.wm_overrideredirect(True)
                    tip.geometry(
                        f"280x28+{self.root.winfo_rootx()+self.root.winfo_width()-290}+{self.root.winfo_rooty()+self.root.winfo_height()-60}")
                    tk.Label(tip, text=f"ð¡ ä¸æ¬¡å®¡æ ¸å°ç¬¬ {saved_row} è¡ï¼ç¹å»ãæ¢å¤è¿åº¦ãç»§ç»­",
                             font=("Microsoft YaHei", 9), bg="#1a1a2e", fg="white",
                             padx=8, pady=4).pack()
                    tip.after(4000, tip.destroy)
            # éèè¿åº¦æ¡
            if hasattr(self, 'progress_bar'):
                self.progress_bar.stop()
                self.progress_bar.pack_forget()
            if hasattr(self, 'set_status'):
                self.set_status(f"å è½½å®æï¼å± {total} è¡")
            # èªå¨è®¾ç½®æ¥æç­éæ§ä»¶çé»è®¤èå´å¹¶èªå¨ç­é
            if hasattr(self, 'date_start_de') and hasattr(self, 'date_end_de') and 'è®¢åæ¥æ' in self.audit_data.columns:
                try:
                    dates = pd.to_datetime(
                        self.audit_data['è®¢åæ¥æ'], errors='coerce')
                    min_date = dates.min()
                    max_date = dates.max()
                    if pd.notna(min_date) and pd.notna(max_date):
                        self.date_start_de.set_date(min_date.date())
                        self.date_end_de.set_date(max_date.date())
                        self.log(
                            f"æ¥æç­éæ§ä»¶å·²è®¾ç½®ä¸ºæ°æ®èå´ï¼{min_date.date()} è³ {max_date.date()}", "info")
                        # èªå¨è§¦åæ¥æç­é
                        self._on_filter_changed('order_date')
                except Exception as e:
                    self.log(f"è®¾ç½®é»è®¤æ¥æèå´å¤±è´¥ï¼{e}", "warn")

            # èªå¨ AI å®¡æ ¸ + èªå¨ç»æ¡
            if not getattr(self, '_auto_processed', False):
                self.root.after(500, self._auto_audit_and_close)
        except Exception as e:
            if hasattr(self, 'progress_bar'):
                self.progress_bar.stop()
                self.progress_bar.pack_forget()
            raise

        # --- å è½½æååæ¸çä¸´æ¶ Excel æä»¶ï¼å»¶åå°ç¨åºéåºï¼---
        # ä¸åç«å³å é¤ï¼æ¹ä¸ºå¨ ZPP011Beautiful.on_closing ä¸­ç»ä¸æ¸ç
        # if hasattr(self, '_analysis_output_path') and self._analysis_output_path:
        #     try:
        #         if os.path.exists(self._analysis_output_path):
        #             os.remove(self._analysis_output_path)
        #             self.log("å·²æ¸çä¸´æ¶æä»¶: " + os.path.basename(self._analysis_output_path), "info")
        #             self._analysis_output_path = None
        #     except Exception as e:
        #         self.log("æ¸çä¸´æ¶æä»¶å¤±è´¥: " + str(e), "warn")

    def _on_load_error(self, error):
        """å¼æ­¥å è½½å¤±è´¥åè°"""
        if hasattr(self, 'progress_bar'):
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
        if hasattr(self, 'set_status'):
            self.set_status("å è½½å¤±è´¥")
        messagebox.showerror(_safe_for_gbk("å è½½éè¯¯"), _safe_for_gbk(str(error)))
        try:
            from core.logger import get_logger
            get_logger().error(f"æ°æ®å è½½å¼æ­¥å¤±è´¥: {error}")
        except Exception:
            pass

    def _filter_by_stat(self, filter_key):
        """æ ¹æ®é¡¶é¨ç»è®¡å¡çç¹å»ï¼ç­éå®¡æ ¸è¡¨æ ¼ï¼å§æPresenterï¼"""
        if self.audit_data is None or len(self.audit_data) == 0:
            return
        # å§æ Presenter æ§è¡çº¯ä¸å¡ç­éé»è¾
        filtered = self.audit_presenter.filter_audit_data({'stat': filter_key})
        # å·æ°å®¡æ ¸è¡¨æ ¼ï¼Viewå±ï¼
        self._refresh_audit_tree(filtered)
        self._update_audit_stats(filtered)
        # æ´æ°ç­éæç¤ºæå­
        labels = {'total': 'å¨é¨', 'big_dev': 'åå·®>10%',
                  'no_note': 'éè¡¥å¤æ³¨', 'approved': 'å·²å®¡æ ¸'}
        self.status_filter_label.config(
            text="ç­éï¼{} | å± {} æ¡".format(
                labels.get(filter_key, ''), len(filtered))
        )

    def _s01_start_inventory_check(self, data: pd.DataFrame) -> None:
        """å¯å¨åºå­æµç¨æ£æ¥ï¼å¼æ­¥ï¼ï¼çº¿ç¨åå»ºå¤±è´¥æ¶éæ¾é"""
        if self._is_s01_processing:
            messagebox.showwarning(_safe_for_gbk("æç¤º"), _safe_for_gbk("å·²æåºå­æ£æ¥ä»»å¡è¿è¡ä¸­"))
            return
        if data is None or data.empty:
            messagebox.showwarning(_safe_for_gbk("è­¦å"), _safe_for_gbk("æ æ°æ®å¯æ£æ¥"))
            return
        self._is_s01_processing = True
        try:
            self._s01_disable_ui()
            self.progress_bar.configure(mode='determinate', maximum=100)
            self.progress_bar['value'] = 0
            self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)
            self.set_status("æ­£å¨æ§è¡åºå­æ£æ¥...")
            data_copy = data.copy(deep=True)
            self._s01_cancel_flag = threading.Event()
            self._s01_thread = threading.Thread(
                target=self._s01_inventory_worker,
                args=(data_copy, self._s01_cancel_flag)
            )
            self._s01_thread.daemon = True
            self._s01_thread.start()
        except Exception as e:
            self._s01_cleanup()
            messagebox.showerror(_safe_for_gbk("éè¯¯"), _safe_for_gbk(f"ä»»å¡å¯å¨å¤±è´¥: {e}"))

    def _s01_inventory_worker(self, df, cancel_flag):
        """å¼æ­¥ä»»å¡ï¼çº¯æ°æ®å¤çï¼ä½¿ç¨ itertuplesï¼æ¯ 50 è¡æ£æ¥åæ¶å¹¶åè°è¿åº¦"""
        total = len(df)
        status_col = None
        for col in ['åºå­ç¶æ', 'inventory_status']:
            if col in df.columns:
                status_col = col
                break
        if status_col is None:
            df['åºå­ç¶æ'] = ''
            status_col = 'åºå­ç¶æ'
        try:
            for idx, row in enumerate(df.itertuples(index=False)):
                if cancel_flag.is_set():
                    self._s01_clean_temp_files()
                    raise InterruptedError("æä½å·²åæ¶")
                row_dict = row._asdict()
                # ä¸å¡é»è¾å¾è¡¥å
                if (idx + 1) % 50 == 0 or idx + 1 == total:
                    self.root.after(self.config.get(
                        'analysis.cleanup_delay_ms', 0), self._s01_progress_callback, idx + 1, total)
            self.root.after(self.config.get(
                'analysis.cleanup_delay_ms', 0), self._s01_on_success, df)
        except InterruptedError as e:
            self.root.after(self.config.get(
                'analysis.cleanup_delay_ms', 0), self._s01_on_error, e)
        except Exception as e:
            self.root.after(self.config.get(
                'analysis.cleanup_delay_ms', 0), self._s01_on_error, e)

    def _s01_progress_callback(self, current: int, total: int) -> None:
        """è¿åº¦åè°"""
        percent = int(current / total * 100) if total else 0
        self.progress_bar['value'] = percent
        self.set_status(f"åºå­æ£æ¥ä¸­: {current}/{total} ({percent}%)")

    def _s01_on_success(self, result_df: pd.DataFrame) -> None:
        """æååè°ï¼è°ç¨é«äº®ç _s01_populate_table"""
        self.audit_data = result_df
        self._s01_populate_table(result_df)
        self._s01_cleanup()

    def _s01_on_error(self, error: Exception) -> None:
        """éè¯¯/åæ¶åè°"""
        if isinstance(error, InterruptedError):
            self.set_status("æä½å·²åæ¶")
            messagebox.showwarning(_safe_for_gbk("å·²åæ¶"), _safe_for_gbk(str(error)))
        else:
            self.set_status("æä½å¤±è´¥")
            messagebox.showerror(_safe_for_gbk("éè¯¯"), _safe_for_gbk(f"åºå­æ£æ¥å¤±è´¥: {error}"))
        self._s01_cleanup()

    def _s01_cleanup(self) -> None:
        """ç»ä¸æ¸ç"""
        self._is_s01_processing = False
        self._s01_cancel_flag = None
        self._s01_thread = None
        self._s01_enable_ui()
        self.progress_bar['value'] = 0
        self.progress_bar.pack_forget()
        self.set_status("å°±ç»ª")
        self._s01_clean_temp_files()

    def _s01_cancel_inventory_check(self) -> None:
        if self._s01_cancel_flag:
            self._s01_cancel_flag.set()

    def _s01_disable_ui(self) -> None:
        """ç¦ç¨ç¸å³æé®"""
        pass

    def _s01_enable_ui(self) -> None:
        """æ¢å¤æé®"""
        pass
    # ââ S01 å¼å¸¸é«äº®æ¹æ³ âââââââââââââââââââââââââââââââââââââââââââââââââ

    def _evaluate_condition(self, row: dict, condition_str: str) -> bool:
        """å®å¨è¯ä¼°æ¡ä»¶è¡¨è¾¾å¼ï¼ä»ç¨äº S01 é«äº®ï¼"""
        dangerous = ['__', 'import', 'exec', 'eval', 'compile', 'open', 'file']
        if any(d in condition_str for d in dangerous):
            return False
        allowed = {k: v for k, v in row.items() if isinstance(
            v, (int, float, str, bool))}
        try:
            return bool(eval(condition_str, {"__builtins__": {}}, allowed))
        except Exception:
            return False

    def _s01_setup_treeview_tags(self, tree):
        """ä¸º Treeview éç½®é«äº® tagï¼ä»é¦æ¬¡ï¼"""
        cache_key = f"_s01_tags_configured_{id(tree)}"
        if getattr(self, cache_key, False):
            return
        config = getattr(self, 's01_display_config', {})
        for rule in config.get('rules', []):
            tag = rule.get('tag')
            color = rule.get('color')
            if tag and color:
                try:
                    tree.tag_configure(tag, background=color)
                except Exception:
                    pass
        setattr(self, cache_key, True)

    def _s01_populate_table(self, df):
        """å° DataFrame å¡«åå° audit_treeï¼å¹¶æ ¹æ® s01_display_config åºç¨é«äº®"""
        tree = getattr(self, 'audit_tree', None)
        if tree is None:
            self._refresh_audit_tree(df)
            return
        self._s01_setup_treeview_tags(tree)
        config = getattr(self, 's01_display_config', {})
        rules = config.get('rules', [])
        default_color = config.get('default_color', '#FFFFFF')
        try:
            tree.tag_configure('s01_normal', background=default_color)
        except Exception:
            pass
        # æ¸ç©ºæ§æ°æ®
        for item in tree.get_children():
            tree.delete(item)
        cols = list(tree['columns']) if tree['columns'] else (
            list(df.columns) if not df.empty else [])
        for row_tuple in df.itertuples():
            row_dict = row_tuple._asdict()
            row_dict.pop('Index', None)
            tag = 's01_normal'
            for rule in rules:
                condition = rule.get('condition', '')
                if condition and self._evaluate_condition(row_dict, condition):
                    tag = rule.get('tag', 's01_normal')
                    break
            values = [row_dict.get(c, '') for c in cols]
            tree.insert('', 'end', values=values, tags=(tag,))
    # ---------- View æ¥å£æ¹æ³ï¼ä¾ AuditPresenter è°ç¨ï¼----------

    # _update_filter_options å·²ç§»å¥ table_events.pyï¼ä¼åçº§æ´é«ç MRO ä½ç½®ï¼
    # å¨é£éç»ä¸å¤çå¨é¨ä¸ææ¡ï¼åæ¬ material_categoryï¼åºäº full_audit_dataï¼
