# -*- coding: utf-8 -*-
"""导出、PPT生成、备份等事件"""

import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import pandas as pd
from ppt_generator import run_ppt_generation as _run_ppt
from core.decorators import with_feedback
from widgets import C
from core.exporter import ExcelExporter
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook


class ExportEvents:
    """导出、PPT生成、备份等事件"""

    def generate_ppt(self):
        """选择 Excel 分析结果，生成 PPT 报告"""

        excel_path = filedialog.askopenfilename(
            title="选择 zpp011 偏差分析 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        )

        if not excel_path:
            return

        out_dir = self.output_dir.get() or os.path.dirname(excel_path)

        base = os.path.splitext(os.path.basename(excel_path))[0]

        if self.audit_data is None or self.audit_data.empty:
            messagebox.showwarning("提示", "请先加载并分析数据后再生成 PPT！")

            return

        import datetime

        default_name = (
            self.config.get(
                "export.default_ppt_filename", "ZPP011偏差分析_{datetime}.pptx"
            ).format(datetime=datetime.datetime.now().strftime("%Y%m%d_%H%M"))
            + ".pptx"
        )

        ppt_output = filedialog.asksaveasfilename(
            initialdir=out_dir,
            initialfile=default_name,
            defaultextension=".pptx",
            filetypes=[("PPT 文件", "*.pptx"), ("所有文件", "*")],
            title="保存 PPT 报告",
        )

        if not ppt_output:
            return

        self.log("📊 开始生成 PPT...", "info")

        self.ppt_btn.configure(state="disabled", text="生成中...")

        self.status_lbl.configure(text="正在生成 PPT...", fg=C["purple"])

        def worker():

            try:

                def progress(pct):

                    self.root.after(0, lambda p=pct: self._update_ppt_progress(p))

                self.audit_presenter.generate_ppt(
                    output_path=ppt_output, progress_callback=progress
                )

                self.root.after(0, lambda: self._on_ppt_done(ppt_output))

            except Exception as e:
                self.root.after(0, lambda e=e: self._on_ppt_error(str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _update_ppt_progress(self, pct):
        """更新 PPT 生成进度显示"""
        self.status_lbl.configure(text=f"正在生成 PPT... {int(pct)}%", fg=C["purple"])

    def _on_ppt_done(self, output_path):

        self.ppt_btn.configure(state="normal", text="📊 生成PPT")

        self.status_lbl.configure(
            text=f"PPT 已生成 — {os.path.basename(output_path)}", fg=C["green"]
        )

        self.log(f"✅ PPT 已保存：{output_path}", "success")

        try:
            os.startfile(output_path)

        except Exception:
            pass

    def _on_ppt_error(self, msg):

        self.ppt_btn.configure(state="normal", text="📊 生成PPT")

        self.status_lbl.configure(text="PPT 生成出错", fg=C["danger"])

        self.log(f"❌ PPT 生成失败：{msg}", "error")

        messagebox.showerror("PPT 生成出错", msg)

    def generate_excel_direct(self):
        """弹出另存为对话框，生成 Excel 分析表格"""

        input_path = self.input_file.get()

        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("错误", "请先选择输入文件！")

            return

        # ── 尝试从原始数据中获取日期范围，用于默认文件名 ──

        date_tag = ""

        try:
            df_raw = pd.read_excel(input_path, sheet_name="Data")

            # 查找可能的日期列

            date_col = None

            for col in ["订单开始日期", "订单日期", "日期"]:
                if col in df_raw.columns:
                    date_col = col

                    break

            if date_col is not None:
                dates = pd.to_datetime(df_raw[date_col], errors="coerce").dropna()

                if not dates.empty:
                    d_min = dates.min().strftime("%Y%m%d")

                    d_max = dates.max().strftime("%m%d")

                    date_tag = f"{d_min}-{d_max}"

        except Exception:
            pass

        # 如果从数据中获取失败，回退到手动输入或当前月份

        if not date_tag:
            start_str = self.start_date.get().strip()

            end_str = self.end_date.get().strip()

            if start_str and end_str:
                date_tag = (
                    f"{start_str.replace('-', '')[:8]}-{end_str.replace('-', '')[-4:]}"
                )

            else:
                now = datetime.now()

                first_day = now.replace(day=1).strftime("%Y%m%d")

                # 使用 calendar 计算月末

                import calendar

                last_day_num = calendar.monthrange(now.year, now.month)[1]

                last_day = now.replace(day=last_day_num).strftime("%Y%m%d")

                date_tag = f"{first_day}-{last_day[-4:]}"

        # ── 构建默认路径和文件名 ──

        default_dir = os.path.join(os.path.expanduser("~"), "ZPP011偏差分析")

        os.makedirs(default_dir, exist_ok=True)

        default_name = f"ZPP011偏差分析_{date_tag}.xlsx"

        # ── 弹出另存为对话框 ──

        file_path = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            title="保存偏差分析表格",
        )

        if not file_path:
            return  # 用户取消

        self.excel_btn.configure(state="disabled", text="生成中...")

        self.status_lbl.configure(text="正在生成表格...", fg=C["purple"])

        self.log("📋 开始生成偏差分析表格...", "info")

        threading.Thread(
            target=self._generate_excel_thread, args=(file_path,), daemon=True
        ).start()

    def _generate_excel_thread(self, output_path: str):
        """后台线程：委托 AuditPresenter 生成偏差分析 Excel，带详细错误日志"""

        import traceback

        try:
            result = self.audit_presenter.generate_excel_direct(
                input_file=self.input_file.get(),
                output_path=output_path,
                alt_pairs=self.alt_pairs,
                start_date=self.start_date.get() or None,
                end_date=self.end_date.get() or None,
                material_search=self.material_search.get() or None,
                log_cb=self.log,
            )

            def on_success():

                self.log(f"\u2705 表格生成：{os.path.basename(result)}", "success")

                self.excel_btn.configure(state="normal", text="\U0001f4cb 生成表格")

                self.status_lbl.configure(
                    text=f"已生成：{os.path.basename(result)}", fg=C["green"]
                )

                _name = os.path.basename(result)

                _msg = "表格已生成：" + _name + chr(10) + chr(10) + "是否立即打开？"

                if messagebox.askyesno("生成成功", _msg):
                    try:
                        os.startfile(result)

                        self.log(f"\u2705 已打开文件：{result}", "info")

                    except Exception as oe:
                        self.log(f"\u26a0\ufe0f 无法打开文件：{oe}", "warning")

                        messagebox.showwarning("打开失败", f"无法打开文件：{oe}")

            self.root.after(0, on_success)

        except Exception as e:
            tb = traceback.format_exc()

            error_msg = str(e)

            if "Permission denied" in error_msg or "PermissionError" in error_msg:
                _warn_msg = (
                    "\u26a0\ufe0f 无法保存文件"
                    + chr(10)
                    + chr(10)
                    + "\u53ef\u80fd\u7684\u539f\u56e0\uff1a"
                    + chr(10)
                    + "  • 文件已用 Excel 打开"
                    + chr(10)
                    + "  • 文件被 WPS 或其他程序占用"
                    + chr(10)
                    + chr(10)
                    + "解决方法："
                    + chr(10)
                    + "  1. 关闭 Excel 中打开的这个文件"
                    + chr(10)
                    + "  2. 点击「生成表格」重试"
                    + chr(10)
                    + "  3. 或者另存为其他文件名"
                )

                self.root.after(
                    0, lambda: messagebox.showwarning("文件被占用", _warn_msg)
                )

                self.root.after(
                    0,
                    lambda: self.log(
                        "\u26a0\ufe0f 文件被占用，请关闭 Excel 后重试", "warning"
                    ),
                )

            else:
                _err = str(e)

                self.root.after(
                    0,
                    lambda: messagebox.showerror("生成失败", f"生成表格时出错：{_err}"),
                )

                self.root.after(
                    0, lambda: self.log(f"\u274c 表格生成失败：{_err}", "error")
                )

            self.root.after(0, lambda: self.log(tb, "error"))

            self.root.after(
                0,
                lambda: self.excel_btn.configure(
                    state="normal", text="\U0001f4cb 生成表格"
                ),
            )

    # ── 树形视图（v31 新增）────────────────────────

    @with_feedback("导出成功", "导出失败")
    @with_feedback("Exporting Excel...", show_progress=True)
    def _export_audit_excel(self, cancel_flag=None, progress_callback=None):
        """Async export audit result (launcher)"""

        if self.audit_data is None or len(self.audit_data) == 0:
            messagebox.showwarning("Tip", "No data to export")

            return

        out_path = (
            self.output_dir.get()
            or os.path.dirname(self.input_file.get())
            or os.getcwd()
        )

        file_path = filedialog.asksaveasfilename(
            initialdir=out_path,
            initialfile="ZPP011_Audit_Result.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )

        if not file_path:
            return

        # Start async task

        self.is_exporting = True

        self.task_manager.run(
            lambda c, p: ExcelExporter.export(
                self.audit_data.copy(), file_path, cancel_flag=c, progress_callback=p
            ),
            task_id="export_audit_excel",
            on_progress=self._on_export_progress,
            on_done=self._on_export_done,
            on_error=self._on_export_error,
        )

    def _on_export_progress(self, current, total, **kwargs):
        """Export progress callback"""

        if hasattr(self, "progress_bar") and self.progress_bar:
            self.progress_bar["value"] = current

            self.progress_bar["maximum"] = total

            percent = int(current / total * 100) if total > 0 else 0

            eta_seconds = kwargs.get("eta_seconds", 0)

            eta_str = f"ETA {int(eta_seconds)}s" if eta_seconds > 0 else ""

            self.progress_bar.configure(
                text=f"Exporting {current}/{total} ({percent}%) {eta_str}"
            )

            self.root.update_idletasks()

    def _on_export_done(self, result):
        """Export success callback"""

        self.is_exporting = False

        if hasattr(self, "progress_bar") and self.progress_bar:
            self.progress_bar.stop()

            self.progress_bar.pack_forget()

        file_path = result.get("file_path", "")

        if file_path and os.path.exists(file_path):
            if messagebox.askyesno(
                "Success", f"Exported to:\n{file_path}\n\nOpen folder?"
            ):
                os.startfile(os.path.dirname(file_path))

            self.log(f"Audit result exported (async): {file_path}", "success")

        else:
            messagebox.showinfo("Success", "Export completed!")

    def _on_export_error(self, error):
        """Export error callback"""

        self.is_exporting = False

        if hasattr(self, "progress_bar") and self.progress_bar:
            self.progress_bar.stop()

            self.progress_bar.pack_forget()

        # Clean temp files

        temp_dir = os.path.join(os.path.dirname(__file__), "..", "temp")

        if os.path.exists(temp_dir):
            for f in os.listdir(temp_dir):
                if f.endswith(".tmp.xlsx"):
                    try:
                        os.remove(os.path.join(temp_dir, f))

                    except:
                        pass

        messagebox.showerror("Error", f"Export failed: {error}")

        self.log(f"Export failed: {error}", "error")

    def _export_audit_backup(self):

        default_name = f"审核记录备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("ZIP 压缩文件", "*.zip")],
            initialfile=default_name,
            title="导出审核记录备份",
        )

        if not file_path:
            return

        try:
            storage.export_audit_backup(file_path, log_cb=self.log)

            self.log(f"✅ 审核记录已导出：{file_path}", "success")

            messagebox.showinfo("导出成功", f"备份文件已保存到：\n{file_path}")

        except FileNotFoundError:
            self.log("❌ 导出失败：审核数据库不存在", "error")

            messagebox.showerror(
                "导出失败", "审核数据库不存在，请先保存审核记录后再导出。"
            )

        except Exception as e:
            self.log(f"❌ 导出失败：{e}", "error")

            messagebox.showerror("导出失败", str(e))

    def _import_audit_backup(self):

        file_path = filedialog.askopenfilename(
            title="选择审核记录备份文件",
            filetypes=[("ZIP 压缩文件", "*.zip"), ("所有文件", "*.*")],
        )

        if not file_path:
            return

        try:
            storage.import_audit_backup(file_path, log_cb=self.log)

            self.log("✅ 审核记录已从备份恢复，下次加载时生效", "success")

            messagebox.showinfo(
                "导入成功", "审核记录已恢复。\n重新加载数据时将自动匹配历史审核。"
            )

        except Exception as e:
            self.log(f"❌ 导入失败：{e}", "error")

            messagebox.showerror("导入失败", str(e))

    # ── 版本日志已迁移到 utils/version_history.py ──────────

    # 原 _CHANGELOG_EMBEDDED 硬编码已删除，通过导入动态读取

    _CHANGELOG_EMBEDDED = None  # 占位，保持向后兼容

    # ── 公共函数：获取 changelog.json 路径 ──────────

    
    def _export_audit_log(self):
        """导出审计日志为 CSV（异步）"""
        from tkinter import filedialog
        import os
        
        output_path = filedialog.asksaveasfilename(
            title="导出审计日志",
            initialdir=os.path.expanduser("~"),
            initialfile=f"audit_log_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")]
        )
        
        if not output_path:
            return
        
        if hasattr(self, 'audit_logger'):
            self.audit_logger.export_csv_async(
                output_path,
                callback=lambda path, error: self.root.after(0, lambda: messagebox.showinfo("导出完成", f"审计日志已导出到：\n{path}"))
            )

    def _export_log(self):
        """导出日志到文件"""

        try:
            import datetime as _dt

            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")

            default_name = self.config.get(
                "export.default_log_filename", "ZPP011日志_{ts}.txt"
            ).format(ts=ts)

            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                initialfile=default_name,
                title="导出日志",
            )

            if not path:
                return

            with open(path, "w", encoding="utf-8") as f:
                f.write(self.log_text.get("1.0", "end"))

            self.log(f"日志已导出：{os.path.basename(path)}", "success")

        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _export_changelog(self):
        """导出版本日志"""

        try:
            # 使用公共函数获取 changelog.json 路径

            cl_path = self._get_changelog_path() or ""

            # 读取 changelog.json

            cl_data = {}

            if os.path.isfile(cl_path):
                with open(cl_path, "r", encoding="utf-8") as f:
                    cl_data = json.load(f)

            # 格式化日志

            lines = []

            app_name = cl_data.get("app_name", "云南达利ZPP011生产偏差分析器")

            author = cl_data.get("author", "裴盛清")

            versions = cl_data.get("versions", [])

            lines.append(app_name)

            lines.append(f"作者：{author}")

            lines.append("=" * 50)

            for v in versions:
                lines.append(f"\n【{v.get('version', '')}】{v.get('date', '')}")

                for feat in v.get("features", []):
                    lines.append(f"  ✦ {feat}")

                for fix in v.get("fixes", []):
                    lines.append(f"  🔧 {fix}")

                for opt in v.get("optimizations", []):
                    lines.append(f"  ⚡ {opt}")

                for les in v.get("lessons", []):
                    lines.append(f"  📌 {les}")

            changelog_text = "\n".join(lines)

            # 弹出保存对话框

            import datetime as _dt

            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")

            default_name = self.config.get(
                "export.default_changelog_filename", "ZPP011版本日志_{ts}.txt"
            ).format(ts=ts)

            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                initialfile=default_name,
                title="导出版本日志",
            )

            if not path:
                return

            with open(path, "w", encoding="utf-8") as f:
                f.write(changelog_text)

            self.log(f"版本日志已导出：{os.path.basename(path)}", "success")

            messagebox.showinfo("导出成功", f"版本日志已保存到：\n{path}")

        except Exception as e:
            self.log(f"导出版本日志失败：{e}", "error")

            messagebox.showerror("导出失败", str(e))

    def _save_audit_back(self):
        """保存审核结果到原 Excel，同时同步到本地数据库"""

        # ── 1. 前置检查 ──

        self.log("💾 正在保存审核结果...", "info")

        if self.audit_data is None or self.audit_data.empty:
            messagebox.showwarning("提示", "没有审核数据可保存")

            return

        src_path = self.input_file.get()

        if not src_path or not os.path.exists(src_path):
            messagebox.showerror("错误", "原始文件不存在，请先选择正确的输入文件")

            return

        try:
            # ── 2. 备份原文件 ──

            self.log("📦 正在备份原文件...", "info")

            backup_dir = os.path.dirname(src_path)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")

            backup_name = f"{os.path.splitext(os.path.basename(src_path))[0]}_审核前备份_{ts}.xlsx"

            backup_path = os.path.join(backup_dir, backup_name)

            shutil.copy2(src_path, backup_path)

            self.log(f"📦 已备份原文件：{backup_name}", "info")

        except Exception as e:
            self.log(f"⚠ 备份失败：{e}", "warning")

            if not messagebox.askyesno(
                "备份失败", f"无法备份原文件：{e}\n是否不备份直接保存？"
            ):
                return

        # ── 3. 打开 Excel 写入审核结果 ──

        try:
            self.log("📂 正在打开 Excel 文件...", "info")

            wb = load_workbook(src_path)

            ws = wb["Data"]

            self.log(
                f"   Excel 已打开，共 {ws.max_row} 行 x {ws.max_column} 列", "info"
            )

        except Exception as e:
            self.log(f"❌ 打开 Excel 失败：{e}", "error")

            messagebox.showerror(
                "保存失败", f"无法打开原始文件：\n{e}\n请关闭 Excel 后重试。"
            )

            return

        # ── 4. 构建待写入的数据列表 ──

        save_list = []

        # 订单列查找（支持多种列名）

        order_col = None

        for possible in [
            "流程订单",
            "订单号",
            "订单编号",
            "订单号码",
            "订单No",
            "Order No",
            "生产订单",
        ]:
            if possible in self.audit_data.columns:
                order_col = possible

                break

        if order_col is None:
            wb.close()

            self.log(
                f"❌ 审核数据中缺少订单列，实际列名: {list(self.audit_data.columns)[:10]}",
                "error",
            )

            messagebox.showerror(
                "保存失败",
                f"审核数据中缺少订单号列，无法定位原表行。\n实际列名: {list(self.audit_data.columns)[:10]}",
            )

            return

        for _, row in self.audit_data.iterrows():
            work_date = str(row.get("订单日期", ""))[:10]

            order_no = str(row.get(order_col, ""))

            mat_code = str(row.get("组件物料号", ""))

            if not work_date or not order_no or not mat_code:
                continue

            # 取最终备注：优先取已有备注，其次取 AI 建议

            remark = str(row.get("备注原因", "") or row.get("AI建议", "") or "").strip()

            save_list.append((work_date, order_no, mat_code, remark))

        self.log(f"   待保存记录：{len(save_list)} 条", "info")

        # 调试：打印 save_list 前3条

        if len(save_list) > 0:
            self.log(f"[DEBUG] save_list 前3条: {save_list[:3]}", "debug")

        # ── 5. 匹配 Excel 表头列索引 ──

        headers = {}

        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(1, col_idx).value

            if val:
                headers[val.strip()] = col_idx

        date_col = headers.get("订单开始日期") or headers.get("订单日期")

        order_col_excel = headers.get("流程订单") or headers.get("订单号")

        mat_col = headers.get("组件物料号") or headers.get("物料编码")

        if not all([date_col, order_col_excel, mat_col]):
            wb.close()

            missing = []

            if not date_col:
                missing.append("日期")

            if not order_col_excel:
                missing.append("订单号")

            if not mat_col:
                missing.append("物料编码")

            self.log(f"❌ Excel 中缺少关键列：{', '.join(missing)}", "error")

            messagebox.showerror(
                "列匹配失败",
                f"原始文件中未找到以下关键列：\n{', '.join(missing)}\n无法定位写入位置。",
            )

            return

        self.log(
            f"   定位到的列：日期={date_col}, 订单={order_col_excel}, 物料={mat_col}",
            "info",
        )

        # ── 6. 写入审核状态列 ──

        audit_cols = {
            "审核状态": None,
            "审核备注": None,
            "审核人": None,
            "审核时间": None,
        }

        next_col = ws.max_column + 1

        for col_name in audit_cols:
            if col_name in headers:
                audit_cols[col_name] = headers[col_name]

            else:
                ws.cell(1, next_col, col_name).font = Font(bold=True)

                audit_cols[col_name] = next_col

                next_col += 1

        self.log(
            f"   审核列：审核状态={audit_cols['审核状态']}, 审核备注={audit_cols['审核备注']}",
            "info",
        )

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        auditor = os.getlogin()

        saved_count = 0

        match_count = 0

        for r_idx in range(2, ws.max_row + 1):
            r_date = (
                str(ws.cell(r_idx, date_col).value)[:10]
                if ws.cell(r_idx, date_col).value
                else ""
            )

            r_order = str(ws.cell(r_idx, order_col_excel).value).strip()

            r_mat = str(ws.cell(r_idx, mat_col).value).strip()

            match = next(
                (
                    item
                    for item in save_list
                    if item[0] == r_date and item[1] == r_order and item[2] == r_mat
                ),
                None,
            )

            if match:
                match_count += 1

                _, _, _, remark = match

                # 调试：打印匹配到的行号和匹配键

                self.log(
                    f"[DEBUG] 匹配到行 {r_idx}: date={r_date}, order={r_order}, mat={r_mat}, 写入备注={remark[:20] if remark else '空'}",
                    "debug",
                )

                ws.cell(r_idx, audit_cols["审核状态"], "已备注" if remark else "未审核")

                ws.cell(r_idx, audit_cols["审核备注"], remark)

                ws.cell(r_idx, audit_cols["审核人"], auditor)

                ws.cell(r_idx, audit_cols["审核时间"], now_str)

                saved_count += 1

        self.log(f"   匹配到 {match_count} 行，待写入 {saved_count} 行", "info")

        # ── 7. 保存 Excel 文件 ──

        try:
            self.log("💾 正在保存 Excel 文件...", "info")

            wb.save(src_path)

            wb.close()

            self.log("   Excel 保存成功", "info")

        except PermissionError:
            wb.close()

            self.log("❌ 文件被占用，无法保存。请关闭 Excel 后重试。", "error")

            messagebox.showerror(
                "保存失败", "文件被其他程序占用，请关闭 Excel 后重试。"
            )

            return

        except Exception as e:
            wb.close()

            self.log(f"❌ 写入 Excel 失败：{e}", "error")

            messagebox.showerror("保存失败", f"写入文件时出错：\n{e}")

            return

        finally:
            try:
                wb.close()

            except:
                pass

        # ── 8. 同步到 SQLite 审核数据库 ──

        try:
            self.log("📊 正在同步到审核数据库...", "info")

            # 构造适合 storage 模块的 DataFrame（确保有订单号列）

            save_df = self.audit_data.copy()

            if "订单号" not in save_df.columns and "流程订单" in save_df.columns:
                save_df["订单号"] = save_df["流程订单"]

            if "订单日期" not in save_df.columns:
                save_df["订单日期"] = save_df.get("订单日期", "")

            self.audit_presenter.save_audit_back(auditor=auditor)

            self.log("   数据库同步成功", "info")

        except Exception as e:
            self.log(f"⚠ 审核数据库同步失败（不影响 Excel 保存）：{e}", "warn")

        # ── 9. 最终反馈 ──

        self.log(
            f"✅ 审核结果已保存：Excel 写入 {saved_count} 行，备份 {backup_name}",
            "success",
        )

        messagebox.showinfo(
            "保存成功",
            f"审核结果已写入原始文件 {saved_count} 行。\n"
            f"新增列：审核状态 / 审核备注 / 审核人 / 审核时间\n"
            f"原始备份：{backup_name}",
        )


        # 记录审计日志（Task 004）
        if hasattr(self, 'audit_logger'):
            self.audit_logger.log(
                action='save_audit',
                extra={'saved_count': saved_count},
                source='manual'
            )

    def _batch_export(self, event=None):
        """批量导出选中行为Excel或CSV"""

        selected = self.audit_tree.selection()

        if not selected:
            messagebox.showwarning("提示", "请先选择要导出的行")

            return

        file_path = filedialog.asksaveasfilename(
            title="导出选中行",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )

        if not file_path:
            return

        columns = self.audit_tree["columns"]

        export_data = []

        for item in selected:
            row_values = [self.audit_tree.set(item, col) for col in columns]

            export_data.append(row_values)

        try:
            if file_path.endswith(".csv"):
                import csv

                with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)

                    writer.writerow(columns)

                    writer.writerows(export_data)

            else:
                pd.DataFrame(export_data, columns=columns).to_excel(
                    file_path, index=False, engine="openpyxl"
                )

            messagebox.showinfo(
                "导出成功", f"已成功导出 {len(export_data)} 行数据到\n{file_path}"
            )

        except Exception as e:
            messagebox.showerror("导出失败", f"导出时发生错误：{str(e)}")

    def _copy_wechat_draft(self):
        """将选中行生成微信草稿并复制到剪贴板"""

        selected = self.audit_tree.selection()

        if not selected:
            messagebox.showwarning("提示", "请先选择要生成草稿的行")

            return

        cols = self.audit_tree["columns"]

        lines = [self.config.get("wechat.draft_title", "【料控指令】"), ""]

        for i, item in enumerate(selected, 1):
            vals = self.audit_tree.item(item, "values")

            data = dict(zip(cols, vals))

            order_no = data.get("order_date", "")

            mat_name = data.get("name", "")

            dev_rate = data.get("dev_rate", "")

            status = data.get("status", "")

            remark = data.get("remark", "")

            code = data.get("code", "")

            # 简洁格式：序号. 物料 偏差率 状态

            line = f"{i}. {mat_name}（{code}）偏差{dev_rate} {status}"

            if remark and remark.strip():
                line += f"\n   备注：{remark}"

            lines.append(line)

        lines.append("")

        lines.append(f"共 {len(selected)} 条 | 请确认后处理")

        draft = "\n".join(lines)

        self.root.clipboard_clear()

        self.root.clipboard_append(draft)

        self.log(f"📋 已复制 {len(selected)} 条微信草稿到剪贴板", "info")

        messagebox.showinfo(
            "复制成功", f"已复制 {len(selected)} 条指令到剪贴板，可直接粘贴到微信"
        )
