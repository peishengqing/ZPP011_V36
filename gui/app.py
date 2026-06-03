#!/usr/bin/env python

# -*- coding: utf-8 -*-

"""

云南达利ZPP011生产偏差分析器 — 多条件智能筛选 v36

基于 v30 增强：

  - 审核结果回写 Excel（用日期+订单号+物料编码定位，不依赖行号）

  - SQLite 本地审核记录库（用户目录，不随程序移动丢失）

  - 导出/导入审核备份（ZIP）

  - 保留原表行号用于 GUI 对照

"""

import pandas as pd

import numpy as np

from openpyxl import Workbook, load_workbook

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from openpyxl.chart import BarChart, LineChart, Reference

from openpyxl.utils import get_column_letter

from datetime import datetime
from core.advanced_ppt_generator import generate_advanced_report

import os
import re
import glob as _glob
import sys as _sys
import sys
import threading
import time
import json
import tempfile
import subprocess
import queue
import traceback

import shutil

import sqlite3

import zipfile

import yaml

from pathlib import Path

import tkinter as tk

from tkinter import scrolledtext, messagebox, filedialog, ttk, simpledialog

from modules.audit.presenters.audit_presenter import AuditPresenter

from modules.audit.models.audit_model import AuditModel

from widgets import C


# ── 模块化组件 ───────────────────────────────────

from storage import init_audit_db, restore_audit_from_db

from domain.alt_material import alt_manager

from analysis.analyzer import do_analysis_v2


from gui.ui_builder import build_ui

import ppt_generator

from domain.alt_material.alt_manager import (
    DEFAULT_ALT_PAIRS,
    save_alt_pairs,
    load_alt_pairs,
)

from core.config_manager import ConfigManager

import core.task_manager

from core.ai_client import AIClient

from gui.events import EventsMixIn

from utils.version_history import (
    get_version_display,
    get_current_version,
    APP_NAME as _VH_APP_NAME,
)


# ── Task 003/004：备份管理器 & 审计日志 ──

from core.audit_logger import AuditLogger
from core.view_manager import ViewManager


# ── 备注标准化（已迁移到 utils/helpers.py）───

from utils.helpers import standardize_remark


# ── 动态阈值计算 ─────────────────────────────


def calc_dynamic_threshold(df, method="robust"):
    """

    根据数据自动计算偏差率阈值（改进版）

    method:

      'robust'  — 中位数 + 1.5×IQR（对极值不敏感，推荐）

      'percentile' — 指定分位数法

      'std'    — 均值 + 2σ

    返回: (threshold, description_string)

    """

    rates = df["偏差率(%)"].dropna().astype(float).abs()

    if rates.empty:
        return 10.0, "数据为空，使用默认阈值10%"

    if method == "robust":
        q75 = rates.quantile(0.75)

        q25 = rates.quantile(0.25)

        iqr = q75 - q25

        thresh = float(q75 + 1.5 * iqr)

        thresh = max(thresh, 5.0)

        thresh = min(thresh, 40.0)

        desc = f"动态阈值（稳健法：Q75+1.5×IQR）：±{thresh:.1f}%"

        return round(thresh, 1), desc

    elif method == "percentile":
        p = 85

        thresh = float(rates.quantile(p / 100))

        thresh = max(thresh, 5.0)

        thresh = min(thresh, 35.0)

        desc = f"动态阈值（{p}分位数法）：±{thresh:.1f}%"

        return round(thresh, 1), desc

    else:
        mu = rates.mean()

        sigma = rates.std()

        thresh = float(mu + 2 * sigma)

        thresh = max(thresh, 5.0)

        thresh = min(thresh, 40.0)

        desc = f"动态阈值（均值+2σ法）：±{thresh:.1f}%"

        return round(thresh, 1), desc


# 美化版 GUI

# ─────────────────────────────────────────────


def _get_mode_config_dir():
    """获取 mode.json 配置目录（兼容开发和 EXE 环境）"""

    if getattr(sys, "frozen", False):
        # EXE 模式：使用用户目录下的 .zpp011_audit（运行时可写）

        config_dir = os.path.join(os.path.expanduser("~"), ".zpp011_audit")

    else:
        # 开发模式：项目根目录下的 .zpp011_audit

        config_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), ".zpp011_audit"
        )

    os.makedirs(config_dir, exist_ok=True)

    return config_dir


class ModeSelector:
    """启动模式选择窗口（使用 Toplevel，避免创建多个 tk.Tk 实例）"""

    def __init__(self, parent):

        self.selected_mode = None

        self.parent = parent

        # 从 utils.version_history 动态读取版本号

        _title = get_version_display()

        # 使用 Toplevel 而非 tk.Tk()，避免多个 Tk 实例导致状态混乱

        self.win = tk.Toplevel(parent)

        self.win.title(_title)

        self.win.geometry("400x300")

        self.win.resizable(False, False)

        self.win.grab_set()  # 模态

        self.win.focus_force()

        # 居中窗口

        self.win.update_idletasks()

        sw = self.win.winfo_screenwidth()

        sh = self.win.winfo_screenheight()

        x = (sw - 400) // 2

        y = (sh - 300) // 2

        self.win.geometry(f"+{x}+{y}")

        # 标题

        tk.Label(
            self.win, text="请选择启动模式", font=("Microsoft YaHei", 12, "bold")
        ).pack(pady=20)

        # 模式按钮

        tk.Button(
            self.win,
            text="📊 生产偏差分析",
            width=30,
            height=2,
            command=lambda: self._start("analysis"),
        ).pack(pady=5)

        tk.Button(
            self.win,
            text="📦 库存流水管理",
            width=30,
            height=2,
            command=lambda: self._start("inventory"),
        ).pack(pady=5)

        # 记住选择

        self.remember_var = tk.BooleanVar(value=False)

        tk.Checkbutton(
            self.win, text="记住我的选择，下次不再询问", variable=self.remember_var
        ).pack(pady=15)

        # 底部按钮

        bottom = tk.Frame(self.win)

        bottom.pack(side="bottom", fill="x", padx=10, pady=10)

        tk.Button(bottom, text="设置默认模式", command=self._set_default).pack(
            side="left"
        )

        tk.Button(bottom, text="关闭", command=self._close).pack(side="right")

        # 关闭窗口时退出

        self.win.protocol("WM_DELETE_WINDOW", self._close)

        # 隐藏父窗口（模式选择期间）

        parent.withdraw()

        # 等待窗口关闭（模拟模态对话框）

        self.win.wait_window(self.win)

    def _start(self, mode):

        if self.remember_var.get():
            self._save_default(mode)

        self.selected_mode = mode

        self.parent.deiconify()  # 恢复父窗口

        self.win.destroy()

    def _close(self):

        self.selected_mode = None

        self.parent.deiconify()  # 恢复父窗口

        self.win.destroy()

    def _save_default(self, mode):
        """保存默认模式到配置文件"""

        config_dir = _get_mode_config_dir()

        config_path = os.path.join(config_dir, "mode.json")

        with open(config_path, "w", encoding="utf-8") as f:
            json.dump({"default_mode": mode}, f, ensure_ascii=False, indent=2)

    def _set_default(self):
        """点击"设置默认模式"按钮时弹出菜单"""

        menu = tk.Menu(self.win, tearoff=0)

        menu.add_command(label="🔄 清除默认设置", command=self._clear_default)

        menu.add_separator()

        menu.add_command(label="❌ 取消", command=lambda: None)

        try:
            menu.tk_popup(self.win.winfo_pointerx(), self.win.winfo_pointery())

        finally:
            menu.grab_release()

    def _clear_default(self):

        config_dir = _get_mode_config_dir()

        config_path = os.path.join(config_dir, "mode.json")

        if os.path.exists(config_path):
            os.remove(config_path)

        messagebox.showinfo("已清除", "默认模式已清除，下次启动将重新询问。")


class ZPP011Beautiful(EventsMixIn):
    def __init__(self, root):

        self.root = root

        self.is_exporting = False  # Export lock (Task 4.3)

        # 从 utils.version_history 动态读取版本号

        self.root.title(get_version_display())

        self.root.geometry("1200x820")

        self.root.minsize(1000, 700)

        self.root.configure(bg=C["bg"])

        # ── 菜单栏 ─────────────────

        menubar = tk.Menu(self.root)

        # Task 009：历史对比菜单

        history_menu = tk.Menu(menubar, tearoff=0)

        menubar.add_cascade(label="历史", menu=history_menu)

        history_menu.add_command(label="历史对比", command=self._show_history_compare)

        history_menu.add_separator()

        history_menu.add_command(
            label="📂 查看历史源码", command=self._open_source_backup
        )

        history_menu.add_separator()

        history_menu.add_command(
            label="📊 管理看板", command=self._show_management_dashboard
        )

        history_menu.add_command(
            label="🔍 AI 归因分析", command=self._show_attribution_standalone
        )

        help_menu = tk.Menu(menubar, tearoff=0)

        menubar.add_cascade(label="帮助", menu=help_menu)

        help_menu.add_command(label="快捷键说明", command=self._show_shortcuts_help)

        help_menu.add_separator()

        help_menu.add_command(label="规则配置", command=self._open_rule_config)

        help_menu.add_separator()

        help_menu.add_command(label="效益报告", command=self._show_benefit_report)

        help_menu.add_separator()

        help_menu.add_command(
            label="生成详细分析报告", command=self._generate_full_report
        )

        help_menu.add_command(
            label="生成详细分析报告(新版V2)", command=self._generate_advanced_report_v2
        )

        help_menu.add_command(
            label="生成ZPP011完整报告(V3)", command=self._generate_zpp011_report_v3
        )

        help_menu.add_command(
            label="生成企业级报告(原版)", command=self._generate_enterprise_ppt
        )

        help_menu.add_command(
            label="生成ZPP011偏差分析报告", command=self._generate_zpp011_report
        )

        help_menu.add_separator()

        help_menu.add_command(label="关于", command=self._show_about)

        self.root.config(menu=menubar)

        # ── 快捷键绑定（仅在窗口内生效）──────────────

        self.root.bind("<Control-s>", lambda e: self._save_audit_back())

        self.root.bind("<Control-e>", lambda e: self._export_audit_excel())

        self.root.bind("<Control-a>", lambda e: self._run_ai_audit())

        self.root.bind("<F1>", lambda e: self._show_shortcuts_help())

        self.root.bind("<Control-q>", lambda e: self.root.quit())

        self.running = False

        self.cancel_req = False

        self.output_path = None

        self.start_time = None

        self.timer_id = None

        self.alt_pairs = list(DEFAULT_ALT_PAIRS)

        self.input_file = tk.StringVar()

        self.output_dir = tk.StringVar(
            value=os.path.join(os.path.expanduser("~"), "Documents", "ZPP011分析报告")
        )

        self.start_date = tk.StringVar()

        self.end_date = tk.StringVar()

        # 日期筛选变量（日历选择器）

        self.date_start_val = None

        self.date_end_val = None

        self._tmp_start = None

        self._tmp_end = None

        self._picker_year = None

        self._picker_month = None

        self.material_search = tk.StringVar()

        self.filter_vars = {}

        self.filter_widgets = {}

        self.custom_statuses = []

        self.material_list = []

        self.code_to_info = {}

        init_audit_db()

        self.audit_model = AuditModel()

        self.audit_presenter = AuditPresenter(self.audit_model, self)

        # ── Task 012：视图管理器 ──
        self.view_manager = ViewManager()

        # ── 侧边栏筛选面板：必须在 build_ui 之前 pack，才能正确占位 ──

        self.config = ConfigManager()

        self.filter_panel = None

        # 读取项目级特性开关（config/defaults.json，非用户配置）

        _feature_enabled = False

        try:
            _defaults_path = os.path.join(
                os.path.dirname(os.path.dirname(__file__)), "config", "defaults.json"
            )

            if os.path.exists(_defaults_path):
                with open(_defaults_path, "r", encoding="utf-8") as _f:
                    _defaults = json.load(_f)

                _feature_enabled = _defaults.get("features", {}).get(
                    "new_filter_panel", False
                )

        except Exception:
            pass

        if _feature_enabled:
            try:
                from gui.filter_panel import FilterPanel

                self.filter_panel = FilterPanel(
                    self.root, on_filter_changed=self._on_sidebar_filter_changed
                )

                self.filter_panel.pack(side="right", fill="y")

            except Exception as e:
                import traceback

                # FilterPanel 初始化失败，忽略
        build_ui(self)
        self._check_and_upgrade_db()  # v37.44 启动时检测并升级旧数据库

        self.config.apply_window_geometry(self.root)

        # 侧边栏加载成功日志（此时 log 控件已就绪）

        if self.filter_panel is not None:
            self.log("✅ 侧边栏筛选面板已加载", "success")

        self._restore_column_widths()

        self._init_column_width_tracking()  # Task 008：列宽变化追踪

        self._init_sort_columns()  # 初始化多列排序系统

        # 菜单栏已在上方创建，无需再次初始化

        self.log("✅ UI 初始化完成，日志系统测试", "success")

        self.alt_pairs = load_alt_pairs(log_cb=self.log)

        self._refresh_alt_view(self._alt_inner)

        self._auto_find()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # 状态栏

        self.status_var = tk.StringVar()

        self.status_var.set("就绪")

        self.status_bar = tk.Label(
            self.root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W
        )

        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # 任务管理器 + 进度条

        self.task_manager = core.task_manager.TaskManager(max_workers=2)

        # 审计日志器（Task 004）

        from core.audit_logger import AuditLogger

        self.audit_logger = AuditLogger()

        self.ai_client = AIClient()

        self.is_auditing = False

        self.unsaved_ai_results = False

        self.progress_bar = ttk.Progressbar(self.root, mode="indeterminate", length=200)

        self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2)

        self.progress_bar.pack_forget()  # 初始隐藏

        # S01 库存检查状态

        self._is_s01_processing = False

        self._s01_cancel_flag = None

        self._s01_thread = None

        self._tab_data_cache = {}

        # S01 异常高亮配置（配置化，从 config/s01_display.yaml 加载）

        self.s01_display_config = self._load_s01_display_config()

        self._ensure_temp_dir()

        # 绑定 Tab 切换事件

        if hasattr(self, "notebook"):
            self.notebook.bind("<<NotebookTabChanged>>", self._s01_on_tab_changed)

        #

        # Task 011: Dashboard quick button (menu already available at 历史 >
        # 管理看板)

        # Button added in _show_management_dashboard via menu

        # 启动回调轮询
        self.task_manager.poll(self.root)

    def _on_sidebar_filter_changed(self, filters):
        """边栏筛选条件变更时，合并顶部栏条件，使用 FilterEngine 统一过滤"""

        if not hasattr(self, "audit_data") or self.audit_data is None:
            return

        # 懒加载 FilterEngine

        if not hasattr(self, "filter_engine") or self.filter_engine is None:
            from modules.audit.filters.filter_engine import FilterEngine

            self.filter_engine = FilterEngine()

        # ── 收集顶部筛选栏的值（不覆盖侧边栏已有的 key）──

        top_keys_used = set(filters.keys())

        # 搜索关键词

        if "search" not in top_keys_used and hasattr(self, "search_var"):
            st = self.search_var.get().strip()

            default_hint = "输入任意关键词，实时过滤全部列..."

            if st and st != default_hint:
                filters["search"] = st

        # 日期范围（从顶部栏的 DateEntry 控件读取，不是 self.date_start_val）

        if "date_start" not in top_keys_used and "date_end" not in top_keys_used:
            if hasattr(self, "filter_widgets") and "order_date" in self.filter_widgets:
                date_widgets = self.filter_widgets["order_date"]

                if isinstance(date_widgets, tuple) and len(date_widgets) == 2:
                    try:
                        if date_widgets[0].get_date():
                            filters["date_start"] = (
                                date_widgets[0].get_date().strftime("%Y-%m-%d")
                            )

                        if date_widgets[1].get_date():
                            filters["date_end"] = (
                                date_widgets[1].get_date().strftime("%Y-%m-%d")
                            )

                    except Exception:
                        pass

        # 顶部栏下拉框值（remark, ai_result, _color, audit_source, remark_check_status）

        top_combos = [
            "remark",
            "ai_result",
            "_color",
            "audit_source",
            "remark_check_status",
        ]

        if hasattr(self, "filter_widgets"):
            for key in top_combos:
                if key not in top_keys_used and key in self.filter_widgets:
                    w = self.filter_widgets[key]

                    if hasattr(w, "get"):
                        val = w.get()

                        if val and val != "全部":
                            filters[key] = val

        df_filtered = self.filter_engine.apply(filters, self.audit_data)

        # 刷新表格和统计

        self._refresh_audit_tree(df_filtered)

        self._update_audit_stats(df_filtered)

        self.status_lbl.configure(text=f"筛选结果：{len(df_filtered)} 条")

    def _on_filter_panel_expand(self, expanded):
        """侧边栏展开/折叠时的回调（消除平移抖动）"""

        try:
            # 动态获取侧边栏宽度（消除硬编码 250）

            sidebar_width = getattr(self.filter_panel, "width", 250)

            new_width = self.root.winfo_width() - sidebar_width

            # 优化：仅在宽度变化超过 5px 时执行 update_idletasks

            current_width = self.table_frame.winfo_width()

            if abs(new_width - current_width) > 5:
                self.table_frame.configure(width=new_width)

                self.root.update_idletasks()

        except Exception as e:
            pass  # 忽略侧边栏展开错误

    def set_status(self, msg):

        self.status_var.set(msg)

        self.root.update_idletasks()

    def clear_status(self):

        self.status_var.set("")

    # ── S01 库存检查相关方法 ─────────────────────────────

    def _load_s01_display_config(self) -> dict:
        """加载 S01 异常高亮配置，文件缺失/损坏时使用内置默认规则"""

        _default = {
            "rules": [
                {
                    "name": "呆滞预警",
                    "condition": "days_in_stock >= 30 and days_in_stock < 60",
                    "color": "#FFFF00",
                    "tag": "warning",
                },
                {
                    "name": "超限告警",
                    "condition": "days_in_stock >= 60",
                    "color": "#FF0000",
                    "tag": "critical",
                },
            ],
            "default_color": "#FFFFFF",
        }

        config_path = Path(__file__).parent.parent / "config" / "s01_display.yaml"

        if not config_path.exists():
            self.log(
                f"⚠ S01 高亮配置不存在 {config_path}，使用内置默认规则", "warning"
            ) if hasattr(self, "log") else None

            return _default

        try:
            with open(config_path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)

            if not isinstance(data, dict) or "rules" not in data:
                raise ValueError("配置文件缺少 'rules' 字段")

            return data

        except Exception as e:
            self.log(
                f"⚠ 加载 S01 高亮配置失败: {e}，使用内置默认规则", "warning"
            ) if hasattr(self, "log") else None

            return _default

    def _ensure_temp_dir(self):
        """确保用户目录下存在 temp/ 文件夹（打包后也可写）"""

        from pathlib import Path

        # 使用用户目录下的 .zpp011_audit/temp，确保可写

        temp_dir = Path.home() / ".zpp011_audit" / "temp"

        temp_dir.mkdir(parents=True, exist_ok=True)

        self._temp_dir = temp_dir

    def _s01_on_tab_changed(self, event=None):
        """切换 Tab 时保存/恢复数据（深拷贝）"""

        current_tab = self.notebook.index(self.notebook.select())

        # 保存当前 Tab 数据

        if hasattr(self, "audit_data") and self.audit_data is not None:
            self._tab_data_cache[current_tab] = self.audit_data.copy(deep=True)

        # 恢复新 Tab 数据

        self.audit_data = self._tab_data_cache.get(current_tab, pd.DataFrame()).copy(
            deep=True
        )

        self._refresh_audit_tree(self.audit_data)

    def _s01_clean_temp_files(self):
        """清理项目 temp/ 目录下的所有 S01 临时文件"""

        import glob

        patterns = ["*.s01.tmp", "*.s01.temp"]

        for pattern in patterns:
            for f in glob.glob(str(self._temp_dir / pattern)):
                try:
                    import os as _os

                    _os.remove(f)

                except OSError:
                    pass

    # ── 排序已由 EventsMixIn._init_sort_columns + _on_tree_sort 接管

    # bind_multi_sort 已禁用，避免两套排序系统冲突（F1修复）

    # def _init_sort_columns(self):

    #     from gui.tree_utils import bind_multi_sort

    #     self._sort_states = {}

    #     audit_sort_key = "audit"

    #     self._sort_states[audit_sort_key] = {}

    #     audit_cols = ("idx", "excel_row", "factory", "admin", "order_date",

    #                   "order_no", "code", "name", "quota", "actual", "dev_rate",

    #                   "is_alt", "status", "remark", "batch_remark", "audit_result",

    #                   "AI建议", "audit_status", "audit_source", "deviation_amount")

    #     def audit_state_ref():

    #         return self._sort_states[audit_sort_key]

    #     bind_multi_sort(self.audit_tree, audit_state_ref, audit_cols)

    # ── 审核数据库相关 ─────────────────────────────

    def _get_changelog_path(self) -> str:
        """获取 changelog.json 的正确路径（兼容开发和 exe 环境）"""

        if getattr(sys, "frozen", False):
            # EXE环境：先找EXE同目录，再找sys._MEIPASS

            _exe_dir = os.path.dirname(sys.executable)

            _cl_path_exe = os.path.join(_exe_dir, "changelog.json")

            if os.path.isfile(_cl_path_exe):
                return _cl_path_exe

            elif hasattr(sys, "_MEIPASS"):
                _cl_path = os.path.join(sys._MEIPASS, "changelog.json")

                if os.path.isfile(_cl_path):
                    return _cl_path

        else:
            # 开发环境：先找workspace_dir，再找main.py同目录

            _cl_path = os.path.join(
                getattr(
                    self, "workspace_dir", os.path.dirname(os.path.abspath(__file__))
                ),
                "changelog.json",
            )

            if os.path.isfile(_cl_path):
                return _cl_path

            _cl_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "changelog.json"
            )

            if os.path.isfile(_cl_path):
                return _cl_path

        return None

    # ── 版本信息（从 changelog.json 动态读取） ──────────

    def _generate_excel_thread(self, output_path: str):
        """后台线程生成完整Excel，带详细错误日志"""

        import traceback

        try:
            self.log(f"[DEBUG] 生成表格线程启动，输出路径：{output_path}", "info")

            input_file = self.input_file.get()

            result = do_analysis_v2(
                input_file=input_file,
                output_dir=os.path.dirname(output_path),
                alt_pairs=self.alt_pairs,
                progress_callback=lambda *a: None,
                cancel_check=None,
                start_date=self.start_date.get() or None,
                end_date=self.end_date.get() or None,
                material_search=self.material_search.get() or None,
                output_path=output_path,
            )

            self.log(f"[DEBUG] do_analysis_v2 返回：{result}", "info")

            if result is None:
                raise RuntimeError("do_analysis_v2 返回了 None，分析过程可能出错")

            # 生成成功，询问用户是否打开文件

            def on_success():
                # ✅ 保存分析结果路径，供PPT生成等功能使用
                self._analysis_output_path = result
                self.log(f"✅ 已保存分析结果路径：{result}", "info")

                self.log(f"✅ 表格生成：{os.path.basename(result)}", "success")

                self.excel_btn.configure(state="normal", text="📋 生成表格")

                self.status_lbl.configure(
                    text=f"已生成：{os.path.basename(result)}", fg=C["green"]
                )

                # 弹出询问框

                if messagebox.askyesno(
                    "生成成功",
                    f"表格已生成：\n{os.path.basename(result)}\n\n是否立即打开？",
                ):
                    try:
                        os.startfile(result)

                        self.log(f"✅ 已打开文件：{result}", "info")

                    except Exception as e:
                        self.log(f"⚠️ 无法打开文件：{e}", "warning")

                        messagebox.showwarning("打开失败", f"无法打开文件：\n{e}")

            self.root.after(0, on_success)

        except Exception as e:
            tb = traceback.format_exc()

            err_file = os.path.join(
                os.path.expanduser("~"), "Desktop", "gen_excel_error.log"
            )

            with open(err_file, "w", encoding="utf-8") as f:
                f.write(f"生成表格失败：{e}\n\n{tb}")

            # 检查是否是文件占用错误

            error_msg = str(e)

            if "Permission denied" in error_msg or "PermissionError" in error_msg:
                # 文件被占用的友好提示

                self.root.after(
                    0,
                    lambda: messagebox.showwarning(
                        "文件被占用",
                        "⚠️ 无法保存文件\n\n"
                        "可能的原因：\n"
                        "  • 文件已用 Excel 打开\n"
                        "  • 文件被 WPS 或其他程序占用\n\n"
                        "解决方法：\n"
                        "  1. 关闭 Excel 中打开的这个文件\n"
                        "  2. 点击「生成表格」重试\n"
                        "  3. 或者另存为其他文件名",
                    ),
                )

                self.root.after(
                    0, lambda: self.log(f"⚠️ 文件被占用，请关闭 Excel 后重试", "warning")
                )

            else:
                # 其他错误的通用提示

                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "生成失败",
                        f"生成表格时出错：\n{e}\n\n详细错误已保存到桌面 gen_excel_error.log",
                    ),
                )

                self.root.after(0, lambda: self.log(f"❌ 表格生成失败：{e}", "error"))

            self.root.after(0, lambda: self.log(tb, "error"))

            self.root.after(
                0, lambda: self.excel_btn.configure(state="normal", text="📋 生成表格")
            )

            self.root.after(
                0, lambda: self.status_lbl.configure(text="就绪", fg=C["text_dim"])
            )

    # 从Excel加载物料列表

    def _load_material_list(self):

        excel_path = self.input_file.get()

        self.material_list = []

        self.code_to_info = {}  # {code: (factory, code, name)}

        if excel_path and os.path.exists(excel_path):
            try:
                df = pd.read_excel(excel_path, sheet_name="Data")

                code_cols = [
                    c
                    for c in df.columns
                    if any(
                        k in str(c).lower()
                        for k in ["组件物料号", "组件编码", "物料编码", "code", "编码"]
                    )
                ]

                name_cols = [c for c in df.columns if c == "组件物料描述"] or [
                    c
                    for c in df.columns
                    if any(
                        k in str(c).lower()
                        for k in ["物料描述", "组件描述", "名称", "name"]
                    )
                ]

                factory_cols = [
                    c
                    for c in df.columns
                    if any(k in str(c).lower() for k in ["工厂名称", "工厂", "factory"])
                ]

                if code_cols:
                    seen = set()

                    for _, row in df.iterrows():
                        code = str(row[code_cols[0]])

                        if code and code != "nan" and code not in seen:
                            seen.add(code)

                            name = str(row[name_cols[0]]) if name_cols else ""

                            factory = str(row[factory_cols[0]]) if factory_cols else ""

                            name = name if name and name != "nan" else ""

                            factory = factory if factory and factory != "nan" else ""

                            self.code_to_info[code] = (factory, code, name)

                            # 下拉框显示: 工厂名称 | 物料编码 | 物料描述

                            display = (
                                f"{factory} | {code} | {name}"
                                if factory
                                else f"{code} | {name}"
                            )

                            self.material_list.append(display)

                    self.material_list.sort()

            except Exception as e:
                pass  # 加载物料列表失败，忽略

    def _load_audit_database(self):
        """加载已审核的偏差数据（从SQLite数据库恢复备注）"""

        if self.audit_data is None or self.audit_data.empty:
            return

        try:
            restore_audit_from_db(self.audit_data, log_cb=self.log)

        except Exception as e:
            self.log(f"加载审核数据库失败：{e}", "warning")

    def _analyze_trend(self):
        """分析审核数据的阶段性趋势"""

        df = self.audit_data.copy()

        if df is None or len(df) == 0:
            return None

        date_col = None

        for c in ["订单开始日期", "订单日期", "日期", "工单日期"]:
            if c in df.columns:
                date_col = c

                break

        if date_col is None:
            return None

        df["_trend_date"] = pd.to_datetime(df[date_col], errors="coerce")

        df = df.sort_values("_trend_date")

        dates = df["_trend_date"].dropna()

        if len(dates) < 3:
            df.drop(columns=["_trend_date"], inplace=True, errors="ignore")

            return None

        split1 = dates.quantile(0.33)

        split2 = dates.quantile(0.66)

        early = df[df["_trend_date"] <= split1]

        mid = df[(df["_trend_date"] > split1) & (df["_trend_date"] <= split2)]

        recent = df[df["_trend_date"] > split2]

        def calc_metrics(data):

            if len(data) == 0:
                return {"偏差率均值": 0, "偏差金额合计": 0, "通过率": 0}

            dev_rate = data["偏差率(%)"].apply(
                lambda x: float(str(x).replace("%", "")) if isinstance(x, str) else x
            )

            dev_rate_mean = dev_rate.mean()

            dev_amount_sum = data.get("偏差金额", pd.Series([0])).sum()

            if "审核状态" in data.columns:
                pass_rate = len(data[data["审核状态"] == "通过"]) / len(data) * 100

            else:
                pass_rate = 0

            return {
                "偏差率均值": round(dev_rate_mean, 1),
                "偏差金额合计": round(dev_amount_sum, 2),
                "通过率": round(pass_rate, 1),
            }

        return {
            "早期": {
                "日期范围": f"{early['_trend_date'].min().date()}~{early['_trend_date'].max().date()}",
                **calc_metrics(early),
            },
            "中期": {
                "日期范围": f"{mid['_trend_date'].min().date()}~{mid['_trend_date'].max().date()}",
                **calc_metrics(mid),
            },
            "近期": {
                "日期范围": f"{recent['_trend_date'].min().date()}~{recent['_trend_date'].max().date()}",
                **calc_metrics(recent),
            },
        }

    # ── P1#14：更新趋势显示 ──

    def _update_trend_display(self):
        """调用分析趋势并更新UI标签"""

        try:
            trend = self._analyze_trend()

            if not trend:
                for period in ["早期", "中期", "近期"]:
                    self.trend_labels[period]["range"].configure(text="数据不足")

                    self.trend_labels[period]["dev_rate"].configure(text="--")

                    self.trend_labels[period]["dev_amount"].configure(text="--")

                    self.trend_labels[period]["pass_rate"].configure(text="--")

                return

            for period in ["早期", "中期", "近期"]:
                p = trend.get(period, {})

                range_txt = p.get("日期范围", "--")

                dev_rate_txt = f"偏差率: {p.get('偏差率均值', 0)}%"

                dev_amount_txt = f"偏差金额: ¥{p.get('偏差金额合计', 0):,.0f}"

                pass_rate_txt = f"通过率: {p.get('通过率', 0)}%"

                self.trend_labels[period]["range"].configure(text=range_txt)

                self.trend_labels[period]["dev_rate"].configure(text=dev_rate_txt)

                self.trend_labels[period]["dev_amount"].configure(text=dev_amount_txt)

                self.trend_labels[period]["pass_rate"].configure(text=pass_rate_txt)

        except Exception as e:
            self.log(f"趋势显示更新失败: {e}", "warn")

    def _run_pre_check_from_excel(self, output_path=None):
        """从分析结果 Excel 读取数据，生成预检报告并弹窗"""

        if not output_path or not os.path.exists(output_path):
            messagebox.showwarning("文件缺失", "请先生成分析结果 Excel。")

            return

            try:
                # 读取完整偏差明细和汇总统计

                dev_df = pd.read_excel(self.output_path, sheet_name="完整偏差明细")

                summary_df = pd.read_excel(self.output_path, sheet_name="汇总统计")

            except Exception as e:
                self.log(f"⚠ 预检报告生成失败：{e}", "warn")

                return

            results = []

            # 1. 列结构检查

            required_cols = [
                "订单日期",
                "工厂",
                "车间",
                "物料编码",
                "物料名称",
                "偏差率",
                "偏差金额(含税)",
                "备注",
            ]

            for col in required_cols:
                if col not in dev_df.columns:
                    results.append(("严重", f"缺失关键列：{col}"))

                elif dev_df[col].isna().all():
                    results.append(("严重", f"关键列全为空：{col}"))

                elif dev_df[col].isna().mean() > 0.5:
                    results.append(
                        (
                            "警告",
                            f"关键列空值过多：{col} ({dev_df[col].isna().mean():.0%})",
                        )
                    )

                else:
                    results.append(("通过", f"关键列检查通过：{col}"))

            # 2. 数据量检查

            if len(dev_df) == 0:
                results.append(("严重", "偏差明细为空，无任何记录"))

            elif len(dev_df) < 10:
                results.append(("警告", f"偏差记录偏少：{len(dev_df)} 条"))

            else:
                results.append(("通过", f"偏差记录：{len(dev_df)} 条"))

            # 3. 无备注比例

            if "备注" in dev_df.columns:
                no_note = (
                    dev_df["备注"].isna().sum()
                    + (dev_df["备注"].astype(str).str.strip() == "").sum()
                )

                note_rate = (
                    (len(dev_df) - no_note) / len(dev_df) if len(dev_df) > 0 else 0
                )

                if note_rate < 0.3:
                    results.append(
                        ("警告", f"备注覆盖率偏低：{note_rate:.0%}，{no_note} 条无备注")
                    )

                else:
                    results.append(("通过", f"备注覆盖率：{note_rate:.0%}"))

            # 4. 重复订单检测（流程订单+物料编码）

            dup_cols = ["流程订单", "物料编码"]

            self._duplicate_records = pd.DataFrame()  # 初始化

            if all(c in dev_df.columns for c in dup_cols):
                dup_mask = dev_df.duplicated(subset=dup_cols, keep=False)

                dup_count = dup_mask.sum()

                if dup_count > 0:
                    results.append(
                        (
                            "警告",
                            f"发现 {dup_count} 条同一订单重复物料记录（流程订单+物料编码）",
                        )
                    )

                    # 保存所有重复行的完整数据

                    self._duplicate_records = dev_df[dup_mask].copy()

                    # 添加重复组编号（按流程订单+物料编码分组）

                    self._duplicate_records["_重复组"] = (
                        self._duplicate_records.groupby(dup_cols).ngroup()
                    )

                else:
                    results.append(("通过", "同一订单中物料无重复"))

            # ── 黄金模板列位自检 ─────────────────────────────

            try:
                gold_cols = self._load_golden_columns()

                if gold_cols:
                    actual_cols = set(dev_df.columns)

                    template_cols = set(gold_cols)

                    missing = template_cols - actual_cols

                    extra = actual_cols - template_cols

                    if missing:
                        results.append(("严重", f"❌ 黄金模板缺失列: {list(missing)}"))

                    if extra:
                        results.append(("警告", f"⚠️ 黄金模板多余列: {list(extra)}"))

                    if not missing and not extra:
                        results.append(("通过", "✅ 列结构完全匹配黄金模板"))

                else:
                    results.append(("警告", "⚠️ 尚未设置黄金模板，跳过列结构检查"))

            except Exception as e:
                results.append(("警告", f"⚠️ 列结构检查失败: {e}"))

            # 输出到日志

            self.log("📋 数据预检报告：", "info")

            for severity, msg in results:
                self.log(
                    f"  [{severity}] {msg}",
                    severity if severity in ("通过", "警告", "严重") else "info",
                )

            # 弹窗

            self._show_pre_check_report(results)

    def log(self, msg, tag="info"):
        """向日志区写入消息（带时间戳，自动处理 state）"""
        if hasattr(self, "log_text") and self.log_text:
            import time as _t

            ts = _t.strftime("%H:%M:%S")
            self.log_text.configure(state="normal")
            self.log_text.insert("end", f"[{ts}] {msg}\n", tag)
            self.log_text.see("end")
            self.log_text.configure(state="disabled")

    def _refresh_alt_view(self, inner):

        for w in inner.winfo_children():
            w.destroy()

        # 建立编码->名称映射（兼容旧格式）

        excel_path = self.input_file.get()

        code_to_name = {}

        if excel_path and os.path.exists(excel_path):
            try:
                df = pd.read_excel(excel_path, sheet_name="Data")

                code_cols = [
                    c
                    for c in df.columns
                    if any(
                        k in str(c).lower()
                        for k in ["组件物料号", "组件编码", "物料编码", "code", "编码"]
                    )
                ]

                name_cols = [c for c in df.columns if c == "组件物料描述"] or [
                    c
                    for c in df.columns
                    if any(
                        k in str(c).lower()
                        for k in ["物料描述", "组件描述", "名称", "name"]
                    )
                ]

                if code_cols and name_cols:
                    for _, row in df.iterrows():
                        code = str(row[code_cols[0]])

                        name = str(row[name_cols[0]])

                        if code and code != "nan":
                            code_to_name[code] = name if name != "nan" else ""

            except Exception:
                pass

        for a, b in self.alt_pairs:
            fr = tk.Frame(inner, bg=C["surface2"])

            fr.pack(fill="x", pady=1)

            # 解析物料A：支持新格式 (factory, code, name) 和旧格式 (code, name)

            if isinstance(a, tuple):
                if len(a) == 3:
                    _, a_code, a_name = a

                elif len(a) == 2:
                    a_code, a_name = a

                else:
                    a_code = str(a)

                    a_name = ""

            else:
                a_code = str(a)

                a_name = code_to_name.get(a_code, "")

            # 解析物料B：支持新格式 (factory, code, name) 和旧格式 (code, name)

            if isinstance(b, tuple):
                if len(b) == 3:
                    _, b_code, b_name = b

                elif len(b) == 2:
                    b_code, b_name = b

                else:
                    b_code = str(b)

                    b_name = ""

            else:
                b_code = str(b)

                b_name = code_to_name.get(b_code, "")

            # 显示: 编码 + 名称（不显示工厂名称）

            a_disp = f"{a_code} {a_name}" if a_name else a_code

            b_disp = f"{b_code} {b_name}" if b_name else b_code

            tk.Label(
                fr,
                text=f"↔ {a_disp}",
                font=("Consolas", 8),
                fg=C["text"],
                bg=C["surface2"],
                anchor="w",
            ).pack(side="left", padx=4)

            tk.Label(
                fr, text="|", font=("Consolas", 8), fg=C["text_dim"], bg=C["surface2"]
            ).pack(side="left")

            tk.Label(
                fr,
                text=f"{b_disp}",
                font=("Consolas", 8),
                fg=C["purple"],
                bg=C["surface2"],
                anchor="w",
            ).pack(side="left", padx=4)

    def _export_duplicate_records(self):
        """导出预检中发现的重复记录，按组加颜色区分"""

        if not hasattr(self, "_duplicate_records") or self._duplicate_records.empty:
            messagebox.showwarning("提示", "没有重复记录可导出")

            return

        default_dir = self.output_dir.get() or os.path.expanduser("~/Desktop")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        default_name = f"重复记录_{ts}.xlsx"

        file_path = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            title="导出重复记录",
        )

        if not file_path:
            return

        try:
            # 按重复组编号排序

            export_df = self._duplicate_records.sort_values("_重复组")

            # 去掉辅助列（导出时不包含重复组编号列）

            cols_to_export = [c for c in export_df.columns if c != "_重复组"]

            # ── 用 openpyxl 直接构建 Excel（支持按组着色）──

            from openpyxl import Workbook

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()

            ws = wb.active

            ws.title = "重复记录"

            # 表头样式

            header_fill = PatternFill(
                start_color="1B5E20", end_color="1B5E20", fill_type="solid"
            )

            header_font = Font(bold=True, size=11, color="FFFFFF")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 按组循环的颜色（8种浅色）

            group_colors = [
                "FFCDD2",  # 浅红
                "FFE0B2",  # 浅橙
                "FFF9C4",  # 浅黄
                "C8E6C9",  # 浅绿
                "BBDEFB",  # 浅蓝
                "E1BEE7",  # 浅紫
                "B2EBF2",  # 浅青
                "F5F5F5",  # 浅灰
            ]

            # 写入表头

            for j, col_name in enumerate(cols_to_export, 1):
                cell = ws.cell(row=1, column=j, value=col_name)

                cell.font = header_font

                cell.fill = header_fill

                cell.alignment = Alignment(horizontal="center", vertical="center")

                cell.border = thin_border

            # 写入数据行（按组着色）

            current_group = None

            color_idx = -1

            for i, (_, row) in enumerate(export_df.iterrows(), 2):
                group_id = row.get("_重复组", -1)

                # 遇到新组时切换颜色

                if group_id != current_group:
                    current_group = group_id

                    color_idx = (color_idx + 1) % len(group_colors)

                row_fill = PatternFill(
                    start_color=group_colors[color_idx],
                    end_color=group_colors[color_idx],
                    fill_type="solid",
                )

                for j, col_name in enumerate(cols_to_export, 1):
                    cell = ws.cell(row=i, column=j, value=row.get(col_name, ""))

                    cell.font = Font(size=10)

                    cell.border = thin_border

                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    cell.fill = row_fill

            # 自动调整列宽（根据内容长度）

            for j, col_name in enumerate(cols_to_export, 1):
                max_width = len(str(col_name)) * 2  # 中文字符宽度估算

                for row in ws.iter_rows(
                    min_row=2, max_row=ws.max_row, min_col=j, max_col=j
                ):
                    for cell in row:
                        if cell.value:
                            max_width = max(max_width, len(str(cell.value)) * 1.2)

                ws.column_dimensions[
                    ws.cell(row=1, column=j).column_letter
                ].width = min(max_width + 4, 40)

            ws.freeze_panes = "A2"

            wb.save(file_path)

            self.log(
                f"📤 重复记录已导出：{file_path}（共 {len(export_df)} 条，{current_group + 1 if current_group is not None else 0} 组）",
                "success",
            )

            messagebox.showinfo(
                "导出成功", f"已导出 {len(export_df)} 条重复记录到：\n{file_path}"
            )

        except Exception as e:
            self.log(f"❌ 导出重复记录失败：{e}", "error")

            messagebox.showerror("导出失败", str(e))

    def _load_golden_columns(self):
        """从黄金模板JSON文件中读取列名列表"""

        config_dir = _get_mode_config_dir()

        config_path = os.path.join(config_dir, "golden_template.json")

        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                data = json.load(f)

                return data.get("columns", [])

        return None

    def _show_pre_check_report(self, results):
        """显示数据预检报告窗口"""

        d = tk.Toplevel(self.root)

        d.title("数据预检报告")

        d.geometry("500x400")

        d.transient(self.root)

        d.grab_set()

        # 居中显示

        d.update_idletasks()

        x = self.root.winfo_rootx() + (self.root.winfo_width() - 500) // 2

        y = self.root.winfo_rooty() + (self.root.winfo_height() - 400) // 2

        d.geometry(f"+{x}+{y}")

        tk.Label(d, text="📋 数据预检报告", font=("Microsoft YaHei", 12, "bold")).pack(
            pady=10
        )

        text = tk.Text(
            d, font=("Microsoft YaHei", 10), wrap="word", height=15, width=60
        )

        text.pack(fill="both", expand=True, padx=10, pady=5)

        text.tag_configure(
            "严重", foreground="#cf222e", font=("Microsoft YaHei", 10, "bold")
        )

        text.tag_configure("警告", foreground="#d29922")

        text.tag_configure("通过", foreground="#1a7f37")

        for severity, msg in results:
            text.insert("end", f"  [{severity}] {msg}\n", severity)

        text.configure(state="disabled")

        # 如果有重复记录，显示导出按钮

        if hasattr(self, "_duplicate_records") and not self._duplicate_records.empty:
            tk.Button(
                d,
                text="📤 导出重复数据",
                command=lambda: [self._export_duplicate_records(), d.destroy()],
                bg="#f59e0b",
                fg="white",
                font=("Microsoft YaHei", 10),
                relief="flat",
                width=14,
            ).pack(side="left", padx=(20, 5), pady=8)

        tk.Button(
            d,
            text="关闭",
            command=d.destroy,
            bg="#d0d7de",
            font=("Microsoft YaHei", 10),
            relief="flat",
            width=10,
        ).pack(side="right", padx=(5, 20), pady=8)

        # 同时将预检结果打印到日志

        self.log("📋 数据预检报告：", "info")

        for severity, msg in results:
            self.log(
                f"  [{severity}] {msg}",
                severity if severity in ("通过", "警告", "严重") else "info",
            )

    def _update_detail_table(self):
        """从Excel读取数据更新统计卡片和明细表格"""

        try:
            import pandas as pd

            # 读取完整偏差明细sheet

            dev_df = pd.read_excel(self.output_path, sheet_name="完整偏差明细")

            if dev_df.empty:
                return

            # 保存完整数据用于筛选

            self._full_dev_df = dev_df

            # 计算偏差率（处理百分比字符串）

            def parse_rate(rate_val):

                if isinstance(rate_val, str):
                    return (
                        float(
                            rate_val.replace("%", "")
                            .replace("＞", ">")
                            .replace(">", "")
                        )
                        / 100
                    )

                return abs(float(rate_val)) if pd.notna(rate_val) else 0

            dev_df["偏差率数值"] = dev_df["偏差率"].apply(parse_rate)

            # 更新统计数字

            total_count = len(dev_df)

            # 偏差>10%

            big_dev_count = len(dev_df[dev_df["偏差率数值"] > 0.10])

            # 需补备注（备注为空或null）

            no_note_count = len(dev_df[dev_df["备注"].isna() | (dev_df["备注"] == "")])

            # 已审核（备注来源为"AI审核"）

            approved_count = len(dev_df[dev_df["备注来源"] == "AI审核"])

            # 统计数字已通过 audit_stat_labels 更新，这里只需更新 unified_result_lbl

            if hasattr(self, "unified_result_lbl"):
                self.unified_result_lbl.configure(
                    text=f"已加载 {total_count} 条 | 偏差>10%: {big_dev_count} | 需补备注: {no_note_count} | 已审核: {approved_count}"
                )

            # 清空并填充表格（默认显示全部）

            self._refresh_audit_tree(self.audit_data)

            self.current_filter = None

            if not hasattr(self, "status_filter_label"):
                from tkinter import ttk

                self.status_filter_label = ttk.Label(self, text="")

                self.status_filter_label.pack(side=tk.BOTTOM, fill=tk.X)

            else:
                self.status_filter_label.config(text="")

            self.log(f"📊 已加载 {total_count} 条偏差记录到表格", "info")

        except Exception as e:
            self.log(f"⚠ 更新表格失败：{e}", "warn")

            import traceback

            traceback.print_exc()

    # ====== Smart Audit Module ======

    def _load_detail_from_output(self):
        """从分析结果Excel加载数据到偏差明细表"""

        output_dir = self.output_dir.get()

        if not output_dir:
            output_dir = os.path.join(os.path.expanduser("~"), "Desktop")

        # 查找最新的分析结果文件

        pattern = os.path.join(output_dir, "ZPP011偏差分析最终版_*.xlsx")

        files = _glob.glob(pattern)

        if not files:
            messagebox.showwarning("提示", "未找到分析结果文件\n请先运行「开始分析」")

            return

        latest = max(files, key=os.path.getmtime)

        try:
            # 读取完整偏差明细sheet

            dev_df = pd.read_excel(latest, sheet_name="完整偏差明细")

            # 解析偏差率

            def parse_rate(rate_val):

                if isinstance(rate_val, str):
                    return (
                        float(
                            rate_val.replace("%", "")
                            .replace("＞", ">")
                            .replace(">", "")
                        )
                        / 100
                    )

                return abs(float(rate_val)) if pd.notna(rate_val) else 0

            dev_df["偏差率数值"] = dev_df["偏差率"].apply(parse_rate)

            self._full_dev_df = dev_df.copy()

            # 清空并填充表格（使用 audit_tree）

            for item in self.audit_tree.get_children():
                self.audit_tree.delete(item)

            dev_df_sorted = dev_df.sort_values("偏差率数值", ascending=False)

            for idx, (_, row) in enumerate(dev_df_sorted.iterrows(), start=1):
                # 判断状态

                note_source = str(row.get("备注来源", ""))

                status = "已审核" if note_source == "AI审核" else "未审核"

                # 偏差方向

                dev_qty = row.get("偏差数量", 0)

                if pd.notna(dev_qty) and dev_qty > 0:
                    dev_dir = "↑"

                elif pd.notna(dev_qty) and dev_qty < 0:
                    dev_dir = "↓"

                else:
                    dev_dir = ""

                # 备注

                note = str(row.get("备注", "")) if pd.notna(row.get("备注")) else ""

                # 判断是否替代料

                is_alt = "是" if note_source == "替代料" else "否"

                # 正确顺序：idx, excel_row, factory, admin, code, name, order_date, quota, actual, dev_rate, is_alt, status, remark, batch_remark

                self.audit_tree.insert(
                    "",
                    "end",
                    values=(
                        idx,  # 序号
                        int(row["原表行号"])
                        if pd.notna(row.get("原表行号"))
                        else "",  # excel_row
                        str(row.get("工厂", ""))[:10],  # factory
                        "",  # admin (生产管理员) - 数据中无此列
                        str(row.get("物料编码", ""))[:15],  # code
                        str(row.get("物料名称", ""))[:25],  # name
                        str(row.get("订单日期", ""))[:12]
                        if pd.notna(row.get("订单日期"))
                        else "",  # order_date
                        f"{row.get('定额', 0):.2f}"
                        if pd.notna(row.get("定额"))
                        else "-",  # quota
                        f"{row.get('实际', 0):.2f}"
                        if pd.notna(row.get("实际"))
                        else "-",  # actual
                        str(row.get("偏差率", "-")),  # dev_rate
                        is_alt,  # is_alt
                        f"{status}{dev_dir}",  # status
                        note[:30] if note else "",  # remark
                        "",  # batch_remark
                    ),
                )

            # 更新统计

            total = len(dev_df)

            high_dev = len(dev_df[dev_df["偏差率数值"] > 0.10])

            no_note = len(dev_df[(dev_df["备注"].isna()) | (dev_df["备注"] == "")])

            approved = len(dev_df[dev_df["备注来源"] == "AI审核"])

            # 统计数字已通过 audit_stat_labels 更新，这里只需更新 unified_result_lbl

            if hasattr(self, "unified_result_lbl"):
                self.unified_result_lbl.configure(
                    text=f"已加载 {total} 条 | 偏差>10%: {high_dev} | 需补备注: {no_note} | 已审核: {approved}"
                )

            if hasattr(self, "unified_result_lbl"):
                self.unified_result_lbl.configure(
                    text=f"已加载 {total} 条 | 偏差>10%: {high_dev} | 需补备注: {no_note} | 已审核: {approved}"
                )

            # 同步数据到 audit_data（供智能审核使用）

            audit_df = dev_df.copy()

            audit_df["excel_row"] = audit_df["原表行号"].apply(
                lambda x: int(x) if pd.notna(x) else 0
            )

            audit_df["组件物料号"] = audit_df["物料编码"]

            audit_df["组件物料描述"] = audit_df["物料名称"]

            audit_df["工厂名称"] = audit_df["工厂"]

            audit_df["生产管理员描述"] = audit_df["车间"]

            audit_df["数量-定额"] = audit_df["定额"]

            audit_df["数量-实际"] = audit_df["实际"]

            audit_df["偏差率(%)"] = audit_df["偏差率数值"] * 100

            audit_df["备注原因"] = audit_df["备注"]

            audit_df["订单日期"] = audit_df["订单日期"]  # 添加订单日期列

            # ── 物料大类列（供筛选用） ──

            mat_category_map = {
                "100": "原辅料",
                "200": "包材",
                "400": "食品辅料/食品半成品",
                "410": "饮料辅料/饮料半成品",
                "500": "食品成品",
                "510": "饮料成品",
                "600": "促销品",
            }

            audit_df["material_category"] = audit_df["物料编码"].apply(
                lambda x: (
                    mat_category_map.get(str(x)[:3], str(x)[:3]) if pd.notna(x) else ""
                )
            )

            self.audit_data = audit_df

            self._update_filter_options()

            self._refresh_audit_tree(self.audit_data)

            # 启用统一按钮

            if hasattr(self, "unified_ai_btn"):
                self.unified_ai_btn.configure(state="normal")

            if hasattr(self, "unified_export_btn"):
                self.unified_export_btn.configure(state="normal")

            self.log(
                f"已加载分析结果：{os.path.basename(latest)} | 共{total}条记录",
                "success",
            )

            # ── P2：数据变更报告 ──

            prev_snapshot = self._load_change_snapshot()

            if prev_snapshot is not None and not prev_snapshot.empty:
                # 构建本次数据的快照

                current_snap = audit_df[
                    ["订单日期", "流程订单", "组件物料号", "备注原因"]
                ].copy()

                current_snap["订单日期"] = current_snap["订单日期"].astype(str).str[:10]

                prev_snap["订单日期"] = prev_snap["订单日期"].astype(str).str[:10]

                # 合并对比

                merged = current_snap.merge(
                    prev_snap,
                    on=["订单日期", "流程订单", "组件物料号"],
                    how="inner",
                    suffixes=("_new", "_old"),
                )

                changed = merged[
                    (
                        merged["备注原因_new"].fillna("")
                        != merged["备注原因_old"].fillna("")
                    )
                ]

                if not changed.empty:
                    total_changes = len(changed)

                    self.log(
                        f"📋 数据变更报告：发现 {total_changes} 条备注变更", "info"
                    )

                    # 弹窗显示前20条

                    d = tk.Toplevel(self.root)

                    d.title("📋 数据变更报告")

                    d.geometry("650x420")

                    d.transient(self.root)

                    d.grab_set()

                    d.update_idletasks()

                    rx = self.root.winfo_rootx() + (self.root.winfo_width() - 650) // 2

                    ry = self.root.winfo_rooty() + (self.root.winfo_height() - 420) // 2

                    d.geometry(f"+{rx}+{ry}")

                    tk.Label(
                        d,
                        text=f"📋 数据变更报告（共 {total_changes} 条变更）",
                        font=("Microsoft YaHei", 12, "bold"),
                    ).pack(pady=10)

                    text = tk.Text(
                        d, font=("Microsoft YaHei", 9), wrap="word", height=18
                    )

                    text.pack(fill="both", expand=True, padx=10, pady=5)

                    text.tag_configure("changed", foreground="#cf222e")

                    # 只显示前20条

                    display_count = min(20, total_changes)

                    for _, row in changed.head(display_count).iterrows():
                        line = (
                            f"订单：{row['流程订单']} | 物料：{row['组件物料号']} | "
                            f"日期：{row['订单日期']}\n"
                            f"  旧备注：{row['备注原因_old']}\n"
                            f"  新备注：{row['备注原因_new']}\n"
                            f"{'─' * 40}\n"
                        )

                        text.insert("end", line, "changed")

                    if total_changes > 20:
                        text.insert(
                            "end",
                            f"\n... 还有 {total_changes - 20} 条变更，详见日志",
                            "changed",
                        )

                    text.configure(state="disabled")

                    tk.Button(
                        d,
                        text="关闭",
                        command=d.destroy,
                        bg="#d0d7de",
                        font=("Microsoft YaHei", 10),
                        relief="flat",
                        width=10,
                    ).pack(pady=8)

                    # 日志中输出完整变更列表

                    for _, row in changed.iterrows():
                        self.log(
                            f"变更：{row['流程订单']}/{row['组件物料号']} "
                            f"「{row['备注原因_old']}」→「{row['备注原因_new']}」",
                            "info",
                        )

                else:
                    self.log("📋 数据变更报告：无备注变更", "info")

            # 无论是否有变更，都保存本次快照供下次对比

            self._save_change_snapshot(audit_df)

        except Exception as e:
            messagebox.showerror("错误", f"加载分析结果失败：\n{e}")

            self.log(f"加载分析结果失败：{e}", "error")

    def _get_lock_state_path(self):
        """返回列宽锁定状态配置文件路径"""

        dir_path = os.path.join(os.path.expanduser("~"), ".zpp011_audit")

        os.makedirs(dir_path, exist_ok=True)

        return os.path.join(dir_path, "column_lock.json")

    def _load_lock_state(self):
        """加载列锁定状态，默认锁定"""

        path = self._get_lock_state_path()

        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)

                return data.get("locked", True)

            except Exception:
                return True

        return True

    def _save_lock_state(self, locked):
        """保存列锁定状态"""

        path = self._get_lock_state_path()

        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"locked": locked}, f)

        except Exception:
            pass

    def _get_remark_freq_path(self):
        """获取备注频率统计文件路径"""

        dir_path = os.path.join(os.path.expanduser("~"), ".zpp011_audit")

        os.makedirs(dir_path, exist_ok=True)

        return os.path.join(dir_path, "remark_freq.json")

    def _record_remark_freq(self, remark):
        """记录一次备注使用，并更新频率文件"""

        if not remark or remark == "(清空备注)":
            return

        path = self._get_remark_freq_path()

        freq = {}

        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    freq = json.load(f)

            except:
                pass

        freq[remark] = freq.get(remark, 0) + 1

        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(freq, f, ensure_ascii=False, indent=2)

        except:
            pass

    def _get_sorted_remarks(self, preset_list):
        """返回按使用频率降序排列的备注列表"""

        path = self._get_remark_freq_path()

        freq = {}

        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    freq = json.load(f)

            except:
                pass

        # 排序：次数高的在前，次数相同的保持原顺序

        sorted_list = sorted(preset_list, key=lambda x: freq.get(x, 0), reverse=True)

        return sorted_list, freq

    def _show_date_picker(self):
        """弹出日历选择窗口（简化版）"""

        import calendar

        from datetime import datetime, timedelta

        from tkinter import Toplevel, Frame, Label, Button

        win = Toplevel(self.root)

        win.title("选择日期范围")

        win.geometry("300x280")

        win.resizable(False, False)

        # 初始化临时变量

        if not hasattr(self, "_temp_start"):
            self._temp_start = None

        if not hasattr(self, "_temp_end"):
            self._temp_end = None

        self._temp_start = None

        self._temp_end = None

        # 当前显示的年月

        today = datetime.today()

        self._picker_year = today.year

        self._picker_month = today.month

        def update_calendar():
            """更新日历显示"""

            # 清空旧控件

            for widget in cal_frame.winfo_children():
                widget.destroy()

            # 月份标题

            Label(
                cal_frame,
                text=f"{self._picker_year}年 {self._picker_month}月",
                font=("Microsoft YaHei", 10, "bold"),
            ).grid(row=0, column=1, columnspan=5)

            # 星期标题

            for i, day in enumerate(["一", "二", "三", "四", "五", "六", "日"]):
                Label(cal_frame, text=day, font=("Microsoft YaHei", 8)).grid(
                    row=1, column=i
                )

            # 日期网格

            cal = calendar.monthcalendar(self._picker_year, self._picker_month)

            for r, week in enumerate(cal, 2):
                for c, day in enumerate(week):
                    if day == 0:
                        continue

                    btn = Button(
                        cal_frame,
                        text=str(day),
                        width=4,
                        command=lambda d=day: self._on_day_click(d),
                    )

                    btn.grid(row=r, column=c, padx=1, pady=1)

        def change_month(delta):
            """切换月份"""

            self._picker_month += delta

            if self._picker_month > 12:
                self._picker_month = 1

                self._picker_year += 1

            elif self._picker_month < 1:
                self._picker_month = 12

                self._picker_year -= 1

            update_calendar()

        # 控件布局

        top_frame = Frame(win)

        top_frame.pack(pady=5)

        Button(top_frame, text="<", command=lambda: change_month(-1)).pack(
            side="left", padx=5
        )

        Label(top_frame, text="日历", font=("Microsoft YaHei", 10)).pack(
            side="left", padx=10
        )

        Button(top_frame, text=">", command=lambda: change_month(1)).pack(
            side="left", padx=5
        )

        cal_frame = Frame(win)

        cal_frame.pack(pady=5)

        btn_frame = Frame(win)

        btn_frame.pack(pady=5)

        Button(btn_frame, text="清除", command=self._clear_date_filter).pack(
            side="left", padx=5
        )

        Button(btn_frame, text="确定", command=self._apply_date_selection).pack(
            side="left", padx=5
        )

        update_calendar()

    def _on_day_click(self, day):
        """点击日期"""

        from datetime import date

        clicked = date(self._picker_year, self._picker_month, day)

        if self._temp_start is None or self._temp_end is not None:
            # 第一次点击或已选完区间，重新开始

            self._temp_start = clicked

            self._temp_end = None

        else:
            # 第二次点击，确定结束日期

            if clicked > self._temp_start:
                self._temp_end = clicked

            else:
                self._temp_end = self._temp_start

                self._temp_start = clicked

    def _apply_date_selection(self):
        """应用日期选择"""

        self.date_start_val = self._temp_start

        self.date_end_val = self._temp_end

        if self.date_start_val and self.date_end_val:
            self.date_range_var.set(f"{self.date_start_val} ~ {self.date_end_val}")

        elif self.date_start_val:
            self.date_range_var.set(f"{self.date_start_val} 起")

        else:
            self.date_range_var.set("全部日期")

        self._on_filter_changed("date_range")

    def _clear_date_filter(self):
        """清除日期筛选"""

        self.date_start_val = None

        self.date_end_val = None

        self.date_range_var.set("全部日期")

        self._on_filter_changed("date_range")

    def _show_history_compare(self):
        """显示历史对比窗口（Task 009）"""

        from gui.history_compare_dialog import HistoryCompareDialog

        HistoryCompareDialog(self.root)

    def get_current_audit_data(self):
        """供看板调用的接口"""

        if hasattr(self, "audit_data") and self.audit_data is not None:
            return self.audit_data.copy()

        return pd.DataFrame()

    def _show_management_dashboard(self):
        """显示管理看板窗口（Task 011）"""

        from gui.management_dashboard import DashboardWindow

        DashboardWindow(self, current_data_func=self.get_current_audit_data)

    def _open_dashboard(self):
        """管理看板按钮（Task 011）"""

        from gui.management_dashboard import DashboardWindow

        DashboardWindow(self, current_data_func=self.get_current_audit_data)

    def set_filter_and_refresh(self, filter_key: str, filter_value: str):
        """供看板下钻调用，设置筛选器并刷新表格（Task 019）"""
        if hasattr(self, "table_events") and self.table_events is not None:
            self.table_events.set_filter_and_refresh(filter_key, filter_value)
        elif hasattr(self, "log"):
            self.log("table_events 未初始化，无法下钻", "error")

    def _show_attribution_standalone(self):
        """独立 AI 归因分析入口（Task 013）"""
        from core.attribution import generate_report_text, get_latest_history_analysis
        from tkinter import filedialog, messagebox
        import traceback

        df = self.get_current_audit_data()
        if df is None or df.empty:
            messagebox.showinfo("提示", "当前无分析数据，请先完成分析")
            return

        try:
            history_df = get_latest_history_analysis()
            report = generate_report_text(df, history_df)
            if not report or report.strip() == "":
                report = "（未生成有效报告，请检查数据完整性或联系管理员）"
        except Exception as e:
            report = f"归因分析失败：{str(e)}\n{traceback.format_exc()}"

        # 弹出报告窗口
        win = tk.Toplevel(self.root)
        win.title("AI 归因分析报告")
        win.geometry("600x400")
        text = tk.Text(win, wrap=tk.WORD, font=("微软雅黑", 10))
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text.insert(tk.END, report)
        text.config(state=tk.DISABLED)

        def save_report():
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt", filetypes=[("Text files", "*.txt")]
            )
            if file_path:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(report)
                messagebox.showinfo("保存成功", f"报告已保存至 {file_path}")

        tk.Button(win, text="保存报告", command=save_report).pack(
            side=tk.LEFT, padx=10, pady=5
        )
        tk.Button(win, text="关闭", command=win.destroy).pack(
            side=tk.RIGHT, padx=10, pady=5
        )

    # ── Task 012：视图管理方法 ──

    def _refresh_view_list(self):
        """刷新视图下拉框列表"""

        views = self.view_manager.list_views()

        self.view_combo["values"] = views

        if views:
            current = self.view_combo.get()

            self.view_combo.set(
                current if current in views else (views[0] if views else "")
            )

        else:
            self.view_combo.set("")

    def _save_current_view(self):
        """保存当前视图状态"""

        name = simpledialog.askstring("保存视图", "请输入视图名称:")

        if name and name.strip():
            state = self._get_current_view_state()

            self.view_manager.save_view(name.strip(), state)

            self._refresh_view_list()

            self.log(f"视图 '{name}' 已保存", "info")

    def _load_selected_view(self):
        """加载选中的视图"""

        name = self.view_combo.get()

        if not name:
            return

        state = self.view_manager.load_view(name)

        if state:
            self._apply_view_state(state)

            self.log(f"已加载视图 '{name}'", "info")

    def _delete_selected_view(self):
        """删除选中的视图"""

        name = self.view_combo.get()

        if not name:
            return

        if messagebox.askyesno("确认删除", f"确定要删除视图 '{name}' 吗？"):
            self.view_manager.delete_view(name)

            self._refresh_view_list()

            self.log(f"视图 '{name}' 已删除", "info")

    def _get_current_view_state(self):
        """收集当前界面状态（筛选、排序、列顺序、列宽）"""

        state = {
            "filters": {},
            "sort_columns": getattr(self, "sort_columns", []),
            "column_order": list(self.audit_tree["displaycolumns"])
            if self.audit_tree["displaycolumns"]
            else list(self.audit_tree["columns"]),
            "column_widths": {
                col: self.audit_tree.column(col, "width")
                for col in self.audit_tree["columns"]
            },
        }

        # 收集筛选条件

        for key, widget in self.filter_widgets.items():
            if key == "order_date" or isinstance(widget, tuple):
                continue

            val = widget.get() if hasattr(widget, "get") else None

            if val and val != "全部":
                state["filters"][key] = val

        # 日期范围

        if "order_date" in self.filter_widgets:
            dw = self.filter_widgets["order_date"]

            if isinstance(dw, tuple) and len(dw) == 2:
                s = dw[0].get_date()

                e = dw[1].get_date()

                if s or e:
                    state["filters"]["order_date"] = {
                        "start": s.strftime("%Y-%m-%d") if s else "",
                        "end": e.strftime("%Y-%m-%d") if e else "",
                    }

        return state

    def _apply_view_state(self, state):
        """应用视图状态到界面"""

        # 恢复筛选条件

        for key, value in state.get("filters", {}).items():
            if key == "order_date":
                dw = self.filter_widgets.get("order_date")

                if dw and isinstance(dw, tuple) and len(dw) == 2:
                    from datetime import date

                    start_str = value.get("start", "")

                    end_str = value.get("end", "")

                    if start_str:
                        try:
                            dw[0].set_date(date.fromisoformat(start_str))

                        except Exception:
                            pass

                    if end_str:
                        try:
                            dw[1].set_date(date.fromisoformat(end_str))

                        except Exception:
                            pass

            else:
                widget = self.filter_widgets.get(key)

                if widget and hasattr(widget, "set"):
                    widget.set(value)

        # 恢复排序

        sort_columns = state.get("sort_columns", [])

        if sort_columns:
            self.sort_columns = sort_columns

            self._apply_sort_and_refresh()

        # 恢复列顺序

        col_order = state.get("column_order")

        if col_order:
            self._reorder_columns(col_order)

        # 恢复列宽

        for col, width in state.get("column_widths", {}).items():
            if col in self.audit_tree["columns"]:
                self.audit_tree.column(col, width=width)

        # 刷新表格

        self._on_filter_changed(None)

    def _open_rule_config(self):
        """打开可视化规则配置窗口"""
        try:
            from gui.rule_config_dialog import RuleConfigDialog
            import os

            rules_path = os.path.join(
                os.path.dirname(__file__), "..", "config", "system", "rules.json"
            )
            rules_path = os.path.abspath(rules_path)

            def on_rules_changed():
                # 重新加载规则引擎
                if hasattr(self, "rule_engine"):
                    self.rule_engine.reload_rules()
                # 刷新表格颜色/状态
                if hasattr(self, "audit_data") and self.audit_data is not None:
                    self._refresh_audit_tree(self.audit_data)

            RuleConfigDialog(
                self, rules_path, on_rules_changed_callback=on_rules_changed
            )
        except Exception as e:
            from tkinter import messagebox, filedialog, filedialog

            messagebox.showerror("错误", f"打开规则配置失败: {e}")

    def _show_benefit_report(self):
        """打开效益报告生成窗口"""
        if self.audit_data is None or self.audit_data.empty:
            messagebox.showinfo("提示", "请先完成分析并加载审核数据")
            return
        from gui.benefit_report_dialog import BenefitReportDialog

        output_dir = self.output_dir.get() if hasattr(self, "output_dir") else None
        BenefitReportDialog(self, self.audit_data, output_dir)


    def _generate_full_report(self):
        """生成详细分析报告（新版，采用样本样式）"""
        from core.advanced_ppt_generator import generate_advanced_report
        from tkinter import filedialog, messagebox
        from datetime import datetime
        import os

        if self.audit_data is None or self.audit_data.empty:
            messagebox.showinfo("提示", "请先完成分析并加载审核数据")
            return
        # 获取分析结果 Excel 路径
        excel_path = getattr(self, '_analysis_output_path', None)
        if not excel_path or not os.path.exists(excel_path):
            excel_path = filedialog.askopenfilename(
                title="请选择分析结果 Excel 文件",
                filetypes=[("Excel 文件", "*.xlsx")]
            )
            if not excel_path:
                return
        output_dir = self.output_dir.get() or os.path.expanduser("~/Documents/ZPP011分析报告")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"ZPP011详细分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pptx")
        try:
            success = generate_advanced_report(excel_path, output_path, log_cb=self.log)
            if success and messagebox.askyesno("生成成功", f"报告已生成：\n{output_path}\n是否立即打开？"):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("错误", f"生成失败：{e}")

    def _generate_advanced_report_v2(self):
        """生成详细分析报告（新版V2，20+页专业PPT）"""
        from core.advanced_ppt_generator_v2 import generate_advanced_report_v2
        from tkinter import filedialog, messagebox
        from datetime import datetime
        import os

        if self.audit_data is None or self.audit_data.empty:
            messagebox.showinfo("提示", "请先完成分析并加载审核数据")
            return
        # 获取分析结果 Excel 路径
        excel_path = getattr(self, '_analysis_output_path', None)
        if not excel_path or not os.path.exists(excel_path):
            excel_path = filedialog.askopenfilename(title="请选择分析结果 Excel 文件", filetypes=[("Excel 文件", "*.xlsx")])
            if not excel_path: return
        output_dir = self.output_dir.get() or os.path.expanduser("~/Documents/ZPP011分析报告")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"ZPP011详细分析报告_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pptx")
        try:
            success = generate_advanced_report_v2(excel_path, output_path, log_cb=self.log)
            if success and messagebox.askyesno("生成成功", f"报告已生成：\n{output_path}\n是否立即打开？"):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("错误", f"生成失败：{e}")


    def _generate_enterprise_ppt(self):
        """生成企业级PPT（原版阿里巴巴案例）"""
        from core.ppt_enterprise_generator import generate_sample_presentation
        from tkinter import filedialog, messagebox
        from datetime import datetime
        import os

        output_dir = self.output_dir.get() or os.path.expanduser("~/Documents/ZPP011分析报告")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"企业级报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pptx")
        try:
            generate_sample_presentation(output_path)
            if messagebox.askyesno("生成成功", f"企业级PPT已生成：\n{output_path}\n是否立即打开？"):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("错误", f"生成失败：{e}")


    def _on_close(self):
        """窗口关闭时的清理工作（Task 004）"""

        try:
            # 关闭审计日志器

            if hasattr(self, "audit_logger"):
                self.audit_logger.shutdown()

            # 清理分析临时 Excel 文件
            if hasattr(self, "_analysis_output_path") and self._analysis_output_path:
                try:
                    if os.path.exists(self._analysis_output_path):
                        os.remove(self._analysis_output_path)
                        print(f"已清理临时文件: {self._analysis_output_path}")
                except Exception as e:
                    print(f"清理临时文件失败: {e}")

        except Exception as e:
            print(f"关闭审计日志时出错: {e}")

        finally:
            self.root.destroy()


    def _generate_zpp011_report_v3(self):
        """生成ZPP011完整报告(V3) - 基于24页模板的专业PPT"""
        from core.ppt_report_generator_v3 import generate_zpp011_report_v3
        from tkinter import filedialog, messagebox
        from datetime import datetime
        import os

        if self.audit_data is None or self.audit_data.empty:
            messagebox.showinfo("提示", "请先加载并审核数据！")
            return
        excel_path = getattr(self, '_analysis_output_path', None)
        if not excel_path or not os.path.exists(excel_path):
            excel_path = filedialog.askopenfilename(
                title="选择审计数据 Excel 文件",
                filetypes=[("Excel 文件", "*.xlsx")]
            )
            if not excel_path:
                return
        output_dir = self.output_dir.get() or os.path.expanduser("~/Documents/ZPP011报告生成")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(
            output_dir,
            'ZPP011完整报告_V3_' + datetime.now().strftime("%Y%m%d_%H%M%S") + ".pptx"
        )
        try:
            success = generate_zpp011_report_v3(excel_path, output_path, log_cb=self.log)
            if success and messagebox.askyesno(
                "生成成功",
                '报告已生成：\n' + output_path + '\n是否立即打开？'
            ):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("错误", f"生成失败：{e}")

    def _generate_zpp011_report(self):
        """生成ZPP011偏差分析报告（基于企业级模板）"""
        from core.ppt_zpp011_adapter import generate_zpp011_report
        from tkinter import filedialog, messagebox
        from datetime import datetime
        import os

        # 每次都让用户选择Excel文件（优先手动选择）
        excel_path = filedialog.askopenfilename(
            title="请选择分析结果 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialdir=os.path.dirname(getattr(self, '_analysis_output_path', '') or os.path.expanduser("~/Documents/ZPP011分析报告"))
        )
        if not excel_path:
            return
        
        # 检查文件是否包含必要的工作表
        try:
            import pandas as pd
            xl = pd.ExcelFile(excel_path)
            if '汇总统计' not in xl.sheet_names or '完整偏差明细' not in xl.sheet_names:
                messagebox.showerror("错误", "选择的Excel文件缺少必要的工作表（汇总统计、完整偏差明细）\n\n请确保选择的是ZPP011分析结果文件（不是原始数据文件）")
                return
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败：{e}")
            return

        output_dir = self.output_dir.get() or os.path.expanduser("~/Documents/ZPP011分析报告")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"ZPP011偏差分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pptx")

        try:
            generate_zpp011_report(excel_path, output_path)
            if messagebox.askyesno("生成成功", f"报告已生成：\n{output_path}\n是否立即打开？"):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("错误", str(e))

