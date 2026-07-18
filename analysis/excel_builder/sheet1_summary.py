#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
sheet1_summary.py — Sheet1 汇总统计（v36 抽取，增加预警列颜色导出）
"""
import pandas as pd
from analysis.excel_builder.write_sheet_util import ensure_numeric_cols

# 尝试导入 openpyxl（用于颜色填充）
try:
    from openpyxl.styles import PatternFill
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def build_sheet1(df, report_progress, progress_idx=1):
    """
    构建 Sheet1 汇总统计 DataFrame
    参数:
        df: 主数据 DataFrame
        report_progress: 进度回调函数
        progress_idx: 进度索引（默认1）
    返回:
        summary_df: 汇总统计 DataFrame
    """
    report_progress(progress_idx, "Sheet1-汇总统计", 0)
    print("[DEBUG do_analysis_v2] 开始生成Sheet1")

    # 确保数值列为数值类型（防止字符串导致比较错误）
    ensure_numeric_cols(df, ["材料偏差", "偏差率(%)", "偏差金额", "偏差金额(含税)", "数量-实际", "数量-定额"])

    summary_rows = []
    idx = 1
    col_p = '偏差率(%)'

    for (factory, ws_name, mat_cat), grp in df.groupby(['工厂', '车间', '物料分类']):
        pos_dev = grp[grp['材料偏差'] > 0]
        neg_dev = grp[grp['材料偏差'] < 0]
        has_note = grp['备注原因'].notna() & (grp['备注原因'] != '')
        note_rate = has_note.sum() / len(grp) if len(grp) > 0 else 0
        # ③ 预警改为反映真实偏差风险：超 ±10% 行占比（排除无定额假性偏差）
        #    备注覆盖率仍保留为独立信息列，预警不再只看备注是否填写
        grp_valid = grp[~grp['_no_quota']] if '_no_quota' in grp.columns else grp
        over_cnt = (grp_valid['偏差率(%)'].abs() > 10).sum() if len(grp_valid) > 0 else 0
        risk_rate = over_cnt / len(grp_valid) if len(grp_valid) > 0 else 0
        warning = '红色预警' if risk_rate >= 0.1 else ('黄色预警' if risk_rate >= 0.03 else '绿色预警')
        pos_amt = round(pos_dev['偏差金额(含税)'].sum(), 2) if len(pos_dev) > 0 else 0
        neg_amt = round(neg_dev['偏差金额(含税)'].sum(), 2) if len(neg_dev) > 0 else 0
        total_amt = round(grp['偏差金额(含税)'].sum(), 2)
        summary_rows.append({
            '序号': idx,
            '工厂': factory,
            '工厂名称': grp['工厂名称'].iloc[0],
            '车间': ws_name,
            '物料分类': mat_cat,
            '正偏差条数': len(pos_dev),
            '正偏差数量': round(pos_dev['材料偏差'].sum(), 2),
            '正偏差金额(含税)': pos_amt,
            '负偏差条数': len(neg_dev),
            '负偏差数量': round(neg_dev['材料偏差'].sum(), 2),
            '负偏差金额(含税)': neg_amt,
            '总条数': len(grp),
            '总数量': round(grp['材料偏差'].sum(), 2),
            '总偏差金额(含税)': total_amt,
            '备注覆盖率': f"{note_rate:.0%}",
            '预警': warning,
        })
        idx += 1

    summary_df = pd.DataFrame(summary_rows)
    print(f"[DEBUG do_analysis_v2] Sheet1完成，{len(summary_df)} 行")
    report_progress(progress_idx, "Sheet1-汇总统计", 100)
    return summary_df


def write_sheet1_with_colors(writer, summary_df):
    """
    将 summary_df 写入 writer，并为"预警"列添加颜色填充
    参数:
        writer: pd.ExcelWriter 对象（engine='openpyxl'）
        summary_df: 汇总统计 DataFrame
    """
    summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
    if not HAS_OPENPYXL:
        return
    try:
        workbook = writer.book
        worksheet = workbook['汇总统计']
        # 找到"预警"列的列号
        warning_col = None
        for col_idx, cell in enumerate(worksheet[1], 1):  # 第1行是表头
            if cell.value == '预警':
                warning_col = col_idx
                break
        if warning_col is None:
            return
        # 定义颜色
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        # 遍历数据行（第2行开始）
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, warning_col)
            val = str(cell.value).strip()
            if '红色预警' in val:
                cell.fill = red_fill
            elif '黄色预警' in val:
                cell.fill = yellow_fill
            elif '绿色预警' in val:
                cell.fill = green_fill
    except Exception as e:
        print(f"[WARN] 汇总统计上色失败: {e}")
