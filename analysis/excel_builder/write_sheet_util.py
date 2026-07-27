#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
write_sheet_util.py — Excel 写入工具函数（v36 抽取，未修改逻辑）
"""
from copy import copy

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from config.settings import COLORS, FONTS


def get_default_styles():
    """返回默认样式对象（每次调用生成新实例，避免样式对象复用问题）"""
    c = COLORS
    header_fill = PatternFill(start_color=c['header'], end_color=c['header'], fill_type='solid')
    header_font = Font(bold=FONTS['header_bold'], size=FONTS['header_size'], color=c['white'])
    data_font = Font(size=FONTS['data_size'])
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pos_fill = PatternFill(start_color=c['pos_fill'], end_color=c['pos_fill'], fill_type='solid')
    neg_fill = PatternFill(start_color=c['neg_fill'], end_color=c['neg_fill'], fill_type='solid')
    alt_fill = PatternFill(start_color=c['alt_fill'], end_color=c['alt_fill'], fill_type='solid')
    gx_fill = PatternFill(start_color=c['gx_fill'], end_color=c['gx_fill'], fill_type='solid')
    anomaly_fills = {
        '异常1': PatternFill(start_color=c['anomaly_1'], end_color=c['anomaly_1'], fill_type='solid'),
        '异常2': PatternFill(start_color=c['anomaly_2'], end_color=c['anomaly_2'], fill_type='solid'),
        '异常3': PatternFill(start_color=c['anomaly_3'], end_color=c['anomaly_3'], fill_type='solid'),
        '异常4': PatternFill(start_color=c['anomaly_4'], end_color=c['anomaly_4'], fill_type='solid'),
        '异常5': PatternFill(start_color=c['anomaly_5'], end_color=c['anomaly_5'], fill_type='solid'),
    }
    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'data_font': data_font,
        'border': border,
        'center': center,
        'pos_fill': pos_fill,
        'neg_fill': neg_fill,
        'alt_fill': alt_fill,
        'gx_fill': gx_fill,
        'anomaly_fills': anomaly_fills,
    }


def write_sheet(ws, headers, data_rows, col_widths=None):
    """通用 Sheet 写入函数（用于 Sheet1/2/3/4/5/8）

    性能优化（2026-07-27）：数据区样式改为「探针格算一次 StyleArray，
    其余格逐格拷贝 _style 数组」。原来每格 3 次样式赋值都会触发 openpyxl
    对样式对象的递归 hash + 去重查表（12K 行 × 24 列时 200 万次），
    是导出路径最大热点（cProfile 实测 write_sheet 占导出总耗时 ~2/3）。
    每格持有独立 StyleArray 拷贝，后续仍可安全地单独改某格样式（如涨跌色）。"""
    styles = get_default_styles()
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = styles['header_font']
        c.fill = styles['header_fill']
        c.alignment = styles['center']
        c.border = styles['border']
    if data_rows:
        # 探针格：常规赋值一次，让 openpyxl 注册样式并生成 StyleArray 原型
        probe = ws.cell(row=2, column=1)
        probe.font = styles['data_font']
        probe.border = styles['border']
        probe.alignment = styles['center']
        # 关键：探针格若恰好位于日期列（如各 Sheet 首列“订单日期/订单开始日期”），
        # _bind_value 会为其 _style 套上日期 numFmtId；若直接 copy 给所有格，
        # 数值格会被写成日期格式（读回变 1902-xx-xx 等）。这里强制把原型 numFmtId
        # 清成 0（General），再在下方逐格优先“设样式、后设值”——真日期在设值时
        # 由 _bind_value 重新套用日期格式，数值则保持 General。
        proto = copy(probe._style)
        proto.numFmtId = 0
        for i, row in enumerate(data_rows, 2):
            for j, v in enumerate(row, 1):
                c = ws.cell(row=i, column=j)
                c._style = copy(proto)
                c.value = v  # 真日期→_bind_value 重套日期格式；数值→保持 General
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'


def ensure_numeric_cols(df, cols):
    """将指定列转为数值型（转换失败填 0），原地修改并返回 df。
    用于消除各 sheet builder 中重复的 to_numeric 转换块。
    性能优化（2026-07-27）：已是数值 dtype 的列跳过 to_numeric 全量拷贝，
    仅在含 NaN 时补 fillna(0)，语义与原实现完全一致。"""
    for col in cols:
        if col in df.columns:
            s = df[col]
            if pd.api.types.is_numeric_dtype(s):
                if s.isna().any():
                    df[col] = s.fillna(0)
                continue
            df[col] = pd.to_numeric(s, errors='coerce').fillna(0)
    return df
