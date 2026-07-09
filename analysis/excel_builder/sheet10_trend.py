#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
sheet10_trend.py — Sheet10 趋势分析

按时间跨度三等分为早期/中期/近期，计算各物料各时段平均偏差率并判断趋势。
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def build_sheet10(wb, dev_df, date_min, report_progress, progress_idx=10):
    """
    构建 Sheet10 趋势分析并写入工作表
    """
    report_progress(progress_idx, "Sheet10-趋势分析", 0)
    print("[DEBUG do_analysis_v2] 开始生成Sheet10")

    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    data_font = Font(size=10)
    info_font = Font(size=10, color='666666', italic=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')

    def write_trend_sheet(wb_inner, dev_df_inner, date_min_inner):
        ws = wb_inner.create_sheet('趋势分析（自然日分组）')
        headers = ['物料编码', '物料名称', '物料类型', '单位',
                   '早期偏差率', '中期偏差率', '近期偏差率', '趋势']
        col_widths = [16, 30, 12, 8, 16, 16, 16, 14]

        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        ws.freeze_panes = 'A2'

        if dev_df_inner is None or dev_df_inner.empty:
            ws.cell(row=2, column=1, value='无数据，无法生成趋势')
            return

        dev = dev_df_inner.copy()

        # 解析订单日期
        dev['订单日期'] = pd.to_datetime(dev['订单日期'], errors='coerce')
        dev = dev.dropna(subset=['订单日期'])
        if dev.empty:
            ws.cell(row=2, column=1, value='无有效订单日期，无法生成趋势')
            return

        # 解析偏差率：兼容字符串 "10.5%" 和数值类型
        def parse_dev_rate(x):
            if pd.isna(x):
                return None
            if isinstance(x, str):
                s = x.strip().rstrip('%')
                if s in ('', '-', '—', 'nan', 'None'):
                    return None
                try:
                    return float(s)
                except ValueError:
                    return None
            try:
                return float(x)
            except (ValueError, TypeError):
                return None

        dev['_dev_rate_num'] = dev['偏差率'].apply(parse_dev_rate)

        valid_rate = dev['_dev_rate_num'].notna().sum()
        total = len(dev)
        print(f"[DEBUG Sheet10] 总行数={total}, 有效偏差率={valid_rate}")

        if valid_rate == 0:
            ws.cell(row=2, column=1, value='无有效偏差率数据，无法生成趋势')
            return

        # 按时间跨度三等分
        date_min_real = dev['订单日期'].min()
        date_max_real = dev['订单日期'].max()
        total_span = (date_max_real - date_min_real).total_seconds()

        if total_span <= 0:
            # 只有1天数据，全部归为早期
            early_end = date_max_real
            mid_start = date_max_real
            mid_end = date_max_real
            recent_start = date_max_real
        else:
            third = total_span / 3
            early_end = date_min_real + pd.Timedelta(seconds=third)
            mid_start = early_end
            mid_end = date_min_real + pd.Timedelta(seconds=third * 2)
            recent_start = mid_end

        # 格式化日期范围描述
        def fmt(dt):
            return pd.Timestamp(dt).strftime('%m/%d')

        if total_span <= 0:
            period_desc = f"仅单日数据（{fmt(date_min_real)}），全部归为早期"
        else:
            period_desc = (f"早期：{fmt(date_min_real)}～{fmt(early_end)}  |  "
                          f"中期：{fmt(mid_start)}～{fmt(mid_end)}  |  "
                          f"近期：{fmt(recent_start)}～{fmt(date_max_real)}")

        print(f"[DEBUG Sheet10] {period_desc}")

        # 在表头下方写入日期范围说明
        ws.cell(row=2, column=1, value=period_desc)
        ws.cell(row=2, column=1).font = info_font
        ws.cell(row=2, column=1).alignment = left_align
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

        def assign_period(dt):
            if total_span <= 0:
                return 'early'
            if dt <= early_end:
                return 'early'
            elif dt <= mid_end:
                return 'mid'
            else:
                return 'recent'

        dev['_period'] = dev['订单日期'].apply(assign_period)
        dev = dev.dropna(subset=['_period', '_dev_rate_num'])

        if dev.empty:
            ws.cell(row=3, column=1, value='无有效分组数据，无法生成趋势')
            return

        # 确保分组列存在
        group_cols = ['物料编码', '物料名称', '物料类型', '单位']
        for c in group_cols:
            if c not in dev.columns:
                dev[c] = ''

        period_avg = (
            dev.groupby(group_cols + ['_period'])['_dev_rate_num']
            .mean().round(2).unstack(fill_value=None)
        )

        rows = []
        for idx, row in period_avg.iterrows():
            code, name, typ, unit = idx
            r_recent = row.get('recent', None)
            r_mid = row.get('mid', None)
            r_early = row.get('early', None)

            # 判断趋势
            vals = [v for v in (r_early, r_mid, r_recent) if v is not None]
            if len(vals) < 2:
                arrow = '→'
            else:
                changes = []
                for a, b in [(r_early, r_mid), (r_mid, r_recent)]:
                    if a is not None and b is not None:
                        if a > 0 and b > 0 or a < 0 and b < 0:
                            if abs(b) > abs(a) + 1:
                                changes.append('↑')
                            elif abs(b) < abs(a) - 1:
                                changes.append('↓')
                            else:
                                changes.append('→')
                        elif abs(a) <= 0.01 and b != 0:
                            changes.append('↑' if b > 0 else '↓')
                        elif abs(b) <= 0.01 and a != 0:
                            changes.append('↓' if a > 0 else '↑')
                        elif a > 0 and b <= 0:
                            changes.append('↓')
                        elif a <= 0 and b > 0:
                            changes.append('↑')
                        else:
                            changes.append('→')
                if not changes:
                    arrow = '→'
                elif all(c == '↑' for c in changes):
                    arrow = '↑↑ 持续变差'
                elif all(c == '↓' for c in changes):
                    arrow = '↓↓ 持续改善'
                elif changes[-1] == '↑':
                    arrow = '↑ 近期变差'
                elif changes[-1] == '↓':
                    arrow = '↓ 近期改善'
                else:
                    arrow = '→'

            rows.append([
                code, name, typ, unit,
                f"{r_early:.2f}%" if r_early is not None and pd.notna(r_early) else "-",
                f"{r_mid:.2f}%" if r_mid is not None and pd.notna(r_mid) else "-",
                f"{r_recent:.2f}%" if r_recent is not None and pd.notna(r_recent) else "-",
                arrow
            ])

        def sort_key(r):
            v = r[6]
            if v == "-":
                return -999
            try:
                return -abs(float(v.rstrip('%')))
            except (ValueError, AttributeError):
                return -999

        rows.sort(key=sort_key)

        # 数据从第3行开始（第2行是日期范围说明）
        for i, row in enumerate(rows, 3):
            for j, v in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.font = data_font
                c.border = border
                c.alignment = center
                if j == 8:
                    if '↑' in str(v):
                        c.font = Font(size=10, color='C00000')
                    elif '↓' in str(v):
                        c.font = Font(size=10, color='008000')

        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

        print(f"[DEBUG Sheet10] 写入 {len(rows)} 行趋势数据")

    write_trend_sheet(wb, dev_df, date_min)
    print("[DEBUG do_analysis_v2] Sheet10完成")
    report_progress(progress_idx, "Sheet10-趋势分析", 100)
