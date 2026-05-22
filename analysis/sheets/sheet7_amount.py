#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
sheet7_amount.py — Sheet7 偏差金额分析（v36 抽取，未修改逻辑）

本模块接收 wb 对象，在函数内部创建 ws_amt 工作表并写入数据。
analyzer.py 调用 build_sheet7(wb, df, report_progress) 执行。
"""
import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def build_sheet7(wb, df, report_progress, progress_idx=7):
    """
    构建 Sheet7 偏差金额分析数据并写入工作表
    参数:
        wb: openpyxl Workbook 对象
        df: 主数据 DataFrame
        report_progress: 进度回调函数
        progress_idx: 进度索引（默认7）
    返回:
        无（直接写入 wb['偏差金额分析']）
    """

    # 确保数值列为数值类型（防止字符串导致比较错误）
    numeric_cols = ["偏差金额(含税)", "数量-实际", "数量-定额", "材料偏差", "偏差金额"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    report_progress(progress_idx, "Sheet7-偏差金额分析", 0)
    print("[DEBUG do_analysis_v2] 开始生成Sheet7")

    ws_amt = wb.create_sheet('偏差金额分析', index=6)

    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    data_font = Font(size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_sheet7(ws, headers, data_rows, col_widths=None):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        for i, row in enumerate(data_rows, 2):
            for j, v in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.font = data_font
                c.border = border
                c.alignment = center
        if col_widths:
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = 'A2'

    amt_df = df[df['偏差金额(含税)'] != 0].copy()

    _debug_log = os.path.join(os.environ.get('TEMP', '.'), 'zpp011_sheet7_debug.log')
    with open(_debug_log, 'a', encoding='utf-8') as _f:
        _f.write(f"\n=== Sheet7 Debug ===\n")
        _f.write(f"amt_df.empty: {amt_df.empty}\n")
        _f.write(f"amt_df.columns: {list(amt_df.columns)}\n")
        _f.write(f"amt_df.shape: {amt_df.shape}\n")

    if not amt_df.empty:
        amt_summary = amt_df.groupby(['组件物料号', '组件物料描述', '物料分类', '组件单位']).agg(
            正偏差金额=('偏差金额(含税)', lambda x: round(x[x > 0].sum(), 2)),
            负偏差金额=('偏差金额(含税)', lambda x: round(x[x < 0].sum(), 2)),
            总偏差金额=('偏差金额(含税)', lambda x: round(x.sum(), 2)),
            涉及条数=('偏差金额(含税)', 'size'),
        ).reset_index()

        amt_summary.rename(columns={
            '组件物料号': '物料编码',
            '组件物料描述': '物料名称',
            '物料分类': '物料类型',
            '组件单位': '单位',
        }, inplace=True)

        amt_summary['总偏差金额(含税)'] = amt_summary['总偏差金额']
        amt_summary['正偏差金额(含税)'] = amt_summary['正偏差金额']
        amt_summary['负偏差金额(含税)'] = amt_summary['负偏差金额']
        amt_summary.drop(columns=['正偏差金额', '负偏差金额', '总偏差金额'], inplace=True)

        amt_summary['_abs_total'] = amt_summary['总偏差金额(含税)'].apply(
            lambda x: abs(x) if isinstance(x, (int, float)) else 0)
        amt_summary = amt_summary.sort_values(
            '_abs_total', ascending=False).drop('_abs_total', axis=1)
        amt_rows = amt_summary.to_dict('records')
    else:
        amt_rows = []

    headers_amt = ['物料编码', '物料名称', '物料类型', '单位',
                   '正偏差金额(含税)', '负偏差金额(含税)', '总偏差金额(含税)', '涉及条数']
    write_sheet7(ws_amt, headers_amt,
                 [[r['物料编码'], r['物料名称'], r['物料类型'], r['单位'],
                   r['正偏差金额(含税)'], r['负偏差金额(含税)'],
                   r['总偏差金额(含税)'], r['涉及条数']] for r in amt_rows],
                 [16, 30, 12, 8, 20, 20, 20, 10])

    report_progress(progress_idx, "Sheet7-偏差金额分析", 100)
