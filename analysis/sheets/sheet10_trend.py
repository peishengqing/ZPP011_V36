#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
sheet10_trend.py — Sheet10 趋势分析（v36 抽取，未修改逻辑）

注意：Sheet10 包含内嵌函数 write_trend_sheet，
本模块直接操作已创建的 wb Workbook 对象。
analyzer.py 调用 build_sheet10(wb, dev_df, date_min, report_progress) 执行。
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def build_sheet10(wb, dev_df, date_min, report_progress, progress_idx=10):
    """
    构建 Sheet10 趋势分析并写入工作表
    参数:
        wb: openpyxl Workbook 对象
        dev_df: Sheet5 完整偏差明细 DataFrame
        date_min: 分析起始日期
        report_progress: 进度回调函数
        progress_idx: 进度索引（默认10）
    返回:
        无（直接写入 wb）
    """
    report_progress(progress_idx, "Sheet10-趋势分析", 0)
    print("[DEBUG do_analysis_v2] 开始生成Sheet10")

    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    data_font = Font(size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_trend_sheet(wb_inner, dev_df_inner, date_min_inner):
        ws = wb_inner.create_sheet('趋势分析（自然日分组）')
        headers = ['物料编码', '物料名称', '物料类型', '单位',
                   '早期偏差率', '中期偏差率', '近期偏差率', '趋势']
        col_widths = [16, 30, 12, 8, 16, 16, 16, 10]

        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        ws.freeze_panes = 'A2'

        if dev_df_inner.empty or '订单日期' not in dev_df_inner.columns:
            ws.cell(row=2, column=1, value='无有效订单日期，无法生成趋势')
            return

        dev = dev_df_inner.copy()
        dev['订单日期'] = pd.to_datetime(dev['订单日期'], errors='coerce')
        dev = dev.dropna(subset=['订单日期'])
        if dev.empty:
            ws.cell(row=2, column=1, value='无有效订单日期，无法生成趋势')
            return

        unique_dates = sorted(set(str(x) for x in dev['订单日期'] if pd.notna(x)))
        n = len(unique_dates)
        if n == 0:
            ws.cell(row=2, column=1, value='无有效订单日期，无法生成趋势')
            return

        early_cnt = (n + 2) // 3
        recent_cnt = (n + 2) // 3
        mid_cnt = n - early_cnt - recent_cnt
        if mid_cnt < 0:
            early_cnt -= 1
            recent_cnt -= 1
            mid_cnt = n - early_cnt - recent_cnt

        early_dates = set(unique_dates[:early_cnt])
        mid_dates = set(unique_dates[early_cnt:early_cnt + mid_cnt]) if mid_cnt > 0 else set()
        recent_dates = set(unique_dates[early_cnt + mid_cnt:]) if recent_cnt > 0 else set()

        def assign_period(date):
            if date in recent_dates:
                return 'recent'
            elif date in mid_dates:
                return 'mid'
            elif date in early_dates:
                return 'early'
            else:
                return None

        dev['_period'] = dev['订单日期'].apply(assign_period)
        dev['_dev_rate_num'] = dev['偏差率'].apply(
            lambda x: float(str(x).rstrip('%')) if isinstance(x, str) and x not in ('', '-') else None)

        period_avg = (
            dev.dropna(subset=['_period', '_dev_rate_num'])
            .groupby(['物料编码', '物料名称', '物料类型', '单位', '_period'])['_dev_rate_num']
            .mean().round(2).unstack(fill_value=None)
        )

        rows = []
        for idx, row in period_avg.iterrows():
            code, name, typ, unit = idx[0], idx[1], idx[2], idx[3]
            r_recent = row.get('recent', None)
            r_mid = row.get('mid', None)
            r_early = row.get('early', None)
            vals = [v for v in (r_recent, r_mid, r_early) if v is not None]

            if len(vals) < 2:
                arrow = '→'
            else:
                changes = []
                for a, b in [(r_recent, r_mid), (r_mid, r_early)]:
                    if a is not None and b is not None:
                        if a > 0 and b > 0 or a < 0 and b < 0:
                            if a > b + 1:
                                changes.append('↑')
                            elif a < b - 1:
                                changes.append('↓')
                            else:
                                changes.append('→')
                        elif a > 0 and b <= 0:
                            changes.append('↑')
                        elif a <= 0 and b > 0:
                            changes.append('↓')
                        else:
                            changes.append('→')
                if not changes:
                    arrow = '→'
                elif all(c == '↑' for c in changes):
                    arrow = '↑↑ 持续变差'
                elif all(c == '↓' for c in changes):
                    arrow = '↓↓ 持续改善'
                elif changes[-1] == '↑':
                    arrow = '↑ 近期变差'
                elif changes[-1] == '↓':
                    arrow = '↓ 近期改善'
                else:
                    arrow = '→'

            rows.append([
                code, name, typ, unit,
                f"{abs(r_early):.2f}%" if r_early is not None else "-",
                f"{abs(r_mid):.2f}%" if r_mid is not None else "-",
                f"{abs(r_recent):.2f}%" if r_recent is not None else "-",
                arrow
            ])

        def sort_key(r):
            v = r[6]
            if v == "-":
                return -999
            return -abs(float(v.rstrip('%')))

        rows.sort(key=sort_key)

        for i, row in enumerate(rows, 2):
            for j, v in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.font = data_font
                c.border = border
                c.alignment = center
                if j == 8:
                    if '↑' in str(v):
                        c.font = Font(size=10, color='C00000')
                    elif '↓' in str(v):
                        c.font = Font(size=10, color='008000')

        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    write_trend_sheet(wb, dev_df, date_min)
    print("[DEBUG do_analysis_v2] Sheet10完成")
    report_progress(progress_idx, "Sheet10-趋势分析", 100)
