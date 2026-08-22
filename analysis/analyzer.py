#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ZPP011 偏差分析核心逻辑（v36 抽取）
⚠️ 本文件从 main.py 原样抽取，未修改任何分析逻辑
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re
import glob as _glob
import sys as _sys
import threading
import time
import json
import tempfile
import subprocess
import queue
import traceback
import shutil
import sqlite3
import zipfile
from config.settings import DEFAULT_THRESHOLD

# 模块化组件
from storage import storage
from domain.alt_material import alt_manager

# Sheet 构建函数（第五步抽取）
from analysis.excel_builder.sheet1_summary import build_sheet1
from analysis.excel_builder.sheet2_alt import build_sheet2
from analysis.excel_builder.sheet3_no_note import build_sheet3
from analysis.excel_builder.sheet4_middle import build_sheet4
from analysis.excel_builder.sheet5_full import build_sheet5
from analysis.excel_builder.sheet6_anomaly import build_sheet6
from analysis.excel_builder.sheet7_amount import build_sheet7
from analysis.excel_builder.sheet8_reason_summary import build_sheet8
from analysis.excel_builder.sheet9_reason_detail import build_sheet9
from analysis.excel_builder.sheet10_trend import build_sheet10
from analysis.excel_builder.write_sheet_util import write_sheet
from analysis.net_offset import apply_net_offset

# 缓存最近一次分析的中间结果（worker 快速路径写入），供后台缓存线程复用，避免 Sheet1~5 重算
LATEST_INTERMEDIATES = None



# 通用工具函数
from utils.helpers import standardize_remark
from analysis.debug_util import dprint as _dprint


def infer_material_type(code):
    """根据物料编码前缀推断物料类型（与 SAP 物料大类一致）

    前缀约定（来源：ZPP011 导出数据「组件物料类型描述」实测分布）：
      10    → 原料（SAP 原辅料）
      20    → 包材
      40/41 → 半成品
      60    → 广宣
      其余  → 其他
    注：早期误把「原料」写成 30 开头，但真实数据原料为 10 开头，
    导致主表「物料类型」列永无「原料」值，看板原料筛选统计为 0（已修正）。
    """
    if not isinstance(code, str):
        return '未知'
    code = code.strip()
    if code.startswith('10'):
        return '原料'
    if code.startswith('20'):
        return '包材'
    if code.startswith('40') or code.startswith('41'):
        return '半成品'
    if code.startswith('60'):
        return '广宣'
    return '其他'


def do_analysis_v2(
        input_file,
        output_dir,
        alt_pairs,
        progress_callback=None,
        cancel_check=None,
        start_date=None,
        end_date=None,
        material_search=None,
        output_path=None,
        enable_net_offset=True,
        return_dataframe=False,
        dev_rate_threshold=0.0,
        input_df=None,
        dyn_thresh=None):
    _dprint("[DEBUG do_analysis_v2] 函数开始执行")

    # output_dir 兜底，防止调用方传 None 导致 os.path.join 崩溃
    if output_dir is None:
        output_dir = os.path.dirname(input_file) or '.'

    # ========== 数值列追踪初始化 ==========
    _trace_log = os.path.join(os.environ.get('TEMP', '.'), 'zpp011_trace.log')
    _snapshot = {}
    _dprint(f"[TRACE] 追踪日志将写入: {_trace_log}")

    def check_cancel():
        if cancel_check and cancel_check():
            raise KeyboardInterrupt("用户取消")

    def report_progress(step_idx, step_name, percent):
        """封装进度回调，供5阶段调用。"""
        if progress_callback:
            progress_callback(step_idx, step_name, percent)
            time.sleep(0.01)

    from analysis.excel_builder.write_sheet_util import get_default_styles
    _styles = get_default_styles()
    pos_fill = _styles['pos_fill']
    neg_fill = _styles['neg_fill']
    alt_fill = _styles['alt_fill']
    gx_fill = _styles['gx_fill']
    header_font = _styles['header_font']
    header_fill = _styles['header_fill']
    center = _styles['center']
    border = _styles['border']
    data_font = _styles['data_font']
    anomaly_fills = _styles['anomaly_fills']

    check_cancel()

    src_file = input_file
    
    # 检查文件是否存在
    if not os.path.exists(src_file):
        error_msg = f"文件不存在: {src_file}"
        _dprint(f"❌ {error_msg}")
        report_progress(1, "错误：文件不存在", 0)
        raise FileNotFoundError(error_msg)
    
    # 拒绝 Excel 临时锁文件（~$ 开头），避免误把锁文件当成数据源
    if os.path.basename(src_file).startswith("~$"):
        error_msg = f"该文件是 Excel 临时锁文件，无法分析: {src_file}（请关闭 Excel 中对应的文件后重新选择）"
        _dprint(f"❌ {error_msg}")
        report_progress(1, "错误：Excel 临时锁文件", 0)
        raise PermissionError(error_msg)

    # 检查文件是否被占用
    try:
        with open(src_file, 'r+b'):
            pass
    except PermissionError:
        # r+b 失败可能是 Excel 打开了文件（写锁定），尝试只读模式
        try:
            with open(src_file, 'rb'):
                pass
            _dprint(f"⚠ 文件被写锁定（可能被 Excel 打开），但可读取，继续分析: {src_file}")
        except PermissionError:
            error_msg = f"文件被占用或无权限访问: {src_file}"
            _dprint(f"❌ {error_msg}")
            report_progress(1, "错误：文件被占用", 0)
            raise PermissionError(error_msg)
    except Exception as e:
        _dprint(f"⚠ 文件访问检查失败（可能不影响读取）: {e}")
    
    try:
        _dprint(f"[DEBUG] 开始读取Excel: {src_file}")
        if input_df is not None:
            # 优化：直接使用选中文件时已经读取过的 DataFrame，跳过重复的文件 IO（约 20-30s）
            df = input_df.copy()
            _dprint(f"[DEBUG do_analysis_v2] 复用已读 DataFrame，{len(df)} 行")
        else:
            # 容错：优先读 'Data' 工作表，不存在则取第一个
            xl = pd.ExcelFile(src_file)
            _sheet = 'Data' if 'Data' in xl.sheet_names else xl.sheet_names[0]
            if _sheet != 'Data':
                _dprint(f"[WARN] 工作表 'Data' 不存在，改用 '{_sheet}'")
            df = pd.read_excel(src_file, sheet_name=_sheet)
            _dprint(f"[DEBUG do_analysis_v2] 读取Data表成功，{len(df)} 行")

        # 校验：必须为原始 SAP 导出文件，而非分析报告
        _required_cols = ['订单开始日期', '数量-定额', '数量-实际']
        _missing = [c for c in _required_cols if c not in df.columns]
        if _missing:
            error_msg = (
                f"文件缺少原始 SAP 数据所需列：{', '.join(_missing)}\n"
                f"当前工作表：{_sheet}，列名：{list(df.columns)}\n"
                "请确认选择的是 SAP 原始导出文件（如 ZPP011_YYYYMMDD.xlsx），而不是分析报告。"
            )
            _dprint(f"❌ {error_msg}")
            report_progress(1, f"错误：缺少列 {_missing[0]}", 0)
            raise ValueError(error_msg)

        report_progress(1, "1/5 正在读取 Excel 文件", 10)
        # 强制刷新输出（安全模式，忽略线程 stdout 不可用的情况）
        import sys
        try:
            if sys.stdout is not None:
                sys.stdout.flush()
        except (OSError, ValueError, AttributeError):
            pass
    except Exception as e:
        error_detail = f"读取Excel失败: {e}\n文件路径: {src_file}\n文件存在: {os.path.exists(src_file)}"
        _dprint(f"❌ {error_detail}")
        report_progress(1, f"错误: {str(e)[:50]}", 0)
        raise Exception(error_detail)
    

    # ========== 追踪点1: 读取数据后（原始状态） ==========
    _snapshot['after_read'] = {
        '数量-实际': df['数量-实际'].describe().to_dict() if '数量-实际' in df.columns else 'NOT_FOUND',
        '数量-定额': df['数量-定额'].describe().to_dict() if '数量-定额' in df.columns else 'NOT_FOUND',
        '行数': len(df)
    }
    _dprint(f"[TRACE-1] 读取后: 数量-实际 sum={df['数量-实际'].sum() if '数量-实际' in df.columns else 'N/A'}")

    # ========== 诊断：找出哪个数值列被字符串污染（默认关闭，DEBUG 时启用）==========
    # 默认关闭：该诊断块含逐列 apply(isinstance) 检查，数据量大时显著拖慢分析（详见性能报告）。
    # 排查数据质量问题时设环境变量 ZPP011_DEBUG=1 重新运行即可开启。
    ENABLE_DIAGNOSTIC = os.environ.get('ZPP011_DEBUG', '0') == '1'
    if ENABLE_DIAGNOSTIC:
        # 使用文件日志（避免输出被吞掉）
        _diag_log = os.path.join(os.environ.get('TEMP', '.'), 'zpp011_diagnostic.log')
        with open(_diag_log, 'w', encoding='utf-8') as _f:
            _f.write(f"=== 诊断开始 {pd.Timestamp.now()} ===\n")
            _f.write(f"文件: {src_file}\n")
            _f.write(f"行数: {len(df)}\n")
            _f.write(f"列名: {list(df.columns)}\n\n")
        
        _dprint(f"[诊断] 正在检查数值列，日志写入: {_diag_log}")
        print("[诊断] 检查数值列中的字符串...")
        numeric_cols_check = [
            '数量-定额', '数量-实际', '材料偏差', '偏差率(%)',
            '金额-定额(含税)', '金额-实际(含税)', '实际成本', '产量', '组件数量'
        ]
        for col in numeric_cols_check:
            if col in df.columns:
                try:
                    # 找出非数值的行（包括字符串）
                    mask = ~df[col].apply(lambda x: isinstance(x, (int, float)) or pd.isna(x))
                    if mask.any():
                        bad_vals = df.loc[mask, col].unique()[:5]
                        print(f"⚠️ 列 [{col}] 包含非数值：{bad_vals}")
                        # 同时打印对应的物料名称和订单号，便于定位
                        sample_rows = df.loc[mask].head(3)
                        if all(c in df.columns for c in ['流程订单', '组件物料描述', col]):
                            print(f" 示例行：{sample_rows[['流程订单', '组件物料描述', col]].to_dict(orient='records')}")
                except Exception as e:
                    print(f"⚠️ 检查列 [{col}] 时出错: {e}")
        print("[诊断] 数值列检查完成")
        # 同时写入诊断日志文件
        with open(_diag_log, 'a', encoding='utf-8') as _f:
            _f.write(f"\n=== 数值列检查完成 ===\n")
            _f.write(f"df.shape: {df.shape}\n")
            _f.write(f"数值列检查: 完成\n\n")
            _f.write(f"'组件单位' in df.columns: {'组件单位' in df.columns}\n")
    report_progress(2, "2/5 正在解析生产数据", 30)

    # 保留原始 Excel 行号：用 openpyxl 读取真实行号（避免 pandas read_excel 跳过空行导致偏移）
    try:
        from openpyxl import load_workbook
        _wb = load_workbook(src_file, read_only=True, data_only=True)
        # 容错：优先取 'Data' 工作表，不存在则取第一个
        _sheet_names = _wb.sheetnames
        _ws = _wb['Data'] if 'Data' in _sheet_names else _wb[_sheet_names[0]]
        if 'Data' not in _sheet_names:
            _dprint(f"[WARN] 工作表 'Data' 不存在，改用 '{_sheet_names[0]}'")
        _real_rows = []
        _rn = 0
        for _row in _ws:
            _rn += 1
            if _rn == 1:
                continue  # 跳过表头
            _real_rows.append(_rn)
        _wb.close()
        if len(_real_rows) == len(df):
            df.insert(0, '_excel_row', _real_rows)
        else:
            # 行数不匹配时回退到计算方式
            df.insert(0, '_excel_row', range(2, len(df) + 2))
    except Exception:
        df.insert(0, '_excel_row', range(2, len(df) + 2))

    # ========== 强制转换数值列，防止字符串混入 ==========
    numeric_cols = [
        '数量-定额', '数量-实际', '材料偏差', '偏差率(%)',
        '金额-定额(含税)', '金额-实际(含税)', '实际成本', '产量', '组件数量'
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # ── 固定阈值（公司规定）─────────────────────
    dyn_thresh = DEFAULT_THRESHOLD if dyn_thresh is None else dyn_thresh
    thresh_desc = f"固定阈值（公司规定）：±{dyn_thresh:.0f}%"
    df['_dyn_thresh'] = dyn_thresh

    # 确保日期列是datetime类型
    df['订单开始日期'] = pd.to_datetime(df['订单开始日期'], errors='coerce')

    # 新增：日期范围过滤
    if start_date:
        try:
            sd = pd.to_datetime(start_date)
            df = df[df['订单开始日期'] >= sd]
            if progress_callback:
                progress_callback(-1, f"已按开始日期 {start_date} 过滤", 0)
        except BaseException:
            pass

    if end_date:
        try:
            ed = pd.to_datetime(end_date)
            df = df[df['订单开始日期'] <= ed]
            if progress_callback:
                progress_callback(-1, f"已按结束日期 {end_date} 过滤", 0)
        except BaseException:
            pass

    # 新增：物料搜索过滤（编码或名称）
    if material_search:
        search_lower = material_search.lower()
        # 智能列名匹配
        code_cols = [c for c in df.columns if any(k in str(c).lower() for k in ['组件物料号', '组件编码', '物料编码', 'code', '编码', 'mat', 'material'])]
        name_cols = [c for c in df.columns if any(k in str(c).lower() for k in ['组件描述', '物料描述', '名称', 'name', '描述', 'desc', 'description'])]

        if not code_cols and not name_cols:
            if progress_callback:
                progress_callback(-1, f"⚠ 未找到编码/名称列，跳过搜索", 0)
        else:
            # 构建mask
            mask = None
            if code_cols:
                mask = df[code_cols[0]].astype(str).str.lower().str.contains(search_lower, na=False)
            if name_cols:
                name_mask = df[name_cols[0]].astype(str).str.lower().str.contains(search_lower, na=False)
                mask = name_mask if mask is None else (mask | name_mask)

            df = df[mask]
            if progress_callback:
                used_cols = []
                if code_cols: used_cols.append(f"编码列={code_cols[0]}")
                if name_cols: used_cols.append(f"名称列={name_cols[0]}")
                progress_callback(
                    -1, f"已按物料搜索 '{material_search}' 过滤，剩余 {len(df)} 行（{', '.join(used_cols)}）", 0)

    # 优先使用用户输入的日期范围，否则使用数据中的日期范围
    if start_date and end_date:
        # 用户指定了日期范围
        date_min = pd.to_datetime(start_date)
        date_max = pd.to_datetime(end_date)
    elif start_date:
        date_min = pd.to_datetime(start_date)
        date_max = df['订单开始日期'].max()
    elif end_date:
        date_min = df['订单开始日期'].min()
        date_max = pd.to_datetime(end_date)
    else:
        # 没有用户输入，使用数据中的日期范围
        date_min = df['订单开始日期'].min()
        date_max = df['订单开始日期'].max()

    # 防御：df 为空或日期列全空时，min/max 返回 NaT，strftime 会抛 ValueError
    if pd.isna(date_min) or pd.isna(date_max):
        _today = pd.Timestamp.now().normalize()
        date_min = _today if pd.isna(date_min) else date_min
        date_max = _today if pd.isna(date_max) else date_max

    date_range = f"{pd.Timestamp(date_min).strftime('%Y%m%d')}-{pd.Timestamp(date_max).strftime('%Y%m%d')}"
    _dprint(f"[DEBUG do_analysis_v2] 日期范围：{date_range}")
    

    # 物料分类：向量化 np.select 替代逐行 apply(classify_material)（快 20~50 倍）
    # 判定顺序与原函数一致：Z002/Z009→包材；Z004→原材料；描述含"半成品"→半成品；其余→原材料
    _mt = df['组件物料类型']
    _mtd = df['组件物料类型描述'].astype(str)
    df['物料分类'] = np.select(
        [
            _mt.isin(['Z002', 'Z009']),
            _mt == 'Z004',
            _mtd.str.contains('半成品', na=False),
        ],
        ['包材', '原材料', '半成品'],
        default='原材料'
    )
    # ① 无定额标志：数量-定额==0 时偏差率无意义（SAP 填成假性 ±100%），用于区分假性偏差
    df['_no_quota'] = (df['数量-定额'] == 0)
    df['组件物料号_str'] = df['组件物料号'].astype(str)

    # ② 材料半成品分类：区分食品/饮料厂的原料半成品 vs 成品半成品
    # 数据来源：E:\Users\Administrator\Desktop\半成品重分类.xlsx
    _semi_raw_codes_food = set([
        '40000003', '40000022', '40000142', '40000141', '40000138', '40000091',
        '40000064', '40000068', '40000059', '40000058', '40000094', '40000020', '40000012',
    ])
    _semi_finish_codes_drink = set([
        '41000365', '41000026', '41000024', '41000020', '41000018', '41000016',
        '41000014', '41000012', '41000009', '41000008', '41000007', '41000005',
        '41000022', '41000017', '41000028', '41000039', '41000038', '41000035',
        '41000034', '41000033', '41000032', '41000031', '41000030',
    ])
    _factory_col = next((c for c in ['工厂名称', '工厂'] if c in df.columns), None)
    _has_semi = _mtd.str.contains('半成品', na=False)
    df['_is_semi_raw'] = False  # 默认：否
    if _factory_col:
        # 食品厂原料半成品
        mask_food_raw = (df[_factory_col].astype(str).str.contains('食品', na=False)) & _has_semi & (df['组件物料号_str'].isin(_semi_raw_codes_food))
        df.loc[mask_food_raw, '_is_semi_raw'] = True
        # 饮料厂成品半成品（饮料原料半成品暂未分类）
        mask_drink_finish = (df[_factory_col].astype(str).str.contains('饮料', na=False)) & _has_semi & (df['组件物料号_str'].isin(_semi_finish_codes_drink))
        df.loc[mask_drink_finish, '_is_semi_finish'] = True

    # ③ 半成品重分类列：基于「半成品重分类.xlsx」权威分类表 + 400/410 补空规则
    #    - xlsx 命中（按组件物料号）→ 用表里「半成品分类」原值
    #    - 不在表里、组件物料号以 400 开头 → 食品成品半成品
    #    - 不在表里、组件物料号以 410 开头 → 饮料成品半成品
    #    - 其余 → 空（非半成品）
    _semi_map = _load_semi_classify_map()
    df['半成品重分类'] = ''
    if _semi_map:
        df['半成品重分类'] = df['组件物料号_str'].map(_semi_map).fillna('')
    _empty_cls = df['半成品重分类'] == ''
    df.loc[_empty_cls & df['组件物料号_str'].str.startswith('400'), '半成品重分类'] = '食品成品半成品'
    df.loc[_empty_cls & df['组件物料号_str'].str.startswith('410'), '半成品重分类'] = '饮料成品半成品'

    no_note_mask = ~(df['备注原因'].notna() & (df['备注原因'] != ''))
    # ① 系统无定额自动填充的统一前置条件（2026-08-05 修正）：
    #    产量>0（排除「投了料但没填产量」→ SAP 不推送定额、显示0 的假象）
    #    & 定额=0 & 实际>0 & 无备注 → 才标"系统无定额"。
    #    业务背景：车间投了料却没填产量，定额不被推送(显示0)，被误判为系统无定额。
    qty_prod = df['产量'].fillna(0) if '产量' in df.columns else pd.Series(0, index=df.index)
    qty_act = df['数量-实际'].fillna(0)
    real_no_quota = (qty_prod > 0) & (df['数量-定额'] == 0) & (qty_act > 0) & no_note_mask

    # 广宣：仅 600 开头 + real_no_quota（2026-08-05 由"6开头"收窄为"600开头"）
    gx_auto_fill = df['组件物料号_str'].str.startswith('600') & real_no_quota
    # 确保备注原因为对象类型，防止空值列被推断为 float64
    if '备注原因' in df.columns and df['备注原因'].dtype != object:
        df['备注原因'] = df['备注原因'].astype(object)
    df.loc[gx_auto_fill, '备注原因'] = '系统无定额'

    # 透明胶带：描述含"透明胶带" + real_no_quota（2026-08-05 用户确认保留）
    tape_mask = df['组件物料描述'].str.contains('透明胶带', na=False)
    tape_auto_fill = tape_mask & real_no_quota
    df.loc[tape_auto_fill, '备注原因'] = '系统无定额'

    # 注：已删除原"全部包材+定额0"宽泛规则（pkg_no_quota_mask）——200开头包材不再自动填，
    #     统一改由上面的 real_no_quota（产量>0 & 定额0 & 实际>0 & 无备注）约束，
    #     仅保留广宣(600开头) / 透明胶带 两类（见 2026-08-05 决策）。

    # 数值列已在入口（读取 Excel 后）统一转换一次，此处无需重复 to_numeric
    df['_note_source'] = '人工填写'

    # DeepSeek版：标注标准原因列
    df['标准原因'] = df['备注原因'].apply(standardize_remark) if '备注原因' in df.columns else '未填写'

    df.loc[df['备注原因'].isna() | (df['备注原因'] == ''), '_note_source'] = '无'
    df.loc[gx_auto_fill, '_note_source'] = '系统无定额(广宣)'
    df.loc[tape_auto_fill, '_note_source'] = '自动填充'

    df['车间'] = df['生产管理员描述'].apply(lambda x: str(x).strip())

    # 数值列（含金额-实际(含税)/金额-定额(含税)）已在入口统一转换一次，此处无需重复

    # ========== 偏差金额计算（优先使用含税金额直接相减） ==========
    if '金额-实际(含税)' in df.columns and '金额-定额(含税)' in df.columns:
        # 方法1：直接相减（推荐，最准确）
        df['偏差金额(含税)'] = (df['金额-实际(含税)'] - df['金额-定额(含税)']).round(2)
        report_progress(3, "3/5 正在计算偏差金额和偏差率", 50)
        print(f"[偏差金额计算] 使用含税金额直接相减，非零偏差行数: {(df['偏差金额(含税)'] != 0).sum()}")
    else:
        # 方法2：降级使用材料偏差 × 单价（兼容旧格式）
        for col in ['金额-实际(含税)', '金额-定额(含税)']:
            if col not in df.columns:
                print(f"⚠️ 当前文件缺少[{col}]列，相关计算将按0处理")
                df[col] = 0.0
        df['_unit_price_tax'] = 0.0
        valid_mask_actual = (df['数量-实际'] > 0) & (df['金额-实际(含税)'] > 0)
        valid_mask_quota = (df['数量-定额'] > 0) & (df['金额-定额(含税)'] > 0)
        df.loc[valid_mask_actual, '_unit_price_tax'] = (
            df.loc[valid_mask_actual, '金额-实际(含税)'] /
            df.loc[valid_mask_actual, '数量-实际']
        )
        missing_mask = (~valid_mask_actual) & valid_mask_quota
        df.loc[missing_mask, '_unit_price_tax'] = (
            df.loc[missing_mask, '金额-定额(含税)'] /
            df.loc[missing_mask, '数量-定额']
        )
        df['偏差金额(含税)'] = (df['材料偏差'] * df['_unit_price_tax']).round(2)
        print(f"[偏差金额计算] 使用单价计算，成功计算 {(df['_unit_price_tax'] > 0).sum()}/{len(df)} 行的单价")

    check_cancel()
    # Sheet1（第五步抽取 → analysis/sheets/sheet1_summary.py）
    summary_df = build_sheet1(df, report_progress)
    check_cancel()

    # Sheet2（第五步抽取 → analysis/sheets/sheet2_alt.py）
    # 彻底清理 alt_pairs，只保留纯物料编码字符串
    cleaned_pairs = []
    for pair in alt_pairs:
        if not isinstance(pair, (list, tuple)) or len(pair) != 2:
            continue
        a, b = pair
        # 提取编码和描述：兼容三元组、二元组、纯字符串
        def get_code_and_desc(item):
            if isinstance(item, (list, tuple)):
                if len(item) >= 3:
                    code, desc = item[1], item[2]
                elif len(item) == 2:
                    code, desc = item[0], item[1]
                else:
                    code, desc = item[0], ''
            else:
                code, desc = '', str(item)
            if code is None or code == 'None': code = ''
            if desc is None or desc == 'None': desc = ''
            return str(code).strip(), str(desc).strip()

        a_code, a_desc = get_code_and_desc(a)
        b_code, b_desc = get_code_and_desc(b)
        a_match = a_desc if a_desc else a_code
        b_match = b_desc if b_desc else b_code
        if a_match and b_match:
            cleaned_pairs.append((a_match, b_match))
    # 使用清理后的配对

    # ========== 追踪点2: 预处理后（build_sheet2 前） ==========
    _snapshot['after_preprocess'] = {
        '数量-实际': df['数量-实际'].describe().to_dict() if '数量-实际' in df.columns else 'NOT_FOUND',
        '行数': len(df)
    }
    _dprint(f"[TRACE-2] 预处理后: 数量-实际 sum={df['数量-实际'].sum() if '数量-实际' in df.columns else 'N/A'}")

    alt_df, alt_order_mat = build_sheet2(df, cleaned_pairs, report_progress)
    check_cancel()

    # 基于 Sheet2 结果构建订单级替代料标记集合（仅同订单内出现配对物料才标记）
    # 向量化：用 zip + set 推导替代 iterrows（约 12K 行下从逐行 Python 循环改为 C 级操作）
    if not alt_df.empty and all(c in alt_df.columns for c in ('订单号', '物料A', '物料B')):
        _alt_order_s = alt_df['订单号'].astype(str)
        alt_order_mat = set(zip(_alt_order_s, alt_df['物料A'].astype(str)))
        alt_order_mat |= set(zip(_alt_order_s, alt_df['物料B'].astype(str)))
    else:
        alt_order_mat = set()

    # 订单级替代料标记（基于 alt_order_mat，仅同订单内同时存在配对物料才标记）
    # 向量化：构建 (流程订单, 组件物料描述) 键 Series，用 isin 一次性标记
    # 关键：Series 必须带 df.index。筛选后 df 的索引是非连续的(如 0,1,5,9…)，
    # 若用默认 RangeIndex，df.loc[bool_mask, col] 会把 mask 当作标签去对齐，
    # 在 pandas 3.x 下触发 TypeError: unhashable type: 'Series'（日期筛选必崩）。
    _order_keys = pd.Series(
        zip(df['流程订单'].astype(str), df['组件物料描述'].astype(str)),
        index=df.index,
    )
    df.loc[_order_keys.isin(alt_order_mat), '_note_source'] = '替代料'

    # 更新标准原因
    df.loc[df['_note_source'] == '替代料', '标准原因'] = '替代料'

    # 重新计算 _is_alt 标志（仅基于订单级匹配，同一订单内同时存在配对物料才标记）
    report_progress(4, "4/5 正在匹配替代料信息", 70)
    # 向量化：tuple Series + isin 替代逐行 apply（in alt_order_mat）；同样带 df.index
    _order_alt_keys = pd.Series(
        zip(df['流程订单'].astype(str), df['组件物料描述'].astype(str)),
        index=df.index,
    )
    df['_is_alt'] = _order_alt_keys.isin(alt_order_mat)

    check_cancel()

    # Sheet3（第五步抽取 → analysis/sheets/sheet3_no_note.py）
    no_note_df = build_sheet3(df, report_progress, dyn_thresh=dyn_thresh)
    check_cancel()

    # Sheet4（第五步抽取 → analysis/sheets/sheet4_middle.py）
    middle_df = build_sheet4(df, alt_df, alt_pairs, report_progress, dyn_thresh=dyn_thresh)
    check_cancel()

    # Sheet5（第五步抽取 → analysis/sheets/sheet5_full.py）
    dev_df = build_sheet5(df, report_progress, threshold=dev_rate_threshold)

    # 补齐"是否替代料"列
    # 优先使用 _is_alt 标志（已做订单级匹配：同一订单内同时存在配对物料才标记）
    if '_is_alt' in df.columns:
        # 向量化：收集所有 _is_alt 为 True 的 (流程订单, 组件物料描述) 键集合，dev_df 用 isin 匹配
        # 防御：日期筛选可能使 dev_df 为空(无列)，此时直接标 '否'，避免 KeyError: '流程订单'
        if not dev_df.empty and '流程订单' in dev_df.columns and '物料名称' in dev_df.columns:
            _alt_true_keys = set(zip(
                df.loc[df['_is_alt'], '流程订单'].astype(str),
                df.loc[df['_is_alt'], '组件物料描述'].astype(str)
            ))
            _dev_alt_keys = pd.Series(
                zip(dev_df['流程订单'].astype(str), dev_df['物料名称'].astype(str)),
                index=dev_df.index,
            )
            dev_df['是否替代料'] = _dev_alt_keys.isin(_alt_true_keys).map({True: '是', False: '否'})
        else:
            dev_df['是否替代料'] = '否'
    elif '是否替代料' not in dev_df.columns and alt_pairs:
        # 回退：仅在 dev_df 上做订单级匹配（向量化）
        if not alt_df.empty and all(c in alt_df.columns for c in ('订单号', '物料A', '物料B')):
            _ao_s = alt_df['订单号'].astype(str)
            alt_order_mat = set(zip(_ao_s, alt_df['物料A'].astype(str)))
            alt_order_mat |= set(zip(_ao_s, alt_df['物料B'].astype(str)))
        else:
            alt_order_mat = set()
        name_col = next((c for c in ['物料名称', '物料描述', '组件物料描述'] if c in dev_df.columns), None)
        if name_col:
            _dev_fb_keys = pd.Series(zip(dev_df['流程订单'].astype(str), dev_df[name_col].astype(str)))
            dev_df['是否替代料'] = _dev_fb_keys.isin(alt_order_mat).map({True: '是', False: '否'})
        else:
            dev_df['是否替代料'] = '否'
    elif '是否替代料' not in dev_df.columns:
        dev_df['是否替代料'] = '否'

    # ========== 替代料净偏差自动抵消 ==========
    # 使用传入的 enable_net_offset 参数（由调用方根据配置决定）
    if enable_net_offset and alt_pairs and len(alt_pairs) > 0:
        _dprint("[净偏差抵消] 开始计算替代料净偏差...")
        dev_df = apply_net_offset(dev_df, alt_pairs, enable=enable_net_offset, group_key=['订单日期', '流程订单'])
        _dprint(f"[净偏差抵消] 完成，影响行数: {len(dev_df[dev_df['净偏差金额'] != dev_df.get('偏差金额(含税)', dev_df.get('偏差金额', 0))])}")
    else:
        # 确保净偏差列存在（即使没有配对，也添加原始偏差金额作为净偏差）
        dev_df['净偏差数量'] = pd.to_numeric(dev_df.get('偏差数量', 0), errors='coerce').round(2)
        if '偏差金额(含税)' in dev_df.columns:
            dev_df['净偏差金额'] = pd.to_numeric(dev_df['偏差金额(含税)'], errors='coerce').round(2)
        elif '偏差金额' in dev_df.columns:
            dev_df['净偏差金额'] = pd.to_numeric(dev_df['偏差金额'], errors='coerce').round(2)
        else:
            dev_df['净偏差金额'] = 0.0

        # 计算净偏差率
        quota_col = None
        for c in ['数量-定额', '定额']:
            if c in dev_df.columns:
                quota_col = c
                break
        if quota_col:
            dev_df['净偏差率(%)'] = (pd.to_numeric(dev_df['净偏差数量'], errors='coerce').fillna(0) /
                               pd.to_numeric(dev_df[quota_col], errors='coerce').replace(0, np.nan) * 100).fillna(0).round(2)
        else:
            dev_df['净偏差率(%)'] = 0.0

    # 调整列顺序：净偏差数量/净偏差金额移到偏差率后面、偏差金额前面
    desired_order = []
    for col in dev_df.columns:
        desired_order.append(col)
        if col == '偏差率':
            desired_order.append('净偏差数量')
            desired_order.append('净偏差金额')
            desired_order.append('净偏差率(%)')
    # 去重（净偏差列已通过 desired_order 插入，移除末尾的）
    seen = set()
    ordered = []
    for col in desired_order:
        if col not in seen:
            seen.add(col)
            ordered.append(col)
    dev_df = dev_df[ordered]
    # 数值统一保留2位小数
    for col in ['净偏差数量', '净偏差金额', '净偏差率(%)', '偏差数量', '偏差金额']:
        if col in dev_df.columns:
            dev_df[col] = pd.to_numeric(dev_df[col], errors='coerce').round(2)

    # 补充净偏差率（apply_net_offset 可能已返回，此处确保存在）
    if '净偏差率(%)' not in dev_df.columns:
        quota_col = None
        for c in ['数量-定额', '定额']:
            if c in dev_df.columns:
                quota_col = c
                break
        if quota_col:
            dev_df['净偏差率(%)'] = (pd.to_numeric(dev_df['净偏差数量'], errors='coerce').fillna(0) /
                               pd.to_numeric(dev_df[quota_col], errors='coerce').replace(0, np.nan) * 100).fillna(0).round(2)
        else:
            dev_df['净偏差率(%)'] = 0.0

    # ========== 追踪点3: build_sheet5 后 ==========
    _snapshot['after_sheet5'] = {
        '数量-实际': df['数量-实际'].describe().to_dict() if '数量-实际' in df.columns else 'NOT_FOUND',
        '行数': len(df)
    }
    _dprint(f"[TRACE-3] build_sheet5后: 数量-实际 sum={df['数量-实际'].sum() if '数量-实际' in df.columns else 'N/A'}")

    check_cancel()

    # ========== 确保主表所需列存在（原本在 wb 构建之后，现上移到此处） ==========
    # 主表快速路径（return_dataframe=True）在下方会直接 return，这几列必须先备齐
    if '车间' not in dev_df.columns:
        dev_df['车间'] = '未知车间'
    if '物料类型' not in dev_df.columns:
        dev_df['物料类型'] = dev_df['物料编码'].apply(infer_material_type)
    if '订单日期' in dev_df.columns and '周' not in dev_df.columns:
        dev_df['订单日期'] = pd.to_datetime(dev_df['订单日期'], errors='coerce')
        dev_df['周'] = dev_df['订单日期'].dt.strftime('%Y-W%W')
    if '替代料组' not in dev_df.columns:
        dev_df['替代料组'] = ''

        # ========== 暂存中间结果：供后台缓存线程复用，避免 Sheet1~5 重算 ==========
    _inter = {
        'df': df, 'dev_df': dev_df, 'alt_df': alt_df, 'summary_df': summary_df,
        'no_note_df': no_note_df, 'middle_df': middle_df,
        'alt_order_mat': alt_order_mat, 'date_min': date_min, 'date_max': date_max,
        'dyn_thresh': dyn_thresh, 'thresh_desc': thresh_desc, 'src_file': src_file,
        'dev_rate_threshold': dev_rate_threshold,
        '_trace_log': _trace_log, '_snapshot': _snapshot,
    }

# ========== 主表快速路径：return_dataframe=True 时 dev_df 已齐活，直接返回 ==========
    # 后续 build_sheet6~10 与整本 wb 仅为「导出 Excel 报告」服务，主表用不上，跳过以加速加载
    if return_dataframe:
        _snapshot['after_sheet5'] = {
            '数量-实际': df['数量-实际'].describe().to_dict() if '数量-实际' in df.columns else 'NOT_FOUND',
            '行数': len(df)
        }
        report_progress(5, "5/5 主表计算完成", 90)
        report_progress(5, "5/5 分析完成", 100)
        _dprint("[DEBUG do_analysis_v2] 主表快速路径：跳过导出专用 sheet，直接返回 dev_df")
        try:
            with open(_trace_log, 'a', encoding='utf-8') as f:
                f.write(f"\n=== Trace Log {datetime.now()} ===\n")
                f.write(f"Input file: {src_file}\n")
                f.write(json.dumps(_snapshot, indent=2, ensure_ascii=False, default=str))
                f.write('\n')
            _dprint(f"[TRACE] 日志已保存到: {_trace_log}")
        except Exception as e:
            _dprint(f"[TRACE] 保存日志失败: {e}")
        global LATEST_INTERMEDIATES
        LATEST_INTERMEDIATES = _inter
        return dev_df

    return export_full_report_from_intermediates(_inter, output_path=output_path, output_dir=output_dir, progress_callback=progress_callback, cancel_check=cancel_check)



def export_full_report_from_intermediates(intermediates, output_path=None, output_dir=None,
                                         progress_callback=None, cancel_check=None):
    """复用 do_analysis_v2 的计算中间结果，仅生成 Sheet6~10 + 整本 wb 并保存。
    供后台缓存线程(_FullCacheWorker)调用，避免 Sheet1~5 被重复计算。"""
    df = intermediates['df']
    dev_df = intermediates['dev_df']
    alt_df = intermediates['alt_df']
    summary_df = intermediates['summary_df']
    no_note_df = intermediates['no_note_df']
    middle_df = intermediates['middle_df']
    alt_order_mat = intermediates['alt_order_mat']
    date_min = intermediates['date_min']
    date_max = intermediates['date_max']
    dyn_thresh = intermediates['dyn_thresh']
    thresh_desc = intermediates['thresh_desc']
    src_file = intermediates['src_file']
    dev_rate_threshold = intermediates['dev_rate_threshold']
    _trace_log = intermediates.get('_trace_log')
    _snapshot = intermediates.get('_snapshot', {})
    date_range = f"{pd.Timestamp(date_min).strftime('%Y%m%d')}-{pd.Timestamp(date_max).strftime('%Y%m%d')}"

    def check_cancel():
        if cancel_check and cancel_check():
            raise KeyboardInterrupt("用户取消")

    def report_progress(step_idx, step_name, percent):
        if progress_callback:
            progress_callback(step_idx, step_name, percent)
            time.sleep(0.01)

    from analysis.excel_builder.write_sheet_util import get_default_styles
    _styles = get_default_styles()
    pos_fill = _styles['pos_fill']
    neg_fill = _styles['neg_fill']
    alt_fill = _styles['alt_fill']
    gx_fill = _styles['gx_fill']
    header_font = _styles['header_font']
    header_fill = _styles['header_fill']
    center = _styles['center']
    border = _styles['border']
    data_font = _styles['data_font']
    anomaly_fills = _styles['anomaly_fills']

    # 构建净偏差查找表：(流程订单, 物料编码) -> (净偏差数量, 净偏差金额)
    # 向量化：dict(zip(键, zip(数量, 金额))) 替代 iterrows（约 12K 行下消除逐行 Python 循环）
    net_offset_map = {}
    if '净偏差数量' in dev_df.columns:
        _fk = dev_df['流程订单'].astype(str) if '流程订单' in dev_df.columns else pd.Series([''] * len(dev_df), index=dev_df.index)
        _mc = dev_df['物料编码'].astype(str) if '物料编码' in dev_df.columns else pd.Series([''] * len(dev_df), index=dev_df.index)
        net_offset_map = dict(zip(
            zip(_fk, _mc),
            zip(dev_df['净偏差数量'], dev_df['净偏差金额'])
        ))

    # Sheet6（第五步抽取 → analysis/sheets/sheet6_anomaly.py）
    anomaly_df = build_sheet6(df, alt_order_mat, report_progress, net_offset_map=net_offset_map)
    check_cancel()

    # Sheet7（第五步抽取 → analysis/sheets/sheet7_amount.py）
    wb = Workbook()   # 原 Sheet7 代码块中创建（必需，供后续 Sheet 使用）
    build_sheet7(wb, df, report_progress)
    check_cancel()

    # Sheet8（第五步抽取 → analysis/sheets/sheet8_reason_summary.py）
    reason_summary_df = build_sheet8(df, report_progress)
    check_cancel()

    # Sheet9（第五步抽取 → analysis/sheets/sheet9_reason_detail.py）
    reason_analysis_df = build_sheet9(df, report_progress)
    check_cancel()

    # Sheet10（第五步抽取 → analysis/sheets/sheet10_trend.py）
    build_sheet10(wb, dev_df, date_min, report_progress)

    ws1 = wb.active
    ws1.title = '汇总统计'
    headers1 = ['序号', '工厂', '工厂名称', '车间', '物料分类',
                '正偏差条数', '正偏差数量', '正偏差金额(含税)',
                '负偏差条数', '负偏差数量', '负偏差金额(含税)',
                '总条数', '总数量', '总偏差金额(含税)', '备注覆盖率', '预警']
    rows1 = [[r['序号'], r['工厂'], r['工厂名称'], r['车间'], r['物料分类'],
              r['正偏差条数'], r['正偏差数量'], r['正偏差金额(含税)'],
              r['负偏差条数'], r['负偏差数量'], r['负偏差金额(含税)'],
              r['总条数'], r['总数量'], r['总偏差金额(含税)'],
              r['备注覆盖率'], r['预警']] for r in summary_df.to_dict('records')]
    write_sheet(ws1, headers1, rows1,
                [8, 10, 10, 10, 10, 12, 14, 16, 12, 14, 16, 10, 14, 16, 12, 8])

    ws2 = wb.create_sheet('替代料明细')
    headers2 = ['订单日期', '车间', '订单号', '物料A编码', '物料A', '单位', '偏差A', '偏差率A',
                '物料B编码', '物料B', '偏差B', '偏差率B', '净偏差数量', '净偏差金额', '净偏差率', '备注']
    rows2 = []
    for r in alt_df.to_dict('records'):
        material_a_name = str(r['物料A']).strip() if pd.notna(r.get('物料A')) else ''
        material_b_name = str(r['物料B']).strip() if pd.notna(r.get('物料B')) else ''
        code_a = ''
        if material_a_name:
            mask = df['组件物料描述'].astype(str).str.strip() == material_a_name
            if mask.any():
                code_a = str(df.loc[mask, '组件物料号'].iloc[0])
        code_b = ''
        if material_b_name:
            mask = df['组件物料描述'].astype(str).str.strip() == material_b_name
            if mask.any():
                code_b = str(df.loc[mask, '组件物料号'].iloc[0])
        rows2.append([
            r['订单日期'], r['车间'], r['订单号'],
            code_a, r['物料A'], r['单位'], r['偏差A'], r.get('偏差率A', ''),
            code_b, r['物料B'], r['偏差B'], r.get('偏差率B', ''),
            r.get('净偏差数量', ''), r.get('净偏差金额', ''), r.get('净偏差率', ''), r['备注']
        ])
    write_sheet(ws2, headers2, rows2,
                [14, 10, 14, 14, 28, 8, 12, 12, 14, 28, 12, 12, 12, 14, 12, 20])

    ws3 = wb.create_sheet('无备注预警')
    headers3 = ['订单日期', '工厂', '车间', '物料名称', '物料类型', '单位',
                '定额', '实际', '偏差数量', '偏差率', '偏差金额(含税)', '备注']
    rows3 = [[r['订单日期'], r['工厂'], r['车间'], r['物料名称'], r['物料类型'],
              r['单位'], r['定额'], r['实际'], r['偏差数量'], r['偏差率'],
              r['偏差金额(含税)'] if isinstance(r.get('偏差金额(含税)'), (int, float)) and r['偏差金额(含税)'] != 0 else '-',
              r['备注']] for r in no_note_df.to_dict('records')]
    write_sheet(ws3, headers3, rows3,
                [14, 10, 10, 28, 10, 8, 12, 12, 12, 10, 16, 20])

    ws4 = wb.create_sheet('中间地带明细')
    headers4 = ['订单日期', '工厂', '车间', '物料名称', '物料类型', '单位',
                '定额', '实际', '偏差数量', '偏差率', '备注']
    rows4 = [[r['订单日期'], r['工厂'], r['车间'], r['物料名称'], r['物料类型'],
              r['单位'], r['定额'], r['实际'], r['偏差数量'], r['偏差率'],
              r['备注']] for r in middle_df.to_dict('records')]
    write_sheet(ws4, headers4, rows4,
                [14, 10, 10, 28, 10, 8, 12, 12, 12, 10, 20])

    ws5 = wb.create_sheet('完整偏差明细')
    headers5 = ['订单日期', '订单类型', '流程订单', '工厂', '车间', '物料类型', '原表行号',
                '产品物料号码', '产品物料描述',
                '物料编码', '物料名称', '单位', '定额', '实际',
                '偏差数量', '偏差率', '偏差金额', '净偏差数量', '净偏差金额', '净偏差率', '是否替代料', '备注', '备注来源', '偏差区间']
    rows5 = [[r['订单日期'], r.get('订单类型', ''), r.get('流程订单', ''), r['工厂'], r['车间'], r['物料类型'], r['原表行号'],
              r.get('产品物料号码', ''), r.get('产品物料描述', ''),
              r['物料编码'], r['物料名称'], r['单位'], r['定额'], r['实际'],
              r['偏差数量'], r['偏差率'], r['偏差金额'],               r.get('净偏差数量', ''), r.get('净偏差金额', ''),
              (f"{r['净偏差率(%)']:.2f}%" if pd.notna(r.get('净偏差率(%)')) else ''),
              r.get('是否替代料', '否'),
              r['备注'], r['备注来源'], r['偏差区间']] for r in dev_df.to_dict('records')]
    write_sheet(ws5, headers5, rows5,
                [14, 10, 16, 10, 10, 10, 10, 18, 30, 16, 28, 8, 12, 12, 12, 10, 14, 14, 12, 10, 20, 16, 10])

    # 性能优化（2026-07-27）：预注册 fill 拿到索引，循环内只改 _style.fillId 整数，
    # 避免每格 .fill= 赋值触发 openpyxl 样式对象递归 hash + 去重查表（12K 行时为大热点）
    _pos_fid = wb._fills.add(pos_fill)
    _neg_fid = wb._fills.add(neg_fill)
    _alt_fid = wb._fills.add(alt_fill)
    _gx_fid = wb._fills.add(gx_fill)
    for i, r in enumerate(dev_df.to_dict('records'), 2):
        dev_qty = r['偏差数量']
        if isinstance(dev_qty, (int, float)) and dev_qty != 0:
            _fid = _pos_fid if dev_qty > 0 else _neg_fid
            for j in range(1, len(headers5) + 1):
                ws5.cell(row=i, column=j)._style.fillId = _fid
        src = r['备注来源']
        if src == '替代料':
            ws5.cell(row=i, column=21)._style.fillId = _alt_fid  # 第21列 = 是否替代料
        elif src in ('系统无定额(广宣)', '自动填充'):
            ws5.cell(row=i, column=21)._style.fillId = _gx_fid  # 第21列 = 是否替代料

    ws6 = wb.create_sheet('异常预警')
    headers6 = ['订单开始日期', '订单类型', '订单号', '异常类型', '工厂', '车间',
                '原表行号', '产品物料号码', '产品物料描述', '物料编码', '物料名称',
                '单位', '定额', '实际',
                '偏差数量', '净偏差数量', '净偏差金额', '净偏差率', '偏差率', '备注', '处理建议', '替代料']
    for j, h in enumerate(headers6, 1):
        c = ws6.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws6.cell(row=2, column=1, value='颜色说明：').font = Font(size=10, bold=True)
    legend = [
        ('异常1', '浅红', 'FFCDD2'),
        ('异常2', '浅橙', 'FFE0B2'),
        ('异常3', '浅紫', 'E1BEE7'),
        ('异常4', '浅蓝', 'B3E5FC'),
        ('异常5', '浅黄', 'FEFFD6'),
    ]
    for k, (key, label, color) in enumerate(legend, 2):
        c = ws6.cell(row=2, column=k, value=f" {label}")
        c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        c.font = Font(size=10, bold=True, color='FF000000')
        c.alignment = center
        ws6.column_dimensions[get_column_letter(k)].width = 10
    # ④ 阈值说明：主表明细 ±10%（业务口径）；异常预警替代料残差全部列示（不设阈值）
    note_c = ws6.cell(row=2, column=7,
                      value=f'阈值说明：主表明细±{dyn_thresh:.0f}%（业务口径）；异常预警：替代料残差全部列示（不设阈值）')
    note_c.font = Font(size=9, italic=True, color='666666')
    # 性能优化（2026-07-27）：按 5 种异常类型预构建 StyleArray 原型，
    # 数据格只做 _style 数组拷贝，避免每格 4 次样式赋值的 hash/查表开销
    from copy import copy as _copy
    _proto6 = {}
    if len(anomaly_df):
        _probe6 = ws6.cell(row=3, column=1)
        _probe6.font = data_font
        _probe6.border = border
        _probe6.alignment = center
        for _atype, _afill in anomaly_fills.items():
            _probe6.fill = _afill
            _proto6[_atype] = _copy(_probe6._style)
    r_row = 3
    for r in anomaly_df.to_dict('records'):
        proto = _proto6.get(r['row_type'], _proto6['异常1'])
        row_vals = [
            r['订单开始日期'], r['订单类型'], r['流程订单'], r['异常类型'], r['工厂'], r['车间'],
            r['原表行号'],
            r.get('产品物料号码', ''), r.get('产品物料描述', ''),
            r['物料编码'], r['物料名称'], r['单位'],
            r['定额'], r['实际'], r['偏差数量'], r.get('净偏差数量', ''), r.get('净偏差金额', ''),
            r.get('净偏差率', ''),
            r['偏差率'], r.get('备注', ''), r.get('处理建议', ''), r.get('替代料', '否')]
        for j, v in enumerate(row_vals, 1):
            c = ws6.cell(row=r_row, column=j)
            c._style = _copy(proto)
            c.value = v  # 先于样式设值会被 _style 覆盖日期格式；改为先样式后设值，真日期保留日期格式
        r_row += 1
    for j, w in enumerate([14, 10, 18, 10, 10, 10, 10, 16, 28, 8, 12, 12, 12, 12, 14, 10, 30, 10, 30, 10], 1):
        ws6.column_dimensions[get_column_letter(j)].width = w
    check_cancel()

    ws7 = wb.create_sheet('偏差原因汇总')
    ws7.merge_cells('A1:H1')
    tc = ws7.cell(row=1, column=1,
                  value=(f'ZPP011 偏差原因汇总（'
                         f'{pd.Timestamp(date_min).strftime("%Y-%m-%d")} ~ '
                         f'{pd.Timestamp(date_max).strftime("%Y-%m-%d")}）'))
    tc.font = Font(bold=True, size=12)
    tc.alignment = Alignment(horizontal='center')
    headers7 = ['工厂', '车间', '多耗', '少耗', '净偏差数量', '原因数',
                '原料主要原因（Top5）', '包材主要原因（Top5）']
    for j, h in enumerate(headers7, 1):
        c = ws7.cell(row=2, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    # 对原因文本内部添加序号
    def _add_ordinal(text):
        if pd.isna(text) or not text:
            return ''
        lines = str(text).split('\n')
        circles = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩']
        numbered = []
        for i, line in enumerate(lines):
            if not line.strip():
                continue
            prefix = circles[i] if i < 10 else f'{i+1}.'
            numbered.append(f'{prefix} {line.strip()}')
        return '\n'.join(numbered)

    for i, r in enumerate(reason_summary_df.to_dict('records'), 3):
        for j, v in enumerate([r['工厂'], r['车间'], r['多耗'], r['少耗'],
                               r.get('净偏差数量', 0), r['原因数']], 1):
            c = ws7.cell(row=i, column=j, value=v)
            c.border = border
            c.font = Font(size=11)
            # 多耗/少耗/净偏差数量为「原料/包材×单位」多行文本，需换行+左对齐
            if j in (3, 4, 5):
                c.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            else:
                c.alignment = Alignment(vertical='top', horizontal='center')
        for col, key in [(7, '原料主要原因（Top5）'), (8, '包材主要原因（Top5）')]:
            c = ws7.cell(row=i, column=col, value=_add_ordinal(r[key]))
            c.border = border
            c.font = Font(size=11)
            c.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        raw_t = str(r['原料主要原因（Top5）']) if pd.notna(r['原料主要原因（Top5）']) else ''
        pkg_t = str(r['包材主要原因（Top5）']) if pd.notna(r['包材主要原因（Top5）']) else ''
        lines = sum(max(1, (len(p.strip()) + 19) // 20) for p in raw_t.split('\n') if p.strip())
        lines += sum(max(1, (len(p.strip()) + 19) // 20) for p in pkg_t.split('\n') if p.strip())
        # 多耗/少耗/净偏差数量 也是多行文本（原料/包材各一行），行高取两者较大值
        dev_lines = max(
            (str(r.get(k, '')).count('\n') + 1) for k in ('多耗', '少耗', '净偏差数量'))
        ws7.row_dimensions[i].height = max(lines * 16, dev_lines * 32, 67)
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                      [14, 10, 26, 26, 30, 10, 55, 55]):
        ws7.column_dimensions[col].width = w

    ws8 = wb.create_sheet('偏差原因分析')
    headers8 = ['工厂', '车间', '组件物料类型', '组件物料类型描述', '组件物料描述', '备注原因', '原始备注示例',
                '多耗', '少耗', '净偏差数量', '占车间偏差比%']
    rows8 = [[r['工厂'], r['车间'], r['组件物料类型'], r['组件物料类型描述'], r.get('组件物料描述', ''),
              r['备注原因'], r['原始备注示例'],
              r['多耗'], r['少耗'], r.get('净偏差数量', 0), r.get('占车间偏差比%', 0)]
             for r in reason_analysis_df.to_dict('records')]
    write_sheet(ws8, headers8, rows8,
                [14, 10, 12, 18, 30, 20, 25, 14, 14, 14, 14])

    # 如果用户指定了输出路径，直接使用；否则自动生成
    if output_path:
        final_output_path = output_path
    else:
        pattern = os.path.join(output_dir, f'ZPP011偏差分析最终版_{date_range}_v*.xlsx')
        existing = _glob.glob(pattern)
        versions = [int(re.search(r'_v(\d+)\.xlsx$', os.path.basename(f)).group(1))
                    for f in existing if re.search(r'_v(\d+)\.xlsx$', os.path.basename(f))]
        next_ver = max(versions) + 1 if versions else 1
        final_output_path = os.path.join(
            output_dir,
            f'ZPP011偏差分析最终版_{date_range}_v{next_ver:02d}.xlsx')

    report_progress(5, "5/5 正在生成审核表格", 90)

    # ── 分析说明 sheet ────────────────────────────
    ws_info = wb.create_sheet('📋 分析说明', index=0)
    info_rows = [
        ['ZPP011 偏差分析器 · 分析说明', ''],
        ['', ''],
        ['分析日期范围', f"{pd.Timestamp(date_min).strftime('%Y-%m-%d')} ～ {pd.Timestamp(date_max).strftime('%Y-%m-%d')}"],
        ['动态阈值方法', thresh_desc],
        ['动态阈值数值', f"±{dyn_thresh:.1f}%"],
        ['', ''],
        ['各Sheet说明', ''],
        ['汇总统计', '按车间×物料分类统计偏差条数、数量、金额、备注覆盖率'],
        ['替代料明细', '识别到的替代料配对及净偏差'],
        ['无备注预警', f'偏差率超过 ±{dyn_thresh:.1f}% 但未填备注的记录（按偏差金额降序）'],
        ['中间地带明细', f'偏差率在 ±{dyn_thresh:.1f}% 区间内（不纳入汇总统计）'],
        ['完整偏差明细', ('所有偏差记录（剔除替代料）' if dev_rate_threshold <= 0
                          else f'所有偏差率超 ±{dev_rate_threshold:.1f}% 的记录（剔除替代料）')],
        ['异常预警', '5类异常：系统无定额、实际为0/负数、BOM问题、包材负偏差、替代料残差'],
        ['偏差金额分析', '按物料汇总正/负偏差金额（含税）'],
        ['偏差原因汇总', '按车间汇总原因 Top5（备注已自动标准化）'],
        ['偏差原因分析', '按标准原因类别汇总分析'],
        ['趋势分析', '按自然日均分三段（早期/中期/近期），计算各段平均偏差率并判断趋势'],
    ]
    for i, row in enumerate(info_rows, 1):
        for j, v in enumerate(row, 1):
            c = ws_info.cell(row=i, column=j, value=v)
            if i == 1:
                c.font = Font(bold=True, size=14, color='1B5E20')
                c.alignment = Alignment(horizontal='left', vertical='center')
            elif i == 7:
                c.font = Font(bold=True, size=11)
            elif i > 7:
                c.font = Font(size=11)
                if j == 1:
                    c.font = Font(bold=True, size=11)
        ws_info.row_dimensions[i].height = 22
        ws_info.column_dimensions['A'].width = 28
        ws_info.column_dimensions['B'].width = 62

    _dprint(f"[DEBUG do_analysis_v2] 准备保存到：{final_output_path}")
    report_progress(5, "5/5 分析完成", 100)

    # 否则保存文件
    _out_dir = os.path.dirname(final_output_path)
    if _out_dir:
        os.makedirs(_out_dir, exist_ok=True)
    # 为汇总统计预警列上色
    _apply_warning_colors(wb)

    # 设置标签栏占更大比例，确保所有Sheet标签可见
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = False
    wb.active = 0  # 默认打开第一个Sheet（分析说明）
    if wb.worksheets:
        wb.worksheets[0].sheet_view.tabSelected = True
    # 标签栏占水平滚动条的 90%（默认 60%），确保所有标签可见
    if wb.views:
        wb.views[0].tabRatio = 900
        wb.views[0].activeTab = 0
        wb.views[0].showSheetTabs = True

    try:
        wb.save(final_output_path)
    except PermissionError as e:
        raise PermissionError(
            f"无法保存文件，可能被其他程序占用！\n\n"
            f"文件路径：{final_output_path}\n\n"
            f"可能的原因：\n"
            f"  1. 文件已用 Excel 打开，请先关闭 Excel 中的这个文件\n"
            f"  2. 文件被其他程序占用（如 WPS、预览窗口等）\n"
            f"  3. 没有写入权限\n\n"
            f"解决方法：\n"
            f"  • 关闭 Excel 中打开的这个文件，然后重试\n"
            f"  • 或者换一个输出文件名（在弹出的另存为对话框中修改文件名）"
        ) from e
    _dprint(f"[DEBUG do_analysis_v2] 保存完成，返回：{final_output_path}")
    # 保存追踪日志
    try:
        with open(_trace_log, 'a', encoding='utf-8') as f:
            f.write(f"\n=== Trace Log {datetime.now()} ===\n")
            f.write(f"Input file: {src_file}\n")
            f.write(json.dumps(_snapshot, indent=2, ensure_ascii=False, default=str))
            f.write('\n')
        _dprint(f"[TRACE] 日志已保存到: {_trace_log}")
    except Exception as e:
        _dprint(f"[TRACE] 保存日志失败: {e}")

    return final_output_path

def _apply_warning_colors(wb):
    """为汇总统计sheet的预警列添加颜色填充"""
    try:
        from openpyxl.styles import PatternFill
        if '汇总统计' not in wb.sheetnames:
            return
        ws = wb['汇总统计']
        warning_col = None
        for cell in ws[1]:
            if cell.value == '预警':
                warning_col = cell.column
                break
        if warning_col is None:
            return
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=warning_col)
            value = str(cell.value).strip() if cell.value else ''
            if '红色预警' in value:
                cell.fill = red_fill
            elif '黄色预警' in value:
                cell.fill = yellow_fill
            elif '绿色预警' in value:
                cell.fill = green_fill
    except Exception as e:
        print(f"[WARN] 预警列上色失败: {e}")


def _build_deviation_summary(dev_df, orig_df):
    """
    构建偏差金额汇总表（Sheet: 偏差金额分析）
    dev_df: 完整偏差明细 DataFrame
    orig_df: 原始数据 DataFrame（用于获取单价等额外信息，可选）
    """
    # 按物料编码汇总偏差金额
    if '偏差金额' not in dev_df.columns:
        # 如果没有偏差金额列，尝试计算
        if '数量-实际' in dev_df.columns and '数量-定额' in dev_df.columns:
            # 计算单价（从原始数据获取，这里简单处理）
            dev_df['偏差金额'] = (dev_df['数量-实际'] - dev_df['数量-定额']) * dev_df.get('单价', 0)
        else:
            dev_df['偏差金额'] = 0.0
    
    # 按物料编码、物料名称、物料类型分组
    group_cols = []
    for col in ['物料编码', '物料名称', '物料类型']:
        if col in dev_df.columns:
            group_cols.append(col)
    if not group_cols:
        group_cols = ['物料编码']
    
    summary = dev_df.groupby(group_cols).agg(
        正偏差金额=('偏差金额', lambda x: x[x > 0].sum()),
        负偏差金额=('偏差金额', lambda x: x[x < 0].sum()),
        总偏差金额=('偏差金额', 'sum'),
        涉及条数=('偏差金额', 'count')
    ).reset_index()
    
    # 格式化金额（保留两位小数）
    for col in ['正偏差金额', '负偏差金额', '总偏差金额']:
        summary[col] = summary[col].round(2)
    
    # 添加单位（如果有）
    if '单位' in dev_df.columns:
        unit_map = dev_df.groupby('物料编码')['单位'].first().to_dict()
        summary['单位'] = summary['物料编码'].map(unit_map)
    
    # 按总偏差金额绝对值排序（降序）
    summary = summary.sort_values('总偏差金额', key=abs, ascending=False)
    
    return summary


# ---------------------------------------------------------------------------
# 半成品重分类辅助：读取「半成品重分类.xlsx」权威分类表
# 返回 {组件物料号(str): 半成品分类值(str)}
# 查找优先级：
#   1) 打包后资源目录 sys._MEIPASS/config/半成品重分类.xlsx
#   2) 工程内 config/半成品重分类.xlsx
# 找不到或读取失败返回空 dict（不影响主流程，仅半成品重分类列留空 + 400/410 补空规则生效）
# ---------------------------------------------------------------------------
def _load_semi_classify_map():
    import os
    import sys
    candidates = []
    # 1) 打包资源
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        candidates.append(os.path.join(sys._MEIPASS, 'config', '半成品重分类.xlsx'))
    # 2) 工程内
    _here = os.path.dirname(os.path.abspath(__file__))
    candidates.append(os.path.join(_here, '..', 'config', '半成品重分类.xlsx'))
    _path = next((p for p in candidates if os.path.exists(p)), None)
    if not _path:
        return {}
    try:
        _xls = pd.read_excel(_path, sheet_name=None)
        _map = {}
        for _sh, _d in _xls.items():
            if '组件物料号' in _d.columns and '半成品分类' in _d.columns:
                for _, _r in _d.iterrows():
                    _code = str(_r['组件物料号']).strip()
                    _cls = str(_r['半成品分类']).strip()
                    if _code and _cls and _cls not in ('nan', 'None'):
                        _map[_code] = _cls
        return _map
    except Exception:
        return {}
